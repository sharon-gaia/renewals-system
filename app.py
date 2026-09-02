import sys
import io
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)
except Exception:
    pass

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_cors import CORS
import sqlite3
import os
import io
import json
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from openpyxl import Workbook as NewWorkbook
import datetime
import imaplib
import socket
# Global network backstop: no socket recv (IMAP fetch/search, HTTP) can block longer than this.
# Applies per-recv, so legitimate streaming transfers are unaffected — only a genuinely stuck
# read is cut. This is what actually kills the ~20-min OS-TCP-timeout hangs that stalled the
# scanner (imaplib's own timeout param didn't cover fetch reliably). Set BEFORE any socket opens.
socket.setdefaulttimeout(60)
import email as email_lib
from email.header import decode_header
import threading
import time
import re
import codecs
import secrets
import pyotp
import pdfplumber
from bidi.algorithm import get_display
from dotenv import load_dotenv

load_dotenv()

# ── Email polling config ─────────────────────────────────────
EMAIL_CONFIG = {
    'imap_server': 'imap.gmail.com',
    'imap_port': 993,
    'username': os.environ['EMAIL_USERNAME'],
    'password': os.environ['EMAIL_PASSWORD'],
    'sender_filter': 'onboarding@resend.dev',
    'subject_filter': '',
    'check_interval': 180,
    'enabled': True,
}

app = Flask(__name__)
app.secret_key = os.environ['FLASK_SECRET_KEY']
# Session-cookie hardening. Secure defaults on (Railway serves HTTPS); set COOKIE_INSECURE=1
# for local http development so the login cookie still works there.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=(os.environ.get('COOKIE_INSECURE') != '1'),
)

from health_check import health_bp
app.register_blueprint(health_bp)
CORS(app, resources={r"/api/*": {"origins": [
    "https://www.winner-ins.co.il",
    "https://winner-ins.co.il",
    "https://www.gaia-ins.co.il",
    "https://gaia-ins.co.il"
]}})
DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'renewals.db')).strip()
_db_dir = os.path.dirname(DB_PATH)
if _db_dir:
    os.makedirs(_db_dir, exist_ok=True)
print(f'[startup] DB_PATH={DB_PATH}')

@app.template_filter('fdate')
def format_date(value):
    if not value:
        return '—'
    s = str(value).strip()
    # YYYY-MM-DD HH:MM or YYYY-MM-DD HH:MM:SS
    if len(s) >= 10 and s[4] == '-':
        parts = s.split(' ', 1)
        d = parts[0].split('-')
        if len(d) == 3:
            result = f"{d[2]}/{d[1]}/{d[0]}"
            if len(parts) > 1:
                result += ' ' + parts[1][:5]
            return result
    return s

@app.template_filter('mask_card')
def mask_card(value):
    """Show only the last 4 digits of a card number (•••• 1234). Full number is revealed
    on demand via /reveal-card, which logs who revealed it."""
    d = re.sub(r'\D', '', str(value or ''))
    if not d:
        return ''
    return '•••• ' + d[-4:] if len(d) >= 4 else '••••'

STATUSES = ['', 'טופס התקבל', 'חודש', 'חודש - בוצעה שיחת מכירה', 'התקבל חידוש - כ.א לא תקין', 'לא רוצים לחדש', 'נוצר קשר עם לקוח']
# New-business import pipelines — never renewals, so excluded from the renewal funnel by SOURCE
# (in addition to the status-based exclusion) whatever their work status.
NEW_BUSINESS_SOURCES = ('new_policy', 'join_form', 'harel_proposal')
BRANDS = ['גאיה', 'ווינר', 'אופיר']

# Work-queue states for /admin/other-forms. 'ממתין' is the stored default from intake;
# it is displayed as "ממתין לטיפול".
FORM_QUEUE_STATUSES = ('ממתין', 'בטיפול', 'טופל')
FORM_QUEUE_LABELS = {'ממתין': 'ממתין לטיפול', 'בטיפול': 'בטיפול', 'טופל': 'טופל'}

# Identity/contact fields whose every edit is written to the field_changes audit log.
AUDITED_FIELDS = ('name', 'id_number', 'phone', 'email', 'address', 'policy_number', 'brand')
AUDIT_LABELS = {'name': 'שם', 'id_number': 'ת.ז', 'phone': 'טלפון', 'email': 'אימייל',
                'address': 'כתובת', 'policy_number': 'פוליסה', 'brand': 'סוכנות'}

# Status dropdowns differ per agency. Gaia/Winner keep the renewals workflow; Ofir
# (Meir's elementary book) has its own pipeline. Each entry is (stored value, label);
# '' is the default/unstarted state.
GW_STATUS_OPTIONS = [
    ('', 'ממתין לטיפול'),
    ('ממתין להפקה', '📝 ממתין להפקה (טופס חדש)'),
    ('טופס התקבל', '📋 טופס התקבל'),
    ('בוטל', '❌ בוטל'),
    ('הופק', 'הופק ✓'),
    ('חודש', 'חודש ✓'),
    ('חודש - בוצעה שיחת מכירה', 'חודש - בוצעה שיחת מכירה ✓'),
    ('התקבל חידוש - כ.א לא תקין', '⚠️ התקבל חידוש - כ.א לא תקין'),
    ('נוצר קשר עם לקוח', 'נוצר קשר עם לקוח'),
    ('ממתין לחידוש', 'ממתין לחידוש'),
    ('המשך טיפול בוואטסאפ', 'המשך טיפול בוואטסאפ'),
    ('ביקשו לחדש לבד', 'ביקשו לחדש לבד'),
    ('ביקש לחשוב/לבדוק', 'ביקש לחשוב/לבדוק'),
    ('ממתין לאישור מיילדות', 'ממתין לאישור מיילדות'),
    ('לא רוצים לחדש', 'לא רוצים לחדש'),
]
# Ofir uses the same status set as Gaia/Winner (its own list was retired 2026-08-31 per Sharon —
# initial status for everyone is 'ממתין לטיפול').
OFIR_STATUS_OPTIONS = GW_STATUS_OPTIONS

def status_options_for(brand):
    return GW_STATUS_OPTIONS

# Ofir renewal categories, split by the ענף (sector) column. Dashboard shows renewal
# % per category. Each entry is (label, [aliases]); a row matches if any alias is a
# substring of its sector (so 'דירות' lands in 'דירה', 'עסק'→'עסקים', etc.).
OFIR_CATEGORIES = [
    ('רכב',         ['רכב']),
    ('דירה',        ['דירה', 'דירות']),
    ('עסקים',       ['עסק']),
    ('משכנתא',      ['משכנת']),
    ('חובה',        ['חובה']),
    ('חבויות',      ['חבוי']),
    ('בריאות',      ['בריאות']),
    ('עובדים זרים', ['עובדים']),
]

# Optional elementary/car fields (mainly the Ofir/Meir book). Ordered (column, Hebrew
# label). Stored on both customers and insureds; the UI renders each only when it has a
# value, so Gaia/Winner records simply don't show them.
EXTRA_FIELD_DEFS = [
    ('insurer',            'חברה'),
    ('sector',             'ענף'),
    ('license_number',     'רישוי'),
    ('secondary_status',   'סטטוס משני'),
    ('cover_third_party',  "צד ג'"),
    ('cover_compulsory',   'חובה'),
    ('cover_comprehensive','מקיף'),
    ('cover_riders',       'ריידרים'),
    ('sum_insured',        'ס/מ'),
    ('offer_company',      'חברת ההצעה'),
    ('done_company',       'חברה שנעשה'),
    ('handler',            'מטפל'),
    ('sub_agent',          'סוכן מטפל'),
]
EXTRA_FIELDS = [c for c, _ in EXTRA_FIELD_DEFS]


@app.template_filter('form_fields')
def _form_fields(raw):
    """The stored submission JSON as (label, value) pairs for the two-column view.
    Returns [] when there is nothing usable, so the template can fall back."""
    if not raw:
        return []
    try:
        d = json.loads(raw)
    except (ValueError, TypeError):
        return []
    if not isinstance(d, dict):
        return []
    return [(k, v) for k, v in d.items() if str(v).strip() and str(v).strip() != '—']


@app.context_processor
def inject_extra_fields():
    """Make the optional-field defs and per-agency status sets available to every
    template (e.g. the customers list renders each row's dropdown by its brand)."""
    return {'extra_field_defs': EXTRA_FIELD_DEFS,
            'gw_status_options': GW_STATUS_OPTIONS,
            'ofir_status_options': OFIR_STATUS_OPTIONS}

def normalize_id_number(s):
    """Israeli ID numbers are 9 digits — left-pad short numeric IDs with zeros
    (e.g. 33775065 → 033775065). Leaves non-numeric or 9+ digit values untouched."""
    s = str(s or '').strip()
    if s.isdigit() and len(s) < 9:
        return s.zfill(9)
    return s

def is_israeli_id(s):
    """Validate an Israeli national-ID check digit (זהות). Used to tell the real
    insured ת.ז apart from Harel's internal 'מס' מזהה', which shares the same PDF row
    but does not satisfy the ID checksum."""
    d = re.sub(r'\D', '', str(s or ''))
    if not d or len(d) > 9:
        return False
    d = d.zfill(9)
    total = 0
    for i, ch in enumerate(d):
        v = int(ch) * (1 if i % 2 == 0 else 2)
        total += v if v < 10 else v - 9
    return total % 10 == 0

def parse_dmy(s):
    """Parse a DD/MM/YYYY date string (as extracted from Harel PDFs) to a date. None on failure."""
    s = str(s or '').strip()
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})', s)
    if not m:
        return None
    try:
        return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None

def brand_from_agency(agency):
    """Derive the brand (גאיה/ווינר) from the agency name on the policy. Historical
    'אופיר' agency names map to 'ווינר' — Ofir was mislabeled Winner business. Ofir
    remains a selectable agency (BRANDS) for permissions/manual assignment, just not
    auto-derived here, so the one-time relabel is not undone on the next rebuild."""
    a = str(agency or '')
    if 'גאיה' in a:
        return 'גאיה'
    if 'ווינר' in a or 'וינר' in a or 'אופיר' in a:
        return 'ווינר'
    return ''


def allowed_brands():
    """Agencies the current user may access. Returns None for super-admins (= everything);
    managers and agents are limited to their granted brands (possibly empty → sees nothing)."""
    if session.get('role') == 'superadmin':
        return None
    if 'brands' not in session:
        uid = session.get('user_id')
        if not uid:
            return []
        conn = get_db()
        rows = conn.execute("SELECT brand FROM user_brands WHERE user_id=?", (uid,)).fetchall()
        conn.close()
        session['brands'] = [r['brand'] for r in rows]
    return session['brands']


def brand_clause(col='brand'):
    """SQL fragment + params limiting `col` to the user's agencies. ('', []) for admins;
    a never-true clause when a non-admin has no agencies granted."""
    ab = allowed_brands()
    if ab is None:
        return '', []
    if not ab:
        return ' AND 1=0', []
    return f" AND {col} IN ({','.join('?' * len(ab))})", list(ab)


def can_access_brand(brand):
    """Whether the current user may see a record with the given brand."""
    ab = allowed_brands()
    return ab is None or (brand in ab)


def event_key(id_number, fallback):
    """Key for the unified activity log: normalised ת.ז, or a per-record fallback."""
    return (str(id_number or '').lstrip('0')) or fallback


def log_event(conn, idkey, note, who, kind='note'):
    if not (note or '').strip():
        return
    conn.execute(
        "INSERT INTO client_events (idkey, kind, note, created_by, created_at) VALUES (?,?,?,?,?)",
        (idkey, kind, note.strip(), who, datetime.datetime.now().strftime('%Y-%m-%d %H:%M')))


def get_events(conn, idkey, limit=100):
    return conn.execute(
        "SELECT * FROM client_events WHERE idkey=? ORDER BY id DESC LIMIT ?", (idkey, limit)
    ).fetchall()


def _name_search(col, search, like):
    """SQL condition + params for an order-independent name search: the whole string
    as a substring, OR every word appearing in the name in any order — so 'כהן עדן'
    also finds 'עדן כהן'. Returns (sql_fragment, params)."""
    words = [w for w in search.split() if w]
    clauses = [f'{col} LIKE ?']
    params = [like]
    if len(words) > 1:
        clauses.append('(' + ' AND '.join(f'{col} LIKE ?' for _ in words) + ')')
        params.extend(f'%{w}%' for w in words)
    return '(' + ' OR '.join(clauses) + ')', params

def compute_active_status(period_end):
    """Active if today is on/before the policy end date; inactive once it has passed."""
    end = parse_dmy(period_end)
    if not end:
        return 'פעיל'  # unknown end date — assume active until told otherwise
    return 'פעיל' if datetime.date.today() <= end else 'לא פעיל'

def recompute_insured_statuses(conn):
    """Weekly job: refresh פעיל/לא פעיל by date. Never touches admin-overridden rows
    or ones already marked בוטל."""
    changed = 0
    for r in conn.execute(
        "SELECT id, period_end, status FROM insureds WHERE status_override=0 AND status != 'בוטל'"
    ).fetchall():
        new_status = compute_active_status(r['period_end'])
        if new_status != r['status']:
            conn.execute("UPDATE insureds SET status=?, updated_at=? WHERE id=?",
                         (new_status, datetime.datetime.now().isoformat(), r['id']))
            changed += 1
    conn.commit()
    return changed

def _event_sort_key(r):
    """Order a policy_records row on the timeline: prefer the document date, then period start."""
    d = str(r['doc_date'] or '')
    ps = parse_dmy(r['period_start'])
    return (d, ps.isoformat() if ps else '')

def rebuild_insureds(conn):
    """Build/refresh the insureds master from policy_records — one row per ID number,
    using each person's LATEST policy event (by document date). If that latest event is
    a cancellation (ביטול) the insured is marked 'בוטל'; otherwise status is by policy
    period. A stand-alone cancellation with no prior policy still creates the insured
    from the cancellation's own details. Preserves existing activity and admin overrides."""
    # Group policy_records by normalized ID, keep the latest event on the timeline
    best = {}
    for r in conn.execute(
        "SELECT * FROM policy_records WHERE insured_id IS NOT NULL AND insured_id != ''"
    ).fetchall():
        idn = normalize_id_number(r['insured_id'])
        if not idn:
            continue
        k = _event_sort_key(r)
        if idn not in best or k >= best[idn][0]:
            best[idn] = (k, r)

    now = datetime.datetime.now().isoformat()
    upserted = 0
    for idn, (_, r) in best.items():
        agency = r['agent_name'] or ''
        brand = brand_from_agency(agency)
        wa_source = 'ווינר' if brand in ('ווינר', 'אופיר') else None
        existing = conn.execute("SELECT id, status_override FROM insureds WHERE id_number=?", (idn,)).fetchone()
        # Cancellation wins when it is the latest event; otherwise status by period
        if 'ביטול' in str(r['doc_type_label'] or ''):
            status = 'בוטל'
        else:
            status = compute_active_status(r['period_end'])
        if existing:
            # Refresh policy/contact facts but never clobber activity or an admin override
            keep_status = existing['status_override'] == 1
            conn.execute(
                """UPDATE insureds SET name=?, agency=?, brand=?, phone=?, email=?, address=?,
                   policy_number=?, period_start=?, period_end=?, whatsapp_source=COALESCE(whatsapp_source, ?),
                   status=CASE WHEN status_override=1 THEN status ELSE ? END, updated_at=?
                   WHERE id=?""",
                (r['insured_name'], agency, brand, r['phone_mobile'] or r['phone_home'] or '',
                 r['email'], r['address'], r['policy_number'], r['period_start'], r['period_end'],
                 wa_source, status, now, existing['id'])
            )
        else:
            conn.execute(
                """INSERT INTO insureds
                   (id_number, name, agency, brand, phone, email, address, policy_number,
                    period_start, period_end, status, whatsapp_source, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (idn, r['insured_name'], agency, brand, r['phone_mobile'] or r['phone_home'] or '',
                 r['email'], r['address'], r['policy_number'], r['period_start'], r['period_end'],
                 status, wa_source, now, now)
            )
        upserted += 1
    conn.commit()
    return upserted

def promote_customers_to_insureds(conn, month_id, brands=None):
    """Move a renewal month's customers into the insureds master (req 4), preserving
    all activity (calls, notes, VIP, rep credit). Renewed → פעיל, otherwise → לא פעיל
    (req 5). Non-destructive: the original customers rows are left intact as history.
    `brands` limits promotion to specific agencies (used when loading one agency)."""
    now = datetime.datetime.now().isoformat()
    promoted = 0
    q = "SELECT * FROM customers WHERE month_id=?"
    p = [month_id]
    if brands:
        q += f" AND brand IN ({','.join('?' * len(brands))})"
        p += list(brands)
    for cst in conn.execute(q, p).fetchall():
        idn = normalize_id_number(cst['id_number'])
        if not idn:
            continue
        status = 'פעיל' if cst['status'] in ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק') else 'לא פעיל'
        existing = conn.execute("SELECT * FROM insureds WHERE id_number=?", (idn,)).fetchone()
        if existing:
            # Fill blanks and set renewal-based status; never wipe existing activity.
            insured_has_calls = bool(existing['call_status_1'] or existing['call_status_2'] or existing['call_status_3'])
            conn.execute(
                """UPDATE insureds SET
                   name=COALESCE(NULLIF(name,''), ?),
                   phone=COALESCE(NULLIF(phone,''), ?),
                   brand=COALESCE(NULLIF(brand,''), ?),
                   whatsapp_source=COALESCE(whatsapp_source, ?),
                   agent_notes=COALESCE(NULLIF(agent_notes,''), ?),
                   is_vip=MAX(COALESCE(is_vip,0), ?),
                   is_midwife=MAX(COALESCE(is_midwife,0), ?),
                   handled_by=COALESCE(NULLIF(handled_by,''), ?),
                   policy_number=COALESCE(NULLIF(policy_number,''), ?),
                   status=?, updated_at=?
                   WHERE id=?""",
                (cst['name'], cst['phone'], cst['brand'], cst['whatsapp_source'],
                 cst['agent_notes'], cst['is_vip'] or 0, cst['is_midwife'] or 0,
                 cst['handled_by'], cst['policy_number'],
                 status, now, existing['id'])
            )
            if not insured_has_calls:
                conn.execute(
                    """UPDATE insureds SET call_date_1=?, call_status_1=?, call_by_1=?,
                       call_date_2=?, call_status_2=?, call_by_2=?,
                       call_date_3=?, call_status_3=?, call_by_3=? WHERE id=?""",
                    (cst['call_date_1'], cst['call_status_1'], cst['call_by_1'],
                     cst['call_date_2'], cst['call_status_2'], cst['call_by_2'],
                     cst['call_date_3'], cst['call_status_3'], cst['call_by_3'], existing['id'])
                )
        else:
            wa_source = cst['whatsapp_source'] or ('ווינר' if cst['brand'] in ('ווינר', 'אופיר') else None)
            conn.execute(
                """INSERT INTO insureds
                   (id_number, name, phone, brand, whatsapp_source, agent_notes, is_vip, is_midwife, handled_by,
                    policy_number, status,
                    call_date_1, call_status_1, call_by_1, call_date_2, call_status_2, call_by_2,
                    call_date_3, call_status_3, call_by_3, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (idn, cst['name'], cst['phone'], cst['brand'], wa_source, cst['agent_notes'],
                 cst['is_vip'] or 0, cst['is_midwife'] or 0, cst['handled_by'], cst['policy_number'], status,
                 cst['call_date_1'], cst['call_status_1'], cst['call_by_1'],
                 cst['call_date_2'], cst['call_status_2'], cst['call_by_2'],
                 cst['call_date_3'], cst['call_status_3'], cst['call_by_3'], now, now)
            )
        # Carry the optional elementary fields (+ email) into the master, filling blanks
        # only so a re-promotion never wipes a value already curated on the insured.
        iid_row = conn.execute("SELECT id FROM insureds WHERE id_number=?", (idn,)).fetchone()
        if iid_row:
            ckeys = cst.keys()
            for colname in EXTRA_FIELDS + ['email']:
                val = cst[colname] if colname in ckeys else None
                if val:
                    conn.execute(
                        f"UPDATE insureds SET {colname}=COALESCE(NULLIF({colname},''), ?) WHERE id=?",
                        (val, iid_row['id']))
        promoted += 1
    conn.commit()
    return promoted

def _sync_customer_to_insured(conn, cid, active=True):
    """Upsert one customer into the insureds master (כל הלקוחות) — used when a renewal
    ('חודש') or a new-business issuance ('הופק') should make the person active there.
    Non-destructive: fills blanks only, never wipes curated master data."""
    c = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    if not c:
        return
    idn = normalize_id_number(c['id_number'])
    if not idn:
        return
    ck = c.keys()
    def cv(k):
        return c[k] if k in ck else None
    now = datetime.datetime.now().isoformat()
    status = 'פעיל' if active else 'לא פעיל'
    ex = conn.execute("SELECT id FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?",
                      (idn.lstrip('0'),)).fetchone()
    if ex:
        conn.execute(
            """UPDATE insureds SET name=COALESCE(NULLIF(name,''),?), phone=COALESCE(NULLIF(phone,''),?),
               brand=COALESCE(NULLIF(brand,''),?), email=COALESCE(NULLIF(email,''),?),
               address=COALESCE(NULLIF(address,''),?), occupation=COALESCE(NULLIF(occupation,''),?),
               policy_number=COALESCE(NULLIF(policy_number,''),?), status=?, status_override=1, updated_at=?
               WHERE id=?""",
            (cv('name'), cv('phone'), cv('brand'), cv('email'), cv('address'), cv('occupation'),
             cv('policy_number'), status, now, ex['id']))
    else:
        wa_source = cv('whatsapp_source') or ('ווינר' if cv('brand') in ('ווינר', 'אופיר') else None)
        conn.execute(
            """INSERT INTO insureds (id_number, name, phone, brand, whatsapp_source, email, address,
               occupation, policy_number, status, status_override, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (idn, cv('name'), cv('phone'), cv('brand'), wa_source, cv('email'), cv('address'),
             cv('occupation'), cv('policy_number'), status, 1, now, now))

def _resolve_form_queue(conn, idn, escalations=False):
    """Mark this ת"ז's website-form submissions 'טופל' (drop off /admin/other-forms). Called
    when a lead is ingested (dedupe) and when a customer is issued/renewed. With escalations=True
    (issuance) also close the person's 'דורש בירור' admin-queue escalations."""
    z = (idn or '').lstrip('0')
    if not z:
        return
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    conn.execute("UPDATE unmatched_submissions SET status='טופל', handled_at=? "
                 "WHERE ltrim(COALESCE(id_number,''),'0')=? AND status IN ('ממתין','בטיפול')", (now, z))
    if escalations:
        conn.execute("UPDATE unmatched_submissions SET status='resolved', handled_at=? "
                     "WHERE ltrim(COALESCE(id_number,''),'0')=? AND status='pending'", (now, z))

# ── DB helpers ──────────────────────────────────────────────

def get_db():
    # WAL + a busy timeout let the web requests and the background email scan write
    # concurrently without hitting 'database is locked'.
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=5000")
    except Exception:
        pass
    return conn

def init_db():
    # One pre-migration safety backup per day, on the volume — a rollback point taken
    # BEFORE any schema change runs. Cheap insurance; never overwrites an existing one.
    try:
        _bdir = os.path.dirname(DB_PATH) or '.'
        _bpath = os.path.join(_bdir, 'renewals_backup_%s.db' % datetime.date.today().isoformat())
        if os.path.exists(DB_PATH) and not os.path.exists(_bpath):
            import shutil as _sh
            _sh.copy2(DB_PATH, _bpath)
            print('[init] pre-migration backup -> %s' % _bpath)
    except Exception as _e:
        print('[init] backup failed: %s' % _e)
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            display_name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'agent'
        );
        -- Which agencies (brands) each non-admin user may access. Admins see everything.
        CREATE TABLE IF NOT EXISTS user_brands (
            user_id INTEGER NOT NULL,
            brand TEXT NOT NULL,
            PRIMARY KEY (user_id, brand)
        );
        CREATE TABLE IF NOT EXISTS app_meta (key TEXT PRIMARY KEY, value TEXT);
        -- Recycle bin: a deleted customer is copied here (full row as JSON) before
        -- removal, so an accidental delete can be restored.
        -- Audit trail: every edit to an identity/contact field, old → new.
        CREATE TABLE IF NOT EXISTS field_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            field TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            changed_by TEXT,
            changed_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_field_changes_customer ON field_changes(customer_id);
        -- Activity log for a customer file: each saved note/status becomes an event, so
        -- anyone opening the file sees the latest conversation/activity.
        CREATE TABLE IF NOT EXISTS insured_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            insured_id INTEGER NOT NULL,
            kind TEXT DEFAULT 'note',
            note TEXT,
            created_by TEXT,
            created_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_insured_events ON insured_events(insured_id);
        -- Unified activity log per client (keyed by normalised ת.ז), so the same
        -- timeline shows on both the renewals card and the master file.
        CREATE TABLE IF NOT EXISTS client_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            idkey TEXT NOT NULL,
            kind TEXT DEFAULT 'note',
            note TEXT,
            created_by TEXT,
            created_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_client_events ON client_events(idkey);
        CREATE TABLE IF NOT EXISTS deleted_customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            name TEXT,
            brand TEXT,
            data TEXT,
            deleted_at TEXT,
            deleted_by TEXT
        );
        CREATE TABLE IF NOT EXISTS months (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month_id INTEGER NOT NULL,
            policy_number TEXT,
            name TEXT NOT NULL,
            id_number TEXT,
            phone TEXT,
            brand TEXT,
            status TEXT DEFAULT '',
            premium_last_year TEXT,
            whatsapp_sent_date TEXT,
            sharon_notes TEXT,
            requests_to_sharon TEXT,
            contact_date TEXT,
            agent_notes TEXT,
            interested_in_products TEXT,
            FOREIGN KEY (month_id) REFERENCES months(id)
        );
    ''')
    # Add form columns if missing (migration)
    existing = [r[1] for r in conn.execute("PRAGMA table_info(customers)").fetchall()]
    for col, typ in [('form_email','TEXT'), ('form_installments','TEXT'),
                     ('form_payment_method','TEXT'), ('form_received_at','TEXT'),
                     ('form_coverage','TEXT'), ('form_comments','TEXT'),
                     ('is_vip','INTEGER DEFAULT 0'), ('whatsapp_source','TEXT'),
                     ('call_date_1','TEXT'), ('call_status_1','TEXT'), ('call_by_1','TEXT'),
                     ('call_date_2','TEXT'), ('call_status_2','TEXT'), ('call_by_2','TEXT'),
                     ('call_date_3','TEXT'), ('call_status_3','TEXT'), ('call_by_3','TEXT')]:
        if col not in existing:
            conn.execute(f"ALTER TABLE customers ADD COLUMN {col} {typ}")
    # One-time backfill: move any legacy single contact_date into call slot 1
    if 'call_date_1' not in existing:
        conn.execute("""UPDATE customers SET call_date_1=contact_date
                        WHERE contact_date IS NOT NULL AND contact_date != ''""")

    # Table to track processed emails by Message-ID
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS processed_emails (
            message_id TEXT PRIMARY KEY,
            processed_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS processed_leads (
            message_id TEXT PRIMARY KEY,
            processed_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS owner_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            text TEXT NOT NULL,
            created_at TEXT,
            sent_at TEXT
        );
        CREATE TABLE IF NOT EXISTS cert_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket TEXT UNIQUE,
            cust_name TEXT,
            id_number TEXT,
            phone TEXT,
            email TEXT,
            brand TEXT,
            customer_id INTEGER,
            received_at TEXT,
            match_status TEXT,
            pdf_saved INTEGER DEFAULT 0,
            wa_sent_at TEXT,
            email_sent_at TEXT,
            wa_target TEXT,
            message_id TEXT,
            cert_labeled TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS customer_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE TABLE IF NOT EXISTS policy_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            policy_number TEXT,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            received_at TEXT NOT NULL,
            message_id TEXT UNIQUE,
            whatsapp_sent_at TEXT,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE TABLE IF NOT EXISTS policy_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            policy_document_id INTEGER,
            customer_id INTEGER,
            policy_number TEXT,
            doc_type_label TEXT,
            doc_type_code TEXT,
            branch TEXT,
            agent_name TEXT,
            agent_number TEXT,
            insured_name TEXT,
            insured_id TEXT,
            spouse_id TEXT,
            address TEXT,
            phone_mobile TEXT,
            phone_home TEXT,
            email TEXT,
            period_start TEXT,
            period_end TEXT,
            premium TEXT,
            total_payment TEXT,
            doc_date TEXT,
            extracted_at TEXT NOT NULL,
            FOREIGN KEY (policy_document_id) REFERENCES policy_documents(id),
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE TABLE IF NOT EXISTS insureds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_number TEXT UNIQUE,
            name TEXT,
            agency TEXT,
            brand TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            policy_number TEXT,
            period_start TEXT,
            period_end TEXT,
            status TEXT DEFAULT 'פעיל',
            status_override INTEGER DEFAULT 0,
            whatsapp_source TEXT,
            agent_notes TEXT,
            is_vip INTEGER DEFAULT 0,
            handled_by TEXT,
            call_date_1 TEXT, call_status_1 TEXT, call_by_1 TEXT,
            call_date_2 TEXT, call_status_2 TEXT, call_by_2 TEXT,
            call_date_3 TEXT, call_status_3 TEXT, call_by_3 TEXT,
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS unmatched_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            received_at TEXT,
            subject TEXT,
            name TEXT, id_number TEXT, phone TEXT, email TEXT,
            brand TEXT, installments TEXT, payment_method TEXT,
            card_number TEXT, card_expiry TEXT, card_holder_id TEXT,
            coverage TEXT, comments TEXT,
            status TEXT DEFAULT 'pending',
            admin_note TEXT,
            message_id TEXT UNIQUE
        );
    ''')

    # Add card + tracking columns if missing
    existing = [r[1] for r in conn.execute("PRAGMA table_info(customers)").fetchall()]
    for col, typ in [('form_card_number','TEXT'), ('form_card_expiry','TEXT'),
                     ('form_id_card_holder','TEXT'), ('handled_by','TEXT'), ('email','TEXT'),
                     ('address','TEXT'), ('status_changed_at','TEXT'),
                     ('is_midwife','INTEGER'),
                     ('lead_form_json','TEXT'), ('marketing_consent','TEXT'),
                     ('lead_doc_path','TEXT'), ('lead_doc_saved','TEXT'),
                     ('end_reminder_sent_date','TEXT'), ('group_owner','TEXT'),
                     ('lr25_sent_at','TEXT'), ('lreom_sent_at','TEXT')]:
        if col not in existing:
            conn.execute(f"ALTER TABLE customers ADD COLUMN {col} {typ}")
    if 'is_midwife' not in [r[1] for r in conn.execute("PRAGMA table_info(insureds)").fetchall()]:
        conn.execute("ALTER TABLE insureds ADD COLUMN is_midwife INTEGER")
    # Bulk-send delivery log (bot reports back; upsert by wamid) + marketing opt-outs.
    conn.execute("""CREATE TABLE IF NOT EXISTS send_log (
        wamid TEXT PRIMARY KEY, cust_id INTEGER, send_type TEXT, status TEXT,
        error_code TEXT, sent_at TEXT, updated_at TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS optouts (
        id INTEGER PRIMARY KEY AUTOINCREMENT, phone TEXT, id_number TEXT, reason TEXT, created_at TEXT)""")
    # Bulk-send approval batches — a frozen recipient snapshot Sharon approves before the bot sends.
    conn.execute("""CREATE TABLE IF NOT EXISTS send_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT, send_type TEXT, template TEXT, status TEXT,
        count INTEGER, created_at TEXT, approved_at TEXT, approved_by TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS send_batch_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, cust_id INTEGER, name TEXT,
        phone TEXT, brand TEXT, pnid TEXT, template TEXT, body_params TEXT)""")
    # No-send dates (Jewish holidays / eves) for the bulk-send timing gate — Sharon-editable.
    # Fri/Sat are handled automatically in code; this table adds holidays. Seeded ONCE (Tishrei
    # 5787) so Sharon's later edits/deletes persist across restarts.
    conn.execute("CREATE TABLE IF NOT EXISTS no_send_dates (date TEXT PRIMARY KEY, label TEXT)")
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='no_send_seeded'").fetchone():
        for dt, lbl in [('2026-09-11', 'ערב ראש השנה'), ('2026-09-12', 'ראש השנה א׳'),
                        ('2026-09-13', 'ראש השנה ב׳'), ('2026-09-20', 'ערב יום כיפור'),
                        ('2026-09-21', 'יום כיפור'), ('2026-09-25', 'ערב סוכות'),
                        ('2026-09-26', 'סוכות א׳'), ('2026-10-02', 'הושענא רבה'),
                        ('2026-10-03', 'שמחת תורה')]:
            conn.execute("INSERT OR IGNORE INTO no_send_dates (date, label) VALUES (?,?)", (dt, lbl))
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('no_send_seeded','1') "
                     "ON CONFLICT(key) DO UPDATE SET value='1'")
    if 'group_owner' not in [r[1] for r in conn.execute("PRAGMA table_info(insureds)").fetchall()]:
        conn.execute("ALTER TABLE insureds ADD COLUMN group_owner TEXT")
    # cert_requests: email columns (table shipped before the "both email+WhatsApp" rule)
    _cert_cols = [r[1] for r in conn.execute("PRAGMA table_info(cert_requests)").fetchall()]
    for col in ('email', 'email_sent_at', 'message_id', 'cert_labeled'):
        if col not in _cert_cols:
            conn.execute(f"ALTER TABLE cert_requests ADD COLUMN {col} TEXT")
    # Extra elementary/car fields (mainly from the Ofir/Meir book). All optional — shown
    # in the UI only when populated. Added to both customers and the insureds master.
    for tbl in ('customers', 'insureds'):
        have = [r[1] for r in conn.execute(f"PRAGMA table_info({tbl})").fetchall()]
        for col in EXTRA_FIELDS:
            if col not in have:
                conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} TEXT")

    # Add handled_by to unmatched_submissions if missing
    existing_us = [r[1] for r in conn.execute("PRAGMA table_info(unmatched_submissions)").fetchall()]
    if 'handled_by' not in existing_us:
        conn.execute("ALTER TABLE unmatched_submissions ADD COLUMN handled_by TEXT")
    if 'assigned_to' not in existing_us:  # user_id the rep routed this escalation to
        conn.execute("ALTER TABLE unmatched_submissions ADD COLUMN assigned_to INTEGER")
    if 'handled_at' not in existing_us:  # when the form-queue item was last advanced
        conn.execute("ALTER TABLE unmatched_submissions ADD COLUMN handled_at TEXT")
    if 'insured_id' not in existing_us:  # the customer file this form was attached to
        conn.execute("ALTER TABLE unmatched_submissions ADD COLUMN insured_id INTEGER")
    if 'raw_fields' not in existing_us:  # full submitted form, as JSON, for display
        conn.execute("ALTER TABLE unmatched_submissions ADD COLUMN raw_fields TEXT")
    # Audit rows can belong to a customer (customer_id) or a customer file (insured_id).
    existing_fc = [r[1] for r in conn.execute("PRAGMA table_info(field_changes)").fetchall()]
    if existing_fc and 'insured_id' not in existing_fc:
        conn.execute("ALTER TABLE field_changes ADD COLUMN insured_id INTEGER")
    # Which manager an agent reports to (for the agent-performance view).
    existing_user_cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if 'manager_id' not in existing_user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN manager_id INTEGER")
    # One-time (guarded): the historical 'אופיר' brand was mislabeled Winner business —
    # fold it into 'ווינר' across all data. Ofir stays a selectable agency for permissions
    # but is no longer auto-derived (see brand_from_agency), so this relabel sticks.
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='ofir_to_winner_done'").fetchone():
        for tbl in ('customers', 'insureds', 'unmatched_submissions'):
            conn.execute(f"UPDATE {tbl} SET brand='ווינר' WHERE brand='אופיר'")
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('ofir_to_winner_done', ?)",
                     (datetime.datetime.now().isoformat(),))
    # One-time (guarded): a stray customer row had a garbage brand ('י', an import
    # typo). Assign it to Winner per request so it stops being invisible to agents.
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='fix_stray_brand_done'").fetchone():
        conn.execute("UPDATE customers SET brand='ווינר' WHERE brand NOT IN ('גאיה','ווינר','אופיר') AND brand IS NOT NULL AND brand != ''")
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('fix_stray_brand_done', ?)",
                     (datetime.datetime.now().isoformat(),))
    # One-time (guarded): fold the per-file insured_events into the unified client_events,
    # keyed by the insured's normalised ת.ז (fallback: ins-<id> when there is no ת.ז).
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='events_unified'").fetchone():
        for e in conn.execute(
            "SELECT ev.kind, ev.note, ev.created_by, ev.created_at, ins.id_number, ins.id AS iid "
            "FROM insured_events ev LEFT JOIN insureds ins ON ins.id=ev.insured_id").fetchall():
            key = (e['id_number'] or '').lstrip('0') or ('ins-%s' % e['iid'])
            conn.execute("INSERT INTO client_events (idkey, kind, note, created_by, created_at)"
                         " VALUES (?,?,?,?,?)", (key, e['kind'], e['note'], e['created_by'], e['created_at']))
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('events_unified', ?)",
                     (datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),))
    # One-time (guarded): seed the activity log from existing insured notes, then clear
    # them — from now on the notes box is a scratchpad and saved notes become events.
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='insured_notes_to_events'").fetchone():
        now_s = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        for r in conn.execute("SELECT id, agent_notes FROM insureds WHERE COALESCE(agent_notes,'')!=''").fetchall():
            conn.execute("INSERT INTO insured_events (insured_id, kind, note, created_by, created_at)"
                         " VALUES (?,?,?,?,?)", (r[0], 'note', r[1], '(הערה קודמת)', now_s))
        conn.execute("UPDATE insureds SET agent_notes='' WHERE COALESCE(agent_notes,'')!=''")
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('insured_notes_to_events', ?)", (now_s,))
    # One-time (guarded): introduce the super-admin tier. Sharon becomes 'superadmin';
    # any other existing 'admin' stays a manager (agency-scoped). Mark more super-admins
    # later from the users screen.
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='superadmin_seed_done'").fetchone():
        conn.execute("UPDATE users SET role='superadmin' WHERE username='sharon'")
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('superadmin_seed_done', ?)",
                     (datetime.datetime.now().isoformat(),))
    # One-time (guarded): rename the old "לקוח ענה/ V כחול" status to "נוצר קשר עם לקוח".
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='status_rename_done'").fetchone():
        for tbl in ('customers', 'insureds'):
            conn.execute(f"UPDATE {tbl} SET status='נוצר קשר עם לקוח' WHERE status='לקוח ענה/ V כחול'")
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('status_rename_done', ?)",
                     (datetime.datetime.now().isoformat(),))
    # One-time (guarded): seed agency access for pre-existing agents so nobody is locked
    # out — default to Gaia + Winner (not Ofir), matching the intended baseline.
    if not conn.execute("SELECT 1 FROM app_meta WHERE key='seed_user_brands_done'").fetchone():
        for u in conn.execute("SELECT id FROM users WHERE role != 'admin'").fetchall():
            if not conn.execute("SELECT 1 FROM user_brands WHERE user_id=?", (u[0],)).fetchone():
                for b in ('גאיה', 'ווינר'):
                    conn.execute("INSERT OR IGNORE INTO user_brands (user_id, brand) VALUES (?,?)", (u[0], b))
        conn.execute("INSERT INTO app_meta (key, value) VALUES ('seed_user_brands_done', ?)",
                     (datetime.datetime.now().isoformat(),))
    # One-time cleanup: purge automated morning monitor tests captured before the
    # ingestion-level filter existed. Marker-based rows, plus fully-empty rows (the
    # no-field monitor forms like 'מינוי סוכן' leave no identifying data → not actionable).
    conn.execute(
        "DELETE FROM unmatched_submissions WHERE "
        "COALESCE(id_number,'')='999999999' OR COALESCE(email,'')='monitor-check@example.com' "
        "OR COALESCE(name,'')='MONITOR-CHECK-DO-NOT-PROCESS' OR ("
        "COALESCE(name,'')='' AND COALESCE(id_number,'')='' "
        "AND COALESCE(phone,'')='' AND COALESCE(email,'')='')"
    )
    # Add doc_date to policy_records if missing; backfill from linked document date
    existing_pr = [r[1] for r in conn.execute("PRAGMA table_info(policy_records)").fetchall()]
    if 'doc_date' not in existing_pr:
        conn.execute("ALTER TABLE policy_records ADD COLUMN doc_date TEXT")
        conn.execute("""UPDATE policy_records SET doc_date=(
            SELECT received_at FROM policy_documents WHERE policy_documents.id=policy_records.policy_document_id)
            WHERE doc_date IS NULL""")
    # Auto policy-delivery tracking: per-channel send timestamps on the document.
    existing_pd = [r[1] for r in conn.execute("PRAGMA table_info(policy_documents)").fetchall()]
    if 'email_sent_at' not in existing_pd:
        conn.execute("ALTER TABLE policy_documents ADD COLUMN email_sent_at TEXT")
    if 'gmail_labeled' not in [r[1] for r in conn.execute("PRAGMA table_info(policy_documents)").fetchall()]:
        conn.execute("ALTER TABLE policy_documents ADD COLUMN gmail_labeled TEXT")
    conn.commit()

    # Zero-pad short numeric ID numbers to 9 digits (idempotent — once padded,
    # length is 9 so the row is no longer selected).
    short_ids = conn.execute(
        "SELECT id, id_number FROM customers "
        "WHERE id_number GLOB '[0-9]*' AND id_number NOT GLOB '*[^0-9]*' AND length(id_number) < 9"
    ).fetchall()
    for row in short_ids:
        conn.execute("UPDATE customers SET id_number=? WHERE id=?",
                     (row[1].zfill(9), row[0]))
    if short_ids:
        conn.commit()

    # Add email + 2FA columns to users if missing (email code / TOTP app).
    existing_user_cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    for _col, _ddl in (('email', 'TEXT'), ('totp_secret', 'TEXT'), ('twofa_method', "TEXT DEFAULT 'none'")):
        if _col not in existing_user_cols:
            conn.execute(f"ALTER TABLE users ADD COLUMN {_col} {_ddl}")
    conn.commit()

    # Two-phase month load: staged uploads awaiting an explicit commit ("שינוי חידוש לחודש חדש").
    conn.execute("""CREATE TABLE IF NOT EXISTS pending_imports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT, month_name TEXT, filename TEXT,
        file_blob BLOB, report_json TEXT,
        uploaded_at TEXT, uploaded_by TEXT,
        status TEXT DEFAULT 'pending'
    )""")
    # 'טלפון אחר' (phone from the renewal file when it differs) + import provenance.
    for tbl in ('customers', 'insureds'):
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({tbl})").fetchall()]
        if 'alt_phone' not in cols:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN alt_phone TEXT")
        if 'import_source' not in cols:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN import_source TEXT")
        if 'occupation' not in cols:  # עיסוק המבוטח (extracted from the policy PDF, page דף-2)
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN occupation TEXT")
    # Renewal-campaign tracking (email touch date + do-not-contact opt-out).
    _ccols = [r[1] for r in conn.execute("PRAGMA table_info(customers)").fetchall()]
    if 'email_sent_date' not in _ccols:
        conn.execute("ALTER TABLE customers ADD COLUMN email_sent_date TEXT")
    if 'do_not_contact' not in _ccols:
        conn.execute("ALTER TABLE customers ADD COLUMN do_not_contact INTEGER DEFAULT 0")
    # "Update your payment method" message tracking (status 'התקבל חידוש - כ.א לא תקין'),
    # per channel so email still goes out when WhatsApp is held/down.
    if 'card_update_wa_at' not in _ccols:
        conn.execute("ALTER TABLE customers ADD COLUMN card_update_wa_at TEXT")
    if 'card_update_email_at' not in _ccols:
        conn.execute("ALTER TABLE customers ADD COLUMN card_update_email_at TEXT")
    # Simple key/value store — used for the email-scanner heartbeat ('last_scan_at').
    conn.execute("CREATE TABLE IF NOT EXISTS app_kv (k TEXT PRIMARY KEY, v TEXT)")
    conn.commit()

    # Default admin
    if not conn.execute("SELECT id FROM users WHERE username='sharon'").fetchone():
        conn.execute(
            "INSERT INTO users (username, password_hash, display_name, role) VALUES (?,?,?,?)",
            ('sharon', generate_password_hash('admin123'), 'שרון', 'admin')
        )
    conn.commit()

    # First-time backfill of the insureds master from existing policy PDFs
    have_insureds = conn.execute("SELECT COUNT(*) FROM insureds").fetchone()[0]
    have_records = conn.execute("SELECT COUNT(*) FROM policy_records").fetchone()[0]
    if have_insureds == 0 and have_records > 0:
        try:
            rebuild_insureds(conn)
        except Exception as e:
            print(f'[init] insureds backfill שגיאה: {e}')

    conn.close()

def active_month():
    conn = get_db()
    m = conn.execute("SELECT * FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return m

# ── Auth ────────────────────────────────────────────────────

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Manager-level and above (superadmin + admin). Data is still agency-scoped for
    managers via allowed_brands(); superadmins see everything."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ('superadmin', 'admin'):
            flash('גישה מנהל בלבד', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def superadmin_required(f):
    """Super-admin only — user management, imports, cross-agency structural changes."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'superadmin':
            flash('גישה למנהל-על בלבד', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def _refresh_session_from_db():
    """Keep role + agency grants fresh from the DB so a role/permission change takes
    effect on the next request — a session created before such a change won't get stuck
    with stale (and possibly over-restrictive) permissions."""
    uid = session.get('user_id')
    if not uid:
        return
    conn = get_db()
    u = conn.execute("SELECT role, username, display_name FROM users WHERE id=?", (uid,)).fetchone()
    if u:
        session['role'] = u['role']
        session['username'] = u['username']
        session['display_name'] = u['display_name']
        if u['role'] == 'superadmin':
            session.pop('brands', None)
        else:
            session['brands'] = [r['brand'] for r in
                                 conn.execute("SELECT brand FROM user_brands WHERE user_id=?", (uid,)).fetchall()]
    conn.close()

# ── אישור קיום ביטוחים (נספח א') — data ──────────────────────
# The standard regulated "Certificate of Insurance" (נספח א') that ~9 recurring
# companies require signed. Sharon is authorized by Harel to issue these. The form
# layout is fixed; only the per-company block, the per-customer block, and the policy
# dates change. Every field is editable in the preview before printing (the source of
# truth is the signed print-out, so the tool only pre-fills a best-effort draft).

# Suffix appended to every requesting company's name (exact wording, per the official form).
CERT_RELATED_SUFFIX = 'ו/או חברות אם ו/או חברות בנות ו/או חברות שלובות ו/או חברות קשורות'

# The 9 requesting companies. `codes` = the "פירוט השירותים" service code(s) on page 2.
# `hp_extra` (מובמנט only) = additional ח.פ. numbers shown on a lower line, separated by "/".
INSURANCE_CERT_COMPANIES = [
    {'key': 'target',   'name': 'טרגט קאר בע"מ',
     'hp': '515732162', 'hp_extra': '', 'address': 'הבושם 3, אשדוד', 'codes': '73'},
    {'key': 'movement', 'name': 'מוב וולנס בע"מ ו/או פריפיט בע"מ ו/או מובמנט בע"מ',
     'hp': '514527456', 'hp_extra': '516868965 / 513600528', 'address': 'הברזל 30, תל אביב', 'codes': '94, 95'},
    {'key': 'fami',     'name': 'פמי פרימיום בע"מ',
     'hp': '512676206', 'hp_extra': '', 'address': 'המשביר 1, חולון', 'codes': '73'},
    {'key': 'natali',   'name': 'נטלי החברה לשירותי רפואה דחופה בע"מ',
     'hp': '511441701', 'hp_extra': '', 'address': 'החילזון 4, רמת גן', 'codes': '94'},
    {'key': 'bikurofe', 'name': 'ביקורופא בע"מ',
     'hp': '511657322', 'hp_extra': '', 'address': 'יגאל אלון 90, תל אביב', 'codes': '73'},
    {'key': 'bewell',   'name': 'בי וול פתרונות לאיכות חיים בע"מ',
     'hp': '514163823', 'hp_extra': '', 'address': 'בני גאון 14, נתניה', 'codes': '94, 95'},
    {'key': 'fattal',   'name': 'מלונות פתאל בע"מ',
     'hp': '510678816', 'hp_extra': '', 'address': 'יגאל אלון 94, תל אביב', 'codes': '94'},
    {'key': 'dan',      'name': 'מלונות דן בע"מ',
     'hp': '520023573', 'hp_extra': '', 'address': 'הירקון 111, תל אביב', 'codes': '94'},
    {'key': 'space',    'name': 'ספייס מועדוני כושר בע"מ',
     'hp': '515190866', 'hp_extra': '', 'address': 'רפפורט 3, כפר סבא', 'codes': '31'},
]

# Fixed values shared by every certificate (defaults — editable in the preview).
CERT_CONSTANTS = {
    'cert_number':     '7338-0001',     # מספר אישור — בסיס/גיבוי; בפועל רץ סדרתית (ראה next_cert_number)
    'form_edition':    '01/2022',       # נוסח ומהדורת ביטוח
    'amount':          '1,200,000',     # גבול אחריות / סכום ביטוח (₪)
    'currency':        'ש"ח',
    'codes_main':      '302, 304, 309, 315, 321, 322, 326, 328, 329',  # צד ג' / אחריות מקצועית רפואית
    'codes_supp':      '332',            # אחריות מקצועית רפואה משלימה
    'discovery_codes': '302, 304, 309, 315, 321, 326, 327, 328',       # תקופת גילוי
    'deal_type':       'נדל"ן',          # אופי העסקה (highlighted option)
    'req_status':      'משכיר',          # מעמד מבקש האישור (highlighted option)
    'insurer':         'הראל חברה לביטוח בע"מ',
}

# Who may issue certificates. Restricted to Sharon's user only (until further notice) —
# gates both the route and the visibility of the nav/dashboard links.
INSURANCE_CERT_USERS = {'sharon'}

def can_issue_cert():
    return session.get('username') in INSURANCE_CERT_USERS

@app.context_processor
def _inject_cert_perm():
    """Expose `can_issue_cert` to all templates (nav + dashboard link visibility)."""
    return {'can_issue_cert': can_issue_cert()}


def extract_insured_occupation(pdf_path):
    """Best-effort pull of "העיסוק המבוטח" from page 2 of a stored Harel policy PDF.
    Page 2 lists occupation/institution/year triplets under the header
    `עיסוק המבוטח | מוסד הסמכה | שנת הסמכה`. We want the occupation column only; there
    can be several (e.g. "פילאטיס מזרן, פילאטיס מכשירים"). Returns a comma-joined string,
    or '' when nothing is found (the field is then filled manually in the preview)."""
    if not pdf_path or not os.path.exists(pdf_path):
        return ''
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page_lines = None
            for page in pdf.pages:
                # pdfplumber extracts RTL text reversed → bidi-process BEFORE searching.
                bidi = get_display(page.extract_text() or '')
                if 'עיסוק המבוטח' in bidi and 'הסמכה' in bidi:
                    page_lines = [l.strip() for l in bidi.split('\n')]
                    break
        if not page_lines:
            return ''
    except Exception as e:
        print(f'[cert-occupation] שגיאת קריאת PDF: {e}')
        return ''

    # The block is: header row → dashes → occupation line(s) → institution → year+licence.
    hi = next((i for i, l in enumerate(page_lines)
               if 'עיסוק המבוטח' in l and 'הסמכה' in l), None)
    if hi is None:
        return ''
    INSTITUTION = ('משרד', 'מדינת', 'אוניברסיט', 'מכון', 'בית ספר', 'ביה"ס', 'מכלל',
                   'קולג', 'ארגון', 'המכללה', 'הטכניון', 'הסמכה')
    STOP = ('במקום חריג', 'הביטוח ל', 'תשומת לב', 'הפוליסה', 'תא סוכן', 'גבול', 'הרחב',
            'פרמיה', 'תאריך', 'כפוף', 'למען הסר', 'במסגרת', 'חריג', 'כולל מפגשים',
            'סכום', 'תאור הכיסוי', 'המבטח')
    occupations = []
    for l in page_lines[hi + 1:]:
        if not l or set(l) <= set('- '):        # blank or dashes separator
            continue
        if any(k in l for k in INSTITUTION):     # reached the institution column → stop
            break
        if re.fullmatch(r'[\d\s.\-/]+', l):      # a year / licence-number line → stop
            break
        if any(k in l for k in STOP):            # coverage / boilerplate → stop
            break
        occ = re.sub(r'\s*\d[\d\s]*$', '', l).strip(' -|,\t')  # trailing numbers
        if occ and not re.fullmatch(r'[\W_]+', occ):
            occupations.append(occ)
        if len(occupations) >= 3:                # safety — take the leading occupation lines
            break
    # De-dup, order-preserving.
    seen, uniq = set(), []
    for o in occupations:
        if o not in seen:
            seen.add(o)
            uniq.append(o)
    return ', '.join(uniq)

# ── Bulk occupation fill (server-side, from the stored Harel policy PDFs) ─────
# The PDFs already live on the Railway volume, so extraction runs on the server —
# no OneDrive on-demand hydration (which crashed the earlier local batch).
_occ_fill_state = {'running': False, 'done': False, 'total': 0, 'scanned': 0,
                   'filled': 0, 'no_file': 0, 'no_occ': 0, 'error': None,
                   'started_at': None, 'finished_at': None}

def _run_fill_occupations(month_id, overwrite):
    st = _occ_fill_state
    try:
        conn = get_db()
        # Diagnostics: how many stored PDFs exist at all, and on disk.
        st['docs_with_file'] = conn.execute(
            "SELECT COUNT(*) FROM policy_documents WHERE COALESCE(filepath,'')!=''").fetchone()[0]
        # Match a stored policy PDF to each active-month customer BY ת"ז (a doc's customer_id
        # may point at a previous month's row), via the parsed insured_id on policy_records.
        rows = conn.execute(
            """SELECT c.id AS cid, c.id_number, c.occupation, pd.filepath
               FROM customers c
               JOIN policy_records pr
                 ON ltrim(COALESCE(pr.insured_id,''),'0') = ltrim(COALESCE(c.id_number,''),'0')
               JOIN policy_documents pd ON pd.id = pr.policy_document_id
               WHERE c.month_id=? AND COALESCE(pd.filepath,'')!=''
               ORDER BY c.id, pd.id DESC""", (month_id,)).fetchall()
        best = {}
        for r in rows:                       # newest stored PDF per customer
            if r['cid'] not in best:
                best[r['cid']] = r
        st['total'] = len(best)
        for cid, r in best.items():
            if (r['occupation'] or '') and not overwrite:
                continue
            fp = r['filepath']
            if not fp or not os.path.exists(fp):
                st['no_file'] += 1
                continue
            st['scanned'] += 1
            occ = extract_insured_occupation(fp)
            if not occ:
                st['no_occ'] += 1
                continue
            conn.execute("UPDATE customers SET occupation=? WHERE id=?", (occ, cid))
            idn = normalize_id_number(r['id_number'])
            if idn:
                conn.execute(
                    "UPDATE insureds SET occupation=COALESCE(NULLIF(occupation,''),?) "
                    "WHERE ltrim(COALESCE(id_number,''),'0')=?", (occ, idn.lstrip('0')))
            st['filled'] += 1
            if st['filled'] % 10 == 0:
                conn.commit()
        conn.commit()
        conn.close()
    except Exception as e:
        st['error'] = str(e)
    finally:
        st['running'] = False
        st['done'] = True
        st['finished_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

@app.route('/api/fill-occupations', methods=['POST'])
def api_fill_occupations():
    """Extract 'עיסוק המבוטח' from each stored policy PDF of the active month and fill the
    occupation field (customers + insureds master). Runs in a background thread (token)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    if _occ_fill_state.get('running'):
        return jsonify({'error': 'already running', 'state': _occ_fill_state}), 409
    month = active_month()
    if not month:
        return jsonify({'error': 'no active month'}), 400
    overwrite = bool((request.get_json(silent=True) or {}).get('overwrite'))
    for k in ('total', 'scanned', 'filled', 'no_file', 'no_occ'):
        _occ_fill_state[k] = 0
    _occ_fill_state.update({'running': True, 'done': False, 'error': None,
                            'started_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
                            'finished_at': None})
    threading.Thread(target=_run_fill_occupations, args=(month['id'], overwrite),
                     daemon=True).start()
    return jsonify({'ok': True, 'started': True})

@app.route('/api/fill-occupations-status')
def api_fill_occupations_status():
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    return jsonify(_occ_fill_state)

@app.route('/api/resolve-issued-forms', methods=['POST'])
def api_resolve_issued_forms():
    """One-time backfill: resolve website-form submissions whose ת"ז is already an issued
    ('הופק') / renewed ('חודש') customer, so they drop off /admin/other-forms. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    n = conn.execute(
        "UPDATE unmatched_submissions SET status='טופל', handled_at=? "
        "WHERE status IN ('ממתין','בטיפול') AND ltrim(COALESCE(id_number,''),'0') IN "
        "(SELECT ltrim(COALESCE(id_number,''),'0') FROM customers WHERE status IN ('הופק','חודש','חודש - בוצעה שיחת מכירה'))",
        (datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),)).rowcount
    conn.commit()
    conn.close()
    return jsonify({'resolved': n})

@app.route('/api/policy/doc-types')
def api_policy_doc_types():
    """Distinct doc_type_labels in policy_records + a sample stored doc_id for each (to find
    endorsements/cancellations). Token-authed diagnostic."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    rows = conn.execute(
        "SELECT pr.doc_type_label AS lbl, COUNT(*) AS n, "
        "MAX(CASE WHEN COALESCE(pd.filepath,'')!='' THEN pd.id END) AS sample_doc "
        "FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
        "GROUP BY pr.doc_type_label ORDER BY n DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/policy/pdf-lines')
def api_policy_pdf_lines():
    """Bidi text lines of a stored policy PDF (schedule page) — to locate fields. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    doc_id = request.args.get('doc_id', type=int)
    conn = get_db()
    d = conn.execute("SELECT filepath FROM policy_documents WHERE id=?", (doc_id,)).fetchone()
    conn.close()
    if not d or not d['filepath'] or not os.path.exists(d['filepath']):
        return jsonify({'error': 'no file'})
    return jsonify({'lines': _policy_pdf_lines(d['filepath'], limit=90)})

@app.route('/api/resolve-forms-with-policy', methods=['POST'])
def api_resolve_forms_with_policy():
    """Forms in /admin/other-forms whose ת"ז had a policy ISSUED on/after the form date → the
    request was fulfilled → mark 'טופל', drop from the report, document it in the customer file.
    Runs as a backfill + is safe to re-run. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    forms = conn.execute(
        "SELECT id AS uid, id_number, subject, received_at FROM unmatched_submissions "
        "WHERE status IN ('ממתין','בטיפול') AND COALESCE(id_number,'')!=''").fetchall()
    resolved = []
    for u in forms:
        z = (u['id_number'] or '').lstrip('0')
        if not z:
            continue
        pol = conn.execute(
            "SELECT MAX(COALESCE(pd.received_at, pr.doc_date)) AS prd, MAX(pr.policy_number) AS pn "
            "FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
            "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=?", (z,)).fetchone()
        prd = pol['prd'] if pol else None
        if prd and (not u['received_at'] or prd[:10] >= (u['received_at'] or '')[:10]):
            conn.execute("UPDATE unmatched_submissions SET status='טופל', handled_at=? WHERE id=?",
                         (now, u['uid']))
            cust = conn.execute("SELECT id FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? "
                                "ORDER BY id DESC LIMIT 1", (z,)).fetchone()
            if cust:
                log_event(conn, event_key(u['id_number'], 'cust-%d' % cust['id']),
                          f"טופס טופל — הופקה פוליסה {pol['pn'] or ''} ({u['subject'] or ''})",
                          'system', kind='form_linked')
            resolved.append({'id_number': z, 'policy': pol['pn'], 'subject': u['subject']})
    conn.commit()
    conn.close()
    return jsonify({'resolved': len(resolved), 'items': resolved})

@app.route('/api/dedupe-forms', methods=['POST'])
def api_dedupe_forms():
    """One-time dedupe: any website-form submission whose ת"ז is already an active-month customer
    is a duplicate (the person is tracked as a customer / lead) → mark it 'טופל' so it drops off
    /admin/other-forms and lives only in the customer list. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'resolved': 0})
    conn = get_db()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    rows = conn.execute(
        "SELECT u.id AS uid, u.id_number, u.subject, c.id AS cid FROM unmatched_submissions u "
        "JOIN customers c ON c.month_id=? AND "
        "ltrim(COALESCE(c.id_number,''),'0')=ltrim(COALESCE(u.id_number,''),'0') "
        "WHERE u.status IN ('ממתין','בטיפול') AND COALESCE(u.id_number,'')!=''",
        (month['id'],)).fetchall()
    n = 0
    for r in rows:
        conn.execute("UPDATE unmatched_submissions SET status='טופל', handled_at=? WHERE id=?",
                     (now, r['uid']))
        # Document in the customer file so nothing is lost when it leaves the forms list.
        log_event(conn, event_key(r['id_number'], 'cust-%d' % r['cid']),
                  f"טופס מהאתר קושר לתיק הלקוח: {r['subject'] or ''}", 'system', kind='form_linked')
        n += 1
    conn.commit()
    conn.close()
    return jsonify({'resolved': n})

@app.route('/api/brand-mismatches')
def api_brand_mismatches():
    """Active-month customers whose brand differs between the monthly row and the insureds master,
    with the brand implied by the latest policy's agent number (the source of truth). Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'mismatches': 0, 'items': []})
    conn = get_db()
    rows = conn.execute(
        "SELECT c.id, c.name, c.id_number, c.brand AS cust_brand, i.brand AS master_brand "
        "FROM customers c JOIN insureds i "
        "ON ltrim(COALESCE(i.id_number,''),'0')=ltrim(COALESCE(c.id_number,''),'0') "
        "WHERE c.month_id=? AND COALESCE(c.brand,'')!='' AND COALESCE(i.brand,'')!='' "
        "AND c.brand!=i.brand ORDER BY c.name", (month['id'],)).fetchall()
    out = []
    for r in rows:
        z = (r['id_number'] or '').lstrip('0')
        ag = conn.execute("SELECT pr.agent_number FROM policy_records pr "
                          "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? AND COALESCE(pr.agent_number,'')!='' "
                          "ORDER BY pr.id DESC LIMIT 1", (z,)).fetchone()
        pol_brand = NEW_AGENT_BRAND.get(re.sub(r'\D', '', str(ag['agent_number'])) if ag else '', '')
        out.append({'name': r['name'], 'id_number': z, 'customer_row': r['cust_brand'],
                    'master': r['master_brand'], 'policy_truth': pol_brand})
    conn.close()
    return jsonify({'mismatches': len(out), 'items': out})

@app.route('/api/fix-brand-mismatches', methods=['POST'])
def api_fix_brand_mismatches():
    """Align brand (customer rows + insureds master, by ת"ז) to the latest policy's agent number
    for active-month mismatches. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'fixed': 0, 'items': []})
    conn = get_db()
    rows = conn.execute(
        "SELECT c.id, c.name, c.id_number, c.brand AS cust_brand FROM customers c JOIN insureds i "
        "ON ltrim(COALESCE(i.id_number,''),'0')=ltrim(COALESCE(c.id_number,''),'0') "
        "WHERE c.month_id=? AND COALESCE(c.brand,'')!='' AND COALESCE(i.brand,'')!='' "
        "AND c.brand!=i.brand ORDER BY c.name", (month['id'],)).fetchall()
    fixed = []
    for r in rows:
        z = (r['id_number'] or '').lstrip('0')
        if not z:
            continue
        ag = conn.execute("SELECT pr.agent_number FROM policy_records pr "
                          "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? AND COALESCE(pr.agent_number,'')!='' "
                          "ORDER BY pr.id DESC LIMIT 1", (z,)).fetchone()
        pol_brand = NEW_AGENT_BRAND.get(re.sub(r'\D', '', str(ag['agent_number'])) if ag else '', '')
        if not pol_brand or pol_brand == r['cust_brand']:
            continue
        conn.execute("UPDATE customers SET brand=? WHERE ltrim(COALESCE(id_number,''),'0')=?", (pol_brand, z))
        conn.execute("UPDATE insureds SET brand=? WHERE ltrim(COALESCE(id_number,''),'0')=?", (pol_brand, z))
        log_event(conn, event_key(r['id_number'], 'cust-%d' % r['id']),
                  f"מותג יושר לפי הפוליסה: {r['cust_brand']} → {pol_brand}", 'system', kind='brand_fix')
        fixed.append({'name': r['name'], 'from': r['cust_brand'], 'to': pol_brand})
    conn.commit()
    conn.close()
    return jsonify({'fixed': len(fixed), 'items': fixed})

@app.route('/api/duplicates')
def api_duplicates():
    """Customers sharing the same ת"ז in the active month (data-integrity check). Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'duplicate_ids': 0, 'items': []})
    conn = get_db()
    dups = conn.execute(
        "SELECT ltrim(COALESCE(id_number,''),'0') AS z, COUNT(*) AS n FROM customers "
        "WHERE month_id=? AND COALESCE(id_number,'')!='' GROUP BY z HAVING n>1 ORDER BY n DESC",
        (month['id'],)).fetchall()
    out = []
    for d in dups:
        rows = conn.execute(
            "SELECT id, name, id_number, brand, status, import_source, policy_number, email_sent_date "
            "FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=? ORDER BY id",
            (month['id'], d['z'])).fetchall()
        out.append({'id_number': d['z'], 'count': d['n'], 'rows': [dict(r) for r in rows]})
    conn.close()
    return jsonify({'duplicate_ids': len(out), 'items': out})

@app.route('/api/queue-monitor')
def api_queue_monitor():
    """Work-queue watchdog: run a fresh renewal-form scan and report how many are now in the
    queue + how many renewal requests couldn't be matched (a gap that needs attention).
    The local sender polls this on a schedule and WhatsApps Sharon. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    try:
        scanned, unmatched = check_renewal_forms(days_back=3)
    except Exception as e:
        return jsonify({'error': str(e)})
    month = active_month()
    conn = get_db()
    in_queue = conn.execute(
        "SELECT COUNT(*) FROM customers WHERE month_id=? AND status='טופס התקבל'",
        (month['id'],)).fetchone()[0] if month else 0
    conn.close()
    return jsonify({'month': month['name'] if month else None,
                    'scanned_now': scanned, 'unmatched': unmatched, 'in_queue': in_queue})

@app.route('/api/scan-email-replies')
def api_scan_email_replies():
    """Customers who REPLIED to the campaign email (their address appears as an inbox sender)
    → mark 'נוצר קשר עם לקוח'. Default is a dry run; apply=1 writes. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int(request.args.get('days', 2))
    apply = request.args.get('apply') == '1'
    month = active_month()
    if not month:
        return jsonify({'error': 'no active month'})
    conn = get_db()
    recip = {(r['email'] or '').strip().lower(): dict(r) for r in conn.execute(
        "SELECT id, name, id_number, email, status FROM customers "
        "WHERE month_id=? AND COALESCE(email,'') LIKE '%@%'", (month['id'],)).fetchall()}
    senders = set()
    cfg = EMAIL_CONFIG
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        status, data = mail.search(None, f'SINCE {since}')
        ids = data[0].split() if status == 'OK' else []
        for i in range(0, len(ids), 50):
            _, fetched = mail.fetch(b','.join(ids[i:i + 50]), '(BODY.PEEK[HEADER.FIELDS (FROM)])')
            for part in fetched:
                if isinstance(part, tuple):
                    m = re.search(r'[\w.+-]+@[\w.-]+\.[\w.]+', part[1].decode('utf-8', 'replace'))
                    if m:
                        senders.add(m.group(0).lower())
        mail.logout()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)})
    KEEP = ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'ממתין להפקה', 'טופס התקבל', 'לא רוצים לחדש', 'לא מחדש', 'בוטל')
    results = []
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    for em in sorted(senders & set(recip.keys())):
        r = recip[em]
        results.append({'email': em, 'name': r['name'], 'current_status': r['status']})
        if apply and (r['status'] or '') not in KEEP:
            conn.execute("UPDATE customers SET status='נוצר קשר עם לקוח', status_changed_at=? WHERE id=?",
                         (now, r['id']))
            log_event(conn, event_key(r['id_number'], 'cust-%d' % r['id']),
                      "נוצר קשר — הלקוח השיב למייל החידוש", 'system', kind='status')
    if apply:
        conn.commit()
    conn.close()
    return jsonify({'inbox_senders': len(senders), 'matched': len(results),
                    'applied': apply, 'results': results})

@app.route('/api/scan-bounces')
def api_scan_bounces():
    """Read Delivery-Status-Notification (Failure) bounce emails, extract the failed recipient
    addresses, and (with apply=1) clear that email from the customer + insured records so we
    never email a dead address again. Token-authed. Default is a dry run."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int(request.args.get('days', 3))
    apply = request.args.get('apply') == '1'
    cfg = EMAIL_CONFIG
    bounced = set()
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        status, data = mail.search(None, f'(SUBJECT "Delivery Status Notification" SINCE {since})')
        for mid in (data[0].split() if status == 'OK' else []):
            _, fd = mail.fetch(mid, '(BODY.PEEK[])')
            raw = fd[0][1].decode('utf-8', 'replace')
            for m in re.finditer(r'Final-Recipient:\s*rfc822;\s*<?([^\s<>]+@[^\s<>]+?)>?\s', raw, re.I):
                bounced.add(m.group(1).strip().lower().strip('.'))
        mail.logout()
    except Exception as e:
        return jsonify({'error': str(e)})
    conn = get_db()
    results = []
    for em in sorted(bounced):
        rows = conn.execute(
            "SELECT c.id, c.name, m.name AS month FROM customers c JOIN months m ON m.id=c.month_id "
            "WHERE lower(COALESCE(c.email,''))=?", (em,)).fetchall()
        results.append({'email': em, 'customers': [f"{r['name']} ({r['month']})" for r in rows]})
        if apply and rows:
            for r in rows:
                idn = conn.execute("SELECT id_number FROM customers WHERE id=?", (r['id'],)).fetchone()
                if idn:
                    log_event(conn, event_key(idn['id_number'], 'cust-%d' % r['id']),
                              f"מייל הוסר — נדחה (bounce): {em}", 'system', kind='email_bounce')
            conn.execute("UPDATE customers SET email='' WHERE lower(COALESCE(email,''))=?", (em,))
            conn.execute("UPDATE insureds SET email='' WHERE lower(COALESCE(email,''))=?", (em,))
    if apply:
        conn.commit()
    conn.close()
    return jsonify({'bounced_count': len(bounced), 'applied': apply,
                    'matched_customers': sum(len(r['customers']) for r in results), 'results': results})

@app.route('/api/inbox-forms')
def api_inbox_forms():
    """Diagnostic: recent website-form emails (onboarding@resend.dev) in the inbox — subjects +
    dates, and whether a given ת"ז appears in the body. Shows if renewal-request forms arrive
    and how they're titled. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int(request.args.get('days', 7))
    want_id = re.sub(r'\D', '', request.args.get('id', '')).lstrip('0')
    cfg = EMAIL_CONFIG
    out = []
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        status, data = mail.search(None, f'FROM "{JOIN_FORM_SENDER}" SINCE {since}')
        for mid in (data[0].split() if status == 'OK' else []):
            _, hd = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hd[0][1])
            item = {'subject': decode_str(hdr.get('Subject', '')), 'date': hdr.get('Date', '')}
            if want_id:
                _, fd = mail.fetch(mid, '(BODY.PEEK[])')
                msg = email_lib.message_from_bytes(fd[0][1])
                body = ''
                for part in msg.walk():
                    if part.get_content_type() == 'text/html':
                        try:
                            body = part.get_content()
                        except Exception:
                            pl = part.get_payload(decode=True); body = pl.decode('utf-8', 'replace') if pl else ''
                        break
                digits = re.sub(r'\D', '', body)
                item['has_id'] = (want_id in digits) or (want_id.zfill(9) in digits)
            out.append(item)
        mail.logout()
    except Exception as e:
        return jsonify({'error': str(e)})
    return jsonify({'count': len(out), 'items': out})

@app.route('/api/customer-lookup')
def api_customer_lookup():
    """Diagnostic: every customer row for a ת"ז across months + the insured master, so we can
    see why someone does/doesn't appear in a month or the campaign. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    q = re.sub(r'\D', '', request.args.get('q', '')).lstrip('0')
    if not q:
        return jsonify({'error': 'need q'}), 400
    conn = get_db()
    rows = conn.execute(
        "SELECT c.id, c.month_id, m.name AS month, m.is_active, c.name, c.id_number, c.brand, c.status, "
        "c.email, c.phone, c.is_midwife, c.is_vip, c.email_sent_date, c.import_source, c.policy_number, "
        "c.form_received_at "
        "FROM customers c LEFT JOIN months m ON m.id=c.month_id "
        "WHERE ltrim(COALESCE(c.id_number,''),'0')=? ORDER BY c.month_id DESC", (q,)).fetchall()
    # ALL insured master rows (fetchall) — a duplicate master (same ת"ז twice, or a variant id format)
    # is exactly what makes a person show twice in the global search.
    ins = conn.execute("SELECT id, id_number, name, brand, status, email, phone FROM insureds "
                       "WHERE ltrim(COALESCE(id_number,''),'0')=?", (q,)).fetchall()
    unm = conn.execute("SELECT id, name, id_number, status, subject, received_at "
                       "FROM unmatched_submissions WHERE ltrim(COALESCE(id_number,''),'0')=?",
                       (q,)).fetchall()
    conn.close()
    return jsonify({'customers': [dict(r) for r in rows],
                    'insureds': [dict(r) for r in ins], 'insured_count': len(ins),
                    'unmatched_submissions': [dict(r) for r in unm]})

def _policy_facing_status(cust_status, ins_status):
    """Customer-facing policy status for the bot's personalization answers."""
    if (ins_status or '') == 'בוטל':
        return 'בוטלה'
    if (ins_status or '') == 'לא פעיל':
        return 'לא פעילה'
    cs = cust_status or ''
    if cs in ('חודש', 'חודש - בוצעה שיחת מכירה'):
        return 'חודשה'
    if cs == 'הופק':
        return 'הופקה'
    if cs == 'ממתין להפקה':
        return 'בתהליך הפקה'
    if cs and cs not in ('לא רוצים לחדש', 'לא מחדש'):
        return 'בתהליך חידוש'
    return 'פעילה'

@app.route('/api/reminded-but-renewed')
def api_reminded_but_renewed():
    """Token-authed: active-month customers who got the 25th reminder (lr25_sent_at set) BUT already
    have a DELIVERED renewal policy this cycle — i.e. they'd renewed and shouldn't have been
    reminded (status wasn't flipped). Returns them for the apology + status fix."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    cycle = (datetime.datetime.now() - datetime.timedelta(days=45)).strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    rows = conn.execute(
        """SELECT c.id, c.name, c.id_number, c.brand, c.phone, c.status
           FROM customers c
           WHERE c.month_id=? AND COALESCE(c.lr25_sent_at,'')!=''
             AND EXISTS (SELECT 1 FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id
                         WHERE ltrim(COALESCE(pr.insured_id,''),'0')=ltrim(COALESCE(c.id_number,''),'0')
                           AND pr.doc_type_label LIKE '%חידוש%' AND pd.received_at >= ?
                           AND (COALESCE(pd.whatsapp_sent_at,'')!='' OR COALESCE(pd.email_sent_at,'')!=''))
           ORDER BY c.brand, c.name""",
        (month['id'], cycle)).fetchall()
    conn.close()
    return jsonify({'count': len(rows), 'items': [dict(r) for r in rows]})

def _free_old_pdfs(days=7):
    """Delete policy/attachment PDFs older than `days` from the /data volume (they're delivered +
    backed up on OneDrive). NEVER deletes a PDF whose policy is still PENDING delivery (that would
    strand it with a 404). Best-effort protection: on a DB failure (the full-disk emergency) it skips
    protection and deletes freely. Returns (deleted_count, freed_bytes)."""
    cutoff = time.time() - days * 86400
    # Protect undelivered policies' PDFs (neither channel sent yet).
    protected = set()
    try:
        _c = get_db()
        for r in _c.execute("SELECT filepath FROM policy_documents "
                            "WHERE COALESCE(whatsapp_sent_at,'')='' AND COALESCE(email_sent_at,'')='' "
                            "AND COALESCE(filepath,'')!=''").fetchall():
            try:
                protected.add(os.path.normcase(os.path.abspath(r['filepath'])))
            except Exception:
                pass
        _c.close()
    except Exception:
        protected = set()
    freed = 0
    n = 0
    for base_dir in (POLICY_DOCS_DIR, ATTACHMENTS_DIR):
        if not base_dir or not os.path.isdir(base_dir):
            continue
        for root, _dirs, files in os.walk(base_dir):
            for f in files:
                if not f.lower().endswith('.pdf'):
                    continue
                fp = os.path.join(root, f)
                try:
                    if os.path.normcase(os.path.abspath(fp)) in protected:
                        continue
                    if os.path.getmtime(fp) < cutoff:
                        sz = os.path.getsize(fp)
                        os.remove(fp)
                        freed += sz
                        n += 1
                except Exception:
                    pass
    return n, freed

@app.route('/api/admin/free-disk', methods=['POST'])
def api_free_disk():
    """EMERGENCY: free space on the /data volume by deleting policy PDFs older than N days.
    Filesystem-only — works even when the DB is failing with a full-disk I/O error. Body {days:int=2}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int((request.get_json(silent=True) or {}).get('days', 2))
    n, freed = _free_old_pdfs(days)
    return jsonify({'ok': True, 'deleted': n, 'freed_mb': round(freed / 1024 / 1024, 1), 'days': days})

@app.route('/api/policy-lookup')
def api_policy_lookup():
    """Token-authed personalization lookup for the bot: a customer's OWN policy by ת"ז — returned
    ONLY if the supplied phone matches a phone on file for that ת"ז (so a customer can pull only
    their own policy, never a stranger's). Any mismatch/miss → {found:false}. Never logs the ת"ז."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    idn = re.sub(r'\D', '', request.args.get('id_number', '')).lstrip('0')
    phone_in = re.sub(r'\D', '', request.args.get('phone', ''))
    if not idn or not phone_in:
        return jsonify({'found': False})
    def last9(p):
        return re.sub(r'\D', '', str(p or ''))[-9:]
    conn = get_db()
    ins = conn.execute("SELECT name, brand, phone, status, period_start, period_end, is_midwife "
                       "FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?", (idn,)).fetchone()
    cust = conn.execute(
        "SELECT c.name, c.brand, c.phone, c.status, c.occupation, c.premium_last_year, c.is_midwife "
        "FROM customers c JOIN months m ON m.id=c.month_id "
        "WHERE ltrim(COALESCE(c.id_number,''),'0')=? ORDER BY m.is_active DESC, c.id DESC LIMIT 1", (idn,)).fetchone()
    if not (ins or cust):
        conn.close(); return jsonify({'found': False})
    # SECURITY: the phone must match a phone on file for this ת"ז (compare last 9 digits).
    known = {last9(x) for x in [(ins['phone'] if ins else ''), (cust['phone'] if cust else '')] if x}
    if not last9(phone_in) or last9(phone_in) not in known:
        conn.close(); return jsonify({'found': False})
    pr = conn.execute(
        "SELECT pr.period_start, pr.period_end, pr.premium, pr.policy_document_id "
        "FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
        "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? ORDER BY pd.received_at DESC, pr.id DESC LIMIT 1", (idn,)).fetchone()
    name = (cust['name'] if cust else '') or (ins['name'] if ins else '')
    brand = (cust['brand'] if cust else '') or (ins['brand'] if ins else '')
    period_start = (pr['period_start'] if pr else None) or (ins['period_start'] if ins else None)
    period_end = (pr['period_end'] if pr else None) or (ins['period_end'] if ins else None)
    occ = ((cust['occupation'] if cust else '') or '').strip()
    if not occ and pr and pr['policy_document_id']:
        d = conn.execute("SELECT filepath FROM policy_documents WHERE id=?", (pr['policy_document_id'],)).fetchone()
        if d and d['filepath'] and os.path.exists(d['filepath']):
            occ = extract_insured_occupation(d['filepath']) or ''
    professions = [p.strip() for p in re.split(r'[,،/]', occ) if p.strip()]
    prem = (pr['premium'] if pr else None) or (str(cust['premium_last_year']) if cust and cust['premium_last_year'] else None)
    if prem:
        pn = re.sub(r'[^\d.]', '', str(prem))
        prem = (f"{int(float(pn)):,} ₪" if pn else None)
    status = _policy_facing_status(cust['status'] if cust else None, ins['status'] if ins else None)
    # Midwife flag — a midwife (מיילדת) renews on a dedicated link at a different price, so the bot
    # must answer with the midwife-specific renewal link/amount, never the standard one.
    is_mid = bool((cust['is_midwife'] if cust else 0) or (ins['is_midwife'] if ins else 0))
    r_amt = renewal_amount(is_mid, (cust['premium_last_year'] if cust else None))
    renewal = {'link': renewal_link(brand, is_mid)[0],
               'price': (f"{r_amt:,} ₪" if r_amt else None)}
    conn.close()
    return jsonify({'found': True, 'name': name, 'brand': brand, 'is_midwife': is_mid,
                    'policy': {'status': status, 'period_start': period_start, 'period_end': period_end,
                               'professions': professions, 'premium': prem},
                    'renewal': renewal})

@app.route('/api/policy-pdf')
def api_policy_pdf_lookup():
    """Token-authed: the customer's OWN latest policy PDF by ת"ז — phone-gated (last-9 match), so a
    customer gets only their own policy. Serves the newest deliverable (חדש/חידוש) file still on the
    server; 404 if none (older archived copies live only in OneDrive, not reachable here). No ת"ז log."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    idn = re.sub(r'\D', '', request.args.get('id_number', '')).lstrip('0')
    phone_in = re.sub(r'\D', '', request.args.get('phone', ''))
    if not idn or not phone_in:
        return jsonify({'found': False}), 404
    def last9(p):
        return re.sub(r'\D', '', str(p or ''))[-9:]
    conn = get_db()
    known = set()
    for r in conn.execute("SELECT phone FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?", (idn,)).fetchall():
        if r['phone']:
            known.add(last9(r['phone']))
    for r in conn.execute("SELECT phone FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=?", (idn,)).fetchall():
        if r['phone']:
            known.add(last9(r['phone']))
    if not last9(phone_in) or last9(phone_in) not in known:
        conn.close(); return jsonify({'found': False}), 404
    rows = conn.execute(
        "SELECT pd.filename, pd.filepath FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
        "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? "
        "AND (pr.doc_type_label LIKE '%חדש%' OR pr.doc_type_label LIKE '%חידוש%') "
        "ORDER BY pd.received_at DESC, pr.id DESC", (idn,)).fetchall()
    conn.close()
    for r in rows:
        fp = r['filepath']
        if fp and os.path.exists(fp):
            nm = re.sub(r'[\r\n]+', ' ', (r['filename'] or 'policy.pdf')).strip() or 'policy.pdf'
            return send_file(fp, as_attachment=True, download_name=nm)
    return jsonify({'found': False, 'reason': 'no server-side file'}), 404

def _make_dummy_pdf(title):
    """Build a minimal valid single-page PDF (pure Python, no library — fitz isn't on the server)."""
    t = re.sub(r'[()\\]', ' ', str(title))[:90].encode('latin-1', 'replace')
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents 4 0 R "
        b"/Resources << /Font << /F1 5 0 R >> >> >>",
        None,  # content stream, filled below
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    stream = b"BT /F1 16 Tf 50 780 Td (" + t + b") Tj ET"
    objs[3] = b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"
    out = b"%PDF-1.4\n"
    offsets = []
    for i, o in enumerate(objs, start=1):
        offsets.append(len(out))
        out += str(i).encode() + b" 0 obj\n" + o + b"\nendobj\n"
    xref_pos = len(out)
    out += b"xref\n0 " + str(len(objs) + 1).encode() + b"\n0000000000 65535 f \n"
    for off in offsets:
        out += ("%010d 00000 n \n" % off).encode()
    out += (b"trailer\n<< /Size " + str(len(objs) + 1).encode() + b" /Root 1 0 R >>\nstartxref\n"
            + str(xref_pos).encode() + b"\n%%EOF")
    return out

@app.route('/api/test/policy', methods=['POST'])
def api_test_policy():
    """Token-authed TEST helper: attach a dummy policy (PDF + record) to a ת"ז so policy-lookup /
    policy-pdf return real data — for testing. Body {id_number, name?, brand?, policy_number?,
    period_start, period_end, premium?} (dates DD/MM/YYYY)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    idn = re.sub(r'\D', '', str(d.get('id_number') or ''))
    if not idn:
        return jsonify({'error': 'need id_number'}), 400
    name = (d.get('name') or 'טסט').strip()
    pol = (d.get('policy_number') or ('TEST' + idn[-6:])).strip()
    ps = (d.get('period_start') or '').strip()
    pe = (d.get('period_end') or '').strip()
    prem = (d.get('premium') or '750').strip()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    try:
        test_dir = os.path.join(POLICY_DOCS_DIR, 'test')
        os.makedirs(test_dir, exist_ok=True)
        fp = os.path.join(test_dir, f'test_{idn}_{re.sub(r"[^A-Za-z0-9]", "", pol)}.pdf')
        with open(fp, 'wb') as f:
            f.write(_make_dummy_pdf(f"TEST POLICY  ID {idn}  Policy {pol}  {ps} - {pe}"))
    except Exception as e:
        return jsonify({'error': f'pdf create failed: {e}'}), 500
    conn = get_db()
    did = conn.execute("INSERT INTO policy_documents (filename, filepath, received_at, policy_number) VALUES (?,?,?,?)",
                       (f'פוליסת טסט {pol}.pdf', fp, now, pol)).lastrowid
    conn.execute("INSERT INTO policy_records (policy_document_id, policy_number, doc_type_label, insured_name, "
                 "insured_id, period_start, period_end, premium, extracted_at) VALUES (?,?,?,?,?,?,?,?,?)",
                 (did, pol, 'חידוש', name, idn, ps, pe, prem, now))
    # Optionally set up the serviceable records the bot needs — an insured master (for the phone-match
    # security check in /api/policy-lookup) + a dashboard customer row (status 'הופק', manual). Pass
    # phone (+ brand) to enable, mirroring a real test customer.
    brand = (d.get('brand') or '').strip()
    phone = re.sub(r'\D', '', str(d.get('phone') or ''))
    email = (d.get('email') or '').strip()
    made = {}
    if phone:
        iso = datetime.datetime.now().isoformat()
        ins = conn.execute("SELECT id FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?",
                           (idn.lstrip('0'),)).fetchone()
        if ins:
            conn.execute("UPDATE insureds SET name=COALESCE(NULLIF(name,''),?), phone=?, "
                         "email=COALESCE(NULLIF(email,''),?), brand=COALESCE(NULLIF(brand,''),?), "
                         "status='פעיל', updated_at=? WHERE id=?", (name, phone, email, brand, iso, ins['id']))
            made['insured'] = 'updated'
        else:
            conn.execute("INSERT INTO insureds (id_number, name, brand, phone, email, status, created_at, updated_at) "
                         "VALUES (?,?,?,?,?,?,?,?)", (idn, name, brand, phone, email, 'פעיל', iso, iso))
            made['insured'] = 'created'
        month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
        if month and brand in ('גאיה', 'ווינר', 'אופיר'):
            cust = conn.execute("SELECT id FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=?",
                                (month['id'], idn.lstrip('0'))).fetchone()
            if cust:
                made['customer'] = f"exists (id {cust['id']})"
            else:
                cid = conn.execute(
                    "INSERT INTO customers (month_id, name, id_number, phone, email, brand, status, "
                    "import_source, policy_number, status_changed_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (month['id'], name, idn, phone, email, brand, 'הופק', 'manual', pol, now)).lastrowid
                made['customer'] = f"created (id {cid})"
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'doc_id': did, 'policy_number': pol, 'insured_id': idn,
                    'period_start': ps, 'period_end': pe, 'records': made})

@app.route('/api/renewal-status')
def api_renewal_status():
    """Token-authed: does the phone (an inbound WhatsApp sender — its own verified identity) have an
    OPEN renewal this cycle? is_due=true only for an active-month Gaia/Winner renewal that isn't new
    business/group-owner and isn't already settled (renewed/declined/issued). Returns name/period_end/
    brand when due, else {is_due:false}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    p = re.sub(r'\D', '', request.args.get('phone', ''))
    if len(p) < 9:
        return jsonify({'is_due': False})
    p9 = p[-9:]
    conn = get_db()
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'is_due': False})
    settled = ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'לא רוצים לחדש', 'לא מחדש', 'בוטל', 'ממתין להפקה')
    rows = conn.execute(
        "SELECT name, id_number, brand, status, import_source, group_owner, is_midwife, premium_last_year "
        "FROM customers "
        "WHERE month_id=? AND brand IN ('גאיה','ווינר') "
        "AND REPLACE(REPLACE(COALESCE(phone,''),'-',''),' ','') LIKE ?",
        (month['id'], '%' + p9)).fetchall()
    match = None
    for r in rows:
        if (r['import_source'] or '') in NEW_BUSINESS_SOURCES or (r['group_owner'] or '').strip():
            continue
        if (r['status'] or '') in settled:
            continue
        match = r; break
    if not match:
        conn.close(); return jsonify({'is_due': False})
    idn = re.sub(r'\D', '', match['id_number'] or '').lstrip('0')
    ins = conn.execute("SELECT period_end FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?", (idn,)).fetchone() if idn else None
    pe = (ins['period_end'] if ins and ins['period_end'] else None)
    if not pe:
        today = datetime.date.today()
        last = (datetime.date(today.year, 12, 31) if today.month == 12
                else datetime.date(today.year, today.month + 1, 1) - datetime.timedelta(days=1))
        pe = last.strftime('%d/%m/%Y')
    conn.close()
    # Midwife flag + the midwife-specific renewal link/price, so the bot answers a midwife correctly.
    is_mid = bool(match['is_midwife'])
    r_amt = renewal_amount(is_mid, match['premium_last_year'])
    return jsonify({'is_due': True, 'name': match['name'], 'period_end': pe, 'brand': match['brand'],
                    'is_midwife': is_mid,
                    'renewal': {'link': renewal_link(match['brand'], is_mid)[0],
                                'price': (f"{r_amt:,} ₪" if r_amt else None)}})

@app.route('/api/daily-report')
def api_daily_report():
    """Token-authed: the morning health-report text, computed from THIS (production) DB.
    Lets the local scheduled emailer reflect production instead of a stale local/dev copy."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    today = datetime.date.today().isoformat()
    month = conn.execute("SELECT * FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    month_name = month['name'] if month else '(אין חודש פעיל)'
    mid = month['id'] if month else -1
    unmatched = conn.execute(
        "SELECT name, id_number, received_at FROM unmatched_submissions "
        "WHERE status='pending' ORDER BY received_at DESC").fetchall()
    forms_pending = conn.execute(
        "SELECT name, form_received_at FROM customers WHERE month_id=? "
        "AND status='טופס התקבל' ORDER BY form_received_at DESC", (mid,)).fetchall()
    needs_clarify = conn.execute(
        "SELECT id FROM customers WHERE month_id=? AND status='דורש בירור'", (mid,)).fetchall()
    stats = conn.execute(
        "SELECT COUNT(*) AS total, "
        "SUM(CASE WHEN status='חודש' THEN 1 ELSE 0 END) AS renewed, "
        "SUM(CASE WHEN status='' OR status IS NULL THEN 1 ELSE 0 END) AS pending "
        "FROM customers WHERE month_id=?", (mid,)).fetchone()
    conn.close()

    lines = [f"דוח יומי — מערכת שירות לקוחות | {today}",
             f"חודש פעיל: {month_name}", ""]
    alerts = []
    if unmatched:
        alerts.append(f"⚠️  {len(unmatched)} טפסים בתור אדמין שממתינים לבירור")
    if forms_pending:
        alerts.append(f"📋 {len(forms_pending)} לקוחות עם טופס שהתקבל — ממתין לטיפול נציג")
    if needs_clarify:
        alerts.append(f"❓ {len(needs_clarify)} לקוחות סומנו 'דורש בירור'")
    if alerts:
        lines.append("== דרוש טיפול ==")
        lines.extend(alerts)
        lines.append("")
    if stats:
        total = stats['total'] or 0
        renewed = stats['renewed'] or 0
        pending = stats['pending'] or 0
        pct = round(renewed / total * 100) if total else 0
        lines += ["== סטטיסטיקות חודש ==",
                  f"סה\"כ לקוחות: {total}",
                  f"חידשו: {renewed} ({pct}%)",
                  f"ממתינים לטיפול: {pending}", ""]
    if unmatched:
        lines.append("== תור אדמין — פרטים ==")
        for u in unmatched:
            lines.append(f"  • {u['name'] or '(ללא שם)'} | ת.ז: {u['id_number'] or '-'} | {u['received_at']}")
        lines.append("")
    if forms_pending:
        lines.append("== טפסים שהתקבלו — ממתין לטיפול נציג ==")
        for c in forms_pending:
            lines.append(f"  • {c['name']} | {c['form_received_at']}")
        lines.append("")
    lines.append("כניסה למערכת: https://renewals-system-production.up.railway.app/queue")
    if not alerts:
        lines.insert(2, "✅ הכל תקין — אין פריטים ממתינים לטיפול")
        lines.insert(3, "")
    return jsonify({'report': "\n".join(lines), 'generated_at': today})

@app.route('/api/customer-by-name')
def api_customer_by_name():
    """Diagnostic: customers matching a name, with their pending-handling workflow fields
    (status, manual-send, end-reminder flag, call attempts). Token-authed. ?q=<name>."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'error': 'need q'}), 400
    like = f'%{q}%'
    cond, params = _name_search('c.name', q, like)
    conn = get_db()
    rows = conn.execute(
        f"SELECT c.id, m.name AS month, c.name, c.id_number, c.brand, c.status, "
        f"c.whatsapp_sent_date, c.end_reminder_sent_date, "
        f"c.call_status_1, c.call_date_1, c.call_status_2, c.call_date_2, c.call_status_3, c.call_date_3 "
        f"FROM customers c JOIN months m ON m.id=c.month_id WHERE {cond} "
        f"ORDER BY m.is_active DESC, c.id DESC LIMIT 20", params).fetchall()
    conn.close()
    return jsonify({'count': len(rows), 'items': [dict(r) for r in rows]})

@app.route('/api/end-reminder-marked')
def api_end_reminder_marked():
    """Diagnostic: active-month customers flagged 'תזכורת סיום' (end_reminder_sent_date set),
    newest mark first. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    rows = conn.execute(
        "SELECT name, id_number, brand, status, end_reminder_sent_date FROM customers "
        "WHERE month_id=? AND COALESCE(end_reminder_sent_date,'')!='' "
        "ORDER BY end_reminder_sent_date DESC, id DESC LIMIT 20", (month['id'],)).fetchall()
    conn.close()
    return jsonify({'count': len(rows), 'items': [dict(r) for r in rows]})

_HEB_DOW = ['שני', 'שלישי', 'רביעי', 'חמישי', 'שישי', 'שבת', 'ראשון']  # Python weekday(): Mon=0..Sun=6

def _month_end_hebrew(ref=None):
    """The coverage-end date string for the last-reminder template's {{1}} — the last day of the
    current month, formatted 'יום <weekday> DD/MM/YYYY' (e.g. 'יום שני 31/08/2026')."""
    d = ref or datetime.date.today()
    if d.month == 12:
        last = datetime.date(d.year, 12, 31)
    else:
        last = datetime.date(d.year, d.month + 1, 1) - datetime.timedelta(days=1)
    return f'יום {_HEB_DOW[last.weekday()]} {last.strftime("%d/%m/%Y")}'

# Statuses that mean a renewal is settled (renewed / declined) or is new business / in-process —
# used to build the end-of-month "still open" list (everyone NOT in one of these).
_EOM_SETTLED = ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'ממתין להפקה',
                'לא רוצים לחדש', 'לא מחדש', 'בוטל', 'טופס התקבל', 'הלקוח אישר')

@app.route('/api/end-reminder/queue')
def api_end_reminder_queue():
    """Token-authed: build the recipient list for a last-reminder send. ?kind=25 (default) = the
    rep-flagged 'תזכורת סיום' list; ?kind=eom = everyone still open (not renewed/declined/new/in-
    process). Both: Gaia/Winner only, exclude new-business SOURCES, group-owner, midwives, VIPs,
    the already-renewed/issued, and rows already sent this cycle + rows without a phone. The
    {{1}} template value (coverage-end date) is returned once as `end_date`."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    kind = request.args.get('kind', '25')
    sent_col = 'lreom_sent_at' if kind == 'eom' else 'lr25_sent_at'
    conn = get_db()
    month = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    nb = ','.join('?' * len(NEW_BUSINESS_SOURCES))
    common = (
        f" AND brand IN ('גאיה','ווינר') "
        f" AND COALESCE(import_source,'') NOT IN ({nb}) "
        f" AND COALESCE(group_owner,'')='' AND COALESCE(is_midwife,0)=0 AND COALESCE(is_vip,0)=0 "
        f" AND COALESCE(phone,'')!='' AND COALESCE({sent_col},'')='' ")
    if kind == 'eom':
        st = ','.join('?' * len(_EOM_SETTLED))
        where = f"month_id=? AND COALESCE(status,'') NOT IN ({st})" + common
        params = [month['id']] + list(_EOM_SETTLED) + list(NEW_BUSINESS_SOURCES)
    else:  # '25' — the flagged list, minus already renewed/issued
        where = ("month_id=? AND COALESCE(end_reminder_sent_date,'')!='' "
                 "AND COALESCE(status,'') NOT IN ('חודש','חודש - בוצעה שיחת מכירה','הופק','טופס התקבל')" + common)
        params = [month['id']] + list(NEW_BUSINESS_SOURCES)
    rows = conn.execute(
        f"SELECT id, name, id_number, brand, phone, status FROM customers WHERE {where} ORDER BY brand, name",
        params).fetchall()
    conn.close()
    return jsonify({'kind': kind, 'month': month['name'], 'end_date': _month_end_hebrew(),
                    'count': len(rows), 'items': [dict(r) for r in rows]})

@app.route('/api/end-reminder/sent', methods=['POST'])
def api_end_reminder_sent():
    """Token-authed: mark a last-reminder delivered so it isn't re-sent. Body {id, kind:'25'|'eom'}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    cid = d.get('id'); kind = d.get('kind', '25')
    if not cid:
        return jsonify({'error': 'need id'}), 400
    col = 'lreom_sent_at' if kind == 'eom' else 'lr25_sent_at'
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    conn.execute(f"UPDATE customers SET {col}=? WHERE id=?", (now, cid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

BRAND_PNID = {'גאיה': '1030477631130403', 'ווינר': '103910399338118', 'אופיר': '103910399338118'}
SEND_TEMPLATES = {'reminder_25': 'renewal_reminder', 'renewal_campaign': 'renewal_reminder',
                  'last_reminder_eom': 'renewal_last_reminder'}

def _send_queue_recipients(conn, month, typ):
    """Shared send-ready recipient builder (used by /api/send-queue AND batch creation). Targeting +
    exclusions live here — single source of truth. Returns (template, [items]); each item carries
    phone(972), pnid, template, lang, ordered body_params (+ name/id_number for the approval view)."""
    tpl = SEND_TEMPLATES.get(typ)
    nb = ','.join('?' * len(NEW_BUSINESS_SOURCES))
    common = (f" AND brand IN ('גאיה','ווינר') AND COALESCE(import_source,'') NOT IN ({nb}) "
              f" AND COALESCE(group_owner,'')='' AND COALESCE(is_midwife,0)=0 AND COALESCE(is_vip,0)=0 "
              f" AND COALESCE(phone,'')!='' ")
    if typ == 'last_reminder_eom':
        st = ','.join('?' * len(_EOM_SETTLED))
        where = (f"month_id=? AND COALESCE(status,'') NOT IN ({st}) AND COALESCE(lreom_sent_at,'')=''" + common)
        params = [month['id']] + list(_EOM_SETTLED) + list(NEW_BUSINESS_SOURCES)
    else:  # reminder_25
        where = ("month_id=? AND COALESCE(end_reminder_sent_date,'')!='' "
                 "AND COALESCE(status,'') NOT IN ('חודש','חודש - בוצעה שיחת מכירה','הופק','טופס התקבל') "
                 "AND COALESCE(lr25_sent_at,'')=''" + common)
        params = [month['id']] + list(NEW_BUSINESS_SOURCES)
    rows = conn.execute(
        f"SELECT id, name, id_number, brand, phone, status, is_midwife, premium_last_year "
        f"FROM customers WHERE {where} ORDER BY brand, name", params).fetchall()
    optout = {re.sub(r'\D', '', str(r['phone'] or '')).lstrip('0')
              for r in conn.execute("SELECT phone FROM optouts")}
    end_month = (month['name'] or '').split(' ')[0]
    end_date = _month_end_hebrew()
    items = []
    for r in rows:
        p972 = _policy_to972(r['phone'])
        if not p972 or re.sub(r'\D', '', str(r['phone'])).lstrip('0') in optout:
            continue
        if typ == 'last_reminder_eom':
            body = [end_date]
        else:
            amt = renewal_amount(r['is_midwife'], r['premium_last_year'])
            price = f"{amt:,} ₪" if amt else "750 ₪"
            url = renewal_link(r['brand'], r['is_midwife'])[0]
            body = [r['name'] or '', end_month, price, url]
        items.append({'id': r['id'], 'name': r['name'] or '', 'id_number': r['id_number'] or '',
                      'phone': p972, 'brand': r['brand'],
                      'pnid': BRAND_PNID.get(r['brand'], BRAND_PNID['גאיה']),
                      'template': tpl, 'lang': 'he', 'body_params': body,
                      # Exclusions (midwife/VIP/group-owner/new-business) are already enforced in the
                      # WHERE above, so this is always False here — exposed only so the bot can assert it.
                      'is_midwife': bool(r['is_midwife'])})
    return tpl, items

@app.route('/api/send-queue')
def api_send_queue():
    """Token-authed LIVE preview of send-ready recipients (targeting truth). ?type=reminder_25 |
    last_reminder_eom. This is the PREVIEW; a real blast goes through an approved batch (see
    /api/send-batch/*). Each recipient carries everything the bot needs to POST to Meta."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    typ = request.args.get('type', 'reminder_25')
    if typ not in ('reminder_25', 'last_reminder_eom'):
        return jsonify({'error': 'type not implemented yet', 'type': typ, 'ready': list(SEND_TEMPLATES)}), 501
    conn = get_db()
    month = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    tpl, items = _send_queue_recipients(conn, month, typ)
    conn.close()
    return jsonify({'type': typ, 'month': month['name'], 'template': tpl,
                    'count': len(items), 'recipients': items})

@app.route('/api/send-queue/sent', methods=['POST'])
def api_send_queue_sent():
    """Token-authed: the bot reports a send + its evolving status. Upsert by wamid into send_log
    (sent→delivered→read→failed+error_code) + mark the customer's per-type sent flag on first send."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    wamid = (d.get('wamid') or '').strip()
    if not wamid:
        return jsonify({'error': 'need wamid'}), 400
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    cid, typ, status, ec = d.get('id'), d.get('type'), (d.get('status') or 'sent'), d.get('error_code')
    conn = get_db()
    conn.execute(
        """INSERT INTO send_log (wamid, cust_id, send_type, status, error_code, sent_at, updated_at)
           VALUES (?,?,?,?,?,?,?)
           ON CONFLICT(wamid) DO UPDATE SET status=excluded.status,
             error_code=COALESCE(excluded.error_code, send_log.error_code), updated_at=excluded.updated_at""",
        (wamid, cid, typ, status, (str(ec) if ec is not None else None), now, now))
    # On the first 'sent', stamp the customer's per-type flag so the queue won't re-serve them.
    if cid and status == 'sent' and typ in ('reminder_25', 'last_reminder_eom'):
        col = 'lreom_sent_at' if typ == 'last_reminder_eom' else 'lr25_sent_at'
        conn.execute(f"UPDATE customers SET {col}=COALESCE(NULLIF({col},''),?) WHERE id=?", (now, cid))
        # Also mirror it into whatsapp_sent_date (the dashboard's "נשלח בוואטסאפ" indicator) + the
        # client timeline, so bot-reported reminder sends are VISIBLE in the dashboard like the local
        # tool's sends — otherwise a reminder only sets the internal flag and looks unsent.
        _cr = conn.execute("SELECT id_number, whatsapp_sent_date FROM customers WHERE id=?", (cid,)).fetchone()
        if _cr and not (_cr['whatsapp_sent_date'] or '').strip():
            conn.execute("UPDATE customers SET whatsapp_sent_date=? WHERE id=?",
                         (datetime.date.today().isoformat(), cid))
            try:
                log_event(conn, event_key(_cr['id_number'], f'sq-{cid}'),
                          "נשלחה תזכורת חידוש בוואטסאפ (בוט)", 'וואטסאפ אוטומטי', kind='whatsapp_sent')
            except Exception:
                pass
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/backfill-wa-visibility', methods=['POST'])
def api_backfill_wa_visibility():
    """One-time: make already-sent reminder WhatsApps visible in the dashboard — stamp
    whatsapp_sent_date from a reminder flag (lreom/lr25) for active-month customers who were messaged
    (bot) but not marked, and log the timeline. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    rows = conn.execute(
        "SELECT id, id_number, COALESCE(NULLIF(lreom_sent_at,''), lr25_sent_at) AS rdate "
        "FROM customers WHERE month_id=? AND COALESCE(whatsapp_sent_date,'')='' "
        "AND (COALESCE(lreom_sent_at,'')!='' OR COALESCE(lr25_sent_at,'')!='')", (month['id'],)).fetchall()
    n = 0
    for r in rows:
        d = (r['rdate'] or '')[:10] or datetime.date.today().isoformat()
        conn.execute("UPDATE customers SET whatsapp_sent_date=? WHERE id=?", (d, r['id']))
        try:
            log_event(conn, event_key(r['id_number'], f"bf-{r['id']}"),
                      "נשלחה תזכורת חידוש בוואטסאפ (בוט)", 'וואטסאפ אוטומטי', kind='whatsapp_sent')
        except Exception:
            pass
        n += 1
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'backfilled': n})

@app.route('/api/send-log')
def api_send_log():
    """Token-authed, read-only: a live view of the bot's reported sends (send_log) so we can watch
    the round-trip. Returns totals by status + the most recent N rows. ?type= filter, ?limit= (50)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    typ = request.args.get('type', '')
    limit = min(int(request.args.get('limit', 50)), 500)
    conn = get_db()
    where, params = '', []
    if typ:
        where = ' WHERE send_type=?'; params = [typ]
    by_status = {r['status']: r['n'] for r in conn.execute(
        f"SELECT COALESCE(status,'?') status, COUNT(*) n FROM send_log{where} GROUP BY status", params).fetchall()}
    total = sum(by_status.values())
    recent = [dict(r) for r in conn.execute(
        f"SELECT wamid, cust_id, send_type, status, error_code, sent_at, updated_at "
        f"FROM send_log{where} ORDER BY updated_at DESC, wamid DESC LIMIT ?", params + [limit]).fetchall()]
    for r in recent:
        w = r.get('wamid') or ''
        r['wamid'] = (w[:18] + '…') if len(w) > 18 else w
    conn.close()
    return jsonify({'total': total, 'by_status': by_status, 'recent': recent})

@app.route('/api/optout', methods=['POST'])
def api_optout():
    """Token-authed: the bot reports a marketing opt-out (unsubscribe / 'not now'). Stored in
    optouts and excluded from every /api/send-queue from now on. Body {phone, id_number?, reason?}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    phone = re.sub(r'\D', '', str(d.get('phone') or ''))
    idn = re.sub(r'\D', '', str(d.get('id_number') or ''))
    if not phone and not idn:
        return jsonify({'error': 'need phone or id_number'}), 400
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    conn.execute("INSERT INTO optouts (phone, id_number, reason, created_at) VALUES (?,?,?,?)",
                 (phone, idn, (d.get('reason') or '').strip(), now))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ── Approval batches: freeze a recipient snapshot → Sharon approves → the bot sends only approved ──
@app.route('/api/send-batch/create', methods=['POST'])
def api_send_batch_create():
    """Token-authed (bot, at the gate time): freeze the current send-queue for `type` into a
    PENDING batch + alert Sharon (via owner_alerts → wa-sender email). Refuses if a pending/approved
    batch of that type is already open. Body/query {type}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    typ = (request.get_json(silent=True) or {}).get('type') or request.args.get('type', '')
    if typ not in ('reminder_25', 'last_reminder_eom'):
        return jsonify({'error': 'type not implemented', 'ready': list(SEND_TEMPLATES)}), 501
    conn = get_db()
    month = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    dup = conn.execute("SELECT id, status FROM send_batches WHERE send_type=? AND status IN ('pending','approved')",
                       (typ,)).fetchone()
    if dup:
        conn.close(); return jsonify({'error': 'batch already open', 'batch_id': dup['id'], 'status': dup['status']}), 409
    tpl, items = _send_queue_recipients(conn, month, typ)
    if not items:
        conn.close(); return jsonify({'ok': True, 'count': 0, 'batch_id': None, 'note': 'no recipients'})
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    bid = conn.execute("INSERT INTO send_batches (send_type, template, status, count, created_at) VALUES (?,?,?,?,?)",
                       (typ, tpl, 'pending', len(items), now)).lastrowid
    for it in items:
        conn.execute("INSERT INTO send_batch_items (batch_id, cust_id, name, phone, brand, pnid, template, body_params) "
                     "VALUES (?,?,?,?,?,?,?,?)",
                     (bid, it['id'], it['name'], it['phone'], it['brand'], it['pnid'], it['template'],
                      json.dumps(it['body_params'], ensure_ascii=False)))
    _typ_he = {'reminder_25': 'תזכורת חידוש (25 לחודש)', 'last_reminder_eom': 'אי-חידוש (סוף חודש)'}.get(typ, typ)
    conn.execute("INSERT INTO owner_alerts (text, created_at) VALUES (?,?)",
                 (f"📋 מנת שליחה ממתינה לאישור: {len(items)} נמענים — {_typ_he}.\n"
                  f"לאישור או ביטול היכנס לעמוד האישור בדשבורד:\n"
                  f"https://renewals-system-production.up.railway.app/send-approvals", now))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'batch_id': bid, 'count': len(items), 'status': 'pending'})

@app.route('/api/send-batch/next')
def api_send_batch_next():
    """Token-authed (bot): the oldest APPROVED batch's frozen recipients to send now (or null).
    Optional ?type= filter. The bot sends each, reports via /api/send-queue/sent, then POSTs
    /api/send-batch/<id>/done."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    typ = request.args.get('type', '')
    conn = get_db()
    q = "SELECT * FROM send_batches WHERE status='approved'"
    p = []
    if typ:
        q += " AND send_type=?"; p = [typ]
    b = conn.execute(q + " ORDER BY id ASC LIMIT 1", p).fetchone()
    if not b:
        conn.close(); return jsonify({'batch': None})
    items = [{'id': r['cust_id'], 'name': r['name'], 'phone': r['phone'], 'brand': r['brand'],
              'pnid': r['pnid'], 'template': r['template'], 'lang': 'he',
              'body_params': json.loads(r['body_params'] or '[]')}
             for r in conn.execute("SELECT * FROM send_batch_items WHERE batch_id=?", (b['id'],)).fetchall()]
    conn.close()
    return jsonify({'batch_id': b['id'], 'type': b['send_type'], 'template': b['template'],
                    'count': len(items), 'recipients': items})

@app.route('/api/send-batch/<int:bid>/done', methods=['POST'])
def api_send_batch_done(bid):
    """Token-authed (bot): mark an approved batch fully sent."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    conn.execute("UPDATE send_batches SET status='sent' WHERE id=? AND status='approved'", (bid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/send-approvals')
@login_required
def send_approvals():
    """Sharon's approval page — pending/approved bulk-send batches with count + names + buttons."""
    conn = get_db()
    batches = conn.execute("SELECT * FROM send_batches WHERE status IN ('pending','approved') ORDER BY id DESC").fetchall()
    data = []
    for b in batches:
        items = conn.execute("SELECT name, brand, phone FROM send_batch_items WHERE batch_id=? ORDER BY brand, name",
                             (b['id'],)).fetchall()
        data.append({'b': b, 'items': items})
    today = datetime.date.today().strftime('%Y-%m-%d')
    holidays = conn.execute("SELECT date, label FROM no_send_dates WHERE date>=? ORDER BY date", (today,)).fetchall()
    conn.close()
    return render_template('send_approvals.html', batches=data, holidays=holidays)

@app.route('/send-approvals/holidays/add', methods=['POST'])
@login_required
def send_approvals_holiday_add():
    dt = (request.form.get('date') or '').strip()
    lbl = (request.form.get('label') or 'חג').strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}$', dt):
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO no_send_dates (date, label) VALUES (?,?)", (dt, lbl))
        conn.commit(); conn.close()
        flash(f'נוסף תאריך ללא-שליחה: {dt} ({lbl})', 'success')
    else:
        flash('תאריך לא תקין (YYYY-MM-DD)', 'danger')
    return redirect(url_for('send_approvals'))

@app.route('/send-approvals/holidays/remove', methods=['POST'])
@login_required
def send_approvals_holiday_remove():
    dt = (request.form.get('date') or '').strip()
    conn = get_db()
    conn.execute("DELETE FROM no_send_dates WHERE date=?", (dt,))
    conn.commit(); conn.close()
    flash(f'הוסר: {dt}', 'info')
    return redirect(url_for('send_approvals'))

@app.route('/api/send-window')
def api_send_window():
    """Token-authed: is it OK to send on a given date? Blocks Fri/Sat + any no_send_dates holiday.
    ?date=YYYY-MM-DD (default today). Returns ok, reason, and the next allowed send day."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    ds = (request.args.get('date') or '').strip()
    try:
        d = datetime.datetime.strptime(ds, '%Y-%m-%d').date() if ds else datetime.date.today()
    except ValueError:
        return jsonify({'error': 'bad date (YYYY-MM-DD)'}), 400
    conn = get_db()
    hol = {r['date']: r['label'] for r in conn.execute("SELECT date, label FROM no_send_dates")}
    conn.close()

    def blocked(dt):
        wd = dt.weekday()  # Mon=0 … Fri=4, Sat=5
        if wd == 4:
            return 'שישי'
        if wd == 5:
            return 'שבת'
        lbl = hol.get(dt.strftime('%Y-%m-%d'))
        return ('חג: ' + lbl) if lbl else None
    reason = blocked(d)
    nxt = d
    for _ in range(60):
        if not blocked(nxt):
            break
        nxt += datetime.timedelta(days=1)
    return jsonify({'date': d.strftime('%Y-%m-%d'), 'ok': reason is None,
                    'reason': reason or 'תקין', 'next_send_day': nxt.strftime('%Y-%m-%d')})

@app.route('/send-approvals/<int:bid>/approve', methods=['POST'])
@login_required
def send_approvals_approve(bid):
    conn = get_db()
    conn.execute("UPDATE send_batches SET status='approved', approved_at=?, approved_by=? WHERE id=? AND status='pending'",
                 (datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), session.get('username', '?'), bid))
    conn.commit(); conn.close()
    flash('המנה אושרה — הבוט ישלח אותה.', 'success')
    return redirect(url_for('send_approvals'))

@app.route('/send-approvals/<int:bid>/cancel', methods=['POST'])
@login_required
def send_approvals_cancel(bid):
    conn = get_db()
    conn.execute("UPDATE send_batches SET status='cancelled' WHERE id=?", (bid,))
    conn.commit(); conn.close()
    flash('המנה בוטלה.', 'info')
    return redirect(url_for('send_approvals'))

@app.route('/api/group-owner/set', methods=['POST'])
def api_group_owner_set():
    """Mark customers as belonging to a group owner (e.g. Aviram's therapists): sets group_owner,
    overrides the contact phone to the owner's phone, and notes the shared card last-4. Applies to
    EVERY month's customer row + the insured master, keyed by ת"ז. Token-authed.
    Body: {id_numbers:[...], owner, phone, card4}. Pass owner='' to CLEAR the marking."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(force=True, silent=True) or {}
    owner = (d.get('owner') or '').strip()
    phone = re.sub(r'\D', '', str(d.get('phone') or ''))
    card4 = re.sub(r'\D', '', str(d.get('card4') or ''))[-4:]
    ids = [re.sub(r'\D', '', str(x)).lstrip('0') for x in (d.get('id_numbers') or []) if str(x).strip()]
    if not ids:
        return jsonify({'error': 'need id_numbers'}), 400
    card_disp = ('****' + card4) if card4 else None
    conn = get_db()
    done = []
    for z in ids:
        cust = conn.execute(
            "UPDATE customers SET group_owner=?, phone=COALESCE(NULLIF(?,''),phone), "
            "form_card_number=COALESCE(?,form_card_number) WHERE ltrim(COALESCE(id_number,''),'0')=?",
            (owner or None, phone, card_disp, z)).rowcount
        ins = conn.execute(
            "UPDATE insureds SET group_owner=?, phone=COALESCE(NULLIF(?,''),phone) "
            "WHERE ltrim(COALESCE(id_number,''),'0')=?", (owner or None, phone, z)).rowcount
        done.append({'id_number': z, 'customers': cust, 'insureds': ins})
        if owner:
            log_event(conn, event_key(z, 'grp-%s' % z),
                      f"שויך לקבוצת {owner} (טלפון + אשראי של המרכז)", 'system', kind='group_owner')
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'owner': owner, 'phone': phone, 'card': card_disp, 'results': done})

def auto_mark_midwives(conn):
    """New-business rule (Sharon 2026-08-18): a NEW midwife is always Winner + occupation contains
    'מיילד' (premium ~1200). Auto-set is_midwife=1 for any such customer/insured not yet marked —
    renewals are already flagged in advance, so this only fills in the new ones. Returns counts."""
    c = conn.execute("UPDATE customers SET is_midwife=1 WHERE brand='ווינר' "
                     "AND occupation LIKE '%מיילד%' AND COALESCE(is_midwife,0)=0").rowcount
    i = conn.execute("UPDATE insureds SET is_midwife=1 WHERE brand='ווינר' "
                     "AND occupation LIKE '%מיילד%' AND COALESCE(is_midwife,0)=0").rowcount
    return {'customers': c, 'insureds': i}

@app.route('/api/mark-midwives', methods=['POST', 'GET'])
def api_mark_midwives():
    """Trigger the new-midwife auto-marking sweep. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    res = auto_mark_midwives(conn)
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'marked': res})

@app.route('/api/owner-response', methods=['POST'])
def api_owner_response():
    """A group owner (e.g. Aviram) tapped a renewal-confirm button for one of their therapists.
    Records the decision on the therapist's active-month record + logs it + queues a WhatsApp alert
    to Sharon (picked up by the wa-sender). Token-authed.
    Body: {id_number, decision:'approve'|'decline', therapist_name?}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(force=True, silent=True) or {}
    idn = re.sub(r'\D', '', str(d.get('id_number') or '')).lstrip('0')
    decision = (d.get('decision') or '').strip().lower()
    if not idn or decision not in ('approve', 'decline'):
        return jsonify({'error': 'need id_number + decision approve/decline'}), 400
    status = 'חודש' if decision == 'approve' else 'לא רוצים לחדש'
    conn = get_db()
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    row = conn.execute(
        "SELECT id, name, group_owner FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=?",
        (month['id'], idn)).fetchone() if month else None
    name = (row['name'] if row else '') or d.get('therapist_name') or idn
    owner = (row['group_owner'] if row else '') or 'המרכז'
    if row:
        conn.execute("UPDATE customers SET status=?, status_changed_at=? WHERE id=?",
                     (status, datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), row['id']))
        log_event(conn, event_key(idn, 'cust-%d' % row['id']),
                  f"{owner} {'אישר חידוש' if decision=='approve' else 'ביקש לא לחדש'} (תגובת כפתור)",
                  'system', kind='owner_response')
    verb = 'אישר חידוש ✅' if decision == 'approve' else 'לא מחדש ❌'
    alert = f"{owner} {verb} עבור {name}" + ('' if row else ' (לא נמצא בחודש הפעיל — לבדיקה)')
    conn.execute("INSERT INTO owner_alerts (text, created_at) VALUES (?,?)",
                 (alert, datetime.datetime.now().isoformat()))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'matched': bool(row), 'status': status, 'alert': alert})

@app.route('/api/owner-alerts')
def api_owner_alerts():
    """wa-sender pulls unsent owner-response alerts to WhatsApp Sharon. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    rows = conn.execute("SELECT id, text FROM owner_alerts WHERE sent_at IS NULL ORDER BY id LIMIT 20").fetchall()
    conn.close()
    return jsonify({'count': len(rows), 'items': [dict(r) for r in rows]})

@app.route('/api/owner-alerts/sent', methods=['POST'])
def api_owner_alerts_sent():
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    ids = (request.get_json(force=True, silent=True) or {}).get('ids') or []
    if not ids:
        return jsonify({'ok': True, 'marked': 0})
    conn = get_db()
    conn.executemany("UPDATE owner_alerts SET sent_at=? WHERE id=?",
                     [(datetime.datetime.now().isoformat(), int(i)) for i in ids])
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'marked': len(ids)})

@app.route('/api/campaign/wrong-sends')
def api_campaign_wrong_sends():
    """New-business/lead customers that received the renewal campaign email today (a mistake
    — they should have been excluded). Token-authed diagnostic."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'count': 0, 'items': []})
    conn = get_db()
    today = datetime.date.today().isoformat()
    rows = conn.execute(
        "SELECT name, id_number, brand, email, status, email_sent_date FROM customers "
        "WHERE month_id=? AND status IN ('הופק','ממתין להפקה') AND email_sent_date=? "
        "ORDER BY brand, name", (month['id'], today)).fetchall()
    conn.close()
    return jsonify({'count': len(rows), 'items': [dict(r) for r in rows]})

@app.route('/api/occ-debug')
def api_occ_debug():
    """Dump per-page text of one active-month stored PDF, so we can locate the occupation."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    conn = get_db()
    row = conn.execute(
        """SELECT c.id_number, pd.filepath FROM customers c
           JOIN policy_records pr ON ltrim(COALESCE(pr.insured_id,''),'0')=ltrim(COALESCE(c.id_number,''),'0')
           JOIN policy_documents pd ON pd.id=pr.policy_document_id
           WHERE c.month_id=? AND COALESCE(pd.filepath,'')!='' ORDER BY pd.id DESC LIMIT 1""",
        (month['id'],)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'no stored pdf for active month'})
    fp = row['filepath']
    out = {'id_number': row['id_number'], 'filepath': fp, 'exists': os.path.exists(fp),
           'extracted': extract_insured_occupation(fp), 'pages': []}
    try:
        with pdfplumber.open(fp) as pdf:
            out['page_count'] = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                t = page.extract_text() or ''
                bidi = get_display(t)
                hit = ('עיסוק' in bidi) or ('הסמכה' in bidi)
                out['pages'].append({
                    'page': i + 1,
                    'has_עיסוק_bidi': 'עיסוק' in bidi, 'has_הסמכה_bidi': 'הסמכה' in bidi,
                    'bidi': bidi if hit else bidi[:200]})
    except Exception as e:
        out['error'] = str(e)
    return jsonify(out)

# ── Routes ──────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            if (user['twofa_method'] or 'none') != 'none':
                # Password OK, but 2FA is enabled — hold login until a code is verified.
                session.clear()
                session['pre2fa_uid'] = user['id']
                return redirect(url_for('twofa_verify'))
            _finish_login(user)
            return redirect(url_for('index'))
        flash('שם משתמש או סיסמה שגויים', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── Two-factor authentication (TOTP app / email code · re-verify every 6h) ──
TWOFA_TTL = 6 * 3600  # seconds between required re-verifications

def _finish_login(user_row):
    """Complete a login once password (and any 2FA) passed."""
    session['user_id'] = user_row['id']
    session['username'] = user_row['username']
    session['display_name'] = user_row['display_name']
    session['role'] = user_row['role']
    session['verified_at'] = time.time()
    if user_row['role'] != 'superadmin':
        conn = get_db()
        brows = conn.execute("SELECT brand FROM user_brands WHERE user_id=?", (user_row['id'],)).fetchall()
        conn.close()
        session['brands'] = [b['brand'] for b in brows]

def _send_email_code(to_email, code):
    """Send a one-time login code via Resend. No-op (False) if unconfigured."""
    key = os.environ.get('RESEND_API_KEY', '')
    if not key or not to_email:
        return False
    try:
        import requests
        r = requests.post('https://api.resend.com/emails',
            headers={'Authorization': f'Bearer {key}', 'Content-Type': 'application/json'},
            json={'from': os.environ.get('RESEND_FROM', 'onboarding@resend.dev'),
                  'to': [to_email], 'subject': 'קוד אימות — מערכת חידושים',
                  'text': f'קוד האימות שלך: {code}\nהקוד תקף ל-10 דקות.'}, timeout=10)
        return r.status_code < 300
    except Exception as e:
        print(f'[2fa] email send failed: {e}')
        return False

@app.route('/2fa/verify', methods=['GET', 'POST'])
def twofa_verify():
    """Verify a 2FA code — used both at first login (pre2fa_uid) and for the 6-hour re-auth."""
    uid = session.get('pre2fa_uid') or session.get('user_id')
    if not uid:
        return redirect(url_for('login'))
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    if not u:
        session.clear()
        return redirect(url_for('login'))
    method = u['twofa_method'] or 'none'
    if method == 'none':
        if session.pop('pre2fa_uid', None):
            _finish_login(u)
        return redirect(url_for('index'))
    if request.method == 'GET' and method == 'email':
        code = f'{secrets.randbelow(1000000):06d}'
        session['email_code'] = generate_password_hash(code)
        session['email_code_exp'] = time.time() + 600
        if not _send_email_code(u['email'], code):
            flash('שליחת קוד למייל נכשלה — ודא ש-RESEND_API_KEY מוגדר ושיש מייל לנציג', 'warning')
    if request.method == 'POST':
        entered = re.sub(r'\D', '', request.form.get('code', ''))
        ok = False
        if method == 'app' and u['totp_secret']:
            ok = pyotp.TOTP(u['totp_secret']).verify(entered, valid_window=1)
        elif method == 'email':
            h, exp = session.get('email_code'), session.get('email_code_exp', 0)
            ok = bool(h and entered and time.time() < exp and check_password_hash(h, entered))
        if ok:
            session.pop('email_code', None)
            session.pop('email_code_exp', None)
            if session.pop('pre2fa_uid', None):
                _finish_login(u)
            else:
                session['verified_at'] = time.time()
            return redirect(url_for('index'))
        flash('קוד שגוי או שפג תוקפו', 'danger')
    return render_template('twofa_verify.html', method=method, email=(u['email'] or ''))

@app.route('/2fa/setup', methods=['GET', 'POST'])
@login_required
def twofa_setup():
    """Let a user enable/disable 2FA for their own account (TOTP app or email code)."""
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE id=?", (session['user_id'],)).fetchone()
    if request.method == 'POST':
        method = request.form.get('method')
        if method == 'app':
            secret = session.get('setup_secret')
            code = re.sub(r'\D', '', request.form.get('code', ''))
            if secret and pyotp.TOTP(secret).verify(code, valid_window=1):
                conn.execute("UPDATE users SET totp_secret=?, twofa_method='app' WHERE id=?", (secret, u['id']))
                conn.commit(); conn.close()
                session.pop('setup_secret', None)
                session['verified_at'] = time.time()
                flash('אימות דו-שלבי (אפליקציה) הופעל ✓', 'success')
                return redirect(url_for('index'))
            conn.close()
            flash('קוד שגוי — סרוק שוב ונסה', 'danger')
            return redirect(url_for('twofa_setup'))
        if method == 'email':
            if not u['email']:
                conn.close()
                flash('אין מייל מוגדר לנציג — הוסף מייל במסך הניהול קודם', 'warning')
                return redirect(url_for('twofa_setup'))
            conn.execute("UPDATE users SET twofa_method='email' WHERE id=?", (u['id'],))
            conn.commit(); conn.close()
            session['verified_at'] = time.time()
            flash('אימות דו-שלבי (מייל) הופעל ✓', 'success')
            return redirect(url_for('index'))
        if method == 'disable':
            conn.execute("UPDATE users SET twofa_method='none', totp_secret=NULL WHERE id=?", (u['id'],))
            conn.commit(); conn.close()
            flash('אימות דו-שלבי כובה', 'info')
            return redirect(url_for('twofa_setup'))
        conn.close()
        return redirect(url_for('twofa_setup'))
    conn.close()
    # GET — mint a candidate secret + QR for the app method
    secret = pyotp.random_base32()
    session['setup_secret'] = secret
    uri = pyotp.totp.TOTP(secret).provisioning_uri(name=u['username'], issuer_name='חידושים')
    import io as _io, base64, qrcode
    buf = _io.BytesIO(); qrcode.make(uri).save(buf, format='PNG')
    qr = 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode()
    return render_template('twofa_setup.html', current=(u['twofa_method'] or 'none'),
                           secret=secret, qr=qr, has_email=bool(u['email']))

@app.before_request
def _enforce_2fa():
    """Hold mid-login users at the verify page, and force re-verify every 6h for users
    who enabled 2FA. Users without 2FA are unaffected."""
    ep = request.endpoint or ''
    if session.get('pre2fa_uid'):
        if ep not in ('twofa_verify', 'login', 'logout', 'static'):
            return redirect(url_for('twofa_verify'))
        return
    uid = session.get('user_id')
    if not uid or ep in ('twofa_verify', 'twofa_setup', 'login', 'logout', 'static') or ep.startswith('health'):
        return
    conn = get_db()
    row = conn.execute("SELECT twofa_method FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    if row and (row['twofa_method'] or 'none') != 'none' and \
            time.time() - session.get('verified_at', 0) > TWOFA_TTL:
        return redirect(url_for('twofa_verify'))

RENEWAL_NO_RENEW = ('לא רוצים לחדש', 'לא מחדש', 'בוטל')
RENEWAL_CONTACTED = ('נוצר קשר עם לקוח', 'קיבל פניה', 'הלקוח אישר')

def _renewal_funnel(subset):
    """Renewal funnel counts for a set of customer rows (must carry status, import_source,
    form_received_at, call_status_1..3). New business is NOT a renewal — excluded from the
    total/buckets/% both by status (ממתין להפקה/הופק/בוטל) and by source (NEW_BUSINESS_SOURCES),
    whatever its work status. Single source of truth for the dashboard AND /api/funnel."""
    def _contacted(r):
        return bool(r['call_status_1'] or r['call_status_2'] or r['call_status_3'])
    pending_issue = sum(1 for r in subset if r['status'] == 'ממתין להפקה')
    core = [r for r in subset
            if r['status'] not in ('ממתין להפקה', 'הופק', 'בוטל')
            and (r['import_source'] or '') not in NEW_BUSINESS_SOURCES]
    t = len(core)
    rnw = sum(1 for r in core if r['status'] in ('חודש', 'חודש - בוצעה שיחת מכירה'))
    no_renew = sum(1 for r in core if r['status'] in RENEWAL_NO_RENEW)
    seen = sum(1 for r in core if r['status'] in RENEWAL_CONTACTED)
    forms = sum(1 for r in core if r['status'] == 'טופס התקבל')
    return {
        'total': t, 'renewed': rnw,
        'renewed_from_forms': sum(1 for r in core if r['status'] in ('חודש', 'חודש - בוצעה שיחת מכירה') and r['form_received_at']),
        'forms': forms, 'no_renew': no_renew, 'seen': seen,
        'pending_issue': pending_issue,
        'no_contact': sum(1 for r in core if not r['status'] and not _contacted(r)),
        'pending': t - rnw - no_renew - seen - forms,
        'pct': round(rnw / t * 100, 1) if t else 0,
    }

def _pending_split(rows):
    """'ממתין להפקה' split by lead SOURCE (brand-agnostic — a lead has no brand until issuance, which
    is why counting it inside a single-brand view undercounts). join_form=אתר, harel_proposal=הראל."""
    site = harel = other = 0
    for r in rows:
        if (r['status'] or '') != 'ממתין להפקה':
            continue
        src = (r['import_source'] or '')
        if src == 'join_form':
            site += 1
        elif src == 'harel_proposal':
            harel += 1
        else:
            other += 1
    return {'pending_site': site, 'pending_harel': harel, 'pending_other': other,
            'pending_total': site + harel + other}

@app.route('/')
@login_required
def index():
    month = active_month()
    stats = {}
    views, view_labels = {}, []
    if month:
        conn = get_db()
        bc, bp = brand_clause()
        rows = conn.execute("""SELECT status, brand, sector, form_received_at, import_source,
                               call_status_1, call_status_2, call_status_3
                               FROM customers WHERE month_id=?""" + bc,
                            [month['id']] + bp).fetchall()
        _funnel = _renewal_funnel
        # Per-agency views for the top-of-dashboard toggle (client-side switch).
        # Gaia+Winner are the active book; Ofir is planning-only, so its data is masked.
        present = [b for b in ('גאיה', 'ווינר', 'אופיר') if any(r['brand'] == b for r in rows)]
        active = [b for b in ('גאיה', 'ווינר') if b in present]
        GW = 'גאיה + ווינר'
        views = {}
        view_labels = []
        if len(active) > 1:
            views[GW] = _funnel([r for r in rows if r['brand'] in active])
            view_labels.append(GW)
        for b in present:
            views[b] = _funnel([r for r in rows if r['brand'] == b])
            view_labels.append(b)
        if not view_labels:  # no active month data at all
            views[GW] = _funnel(rows)
            view_labels = [GW]
        masked_views = ['אופיר']  # planning-only — show the layout, hide the numbers
        # 'ממתין להפקה' is a brand-agnostic lead bucket — count it GLOBALLY (all rows) and inject the
        # אתר/הראל split into every view, so the card shows the real total (not just branded rows).
        psplit = _pending_split(rows)
        for _v in views.values():
            _v.update(psplit)
            _v['pending_issue'] = psplit['pending_total']
        # Ofir renewals split by ענף (sector): total vs renewed (חודש) per category → %.
        ofir_rows = [r for r in rows if r['brand'] == 'אופיר']
        ofir_by_category = []
        for cat, aliases in OFIR_CATEGORIES:
            in_cat = [r for r in ofir_rows if any(a in (r['sector'] or '') for a in aliases)]
            t = len(in_cat)
            rnw = sum(1 for r in in_cat if r['status'] in ('חודש', 'חודש - בוצעה שיחת מכירה'))
            if t:
                ofir_by_category.append({'category': cat, 'total': t, 'renewed': rnw,
                                         'pct': round(rnw / t * 100, 1)})
        # 'pending' badge = items a rep escalated to the admin queue (mark_clarify).
        unmatched = conn.execute("SELECT COUNT(*) FROM unmatched_submissions WHERE status='pending'").fetchone()[0]
        conn.close()
        # Initial (server-rendered) numbers reflect the default view = first tab.
        stats = dict(views[view_labels[0]], ofir=len(ofir_rows),
                     ofir_by_category=ofir_by_category, unmatched=unmatched)
    return render_template('dashboard.html', month=month, stats=stats,
                           views=views, view_labels=view_labels,
                           masked_views=(masked_views if month else []),
                           views_json=json.dumps(views, ensure_ascii=False))

@app.route('/customers')
@login_required
def customers():
    # Optional ?month=<id> lets ניהול open a PREVIOUS (archived) month's renewal list so
    # late renewals can still be worked; default is the active month.
    req_month = request.args.get('month', type=int)
    if req_month:
        conn0 = get_db()
        month = conn0.execute("SELECT * FROM months WHERE id=?", (req_month,)).fetchone()
        conn0.close()
    else:
        month = active_month()
    if not month:
        flash('אין חודש פעיל. המנהל צריך לטעון נתונים.', 'warning')
        return redirect(url_for('index'))
    is_archived = not month['is_active']

    brand_filter = request.args.get('brand', '')
    status_filter = request.args.get('status', '')
    midwife_filter = request.args.get('mw', '') == '1'
    src_filter = request.args.get('src', '')  # import_source, e.g. join_form (אתר) / harel_proposal (הראל)
    search = request.args.get('q', '').strip()

    query = "SELECT * FROM customers WHERE month_id=?"
    params = [month['id']]

    # Hard permission fence: non-admins only ever see their granted agencies.
    bc, bp = brand_clause()
    query += bc
    params += bp

    if brand_filter:
        # Each agency is its own brand now — Ofir is no longer merged into Winner.
        query += " AND brand=?"
        params.append(brand_filter)
    if midwife_filter:
        # מיילדות marker — Winner only in practice, but filter is agency-agnostic.
        query += " AND is_midwife=1"
    if status_filter == '__empty__':
        query += " AND (status IS NULL OR status='')"
    elif status_filter:
        query += " AND status=?"
        params.append(status_filter)
    if src_filter:
        query += " AND import_source=?"
        params.append(src_filter)
    # New business belongs ONLY in 'ממתין להפקה' + the other-forms page — keep it out of the renewal
    # work views (ממתין לטיפול / work queue / every other status filter). A name/phone search still
    # finds anyone. (Sharon's rule: the work lists are renewals only.)
    if status_filter != 'ממתין להפקה' and not search:
        _nbph = ','.join('?' * len(NEW_BUSINESS_SOURCES))
        query += f" AND COALESCE(import_source,'') NOT IN ({_nbph})"
        params += list(NEW_BUSINESS_SOURCES)
    if search:
        like = f'%{search}%'
        name_cond, name_params = _name_search('name', search, like)
        query += " AND (" + name_cond + " OR phone LIKE ? OR policy_number LIKE ?)"
        params += name_params + [like, like]

    query += " ORDER BY name"

    conn = get_db()
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('customers.html', customers=rows, month=month,
                           brand_filter=brand_filter, status_filter=status_filter,
                           midwife_filter=midwife_filter, search=search,
                           statuses=STATUSES, is_archived=is_archived)

@app.route('/customer/add', methods=['POST'])
@login_required
def customer_add():
    """Manually add a customer to the ACTIVE month + upsert the insured master. import_source
    'manual' so it's a normal renewal candidate (not new-business). Deduped by ת"ז in the month."""
    month = active_month()
    if not month:
        flash('אין חודש פעיל', 'warning'); return redirect(url_for('index'))
    name = (request.form.get('name') or '').strip()
    idn = re.sub(r'\D', '', request.form.get('id_number') or '')
    phone = re.sub(r'\D', '', request.form.get('phone') or '')
    email = (request.form.get('email') or '').strip()
    brand = (request.form.get('brand') or '').strip()
    status = (request.form.get('status') or '').strip()
    if not name or brand not in ('גאיה', 'ווינר', 'אופיר'):
        flash('חובה שם ומותג תקין', 'danger'); return redirect(url_for('customers'))
    conn = get_db()
    if idn:
        dup = conn.execute("SELECT id FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=?",
                           (month['id'], idn.lstrip('0'))).fetchone()
        if dup:
            conn.close(); flash(f'לקוח עם ת"ז {idn} כבר קיים בחודש', 'warning'); return redirect(url_for('customers'))
    now = datetime.datetime.now()
    conn.execute(
        """INSERT INTO customers (month_id, name, id_number, phone, email, brand, status,
                                  import_source, status_changed_at)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (month['id'], name, idn, phone, email, brand, status, 'manual', now.strftime('%Y-%m-%d %H:%M')))
    if idn:
        ins = conn.execute("SELECT id FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?",
                           (idn.lstrip('0'),)).fetchone()
        iso = now.isoformat()
        if ins:
            conn.execute("UPDATE insureds SET name=COALESCE(NULLIF(name,''),?), phone=COALESCE(NULLIF(phone,''),?), "
                         "email=COALESCE(NULLIF(email,''),?), brand=COALESCE(NULLIF(brand,''),?), updated_at=? WHERE id=?",
                         (name, phone, email, brand, iso, ins['id']))
        else:
            conn.execute("INSERT INTO insureds (id_number, name, brand, phone, email, status, created_at, updated_at) "
                         "VALUES (?,?,?,?,?,?,?,?)", (idn, name, brand, phone, email, 'פעיל', iso, iso))
    conn.commit(); conn.close()
    flash(f'נוסף לקוח: {name} ({brand})', 'success')
    return redirect(url_for('customers'))

@app.route('/search')
@login_required
def search_customers():
    """Global customer search (across all months) by name, phone or policy number."""
    search = request.args.get('q', '').strip()
    rows = []
    if search:
        conn = get_db()
        like = f'%{search}%'
        # Normalised phone match too, so 050-123 finds 0501234567 etc.
        digits = re.sub(r'\D', '', search)
        phone_like = f'%{digits}%' if digits else like
        name_cond, name_params = _name_search('c.name', search, like)
        bc, bp = brand_clause('c.brand')
        crows = conn.execute(
            """SELECT c.*, m.name AS month_name
               FROM customers c
               LEFT JOIN months m ON m.id = c.month_id
               WHERE (""" + name_cond + """
                  OR c.phone LIKE ?
                  OR replace(replace(c.phone,'-',''),' ','') LIKE ?
                  OR c.policy_number LIKE ?
                  OR ltrim(c.id_number,'0') LIKE ?)""" + bc + """
               ORDER BY m.id DESC, c.name""",
            name_params + [like, phone_like, like, like] + bp
        ).fetchall()
        seen = set()
        for r in crows:
            d = dict(r)
            d['link_url'] = f"/customer/{r['id']}"
            rows.append(d)
            z = (normalize_id_number(r['id_number']) or '').lstrip('0')
            if z:
                seen.add(z)
        # ALSO search the whole-book master ('לקוחות קבוצת אופיר' / insureds) — everyone, not just
        # this month's renewal list — so a person who isn't a current customer is still found.
        iname_cond, iname_params = _name_search('name', search, like)
        ibc, ibp = brand_clause('brand')
        irows = conn.execute(
            """SELECT * FROM insureds
               WHERE (""" + iname_cond + """
                  OR phone LIKE ?
                  OR replace(replace(phone,'-',''),' ','') LIKE ?
                  OR policy_number LIKE ?
                  OR ltrim(id_number,'0') LIKE ?)""" + ibc + """
               ORDER BY name""",
            iname_params + [like, phone_like, like, like] + ibp
        ).fetchall()
        for r in irows:
            z = (normalize_id_number(r['id_number']) or '').lstrip('0')
            if z and z in seen:
                continue  # already shown as a current customer
            d = dict(r)
            d['link_url'] = f"/insured/{r['id']}"
            d['month_name'] = 'כל הלקוחות'
            rows.append(d)
        conn.close()
    return render_template('search_results.html', customers=rows, search=search)


@app.route('/customer/<int:cid>', methods=['GET', 'POST'])
@login_required
def customer_detail(cid):
    conn = get_db()
    month = active_month()
    customer = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    # Managers/super-admins a rep can route an escalation to.
    managers = conn.execute(
        "SELECT id, display_name, role FROM users WHERE role IN ('admin','superadmin') ORDER BY role DESC, display_name"
    ).fetchall()
    changes = conn.execute(
        "SELECT * FROM field_changes WHERE customer_id=? ORDER BY id DESC LIMIT 50", (cid,)
    ).fetchall()
    events = get_events(conn, event_key(customer['id_number'] if customer else '', 'cust-%d' % cid)) if customer else []
    conn.close()
    if not customer:
        flash('לקוח לא נמצא', 'danger')
        return redirect(url_for('customers'))
    if not can_access_brand(customer['brand']):
        flash('אין לך הרשאה לצפות בלקוח של סוכנות זו', 'danger')
        return redirect(url_for('customers'))
    wa_link = build_followup_wa_link(customer)
    # Rich underwriting data captured from a website join form (fields not in the policy PDF).
    lead_data = None
    try:
        if customer['import_source'] == 'join_form' and customer['lead_form_json']:
            lead_data = json.loads(customer['lead_form_json'])
    except Exception:
        lead_data = None
    return render_template('customer_detail.html', c=customer, month=month,
                           statuses=STATUSES, status_options=status_options_for(customer['brand']),
                           managers=managers, changes=changes, audit_labels=AUDIT_LABELS,
                           events=events, wa_link=wa_link, lead_data=lead_data)


def build_followup_wa_link(customer):
    """Pre-filled WhatsApp reminder link for a customer who didn't answer calls."""
    from urllib.parse import quote
    phone = re.sub(r'\D', '', str(customer['phone'] or ''))
    if not phone:
        return None
    if phone.startswith('0'):
        phone = phone[1:]
    phone = '972' + phone
    site = 'https://www.winner-ins.co.il/renew' if customer['brand'] in ('ווינר', 'אופיר') \
        else 'https://www.gaia-ins.co.il/renew'
    msg = ('היי, \nניסינו להשיג אותך לחידוש הפוליסה. נשמח אם תוכל ליצור איתנו קשר '
           'לטובת החידוש, או לחדש את הפוליסה אונליין באתר ' + site)
    return f'https://wa.me/{phone}?text={quote(msg)}'

@app.route('/customer/<int:cid>/update', methods=['POST'])
@login_required
def update_customer(cid):
    data = request.json or {}
    # Permission fence: only super-admins skip it; managers and agents may modify
    # customers only within their agencies.
    if session.get('role') != 'superadmin':
        _c = get_db()
        _row = _c.execute("SELECT brand FROM customers WHERE id=?", (cid,)).fetchone()
        _c.close()
        if not _row or not can_access_brand(_row['brand']):
            return jsonify({'ok': False, 'error': 'אין הרשאה לסוכנות זו'}), 403
    allowed = ['status', 'contact_date', 'interested_in_products', 'end_reminder_sent_date', 'group_owner',
                'whatsapp_sent_date', 'sharon_notes', 'requests_to_sharon', 'is_vip', 'is_midwife',
                'whatsapp_source', 'brand', 'phone', 'email', 'address', 'name', 'id_number',
                'call_date_1', 'call_status_1', 'call_by_1',
                'call_date_2', 'call_status_2', 'call_by_2',
                'call_date_3', 'call_status_3', 'call_by_3']
    # הערות נציג is an activity log, not a stored field: a saved note becomes an event
    # and the box clears. (Sharon's private notes stay a normal field.)
    note_text = str(data.get('agent_notes') or '').strip()
    # Agents cannot update sharon fields or brand (manager/super-admin only)
    if session.get('role') not in ('superadmin', 'admin'):
        for f in ['sharon_notes', 'requests_to_sharon', 'brand']:
            data.pop(f, None)

    # Ofir customers are contacted from Winner's WhatsApp number
    if data.get('brand') == 'אופיר' and 'whatsapp_source' not in data:
        data['whatsapp_source'] = 'ווינר'

    agent = session.get('display_name') or session.get('username', '')
    conn = get_db()

    # Auto-capture the rep who logged a call attempt (like the date) — only when
    # that attempt's date is newly set or changed, so it isn't reassigned on every save.
    if agent and any(f'call_date_{n}' in data for n in (1, 2, 3)):
        prev = conn.execute(
            "SELECT call_date_1, call_date_2, call_date_3 FROM customers WHERE id=?", (cid,)
        ).fetchone()
        for n in (1, 2, 3):
            key = f'call_date_{n}'
            if key in data and data[key] and (not prev or data[key] != prev[f'call_date_{n}']):
                data[f'call_by_{n}'] = agent

    # Snapshot audited identity/contact fields before the write, so every change can be
    # logged old → new. Keys are whitelisted against AUDITED_FIELDS, never raw input.
    audit_keys = [k for k in data if k in AUDITED_FIELDS and k in allowed]
    before = {}
    if audit_keys:
        snap = conn.execute(
            f"SELECT {','.join(audit_keys)} FROM customers WHERE id=?", (cid,)).fetchone()
        if snap:
            before = {k: snap[k] for k in audit_keys}

    crow = conn.execute("SELECT status, id_number FROM customers WHERE id=?", (cid,)).fetchone()
    idkey = event_key(crow['id_number'] if crow else '', 'cust-%d' % cid)
    status_changed = False
    sets = ', '.join(f"{k}=?" for k in data if k in allowed)
    vals = [data[k] for k in data if k in allowed]
    if sets:
        # Track who changed the status — only when it actually changes. Saving a note (the
        # form posts the status too) must not reassign the customer to whoever pressed save.
        if 'status' in data and (not crow or (crow['status'] or '') != (data.get('status') or '')):
            status_changed = True
            sets += ', status_changed_at=?'
            vals.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M'))
            if agent:
                sets += ', handled_by=?'
                vals.append(agent)
        vals.append(cid)
        conn.execute(f"UPDATE customers SET {sets} WHERE id=?", vals)
        # 'חודש' (renewal) and 'הופק' (new-business issuance) both make the person active in
        # the master (כל הלקוחות): a late renewal/issue reactivates even a 'לא פעיל' record,
        # and a brand-new issued customer is inserted into the master.
        if data.get('status') in ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק'):
            _sync_customer_to_insured(conn, cid, active=True)
            _resolve_form_queue(conn, (crow['id_number'] if crow else '') or data.get('id_number', ''), escalations=True)
        # Entering the "bad payment method" state clears the send-markers, so the
        # "update your card" message goes out fresh — including on a second collection problem.
        if status_changed and data.get('status') == CARD_UPDATE_STATUS:
            conn.execute("UPDATE customers SET card_update_wa_at=NULL, card_update_email_at=NULL WHERE id=?", (cid,))
    # Write the audit trail for any audited field that actually changed.
    if before:
        now_s = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        for k in audit_keys:
            old_v = '' if before.get(k) is None else str(before[k])
            new_v = '' if data.get(k) is None else str(data[k])
            if old_v != new_v:
                conn.execute(
                    "INSERT INTO field_changes (customer_id, field, old_value, new_value, changed_by, changed_at)"
                    " VALUES (?,?,?,?,?,?)", (cid, k, old_v, new_v, agent, now_s))
    # Activity log: status change + the rep's note become timeline events; the note box clears.
    if status_changed:
        log_event(conn, idkey, 'סטטוס עודכן ל: ' + (data.get('status') or '(ריק)'), agent, kind='status')
    if note_text:
        log_event(conn, idkey, note_text, agent)
        conn.execute("UPDATE customers SET agent_notes='' WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── Admin ───────────────────────────────────────────────────

@app.route('/customer/<int:cid>/delete', methods=['POST'])
@login_required
@admin_required
def delete_customer(cid):
    conn = get_db()
    row = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    if not row:
        conn.close()
        flash('לקוח לא נמצא', 'danger')
        return redirect(url_for('customers'))
    # Managers may only delete within their agencies.
    if not can_access_brand(row['brand']):
        conn.close()
        flash('אין הרשאה למחוק לקוח של סוכנות זו', 'danger')
        return redirect(url_for('customers'))
    # Back up the full row to the recycle bin before removing it.
    conn.execute(
        "INSERT INTO deleted_customers (customer_id, name, brand, data, deleted_at, deleted_by) VALUES (?,?,?,?,?,?)",
        (cid, row['name'], row['brand'], json.dumps(dict(row), ensure_ascii=False),
         datetime.datetime.now().isoformat(), session.get('display_name') or session.get('username', ''))
    )
    conn.execute("DELETE FROM customers WHERE id=?", (cid,))
    conn.execute("DELETE FROM customer_attachments WHERE customer_id=?", (cid,))
    conn.commit()
    conn.close()
    flash('הלקוח נמחק והועבר לסל המיחזור — ניתן לשחזר', 'warning')
    return redirect(url_for('customers'))


@app.route('/admin/trash')
@login_required
@superadmin_required
def trash():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, customer_id, name, brand, deleted_at, deleted_by FROM deleted_customers ORDER BY deleted_at DESC LIMIT 500"
    ).fetchall()
    conn.close()
    return render_template('trash.html', items=rows)


@app.route('/admin/trash/<int:tid>/restore', methods=['POST'])
@login_required
@superadmin_required
def restore_customer(tid):
    conn = get_db()
    t = conn.execute("SELECT * FROM deleted_customers WHERE id=?", (tid,)).fetchone()
    if not t:
        conn.close()
        flash('הפריט לא נמצא בסל המיחזור', 'danger')
        return redirect(url_for('trash'))
    data = json.loads(t['data'])
    # Only restore columns that still exist in the table; drop the old id so it re-inserts.
    cols = {r[1] for r in conn.execute("PRAGMA table_info(customers)").fetchall()}
    data.pop('id', None)
    fields = {k: v for k, v in data.items() if k in cols}
    # If the original month is gone, drop it into the active month so it stays visible.
    if 'month_id' in fields:
        exists = conn.execute("SELECT 1 FROM months WHERE id=?", (fields['month_id'],)).fetchone()
        if not exists:
            am = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
            fields['month_id'] = am['id'] if am else None
    keys = list(fields.keys())
    conn.execute(f"INSERT INTO customers ({','.join(keys)}) VALUES ({','.join('?' * len(keys))})",
                 [fields[k] for k in keys])
    conn.execute("DELETE FROM deleted_customers WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    flash('הלקוח שוחזר', 'success')
    return redirect(url_for('trash'))


@app.route('/admin/trash/<int:tid>/purge', methods=['POST'])
@login_required
@superadmin_required
def purge_customer(tid):
    conn = get_db()
    conn.execute("DELETE FROM deleted_customers WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    flash('נמחק לצמיתות מסל המיחזור', 'warning')
    return redirect(url_for('trash'))


@app.route('/export/customers-excel')
@login_required
@admin_required
def export_customers_excel():
    import openpyxl
    from io import BytesIO
    conn = get_db()
    month = active_month()
    if not month:
        flash('אין חודש פעיל', 'warning')
        return redirect(url_for('customers'))
    bc, bp = brand_clause()
    rows = conn.execute("SELECT * FROM customers WHERE month_id=?" + bc + " ORDER BY id",
                        [month['id']] + bp).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month['name']

    headers = ['פוליסה', 'שם', 'ת.ז', 'טלפון', 'מותג', 'סטטוס',
               'פרמיה שנה שעברה', 'וואטסאפ נשלח', 'תאריך התקשרות',
               'הערות נציג', 'הערות שרון', 'בקשות משרון',
               'טופס התקבל', 'מייל לקוח', 'תשלומים', 'גבייה',
               'מספר כרטיס', 'תוקף כרטיס', 'הערות טופס', 'טיפל']
    ws.append(headers)

    for r in rows:
        ws.append([
            r['policy_number'], r['name'], r['id_number'], r['phone'], r['brand'], r['status'],
            r['premium_last_year'], r['whatsapp_sent_date'], r['contact_date'],
            r['agent_notes'], r['sharon_notes'], r['requests_to_sharon'],
            r['form_received_at'], r['form_email'], r['form_installments'], r['form_payment_method'],
            r['form_card_number'], r['form_card_expiry'], r['form_comments'], r['handled_by'],
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"לקוחות_{month['name']}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin-queue')
@login_required
@admin_required
def admin_queue():
    conn = get_db()
    # Only rep-escalated items ('pending', set by mark_clarify). Raw website intake
    # ('ממתין') belongs to /admin/other-forms — keeping them apart avoids duplication.
    name_col = "(SELECT display_name FROM users WHERE id=assigned_to) AS assigned_name"
    if session.get('role') == 'superadmin':
        # Super-admin sees everything, so nothing routed anywhere gets lost.
        items = conn.execute(
            f"SELECT *, {name_col} FROM unmatched_submissions WHERE status='pending' ORDER BY received_at DESC"
        ).fetchall()
    else:
        # A manager sees items routed to them, plus unassigned items in their agencies.
        bc, bp = brand_clause()
        items = conn.execute(
            f"SELECT *, {name_col} FROM unmatched_submissions WHERE status='pending' "
            "AND (assigned_to=? OR (assigned_to IS NULL" + bc + ")) ORDER BY received_at DESC",
            [session.get('user_id')] + bp
        ).fetchall()
    # For items escalated from a customer (message_id 'queue-cid-<id>'), attach the linked
    # customer's uploaded document (if any) so the queue can offer a "view document" link.
    out = []
    for r in items:
        d = dict(r)
        mid = d.get('message_id') or ''
        if mid.startswith('queue-cid-'):
            try:
                ccid = int(mid.replace('queue-cid-', ''))
                c = conn.execute("SELECT id_number, lead_doc_path, sharon_notes, requests_to_sharon "
                                 "FROM customers WHERE id=?", (ccid,)).fetchone()
                if c:
                    if (c['lead_doc_path'] or '').strip():
                        d['doc_cid'] = ccid
                        d['doc_name'] = 'מסמך ' + (normalize_id_number(c['id_number']) or str(c['id_number'] or ''))
                    # Surface the customer's private notes on the card so the admin doesn't have to
                    # open each client to read what they jotted down.
                    d['sharon_notes'] = (c['sharon_notes'] or '').strip()
                    d['requests_to_sharon'] = (c['requests_to_sharon'] or '').strip()
                    d['cust_id'] = ccid
            except (ValueError, TypeError):
                pass
        out.append(d)
    conn.close()
    return render_template('admin_queue.html', items=out)

@app.route('/lead-doc/<int:cid>')
@login_required
def lead_doc_view(cid):
    """View a customer's uploaded website-form document (session-authed, for the admin queue /
    customers list). The file sits on the ephemeral server; if it's been cleaned up on a redeploy,
    the permanent copy is on OneDrive as 'מסמך <ת"ז>'."""
    conn = get_db()
    r = conn.execute("SELECT id_number, lead_doc_path FROM customers WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not r or not (r['lead_doc_path'] or '').strip():
        return "אין מסמך מצורף ללקוח זה.", 404
    if not os.path.exists(r['lead_doc_path']):
        idn = normalize_id_number(r['id_number']) or (r['id_number'] or '')
        return ("<div dir='rtl' style='font-family:sans-serif;padding:28px;font-size:16px'>"
                "הקובץ כבר לא בשרת (אחסון זמני). העותק הקבוע שמור ב-OneDrive בשם "
                f"<b>מסמך {idn}</b>.</div>"), 404
    return send_file(r['lead_doc_path'], download_name=os.path.basename(r['lead_doc_path']))

def guess_category(subject, source):
    """Rough auto-tag for the 'other forms' catch-all — a hint, not a strict classifier."""
    text = subject or ''
    if any(k in text for k in ['כרטיס אשראי', 'אשראי', 'עדכון פרטי תשלום', 'שינוי אמצעי']):
        return 'עדכון אמצעי תשלום'
    if source == 'policy':
        return 'פוליסה לא משויכת (עסקה חדשה?)'
    if 'חדש' in text:
        return 'הצעה חדשה'
    return 'אחר'

@app.route('/admin/other-forms')
@login_required
@admin_required
def other_forms():
    """Catch-all view: every incoming email that didn't become a matched renewal —
    backup net + light organization, regardless of source table."""
    conn = get_db()
    rows = []

    # Only real website-form submissions here. Harel policy PDFs are intentionally
    # excluded — they already live in the insureds master ("כל הלקוחות"), so showing
    # them here too was double-bookkeeping. Automated morning monitor tests are filtered.
    bc, bp = brand_clause()  # managers see only their agencies' forms
    not_monitor = ("AND COALESCE(id_number,'') != '999999999' "
                   "AND COALESCE(email,'') != 'monitor-check@example.com' "
                   "AND COALESCE(name,'') != 'MONITOR-CHECK-DO-NOT-PROCESS' ")
    show = request.args.get('show', 'active')
    wanted = {'done': ('טופל',), 'all': FORM_QUEUE_STATUSES}.get(show, ('ממתין', 'בטיפול'))
    ph = ','.join('?' * len(wanted))
    for r in conn.execute(
        f"SELECT * FROM unmatched_submissions WHERE status IN ({ph}) " + not_monitor + bc +
        " ORDER BY received_at DESC", list(wanted) + bp
    ).fetchall():
        d = dict(r)
        rows.append({
            'id': d['id'], 'received_at': d['received_at'], 'subject': d['subject'],
            'title': d['name'] or '(ללא שם)', 'detail': d['id_number'] or d['phone'] or '',
            'source': 'טופס', 'category': guess_category(d['subject'], 'form'),
            'status': d['status'], 'handled_by': d['handled_by'], 'handled_at': d['handled_at'],
            'insured_id': d['insured_id'],
            'link': None, 'kind': 'form', 'full': d,
        })
    # Counts per queue state, for the filter tabs.
    counts = {}
    for st in FORM_QUEUE_STATUSES:
        counts[st] = conn.execute(
            "SELECT COUNT(*) FROM unmatched_submissions WHERE status=? " + not_monitor + bc,
            [st] + bp).fetchone()[0]

    rows.sort(key=lambda x: x['received_at'] or '', reverse=True)
    conn.close()
    return render_template('other_forms.html', items=rows, counts=counts, show=show,
                           queue_labels=FORM_QUEUE_LABELS)


@app.route('/admin/insurance-cert', methods=['GET', 'POST'])
@login_required
@admin_required
def insurance_cert():
    """אישור קיום ביטוחים (נספח א'): pick one of the recurring companies + a customer ת.ז,
    and get a print-ready certificate pre-filled from the insureds/policy data (Harel logo
    stamped). Every field on the sheet is editable before printing to PDF."""
    # Restricted to Sharon's user only (until further notice).
    if not can_issue_cert():
        flash('הפקת אישור קיום ביטוח מוגבלת כרגע למשתמש מורשה בלבד', 'warning')
        return redirect(url_for('index'))
    companies = INSURANCE_CERT_COMPANIES

    if request.method == 'GET':
        return render_template('insurance_cert_form.html', companies=companies, C=CERT_CONSTANTS)

    # POST — look up the customer and build the certificate draft.
    company_key = request.form.get('company', '')
    id_raw = request.form.get('id_number', '')
    company = next((c for c in companies if c['key'] == company_key), None)
    digits = re.sub(r'\D', '', id_raw)
    norm = digits.lstrip('0')          # normalized key for matching (leading zeros dropped)
    id_display = digits.zfill(9)       # full 9-digit ת.ז for the certificate (keep leading zeros)

    if not company:
        flash('בחר חברה מבקשת אישור', 'danger')
        return redirect(url_for('insurance_cert'))
    if not norm:
        flash('הזן תעודת זהות של המבוטח', 'danger')
        return redirect(url_for('insurance_cert'))

    conn = get_db()
    ins = conn.execute(
        "SELECT * FROM insureds WHERE ltrim(COALESCE(id_number,''),'0') = ?", (norm,)
    ).fetchone()
    # Newest policy record for this ת.ז (policy number, period, name/address fallback + the PDF).
    pr = conn.execute(
        """SELECT pr.*, pd.filepath AS doc_filepath
           FROM policy_records pr
           LEFT JOIN policy_documents pd ON pr.policy_document_id = pd.id
           WHERE ltrim(COALESCE(pr.insured_id,''),'0') = ?
           ORDER BY pr.extracted_at DESC LIMIT 1""", (norm,)
    ).fetchone()
    # Sequential certificate number 7338-0001, 7338-0002, … (persisted in app_meta).
    _row = conn.execute("SELECT value FROM app_meta WHERE key='cert_seq'").fetchone()
    _seq = (int(_row['value']) if _row and str(_row['value']).isdigit() else 0) + 1
    conn.execute("INSERT INTO app_meta (key, value) VALUES ('cert_seq', ?) "
                 "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(_seq),))
    conn.commit()
    cert_number = f"7338-{_seq:04d}"
    conn.close()

    if not ins and not pr:
        flash(f'לא נמצא מבוטח עם ת.ז {id_raw} במאגר', 'warning')
        return redirect(url_for('insurance_cert'))

    def pick(*vals):
        for v in vals:
            if v:
                return v
        return ''

    ins_name    = pick(ins['name'] if ins else '', pr['insured_name'] if pr else '')
    ins_address = pick(ins['address'] if ins else '', pr['address'] if pr else '')
    policy_num  = pick(pr['policy_number'] if pr else '', ins['policy_number'] if ins else '')
    period_start = pick(pr['period_start'] if pr else '', ins['period_start'] if ins else '')
    period_end   = pick(pr['period_end'] if pr else '', ins['period_end'] if ins else '')
    # Prefer the pre-extracted occupation column (filled in bulk via /api/set-occupations),
    # falling back to on-the-fly extraction from the stored policy PDF.
    occ_col = ins['occupation'] if (ins and 'occupation' in ins.keys()) else ''
    occupation   = pick(occ_col, extract_insured_occupation(pr['doc_filepath']) if pr else '')
    customer_phone = pick(ins['phone'] if ins else '', pr['phone_mobile'] if pr else '')

    # WhatsApp link to the customer (the PDF itself is attached manually — web WhatsApp
    # can't auto-attach a file). Opens the customer's chat with a pre-filled message.
    wa_link = None
    _p = re.sub(r'\D', '', str(customer_phone or ''))
    if _p:
        from urllib.parse import quote
        _p = ('972' + _p[1:]) if _p.startswith('0') else ('972' + _p)
        _msg = f"שלום {ins_name}, מצורף אישור קיום ביטוחים עבור {company['name']}."
        wa_link = f"https://wa.me/{_p}?text={quote(_msg)}"

    C = CERT_CONSTANTS
    cert = {
        'cert_number':  cert_number,
        'issue_date':   datetime.date.today().strftime('%d/%m/%Y'),
        # requesting company
        'req_name':     f"{company['name']} {CERT_RELATED_SUFFIX}",
        'req_hp':       company['hp'],
        'req_hp_extra': company['hp_extra'],
        'req_address':  company['address'],
        'service_codes': company['codes'],
        'deal_type':    C['deal_type'],
        'req_status':   C['req_status'],
        # insured (customer)
        'ins_name':     ins_name,
        'ins_id':       id_display,
        'ins_address':  ins_address,
        'occupation':   occupation,
        # policy + coverage
        'policy_number': policy_num,
        'form_edition':  C['form_edition'],
        'period_start':  period_start,
        'period_end':    period_end,
        'amount':        C['amount'],
        'currency':      C['currency'],
        'codes_main':    C['codes_main'],
        'codes_supp':    C['codes_supp'],
        'discovery_codes': C['discovery_codes'],
        'retro_date':    '',   # ריק כברירת מחדל — ממולא ידנית (הרטרו האמיתי שונה מתאריך התחלה)
        'insurer':       C['insurer'],
    }
    # Warn about anything the operator must fill by hand before signing.
    missing = [lbl for lbl, val in
               [('שם המבוטח', ins_name), ('כתובת המבוטח', ins_address),
                ('מספר פוליסה', policy_num), ('תקופת ביטוח', period_start and period_end),
                ('עיסוק המבוטח', occupation)] if not val]
    return render_template('insurance_cert.html', cert=cert, company=company,
                           missing=missing, matched_master=bool(ins), wa_link=wa_link)


@app.route('/admin/insurance-cert/log', methods=['POST'])
@login_required
@admin_required
def insurance_cert_log():
    """Record in the customer's file (client_events, keyed by ת.ז) that a certificate was
    printed / sent — called by the print & WhatsApp buttons on the certificate page."""
    if not can_issue_cert():
        return jsonify({'ok': False}), 403
    data = request.get_json(silent=True) or {}
    norm = re.sub(r'\D', '', str(data.get('id_number', ''))).lstrip('0')
    if not norm:
        return jsonify({'ok': False, 'error': 'no id'}), 400
    company = (data.get('company') or '').strip()
    cert_number = (data.get('cert_number') or '').strip()
    verb = {'print': 'הופק והודפס', 'whatsapp': 'נשלח בוואטסאפ'}.get(data.get('action'), 'הופק')
    note = f"אישור קיום ביטוח {verb}" + (f" — {company}" if company else '') \
           + (f" (מס' {cert_number})" if cert_number else '')
    who = session.get('display_name') or session.get('username') or 'מערכת'
    conn = get_db()
    log_event(conn, norm, note, who, kind='cert')
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/admin/other-forms/<int:sid>/file', methods=['POST'])
@login_required
@admin_required
def other_forms_open_file(sid):
    """Open the customer file for a form. Reuses the existing client's file when one
    matches (by ת.ז, else phone); otherwise creates the file from the form's details."""
    conn = get_db()
    sub = conn.execute("SELECT * FROM unmatched_submissions WHERE id=?", (sid,)).fetchone()
    if not sub:
        conn.close()
        flash('הפריט לא נמצא', 'danger')
        return redirect(url_for('other_forms'))
    if not can_access_brand(sub['brand']):
        conn.close()
        flash('אין הרשאה לסוכנות זו', 'danger')
        return redirect(url_for('other_forms'))

    # Already attached to a file → go straight there.
    if sub['insured_id']:
        exists = conn.execute("SELECT id FROM insureds WHERE id=?", (sub['insured_id'],)).fetchone()
        if exists:
            conn.close()
            return redirect(url_for('insured_detail', iid=exists['id']))

    idn = normalize_id_number(sub['id_number']) or None
    digits = re.sub(r'\D', '', str(sub['phone'] or ''))
    found = None
    if idn:
        found = conn.execute("SELECT id FROM insureds WHERE id_number=?", (idn,)).fetchone()
    if not found and digits:
        found = conn.execute(
            "SELECT id FROM insureds WHERE replace(replace(COALESCE(phone,''),'-',''),' ','')=?",
            (digits,)).fetchone()

    if found:
        iid = found['id']
    else:
        now = datetime.datetime.now().isoformat()
        # Prefer the form's name; otherwise reuse a name we already hold for this ID
        # (e.g. from an earlier form) before falling back to the placeholder.
        new_name = sub['name'] or name_from_records(conn, idn) or NO_NAME
        # Carry the address over from the submitted form when it has one.
        af = {}
        try:
            af = json.loads(sub['raw_fields']) if sub['raw_fields'] else {}
        except (ValueError, TypeError):
            af = {}
        addr = ', '.join(x for x in [str(af.get('כתובת', '')).strip(),
                                     str(af.get('עיר', '')).strip()] if x)
        conn.execute(
            """INSERT INTO insureds (id_number, name, brand, phone, email, address, status, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (idn, new_name, sub['brand'], sub['phone'], sub['email'] or af.get('אימייל', ''),
             addr, 'לא פעיל', now, now))
        iid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        flash('נפתח תיק לקוח חדש מפרטי הטופס', 'success')
    conn.execute("UPDATE unmatched_submissions SET insured_id=? WHERE id=?", (iid, sid))
    conn.commit()
    conn.close()
    return redirect(url_for('insured_detail', iid=iid))


@app.route('/admin/other-forms/<int:sid>/status', methods=['POST'])
@login_required
@admin_required
def other_forms_status(sid):
    """Advance a form through the work queue: ממתין → בטיפול → טופל (or back)."""
    new = request.form.get('status', '')
    if new not in FORM_QUEUE_STATUSES:
        flash('סטטוס לא תקין', 'danger')
        return redirect(url_for('other_forms'))
    conn = get_db()
    row = conn.execute("SELECT brand, status FROM unmatched_submissions WHERE id=?", (sid,)).fetchone()
    if not row or row['status'] not in FORM_QUEUE_STATUSES:
        conn.close()
        flash('הפריט לא נמצא בתור', 'danger')
        return redirect(url_for('other_forms'))
    if not can_access_brand(row['brand']):
        conn.close()
        flash('אין הרשאה לסוכנות זו', 'danger')
        return redirect(url_for('other_forms'))
    # Returning to the start clears the handler stamp.
    if new == 'ממתין':
        conn.execute("UPDATE unmatched_submissions SET status=?, handled_by=NULL, handled_at=NULL WHERE id=?",
                     (new, sid))
    else:
        conn.execute("UPDATE unmatched_submissions SET status=?, handled_by=?, handled_at=? WHERE id=?",
                     (new, session.get('display_name') or session.get('username', ''),
                      datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), sid))
    conn.commit()
    conn.close()
    flash(f'הפריט סומן כ"{new}"', 'success')
    # Allow advancing the status from inside the customer file and staying there.
    back = request.form.get('back_insured')
    if back:
        return redirect(url_for('insured_detail', iid=back))
    return redirect(url_for('other_forms', show=request.form.get('show', 'active')))

@app.route('/admin/other-forms/delete', methods=['POST'])
@login_required
@admin_required
def other_forms_delete():
    """Bulk-delete selected rows from the other-forms catch-all (form or policy items)."""
    selected = request.form.getlist('selected')
    form_ids = [s.split(':', 1)[1] for s in selected if s.startswith('form:')]
    policy_ids = [s.split(':', 1)[1] for s in selected if s.startswith('policy:')]

    conn = get_db()
    if form_ids:
        placeholders = ','.join('?' * len(form_ids))
        conn.execute(f"DELETE FROM unmatched_submissions WHERE id IN ({placeholders})", form_ids)
    if policy_ids:
        placeholders = ','.join('?' * len(policy_ids))
        conn.execute(f"DELETE FROM policy_documents WHERE id IN ({placeholders})", policy_ids)
    conn.commit()
    conn.close()
    flash(f'{len(form_ids) + len(policy_ids)} פריטים נמחקו', 'success')
    return redirect(url_for('other_forms'))

@app.route('/admin/policy-records')
@login_required
@admin_required
def policy_records():
    """All customers (master) — one row per insured (by ID), built from the Harel
    policy PDFs. Best-effort extraction — some fields may need correction."""
    q = request.args.get('q', '').strip()
    view = request.args.get('view', '').strip()   # agency filter (גאיה + ווינר / גאיה / ווינר / אופיר)
    st = request.args.get('st', '').strip()        # status filter (פעיל / לא פעיל / בוטל)
    conn = get_db()
    recompute_insured_statuses(conn)  # keep פעיל/לא פעיל current on view
    bc, bp = brand_clause()  # managers see only their agencies; super-admins see all

    where = ' FROM insureds WHERE 1=1' + bc
    params = list(bp)
    if view == 'גאיה + ווינר':
        where += " AND brand IN ('גאיה','ווינר')"
    elif view in ('גאיה', 'ווינר', 'אופיר'):
        where += " AND brand=?"
        params.append(view)
    if st in ('פעיל', 'לא פעיל', 'בוטל'):
        where += " AND status=?"
        params.append(st)
    if q:
        like = f'%{q}%'
        name_cond, name_params = _name_search('name', q, like)
        where += ' AND (' + name_cond + ' OR id_number LIKE ? OR policy_number LIKE ? OR phone LIKE ? OR email LIKE ?)'
        params += name_params + [like, like, like, like]

    rows = conn.execute('SELECT *' + where + ' ORDER BY name LIMIT 500', params).fetchall()
    total = conn.execute('SELECT COUNT(*)' + where, params).fetchone()[0]

    # Which agency tabs to show (only ones the user may access) + status tallies.
    ab = allowed_brands()
    accessible = ['גאיה', 'ווינר', 'אופיר'] if ab is None else [b for b in ('גאיה', 'ווינר', 'אופיר') if b in ab]
    view_options = ([] if len([b for b in ('גאיה', 'ווינר') if b in accessible]) < 2 else ['גאיה + ווינר']) + accessible
    scnt = {r['status'] or '—': r['n'] for r in conn.execute(
        'SELECT status, COUNT(*) n FROM insureds' + ' WHERE 1=1' + bc +
        (" AND brand IN ('גאיה','ווינר')" if view == 'גאיה + ווינר'
         else (" AND brand='%s'" % view if view in ('גאיה', 'ווינר', 'אופיר') else '')) +
        ' GROUP BY status', bp)}
    conn.close()
    return render_template('policy_records.html', items=rows, q=q, total=total,
                           view=view, st=st, view_options=view_options, status_counts=scnt,
                           backfill=_backfill_state)

def build_followup_wa_link_generic(phone, brand):
    """Pre-filled WhatsApp reminder link from a phone + brand (works for insureds too)."""
    from urllib.parse import quote
    p = re.sub(r'\D', '', str(phone or ''))
    if not p:
        return None
    if p.startswith('0'):
        p = p[1:]
    p = '972' + p
    site = 'https://www.winner-ins.co.il/renew' if brand in ('ווינר', 'אופיר') \
        else 'https://www.gaia-ins.co.il/renew'
    msg = ('היי, \nניסינו להשיג אותך לחידוש הפוליסה. נשמח אם תוכל ליצור איתנו קשר '
           'לטובת החידוש, או לחדש את הפוליסה אונליין באתר ' + site)
    return f'https://wa.me/{p}?text={quote(msg)}'

@app.route('/insured/<int:iid>')
@login_required
@admin_required
def insured_detail(iid):
    conn = get_db()
    ins = conn.execute("SELECT * FROM insureds WHERE id=?", (iid,)).fetchone()
    if not ins:
        conn.close()
        flash('לקוח לא נמצא', 'danger')
        return redirect(url_for('policy_records'))
    # PDF history for this insured (by ID), newest policy first
    docs = conn.execute(
        """SELECT pd.id AS doc_id, pd.filename, pd.received_at,
                  pr.doc_type_label, pr.period_start, pr.period_end
           FROM policy_records pr JOIN policy_documents pd ON pr.policy_document_id = pd.id
           WHERE ltrim(pr.insured_id,'0') = ltrim(?,'0')
           ORDER BY pr.extracted_at DESC""",
        (ins['id_number'],)
    ).fetchall()
    # Website forms attached to this file — shown with their work-queue status so the
    # whole handling happens here, without bouncing back to the forms list.
    # Every form from this client — the one this file was opened from, plus any other
    # submission carrying the same ת.ז, so richer earlier forms aren't hidden.
    forms = conn.execute(
        "SELECT * FROM unmatched_submissions WHERE insured_id=? "
        "   OR (COALESCE(?,'')<>'' AND ltrim(COALESCE(id_number,''),'0')=?) "
        "ORDER BY received_at DESC",
        (iid, ins['id_number'], (ins['id_number'] or '').lstrip('0'))
    ).fetchall()
    managers = conn.execute(
        "SELECT id, display_name, role FROM users WHERE role IN ('admin','superadmin') ORDER BY role DESC, display_name"
    ).fetchall()
    events = get_events(conn, event_key(ins['id_number'], 'ins-%d' % iid))
    conn.close()
    wa_link = build_followup_wa_link_generic(ins['phone'], ins['brand'])
    return render_template('insured_detail.html', c=ins, docs=docs, wa_link=wa_link,
                           forms=forms, queue_labels=FORM_QUEUE_LABELS, managers=managers,
                           events=events)


@app.route('/insured/<int:iid>/clarify', methods=['POST'])
@login_required
@admin_required
def insured_clarify(iid):
    """Escalate a customer file to the admin queue, optionally routed to a manager."""
    data = request.get_json(silent=True) or {}
    note = (data.get('note') or '').strip()
    if not note:
        return jsonify({'ok': False, 'error': 'נא לפרט את הסיבה להעברה'}), 400
    try:
        assigned_to = int(data.get('assigned_to')) if data.get('assigned_to') else None
    except (ValueError, TypeError):
        assigned_to = None
    conn = get_db()
    ins = conn.execute("SELECT * FROM insureds WHERE id=?", (iid,)).fetchone()
    if not ins:
        conn.close()
        return jsonify({'ok': False, 'error': 'תיק לא נמצא'}), 404
    if not can_access_brand(ins['brand']):
        conn.close()
        return jsonify({'ok': False, 'error': 'אין הרשאה לסוכנות זו'}), 403
    conn.execute(
        """INSERT OR REPLACE INTO unmatched_submissions
           (received_at, subject, name, id_number, phone, email, brand, comments,
            status, handled_by, assigned_to, insured_id, message_id)
           VALUES (?,?,?,?,?,?,?,?,'pending',?,?,?,?)""",
        (datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), 'דורש בירור — תיק לקוח',
         ins['name'], ins['id_number'], ins['phone'], ins['email'], ins['brand'], note,
         session.get('display_name') or session.get('username', ''), assigned_to, iid,
         f'queue-iid-{iid}')
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/insured/<int:iid>/update', methods=['POST'])
@login_required
@admin_required
def insured_update(iid):
    data = request.json or {}
    allowed = ['whatsapp_source', 'is_vip', 'is_midwife',
               'name', 'id_number', 'phone', 'email', 'address', 'policy_number',
               'call_date_1', 'call_status_1', 'call_by_1',
               'call_date_2', 'call_status_2', 'call_by_2',
               'call_date_3', 'call_status_3', 'call_by_3']
    agent = session.get('display_name') or session.get('username', '')
    # A saved note is recorded as an event, not stored back on the file (the box is a
    # scratchpad that clears after each save).
    note_text = str(data.get('agent_notes') or '').strip()
    conn = get_db()
    _ins = conn.execute("SELECT id_number FROM insureds WHERE id=?", (iid,)).fetchone()
    idkey = event_key(_ins['id_number'] if _ins else '', 'ins-%d' % iid)

    # Snapshot audited identity/contact fields so every edit is logged old → new.
    audit_keys = [k for k in data if k in AUDITED_FIELDS and k in allowed]
    before = {}
    if audit_keys:
        snap = conn.execute(
            f"SELECT {','.join(audit_keys)} FROM insureds WHERE id=?", (iid,)).fetchone()
        if snap:
            before = {k: snap[k] for k in audit_keys}

    now_s = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    # Manual status change is an admin override that sticks (req 8) — and is logged.
    if 'status' in data and data['status']:
        prev_st = conn.execute("SELECT status FROM insureds WHERE id=?", (iid,)).fetchone()
        conn.execute("UPDATE insureds SET status=?, status_override=1, updated_at=? WHERE id=?",
                     (data['status'], datetime.datetime.now().isoformat(), iid))
        if not prev_st or (prev_st['status'] or '') != data['status']:
            log_event(conn, idkey, 'סטטוס עודכן ל: ' + data['status'], agent, kind='status')

    # Auto-capture the rep who logged a call attempt (like the renewals page)
    if agent and any(f'call_date_{n}' in data for n in (1, 2, 3)):
        prev = conn.execute(
            "SELECT call_date_1, call_date_2, call_date_3 FROM insureds WHERE id=?", (iid,)
        ).fetchone()
        for n in (1, 2, 3):
            key = f'call_date_{n}'
            if key in data and data[key] and (not prev or data[key] != prev[f'call_date_{n}']):
                data[f'call_by_{n}'] = agent

    sets = ', '.join(f"{k}=?" for k in data if k in allowed)
    if sets:
        vals = [data[k] for k in data if k in allowed]
        vals.append(iid)
        conn.execute(f"UPDATE insureds SET {sets} WHERE id=?", vals)
    # Audit trail for identity/contact edits made on the customer file.
    if before:
        now_s = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        for k in audit_keys:
            old_v = '' if before.get(k) is None else str(before[k])
            new_v = '' if data.get(k) is None else str(data[k])
            if old_v != new_v:
                conn.execute(
                    "INSERT INTO field_changes (customer_id, insured_id, field, old_value,"
                    " new_value, changed_by, changed_at) VALUES (0,?,?,?,?,?,?)",
                    (iid, k, old_v, new_v, agent, now_s))
    # A note becomes a timeline event; the file's notes box is left empty for the next one.
    if note_text:
        log_event(conn, idkey, note_text, agent)
        conn.execute("UPDATE insureds SET agent_notes='' WHERE id=?", (iid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/customer/<int:cid>/clarify', methods=['POST'])
@login_required
def mark_clarify(cid):
    """Move customer to admin queue for clarification. Requires a reason (rep notes)."""
    data = request.get_json(silent=True) or {}
    note = (data.get('note') or '').strip()
    if not note:
        return jsonify({'ok': False, 'error': 'נא לפרט את הסיבה להעברה לאדמין'}), 400
    try:
        assigned_to = int(data.get('assigned_to')) if data.get('assigned_to') else None
    except (ValueError, TypeError):
        assigned_to = None
    agent = session.get('display_name') or session.get('username', '')
    conn = get_db()
    c = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    if c:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        comment_parts = []
        if c['form_comments']: comment_parts.append(c['form_comments'])
        if note: comment_parts.append(note)
        comments = ' | '.join(comment_parts)
        # Use INSERT OR REPLACE so re-clarifying the same customer works
        conn.execute('''INSERT OR REPLACE INTO unmatched_submissions
            (received_at, subject, name, id_number, phone, email, brand, installments,
             payment_method, card_number, card_expiry, card_holder_id, coverage, comments,
             status, handled_by, assigned_to, message_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'pending',?,?,?)''',
            (now, 'דורש בירור', c['name'], c['id_number'], c['phone'],
             c['form_email'] or '', c['brand'], c['form_installments'] or '',
             c['form_payment_method'] or '', c['form_card_number'] or '',
             c['form_card_expiry'] or '', c['form_id_card_holder'] or '',
             c['form_coverage'] or '', comments, agent, assigned_to, f'queue-cid-{cid}'))
        conn.execute("UPDATE customers SET status='דורש בירור', handled_by=?, status_changed_at=? WHERE id=?",
                     (agent, now, cid))
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/admin-queue/<int:sid>/action', methods=['POST'])
@login_required
@admin_required
def admin_queue_action(sid):
    action = request.form.get('action')
    note = request.form.get('admin_note', '')
    conn = get_db()
    if action == 'dismiss':
        conn.execute("UPDATE unmatched_submissions SET status='dismissed', admin_note=? WHERE id=?", (note, sid))
    elif action == 'link':
        cid = request.form.get('customer_id', '')
        if cid:
            sub = conn.execute("SELECT * FROM unmatched_submissions WHERE id=?", (sid,)).fetchone()
            if sub:
                now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                conn.execute("""UPDATE customers SET status='טופס התקבל',
                    form_email=?, form_installments=?, form_payment_method=?,
                    form_received_at=?, form_coverage=?, form_comments=?,
                    form_card_number=?, form_card_expiry=?, form_id_card_holder=?,
                    status_changed_at=?
                    WHERE id=?""",
                    (sub['email'], sub['installments'], sub['payment_method'], now,
                     sub['coverage'], sub['comments'], sub['card_number'],
                     sub['card_expiry'], sub['card_holder_id'], now, cid))
                conn.execute("UPDATE unmatched_submissions SET status='linked', admin_note=? WHERE id=?",
                             (f'שויך ללקוח {cid}', sid))
    elif action == 'resolve':
        # For clarify items — set final status on the linked customer
        new_status = request.form.get('new_status', '')
        sub = conn.execute("SELECT * FROM unmatched_submissions WHERE id=?", (sid,)).fetchone()
        if sub and new_status:
            # Extract customer id from message_id = 'queue-cid-{cid}'
            msg_id = sub['message_id'] or ''
            if msg_id.startswith('queue-cid-'):
                cid = msg_id.replace('queue-cid-', '')
                agent = session.get('display_name') or session.get('username', '')
                conn.execute("UPDATE customers SET status=?, handled_by=?, status_changed_at=? WHERE id=?",
                             (new_status, agent,
                              datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), cid))
            conn.execute("UPDATE unmatched_submissions SET status='resolved', admin_note=? WHERE id=?",
                         (f'סטטוס עודכן: {new_status} | {note}', sid))
    conn.commit()
    conn.close()
    flash('בוצע', 'success')
    return redirect(url_for('admin_queue'))

@app.route('/attachment/<int:att_id>')
@login_required
def download_attachment(att_id):
    conn = get_db()
    att = conn.execute('SELECT * FROM customer_attachments WHERE id=?', (att_id,)).fetchone()
    conn.close()
    if not att:
        return 'לא נמצא', 404
    safe_name = re.sub(r'[\r\n]+', ' ', att['filename']).strip()
    return send_file(att['filepath'], as_attachment=True, download_name=safe_name)

@app.route('/policy-document/<int:doc_id>')
@login_required
def download_policy_document(doc_id):
    conn = get_db()
    doc = conn.execute('SELECT * FROM policy_documents WHERE id=?', (doc_id,)).fetchone()
    conn.close()
    if not doc:
        return 'לא נמצא', 404
    safe_name = re.sub(r'[\r\n]+', ' ', doc['filename']).strip()
    return send_file(doc['filepath'], as_attachment=True, download_name=safe_name)


@app.route('/reveal-card', methods=['POST'])
@login_required
def reveal_card():
    """Return the full credit-card number for a submission/customer, and LOG who revealed
    it (card numbers are masked to last-4 everywhere by default)."""
    data = request.get_json(silent=True) or {}
    typ, rid = data.get('type'), data.get('id')
    conn = get_db()
    if typ == 'submission':
        row = conn.execute("SELECT name, id_number, card_number, card_expiry, card_holder_id "
                           "FROM unmatched_submissions WHERE id=?", (rid,)).fetchone()
    elif typ == 'customer':
        row = conn.execute("SELECT name, id_number, form_card_number AS card_number, "
                           "form_card_expiry AS card_expiry, form_id_card_holder AS card_holder_id "
                           "FROM customers WHERE id=?", (rid,)).fetchone()
    else:
        row = None
    if not row or not row['card_number']:
        conn.close()
        return jsonify({'ok': False}), 404
    who = session.get('display_name') or session.get('username', '')
    idkey = event_key(row['id_number'] or '', f'{typ}-{rid}')
    log_event(conn, idkey, f'חשף מספר אשראי מלא של {row["name"] or ""}'.strip(), who, kind='card_reveal')
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'card_number': row['card_number'],
                    'card_expiry': row['card_expiry'] or '', 'card_holder_id': row['card_holder_id'] or ''})


@app.route('/queue')
@login_required
def queue():
    month = active_month()
    if not month:
        flash('אין חודש פעיל', 'warning')
        return redirect(url_for('index'))
    conn = get_db()
    bc, bp = brand_clause()
    _nbph = ','.join('?' * len(NEW_BUSINESS_SOURCES))
    rows = conn.execute(
        f"SELECT * FROM customers WHERE month_id=? AND status='טופס התקבל' "
        f"AND COALESCE(import_source,'') NOT IN ({_nbph})" + bc +
        " ORDER BY form_received_at DESC",
        [month['id']] + list(NEW_BUSINESS_SOURCES) + bp
    ).fetchall()
    # Fetch attachments per customer
    attachments = {}
    for r in rows:
        atts = conn.execute(
            'SELECT * FROM customer_attachments WHERE customer_id=?', (r['id'],)
        ).fetchall()
        if atts:
            attachments[r['id']] = atts
    conn.close()
    return render_template('queue.html', customers=rows, month=month, attachments=attachments)


@app.route('/admin')
@login_required
@superadmin_required
def admin():
    conn = get_db()
    users = conn.execute("SELECT id, username, display_name, role, manager_id, email FROM users ORDER BY role, display_name").fetchall()
    months = conn.execute("SELECT * FROM months ORDER BY id DESC").fetchall()
    ub_map = {}
    for r in conn.execute("SELECT user_id, brand FROM user_brands").fetchall():
        ub_map.setdefault(r['user_id'], []).append(r['brand'])
    managers = conn.execute(
        "SELECT id, display_name, role FROM users WHERE role IN ('admin','superadmin') ORDER BY role DESC, display_name"
    ).fetchall()
    pending_imports = []
    for row in conn.execute("SELECT * FROM pending_imports WHERE status='pending' ORDER BY id DESC").fetchall():
        d = dict(row)
        try:
            d['report'] = json.loads(row['report_json'] or '{}')
        except Exception:
            d['report'] = {}
        pending_imports.append(d)
    conn.close()
    return render_template('admin.html', users=users, months=months,
                           email_sync_enabled=EMAIL_CONFIG['enabled'],
                           agencies=BRANDS, user_brands=ub_map, managers=managers,
                           backfill=_backfill_state, pending_imports=pending_imports,
                           site123=_site123_state)


@app.route('/admin/users/<int:uid>/manager', methods=['POST'])
@login_required
@superadmin_required
def set_user_manager(uid):
    """Assign the manager an agent reports to (for performance grouping)."""
    mid = request.form.get('manager_id')
    try:
        mid = int(mid) if mid else None
    except (ValueError, TypeError):
        mid = None
    conn = get_db()
    conn.execute("UPDATE users SET manager_id=? WHERE id=?", (mid, uid))
    conn.commit()
    conn.close()
    flash('המנהל האחראי עודכן', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/users/<int:uid>/email', methods=['POST'])
@login_required
@superadmin_required
def set_user_email(uid):
    """Set/update a user's (rep's) email — used for email-based 2FA codes and notifications."""
    email = request.form.get('email', '').strip()
    conn = get_db()
    conn.execute("UPDATE users SET email=? WHERE id=?", (email, uid))
    conn.commit()
    conn.close()
    flash('המייל עודכן', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/performance')
@login_required
@admin_required
def performance():
    """Activity for the active month, attributed by who logged the calls (call_by) and
    who raised escalations. A super-admin sees everyone — agents, managers and other
    super-admins; a manager sees only their own agents (and never themselves, so nobody
    is shown their own activity being measured except a super-admin)."""
    conn = get_db()
    month = active_month()
    mid = month['id'] if month else -1
    is_super = session.get('role') == 'superadmin'
    if is_super:
        people = conn.execute(
            "SELECT id, display_name, role FROM users ORDER BY role DESC, display_name").fetchall()
    else:
        people = conn.execute(
            "SELECT id, display_name, role FROM users WHERE role='agent' AND manager_id=? ORDER BY display_name",
            (session.get('user_id'),)
        ).fetchall()
    role_labels = {'superadmin': 'מנהל על', 'admin': 'מנהל', 'agent': 'נציג'}
    rows = []
    for a in people:
        nm = a['display_name']
        if not nm:
            continue
        # Attribution is by who set the status (handled_by). Counts are per PERSON —
        # DISTINCT by ת.ז — so the same client never counts twice, and a client who
        # both changed status and renewed still counts as one. Digital renewals count
        # as ordinary personal renewals.
        key = "COALESCE(NULLIF(ltrim(COALESCE(id_number,''),'0'),''), 'r'||id)"
        q = conn.execute(
            "SELECT "
            "SUM((CASE WHEN call_by_1=? THEN 1 ELSE 0 END)+(CASE WHEN call_by_2=? THEN 1 ELSE 0 END)+(CASE WHEN call_by_3=? THEN 1 ELSE 0 END)) AS calls, "
            f"COUNT(DISTINCT CASE WHEN handled_by=? THEN {key} END) AS touched, "
            f"COUNT(DISTINCT CASE WHEN handled_by=? AND status IN (?,?) THEN {key} END) AS renewals, "
            f"COUNT(DISTINCT CASE WHEN handled_by=? AND status=? THEN {key} END) AS issued "
            "FROM customers WHERE month_id=?",
            [nm, nm, nm, nm, nm, 'חודש', 'חודש - בוצעה שיחת מכירה', nm, 'הופק', mid]
        ).fetchone()
        # Escalations raised by this person (customer card or customer file), which are
        # the queue-* items — not the website-form queue they merely handled.
        escalations = conn.execute(
            "SELECT COUNT(*) FROM unmatched_submissions WHERE handled_by=? "
            "AND (message_id LIKE 'queue-cid-%' OR message_id LIKE 'queue-iid-%')", (nm,)
        ).fetchone()[0]
        calls, touched, renewals = q['calls'] or 0, q['touched'] or 0, q['renewals'] or 0
        issued = q['issued'] or 0
        rows.append({'name': nm, 'role': role_labels.get(a['role'], a['role']),
                     'calls': calls, 'touched': touched, 'renewals': renewals,
                     'issued': issued, 'escalations': escalations,
                     'rate': round(renewals / touched * 100, 1) if touched else 0})
    rows.sort(key=lambda r: (r['renewals'], r['issued'], r['calls']), reverse=True)
    conn.close()
    return render_template('performance.html', rows=rows, month=month, show_role=is_super)

@app.route('/api/campaign/brand-audit')
def api_campaign_brand_audit():
    """Read-only: the WhatsApp-eligible customers of a brand in the active month, with the fields
    needed to audit premiums/midwife/policy. ?brand=ווינר|גאיה. Token."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    brand = request.args.get('brand', 'ווינר')
    month = active_month()
    if not month:
        return jsonify({'count': 0, 'items': []})
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    items = []
    for r in buckets['whatsapp']:
        if r['brand'] != brand:
            continue
        z = re.sub(r'\D', '', str(r['id_number'] or '')).lstrip('0')
        pol = conn.execute(
            "SELECT pr.policy_number, pr.doc_type_label, pd.received_at, pd.whatsapp_sent_at, pd.email_sent_at "
            "FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
            "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? ORDER BY pd.id DESC LIMIT 1", (z,)).fetchone()
        items.append({'id': r['id'], 'name': r['name'], 'id_number': r['id_number'],
                      'status': r['status'], 'premium_last_year': r['premium_last_year'],
                      'is_midwife': r['is_midwife'],
                      'policy': (dict(pol) if pol else None)})
    conn.close()
    return jsonify({'month': month['name'], 'brand': brand, 'count': len(items), 'items': items})

@app.route('/api/campaign/premium-source')
def api_campaign_premium_source():
    """Read-only: for each no-premium (empty/0) WhatsApp-eligible campaign customer of a brand,
    dump their policy_records (doc_type + extracted premium/total_payment) so the real price can
    be resolved — source (חידוש/חדש) vs endorsement (אינ'). Token."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    brand = request.args.get('brand', 'ווינר')
    month = active_month()
    if not month:
        return jsonify({'count': 0, 'items': []})
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    out = []
    for r in buckets['whatsapp']:
        if r['brand'] != brand or _premium_num(r['premium_last_year']) > 0:
            continue
        z = re.sub(r'\D', '', str(r['id_number'] or '')).lstrip('0')
        recs = conn.execute(
            "SELECT pr.doc_type_label lbl, pr.premium, pr.total_payment, pd.received_at "
            "FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
            "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? ORDER BY pd.id DESC", (z,)).fetchall()
        out.append({'id': r['id'], 'name': r['name'], 'id_number': r['id_number'],
                    'docs': [{'type': x['lbl'], 'premium': x['premium'],
                              'total': x['total_payment'], 'at': x['received_at']} for x in recs]})
    conn.close()
    return jsonify({'brand': brand, 'count': len(out), 'items': out})

@app.route('/api/campaign/resolve-premiums', methods=['POST'])
def api_resolve_premiums_from_docs():
    """Backfill premium_last_year for no-premium campaign customers from their policy docs, per
    Sharon's rules: SOURCE doc (חידוש/חדש) → its premium (+ any PAID endorsement); endorsement
    (אינ') with no price → 750; no doc at all → 750 (flagged). The display band still maps ≤850→750
    and >850→actual. Body {brand, apply}. apply=false → dry-run (shows the plan). Token."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    brand = data.get('brand', 'ווינר')
    apply = bool(data.get('apply'))
    month = active_month()
    if not month:
        return jsonify({'count': 0, 'items': []})
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    out = []
    for r in buckets['whatsapp']:
        if r['brand'] != brand or _premium_num(r['premium_last_year']) > 0:
            continue
        z = re.sub(r'\D', '', str(r['id_number'] or '')).lstrip('0')
        recs = conn.execute(
            "SELECT pr.doc_type_label lbl, pr.premium FROM policy_records pr "
            "JOIN policy_documents pd ON pd.id=pr.policy_document_id "
            "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? ORDER BY pd.id DESC", (z,)).fetchall()
        source = next((_premium_num(x['premium']) for x in recs
                       if x['lbl'] in ('חידוש', 'חדש') and _premium_num(x['premium']) > 0), 0)
        paid_endo = sum(_premium_num(x['premium']) for x in recs
                        if x['lbl'] == "אינ'" and _premium_num(x['premium']) > 0)
        has_endo = any(x['lbl'] == "אינ'" for x in recs)
        if source > 0:
            base, src = source + paid_endo, 'מקור' + (' + תוספת בתשלום' if paid_endo else '')
        elif has_endo:
            base, src = 750 + paid_endo, 'תוספת ללא מחיר → 750'
        else:
            base, src = 750, 'ללא מסמך → 750 (ברירת מחדל)'
        disp = renewal_amount(r['is_midwife'], base)
        flag = (0 < source < 650) or (source == 0 and not has_endo)
        out.append({'id': r['id'], 'name': r['name'], 'id_number': r['id_number'],
                    'source_kind': src, 'resolved_premium': int(base),
                    'will_display': disp, 'flag': flag})
        if apply:
            conn.execute("UPDATE customers SET premium_last_year=? WHERE id=?", (str(int(base)), r['id']))
    if apply:
        conn.commit()
    conn.close()
    return jsonify({'brand': brand, 'applied': apply, 'count': len(out), 'items': out})

@app.route('/api/customer-update', methods=['POST'])
def api_customer_update():
    """Token: bulk-update whitelisted customer fields by customer id. Body {updates:[{id,
    premium_last_year?, is_midwife?, status?}]}. Status changes are logged to the timeline."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    ups = (request.get_json(silent=True) or {}).get('updates') or []
    allowed = ('premium_last_year', 'is_midwife', 'status')
    conn = get_db()
    done = []
    for u in ups:
        cid = u.get('id')
        if not cid:
            continue
        cur = conn.execute("SELECT id_number, status FROM customers WHERE id=?", (cid,)).fetchone()
        if not cur:
            continue
        sets, vals = [], []
        for k in allowed:
            if k in u:
                sets.append(f"{k}=?")
                vals.append(u[k])
        if not sets:
            continue
        vals.append(cid)
        conn.execute(f"UPDATE customers SET {','.join(sets)} WHERE id=?", vals)
        if 'status' in u and u['status'] != (cur['status'] or ''):
            try:
                idk = event_key(normalize_id_number(cur['id_number'] or ''), f"upd-{cid}")
                log_event(conn, idk, f"סטטוס עודכן ל-{u['status']} (עדכון מרוכז)", 'system', kind='status')
            except Exception:
                pass
        done.append(cid)
    conn.commit()
    conn.close()
    return jsonify({'updated': len(done), 'ids': done})

@app.route('/api/fix-master-id', methods=['POST'])
def api_fix_master_id():
    """Token: correct a wrong ת"ז on the insured master (+ any customer rows) — old→new. Used when
    a renewal file carries the VALID ת"ז but the master holds a bad legacy one (same person, matched
    by name+phone). Body {pairs:[{old,new}]}. Guards: 'new' must be free (not held by a DIFFERENT
    insured), 'old' must exist. Returns per-pair result."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    pairs = (request.get_json(silent=True) or {}).get('pairs') or []
    conn = get_db()
    out = []
    for p in pairs:
        old = re.sub(r'\D', '', str(p.get('old', ''))).lstrip('0')
        new = re.sub(r'\D', '', str(p.get('new', '')))
        newz = new.lstrip('0')
        if not old or not new:
            out.append({'old': p.get('old'), 'result': 'skip: missing old/new'})
            continue
        src = conn.execute("SELECT name FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=?", (old,)).fetchone()
        if not src:
            out.append({'old': old, 'result': 'skip: old not in master'})
            continue
        clash = conn.execute("SELECT name FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=? "
                             "AND ltrim(COALESCE(id_number,''),'0')!=?", (newz, old)).fetchone()
        if clash:
            out.append({'old': old, 'new': new, 'result': f'ABORT: new ת"ז held by {clash["name"]}'})
            continue
        ni = conn.execute("UPDATE insureds SET id_number=? WHERE ltrim(COALESCE(id_number,''),'0')=?",
                          (new, old)).rowcount
        nc = conn.execute("UPDATE customers SET id_number=? WHERE ltrim(COALESCE(id_number,''),'0')=?",
                          (new, old)).rowcount
        out.append({'name': src['name'], 'old': old, 'new': new,
                    'result': f'ok: insureds={ni}, customers={nc}'})
    conn.commit()
    conn.close()
    return jsonify({'pairs': out})

@app.route('/api/rep-performance')
def api_rep_performance():
    """Token-authed: per-rep (role='agent') performance for the ACTIVE month + a ready-to-send
    Hebrew WhatsApp message. Feeds the daily 08:00 manager broadcast. Managers/super-admins are
    the RECIPIENTS — their own activity is not included (only agents)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    month = active_month()
    mid = month['id'] if month else -1
    key = "COALESCE(NULLIF(ltrim(COALESCE(id_number,''),'0'),''), 'r'||id)"
    rows = []
    for a in conn.execute("SELECT display_name FROM users WHERE role='agent' "
                          "AND COALESCE(display_name,'')!='' ORDER BY display_name").fetchall():
        nm = a['display_name']
        q = conn.execute(
            "SELECT "
            "SUM((CASE WHEN call_by_1=? THEN 1 ELSE 0 END)+(CASE WHEN call_by_2=? THEN 1 ELSE 0 END)+(CASE WHEN call_by_3=? THEN 1 ELSE 0 END)) AS calls, "
            f"COUNT(DISTINCT CASE WHEN handled_by=? THEN {key} END) AS touched, "
            f"COUNT(DISTINCT CASE WHEN handled_by=? AND status IN (?,?) THEN {key} END) AS renewals, "
            f"COUNT(DISTINCT CASE WHEN handled_by=? AND status=? THEN {key} END) AS issued "
            "FROM customers WHERE month_id=?",
            [nm, nm, nm, nm, nm, 'חודש', 'חודש - בוצעה שיחת מכירה', nm, 'הופק', mid]).fetchone()
        rows.append({'name': nm, 'calls': q['calls'] or 0, 'touched': q['touched'] or 0,
                     'renewals': q['renewals'] or 0, 'issued': q['issued'] or 0})
    conn.close()
    rows.sort(key=lambda r: (r['renewals'], r['issued'], r['calls']), reverse=True)
    today = datetime.date.today().strftime('%d/%m/%Y')
    lines = [f"📊 ביצועי נציגים — {month['name'] if month else '—'}", f"🗓️ {today}", ""]
    if rows:
        for i, r in enumerate(rows, 1):
            lines.append(f"{i}. {r['name']} — חודשו {r['renewals']} | הופקו {r['issued']} | "
                         f"טופלו {r['touched']} | שיחות {r['calls']}")
        lines += ["", (f"סה\"כ: חודשו {sum(r['renewals'] for r in rows)} | "
                       f"הופקו {sum(r['issued'] for r in rows)} | "
                       f"טופלו {sum(r['touched'] for r in rows)} | "
                       f"שיחות {sum(r['calls'] for r in rows)}")]
    else:
        lines.append("אין נציגים מוגדרים במערכת.")
    return jsonify({'month': month['name'] if month else None, 'count': len(rows),
                    'rows': rows, 'message': "\n".join(lines)})

def _extract_renewal_rows(ws, source):
    """Light extraction (id/name/phone) from a renewal sheet — for the pre-load VALIDATION
    only (does not insert). Mirrors the header detection of the real importers."""
    all_rows = list(ws.iter_rows(values_only=True))
    if source == 'ofir':
        hdr_i = next((i for i, r in enumerate(all_rows[:12])
                      if r and any('מבוטח' in str(c or '') for c in r)), 4)
        headers = [str(c).strip() if c else '' for c in all_rows[hdr_i]]
        def cidx(name):
            return next((i for i, h in enumerate(headers) if h == name or name in h), None)
        id_i, nm_i, ph_i = cidx('זהות'), cidx('מבוטח'), cidx('טלפון')
    else:
        hdr_i = next((i for i, r in enumerate(all_rows) if r and 'פוליסה' in str(r[0])), 2)
        headers = [str(c).strip() if c else '' for c in all_rows[hdr_i]]
        def cidx(name):
            return next((i for i, h in enumerate(headers) if name in h), None)
        id_i, nm_i, ph_i = cidx('ת.ז'), cidx('שם'), cidx('טלפון')
    out = []
    for r in all_rows[hdr_i + 1:]:
        if not r:
            continue
        def g(i):
            return r[i] if (i is not None and len(r) > i) else None
        idn = normalize_id_number(g(id_i))
        if not idn or len(idn) < 5:
            continue
        out.append({'id': idn, 'name': str(g(nm_i) or '').strip(),
                    'phone': re.sub(r'\D', '', str(g(ph_i) or ''))})
    return out


def _tok_name(n):
    n = re.sub(r'["\'׳״]', '', str(n or '')); n = re.sub(r'[-]', ' ', n)
    return frozenset(t for t in n.split() if len(t) >= 2)


def _validate_renewal_file(conn, rows):
    """Compare renewal rows against the insureds master (by ת"ז). Categorises renewal vs
    new, counts phone conflicts, and flags likely duplicates (a 'new' row whose NAME matches
    an existing insured under a different id — the wrong-id case)."""
    def z(s):
        return re.sub(r'\D', '', str(s or '')).lstrip('0')
    ins, ins_by_name = {}, {}
    for r in conn.execute("SELECT id_number, name, phone FROM insureds"):
        k = z(r['id_number'])
        ins[k] = {'name': r['name'], 'phone': re.sub(r'\D', '', str(r['phone'] or ''))}
        ins_by_name.setdefault(_tok_name(r['name']), []).append(k)
    file_ids = {z(row['id']) for row in rows if z(row['id'])}
    renewals = new = phone_conflict = 0
    dupes = []
    for row in rows:
        k = z(row['id'])
        if not k:
            continue
        if k in ins:
            renewals += 1
            fp, ip = row['phone'], ins[k]['phone']
            if fp and ip and fp[-9:] != ip[-9:]:
                phone_conflict += 1
        else:
            new += 1
            for c in ins_by_name.get(_tok_name(row['name']), []):
                if c in file_ids:
                    continue
                pm = bool(row['phone'] and ins[c]['phone'] and ins[c]['phone'][-9:] == row['phone'][-9:])
                dupes.append({'name': row['name'], 'file_id': row['id'], 'master_id': c, 'phone_match': pm})
                break
    return {'total': len(rows), 'renewals': renewals, 'new': new,
            'phone_conflicts': phone_conflict, 'potential_dupes': dupes}


def _month_activity_summary(cust, month_name):
    """One-line summary of a customer's month activity (status / WhatsApp / calls / note),
    or None if there was none — used to preserve activity in the by-ת"ז event log."""
    parts = []
    st = (cust['status'] or '').strip()
    if st:
        parts.append(st)
    wa = (cust['whatsapp_sent_date'] or '').strip()
    if wa:
        try:
            wa = datetime.datetime.strptime(wa[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            pass
        parts.append(f"וואטסאפ נשלח {wa}")
    calls = sum(1 for i in (1, 2, 3) if (cust[f'call_status_{i}'] or '').strip())
    if calls:
        parts.append(f"{calls} ניסיונות קשר")
    note = (cust['agent_notes'] or '').strip()
    if not parts and not note:
        return None
    text = f"פעילות {month_name}"
    if parts:
        text += ': ' + ' · '.join(parts)
    if note:
        text += f" — {note}"
    return text


def _backfill_activity_for(conn, month_id, month_name, dry=False):
    """Log a one-line activity-summary event (by ת"ז) for every customer of `month_id` that
    had activity — so the event log follows the person across months. Idempotent (skips a
    customer that already has a 'פעילות <month>' event). Returns (logged, skipped, samples)."""
    logged = skipped = 0
    samples = []
    for cust in conn.execute("SELECT * FROM customers WHERE month_id=?", (month_id,)).fetchall():
        summary = _month_activity_summary(cust, month_name)
        if not summary:
            continue
        z = event_key(cust['id_number'], '')
        if not z:
            continue
        if conn.execute("SELECT 1 FROM client_events WHERE idkey=? AND note LIKE ?",
                        (z, f"פעילות {month_name}%")).fetchone():
            skipped += 1
            continue
        if not dry:
            log_event(conn, z, summary, 'מערכת (סיכום חודש)', kind='month_activity')
        logged += 1
        if len(samples) < 8:
            samples.append({'name': cust['name'], 'summary': summary})
    return logged, skipped, samples


def _apply_import(conn, wb, source, month_name):
    """Commit a staged renewal file as a NEW month (option ב — each load is its own month).
    The current active month's customers are promoted into the master and then KEPT (the month
    is only deactivated, not deleted) so late renewals can still be worked from ניהול; the fresh
    rows load into a brand-new active month."""
    ws = wb.active
    source_brands = ['אופיר'] if source == 'ofir' else ['גאיה', 'ווינר']
    now = datetime.datetime.now().isoformat()
    prev = conn.execute("SELECT * FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    promoted = 0
    if prev:
        # Promote the outgoing month into the master, then archive it (keep its customers).
        promoted = promote_customers_to_insureds(conn, prev['id'], brands=source_brands)
        # Preserve the outgoing month's activity in the by-ת"ז event log (follows the person).
        _backfill_activity_for(conn, prev['id'], prev['name'])
    conn.execute("UPDATE months SET is_active=0")
    conn.execute("INSERT INTO months (name, created_at, is_active) VALUES (?,?,1)",
                 (month_name or 'חודש חדש', now))
    month_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    count = _import_ofir(conn, ws, month_id) if source == 'ofir' else _import_gaia_winner(conn, ws, month_id)
    conn.commit()
    label = 'אופיר' if source == 'ofir' else 'גאיה/ווינר'
    return count, promoted, label


@app.route('/admin/import', methods=['POST'])
@login_required
@superadmin_required
def import_excel():
    """Phase 1 — STAGE a renewal file: parse + validate against the master, store it as a
    pending import. Nothing is applied until the 'שינוי חידוש לחודש חדש' (commit) button."""
    f = request.files.get('file')
    month_name = request.form.get('month_name', '').strip()
    if not f or not month_name:
        flash('חסר קובץ או שם חודש', 'danger')
        return redirect(url_for('admin'))
    source = request.form.get('source', 'gaia_winner')
    try:
        blob = f.read()
        wb = load_workbook(io.BytesIO(blob), data_only=True)
        rows = _extract_renewal_rows(wb.active, source)
        conn = get_db()
        report = _validate_renewal_file(conn, rows)
        conn.execute("""INSERT INTO pending_imports
            (source, month_name, filename, file_blob, report_json, uploaded_at, uploaded_by, status)
            VALUES (?,?,?,?,?,?,?, 'pending')""",
            (source, month_name, f.filename, blob, json.dumps(report, ensure_ascii=False),
             datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
             session.get('display_name') or session.get('username', '')))
        conn.commit()
        conn.close()
        flash(f"הקובץ נטען לבדיקה: {report['total']} חידושים בקובץ · "
              f"{report['renewals']} מוכרים במערכת · {report['new']} חדשים למערכת · "
              f"{report['phone_conflicts']} התנגשויות טלפון · {len(report['potential_dupes'])} כפילויות אפשריות. "
              f"בדוק ולחץ 'שינוי חידוש לחודש חדש'.", 'info')
    except Exception as e:
        flash(f'שגיאה בבדיקת הקובץ: {e}', 'danger')
    return redirect(url_for('admin'))


def _snapshot_db(tag='import'):
    """Take a consistent point-in-time copy of the live DB (a rollback point) into the /data volume,
    then prune old snapshots so they never fill the disk. Uses the SQLite backup API (safe on a live
    DB). Returns the snapshot filename, or None on failure. The DB is small (~6MB) so copies are cheap."""
    try:
        bdir = os.path.dirname(DB_PATH) or '.'
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'snapshot_{tag}_{ts}.db'
        path = os.path.join(bdir, fname)
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(path)
        with dst:
            src.backup(dst)
        dst.close(); src.close()
        # Retention: keep the newest 10 snapshots + the newest 14 daily pre-migration backups.
        def _prune(prefix, keep):
            try:
                files = sorted(f for f in os.listdir(bdir)
                               if f.startswith(prefix) and f.endswith('.db'))
                for old in files[:-keep]:
                    try: os.remove(os.path.join(bdir, old))
                    except OSError: pass
            except OSError:
                pass
        _prune('snapshot_', 10)
        _prune('renewals_backup_', 14)
        return fname
    except Exception as e:
        print(f'[snapshot] failed: {e}')
        return None

def _list_snapshots():
    bdir = os.path.dirname(DB_PATH) or '.'
    out = []
    try:
        for f in os.listdir(bdir):
            if (f.startswith('snapshot_') or f.startswith('renewals_backup_')) and f.endswith('.db'):
                fp = os.path.join(bdir, f)
                st = os.stat(fp)
                out.append({'file': f, 'size_mb': round(st.st_size / 1024 / 1024, 1),
                            'created': datetime.datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M')})
    except OSError:
        pass
    return sorted(out, key=lambda x: x['created'], reverse=True)

@app.route('/api/admin/db-snapshots')
def api_db_snapshots():
    """Token: list DB rollback points (pre-import snapshots + daily backups) on the volume."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    return jsonify({'snapshots': _list_snapshots()})

@app.route('/api/admin/db-snapshot', methods=['POST'])
def api_db_snapshot():
    """Token: take a manual rollback point now. Body {tag?}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    tag = re.sub(r'[^A-Za-z0-9_]', '', str((request.get_json(silent=True) or {}).get('tag', 'manual')))[:24] or 'manual'
    fn = _snapshot_db(tag)
    return (jsonify({'ok': True, 'snapshot': fn}) if fn else (jsonify({'error': 'snapshot failed'}), 500))

@app.route('/api/admin/db-restore', methods=['POST'])
def api_db_restore():
    """Token: restore the live DB from a named snapshot on the volume (a safety snapshot of the
    CURRENT state is taken first). Body {file}. Only *.db files in the DB dir are allowed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    base = os.path.basename(str((request.get_json(silent=True) or {}).get('file', '')))
    bdir = os.path.dirname(DB_PATH) or '.'
    src_path = os.path.join(bdir, base)
    if not base.endswith('.db') or not os.path.exists(src_path):
        return jsonify({'error': 'snapshot not found'}), 404
    safety = _snapshot_db('pre_restore')
    try:
        src = sqlite3.connect(src_path)
        dst = sqlite3.connect(DB_PATH)
        with dst:
            src.backup(dst)
        dst.close(); src.close()
    except Exception as e:
        return jsonify({'error': f'restore failed: {e}'}), 500
    return jsonify({'ok': True, 'restored_from': base, 'safety_snapshot': safety})

@app.route('/admin/import/commit/<int:pid>', methods=['POST'])
@login_required
@superadmin_required
def import_commit(pid):
    """Phase 2 — apply a validated pending import ('שינוי חידוש לחודש חדש')."""
    conn = get_db()
    p = conn.execute("SELECT * FROM pending_imports WHERE id=? AND status='pending'", (pid,)).fetchone()
    if not p:
        conn.close()
        flash('לא נמצאה טעינה ממתינה', 'danger')
        return redirect(url_for('admin'))
    _snapshot_db('before_import')  # rollback point BEFORE the month transition is applied
    try:
        wb = load_workbook(io.BytesIO(p['file_blob']), data_only=True)
        count, promoted, label = _apply_import(conn, wb, p['source'], p['month_name'])
        conn.execute("UPDATE pending_imports SET status='committed' WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        # Auto-enrich the newly-loaded customers (SITE123 + stored contact files, by ת"ז),
        # in the background so the commit returns immediately.
        threading.Thread(target=_run_post_load_enrichment, daemon=True).start()
        msg = f'בוצע: נטענו {count} חידושים ({label})'
        if promoted:
            msg += f' · {promoted} לקוחות קודמים עברו ל"כל הלקוחות"'
        msg += ' · העשרת מיילים (SITE123 + קבצים) רצה ברקע'
        flash(msg, 'success')
    except Exception as e:
        conn.close()
        flash(f'שגיאה בביצוע: {e}', 'danger')
    return redirect(url_for('admin'))


@app.route('/admin/import/cancel/<int:pid>', methods=['POST'])
@login_required
@superadmin_required
def import_cancel(pid):
    conn = get_db()
    conn.execute("UPDATE pending_imports SET status='cancelled' WHERE id=? AND status='pending'", (pid,))
    conn.commit()
    conn.close()
    flash('הטעינה הממתינה בוטלה', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/sample-format')
@login_required
@superadmin_required
def sample_format():
    """A blank Gaia/Winner import template with the expected column headers, plus one
    example row, so a new file can be pasted into the right shape."""
    from io import BytesIO
    headers = ['פוליסה', 'שם', 'ת.ז', 'טלפון', 'מותג', 'סטטוס', 'פרמיה', 'מיילדות']
    example = ['881400123456', 'ישראל ישראלי', '012345678', '0501234567', 'ווינר', '', '1200', 'V']
    wb = NewWorkbook()
    ws = wb.active
    ws.title = 'חידושים'
    ws.append(headers)
    ws.append(example)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(h) + 4)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='פורמט_טעינת_חידושים.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Map raw sheet statuses to the system's canonical status values.
IMPORT_STATUS_MAP = {
    'חודש': 'חודש',
    'לא חודש': '',
    'לא התחיל': '',
    'לא מחדש': 'לא רוצים לחדש',
    'לא רוצים לחדש': 'לא רוצים לחדש',
    'לקוח ענה/ V כחול': 'נוצר קשר עם לקוח',  # legacy sheets
    'נוצר קשר עם לקוח': 'נוצר קשר עם לקוח',
}


def _ofir_status(raw):
    """Ofir has its own status set — keep the sheet value as-is (only the unstarted
    'לא התחיל' collapses to the empty/default state)."""
    raw = (raw or '').strip()
    return '' if raw in ('', 'לא התחיל') else raw


def _import_gaia_winner(conn, ws, month_id):
    """Gaia/Winner export: header row contains 'פוליסה' in col A; brand comes from the
    'מותג' column so a single file may hold both Gaia and Winner rows."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and 'פוליסה' in str(row[0]):
            header_row = i
            break
    if not header_row:
        header_row = 3
    headers = [str(c).strip() if c else '' for c in
               list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]]

    def col(name, row_vals):
        try:
            idx = next(i for i, h in enumerate(headers) if name in h)
            return str(row_vals[idx]).strip() if row_vals[idx] is not None else ''
        except StopIteration:
            return ''

    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        policy = str(row[0]).strip() if row[0] else ''
        if not policy or policy in ('None', ''):
            continue
        # Each year the policy number's last digit advances (5→6, 6→7, …).
        if policy.isdigit():
            policy = str(int(policy) + 1)
        name = col('שם', row)
        if not name or name == 'None':
            continue
        row_brand = col('מותג', row)
        # 'מיילדות' column: any mark (V / ✓ / כן / 1) flags a midwife (Winner only).
        mw = col('מיילדות', row).strip().lower()
        is_midwife = 1 if (mw and mw not in ('0', 'לא', 'no', 'false', 'none', '-', '—')) else 0
        conn.execute("""
            INSERT INTO customers
            (month_id, policy_number, name, id_number, phone, brand, status,
             premium_last_year, whatsapp_sent_date, sharon_notes, requests_to_sharon,
             contact_date, agent_notes, interested_in_products, whatsapp_source, is_midwife)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            month_id, policy, name,
            normalize_id_number(col('ת.ז', row)), col('טלפון', row), row_brand,
            IMPORT_STATUS_MAP.get(col('סטטוס', row), ''),
            col('פרמיה', row), col('וואטסאפ', row), col('הערות שרון', row),
            col('בקשות משרון', row), col('תאריך התקשרות', row),
            col('הערות חידושים', row), col('מתעניין', row),
            'ווינר' if row_brand == 'אופיר' else None,
            is_midwife
        ))
        count += 1
    return count


def _import_ofir(conn, ws, month_id):
    """Ofir/Meir book: header on row 5, elementary/car columns. brand is always 'אופיר';
    the extra fields (insurer, coverage breakdown, license, sector, …) are captured."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        if row and any('מבוטח' in str(c or '') for c in row):
            header_row = i
            break
    if not header_row:
        header_row = 5
    headers = [str(c).strip() if c else '' for c in
               list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]]

    def col(name, row_vals):
        try:
            idx = next(i for i, h in enumerate(headers) if h == name or name in h)
            v = row_vals[idx]
            return str(v).strip() if v is not None else ''
        except (StopIteration, IndexError):
            return ''

    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        name = col('מבוטח', row)
        policy = col('פוליסה', row)
        if (not name or name == 'None') and not policy:
            continue
        conn.execute("""
            INSERT INTO customers
            (month_id, policy_number, name, id_number, phone, email, brand, status,
             premium_last_year, agent_notes,
             insurer, sector, license_number, secondary_status,
             cover_third_party, cover_compulsory, cover_comprehensive, cover_riders,
             sum_insured, offer_company, done_company, handler, sub_agent)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            month_id, policy, name,
            normalize_id_number(col('זהות', row)), col('טלפון', row), col('Email', row),
            'אופיר', _ofir_status(col('סטטוס ראשוני', row)),
            col('פרמיה', row), col('הערות ועדכונים', row),
            col('חברה', row), col('ענף', row), col('רשוי', row), col('סטטוס משני', row),
            col("צד ג'", row), col('חובה', row), col('מקיף', row), col('ריידרים', row),
            col('ס/מ', row), col('חברת ההצעה', row), col('חברה שנעשה', row),
            col('מטפל', row), col('סוכן', row),
        ))
        count += 1
    return count

@app.route('/admin/users/add', methods=['POST'])
@login_required
@superadmin_required
def add_user():
    username = request.form['username'].strip()
    display_name = request.form['display_name'].strip()
    password = request.form['password']
    role = request.form.get('role', 'agent')
    email = request.form.get('email', '').strip()
    brands = [b for b in request.form.getlist('brands') if b in BRANDS]
    try:
        conn = get_db()
        conn.execute("INSERT INTO users (username, password_hash, display_name, role, email) VALUES (?,?,?,?,?)",
                     (username, generate_password_hash(password), display_name, role, email))
        uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Super-admins implicitly see everything; managers and agents are agency-scoped.
        if role != 'superadmin':
            for b in brands:
                conn.execute("INSERT OR IGNORE INTO user_brands (user_id, brand) VALUES (?,?)", (uid, b))
        conn.commit()
        conn.close()
        flash(f'משתמש {display_name} נוצר', 'success')
    except Exception as e:
        flash(f'שגיאה: {e}', 'danger')
    return redirect(url_for('admin'))


@app.route('/admin/users/<int:uid>/brands', methods=['POST'])
@login_required
@superadmin_required
def set_user_brands(uid):
    """Replace a user's agency access with the submitted set."""
    brands = [b for b in request.form.getlist('brands') if b in BRANDS]
    conn = get_db()
    conn.execute("DELETE FROM user_brands WHERE user_id=?", (uid,))
    for b in brands:
        conn.execute("INSERT OR IGNORE INTO user_brands (user_id, brand) VALUES (?,?)", (uid, b))
    conn.commit()
    conn.close()
    # If the edited user is logged in, their session cache refreshes on next login;
    # drop our own cache if we edited ourselves (harmless for admins).
    if uid == session.get('user_id'):
        session.pop('brands', None)
    flash('הרשאות הסוכנות עודכנו', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/users/delete/<int:uid>', methods=['POST'])
@login_required
@superadmin_required
def delete_user(uid):
    if uid == session['user_id']:
        flash('לא ניתן למחוק את עצמך', 'danger')
        return redirect(url_for('admin'))
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    flash('משתמש נמחק', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/users/reset-password/<int:uid>', methods=['POST'])
@login_required
@superadmin_required
def reset_password(uid):
    new_pass = request.form['new_password']
    conn = get_db()
    conn.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pass), uid))
    conn.commit()
    conn.close()
    flash('סיסמה עודכנה', 'success')
    return redirect(url_for('admin'))

@app.route('/export/wasender')
@login_required
@admin_required
def export_wasender():
    month = active_month()
    if not month:
        flash('אין חודש פעיל', 'danger')
        return redirect(url_for('index'))

    brand_filter = request.args.get('brand', '')
    mark_sent = request.args.get('mark_sent', '0') == '1'
    # mode: 'first' = all without whatsapp sent | 'reminder' = didn't renew and don't want to cancel
    mode = request.args.get('mode', 'first')

    conn = get_db()

    if mode == 'first':
        # First send: everyone who hasn't received WhatsApp yet
        query = """SELECT id, name, phone FROM customers
                   WHERE month_id=? AND (whatsapp_sent_date IS NULL OR whatsapp_sent_date='')"""
    else:
        # Reminder: only those who haven't renewed and didn't say they don't want to renew
        query = """SELECT id, name, phone FROM customers
                   WHERE month_id=? AND (status IS NULL OR status='' OR status='נוצר קשר עם לקוח')"""

    params = [month['id']]
    if brand_filter:
        query += " AND brand=?"
        params.append(brand_filter)
    query += " ORDER BY name"

    rows = conn.execute(query, params).fetchall()

    wb = NewWorkbook()
    ws = wb.active
    ws.title = 'WASender'
    ws.append(['phone', 'name'])

    today = datetime.date.today().isoformat()
    ids = []
    for r in rows:
        phone = str(r['phone']).replace('-', '').replace(' ', '')
        if phone.startswith('0'):
            phone = '972' + phone[1:]
        ws.append([phone, r['name']])
        ids.append(r['id'])

    if mark_sent and ids:
        placeholders = ','.join('?' * len(ids))
        conn.execute(f"UPDATE customers SET whatsapp_sent_date=? WHERE id IN ({placeholders})",
                     [today] + ids)
        conn.commit()

    conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    mode_label = 'ראשונה' if mode == 'first' else 'תזכורת'
    filename = f"wasender_{mode_label}_{month['name'].replace(' ','_')}_{today}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Renewal campaign: eligibility, pricing, email rendering + sending, logging ──
GAIA_RENEW = 'https://www.gaia-ins.co.il/renew'
WINNER_RENEW = 'https://www.winner-ins.co.il/renew'
MIDWIFE_RENEW = 'https://www.winner-ins.co.il/renew/midwife'
HEB_MONTHS = ['', 'ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני', 'יולי',
              'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר']

# Statuses that mean "no need to send anymore": renewed / not-renewing / handled manually.
CAMPAIGN_STOP_STATUSES = {
    'חודש', 'חודש - בוצעה שיחת מכירה', 'טופס התקבל', 'הלקוח אישר', 'ביקשו לחדש לבד',
    'לא רוצים לחדש', 'לא מחדש', 'בוטל', 'פרוייקט הסתיים',
    'דורש בירור', 'ממתין לאישור מיילדות', 'המשך טיפול בוואטסאפ', 'ממתין לחידוש',
    # New-business statuses — a fresh purchase / a lead awaiting issuance is NOT a renewal,
    # so it must never receive the renewal-reminder campaign.
    'הופק', 'ממתין להפקה',
    # Contact already made (incl. customers who replied to the campaign email) — don't also
    # hit them with the (WhatsApp) reminder.
    'נוצר קשר עם לקוח',
    # Renewal in progress but stuck on a collection/billing problem — already being handled,
    # so no "please renew" reminder.
    'חידוש בעיות גביה',
    # Form received but the payment method was invalid — the customer is already getting the
    # dedicated "update your card" message, so keep the renewal reminder off.
    'התקבל חידוש - כ.א לא תקין',
}

CAMPAIGN_CROSS_SELL = """
  <hr style="border:none;border-top:1px solid #eee;margin:22px 0">
  <p>אגב, יש לי עוד משהו קטן שכדאי שתכירו.<br>
  מי שמכיר אותי יודע שהייתי מטפל בעצמי — עד שהידיים שלי פשוט הפסיקו לעבוד. באותו רגע גיליתי כמה אני חשוף כלכלית, ומשם נולד <strong>"חוסן למחר"</strong> — כיסוי ייעודי שבנינו עם הראל בדיוק בשביל מטפלים כמוכם.</p>
  <p><strong>מה מיוחד בו:</strong></p>
  <ul>
    <li>במקרה של אובדן כושר עבודה למקצוע שלכם (לא עיסוק כללי), אחרי חצי שנה מקבלים פיצוי חד-פעמי</li>
    <li>בלי צורך להוכיח כמה הרווחתם — קובעים מראש את סכום הביטוח</li>
    <li>נמכר כתוספת לביטוח חיים</li>
    <li>החל מ-9 ₪ לחודש בלבד לכיסוי חוסן למחר</li>
  </ul>
  <p>זה נספח קטן, אבל יכול לעשות הבדל גדול ביום שהכי תצטרכו אותו. אם תרצו לצרף אותו לחידוש — רק תגידו לנו ונחבר אתכם למנהל התחום.</p>"""

def _seasonal_line():
    """Closing line for every service message I send (not the bot). During the High-Holidays
    window (20 Aug–15 Oct) it's a Rosh-Hashana greeting; the rest of the year the regular line.
    Auto-reverts. Shared by all policy/cert/copy/service email + free-text bodies."""
    md = (datetime.date.today().month, datetime.date.today().day)
    return 'שתהיה לך שנה טובה, בטוחה ומתוקה,' if (8, 20) <= md <= (10, 15) else 'המשך יום נפלא,'

def _seasonal_signoff():
    """HTML variant — a <p> line for RTL email bodies."""
    return f'\n  <p>{_seasonal_line()}</p>'

# Signature agency name per brand (name + registration year), shown in the renewal
# message signature. Gaia/Winner are the active campaign brands.
AGENCY_SIGNATURE = {
    'ווינר': 'ווינר סוכנות לביטוח (2009) בע"מ',
    'גאיה': 'גאיה סוכנות לביטוח (2019) בע"מ',
}
def _agency_name(brand):
    return AGENCY_SIGNATURE.get(brand, AGENCY_SIGNATURE['גאיה'])

def _premium_num(v):
    d = re.sub(r'[^\d.]', '', str(v or ''))
    return float(d) if d else 0

def renewal_amount(is_midwife, premium):
    """Renewal price shown in the message. Standard renewal is a flat 750 — any premium in the
    normal band (≤850) maps to it (so noisy last-year premiums like 777/51 still show 750). The
    few genuine exceptions whose renewal price is really higher show their ACTUAL premium
    (Sharon's rule: pull the non-750 few and send the real price). Empty/0 or an implausible
    value (>3000) → None → generic 'like last year'. Midwives keep their tier (1200/1600)."""
    p = _premium_num(premium)
    if p <= 0:
        return None
    if is_midwife:
        return 1600 if 1500 <= p <= 1700 else 1200
    if p <= 850:
        return 750
    if p <= 3000:
        return int(p)
    return None

def renewal_link(brand, is_midwife):
    if is_midwife:
        return MIDWIFE_RENEW, 'חידוש מיילדות'
    if brand == 'גאיה':
        return GAIA_RENEW, 'חידוש גאיה'
    return WINNER_RENEW, 'חידוש ווינר'

def _campaign_email_for(conn, cust, _ins_email=None):
    """Resolve a customer's email — renewal files carry no email, so fall back to the
    insureds master by ת"ז."""
    ce = str(cust['email'] or '').strip() if 'email' in cust.keys() else ''
    if ce and '@' in ce:
        return ce
    idn = re.sub(r'\D', '', str(cust['id_number'] or '')).lstrip('0')
    if not idn:
        return ''
    if _ins_email is not None:
        return _ins_email.get(idn, '')
    r = conn.execute("SELECT email FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=? "
                     "AND COALESCE(email,'') LIKE '%@%' LIMIT 1", (idn,)).fetchone()
    return (r['email'].strip() if r and r['email'] else '')

# A new-business policy received within this many days marks the customer as a current-cycle
# NEW customer → excluded from the renewal campaign (see campaign_eligibility).
NEW_BIZ_EXCLUDE_DAYS = 90

def campaign_eligibility(conn, month_id):
    """Bucket the active month's customers. Auto-send targets Gaia/Winner, non-midwife,
    non-VIP, no stop-status. Midwives wait for manual approval; VIPs/stop-statuses excluded."""
    ins_email = {re.sub(r'\D', '', str(r['id_number'] or '')).lstrip('0'): (r['email'] or '').strip()
                 for r in conn.execute("SELECT id_number, email FROM insureds")}
    # ת"ז that bought a NEW-business policy THIS CYCLE (doc received in the last 90 days) → they
    # are new customers now, not renewals, so they must NEVER get a renewal reminder — regardless
    # of a stale/empty status. Time-bounded so a customer who bought new in a PRIOR cycle and is
    # legitimately renewing now is NOT wrongly excluded (Sharon's choice: option ב).
    nb_cutoff = (datetime.datetime.now() - datetime.timedelta(days=NEW_BIZ_EXCLUDE_DAYS)
                 ).strftime('%Y-%m-%d %H:%M')
    new_biz_ids = set()
    for pr in conn.execute(
            "SELECT pr.insured_id, pr.doc_type_label FROM policy_records pr "
            "JOIN policy_documents pd ON pd.id = pr.policy_document_id "
            "WHERE pd.received_at >= ?", (nb_cutoff,)).fetchall():
        if is_new_doc(pr['doc_type_label']):
            k = re.sub(r'\D', '', str(pr['insured_id'] or '')).lstrip('0')
            if k:
                new_biz_ids.add(k)
    rows = conn.execute("SELECT * FROM customers WHERE month_id=?", (month_id,)).fetchall()
    b = {'email': [], 'whatsapp': [], 'midwife_pending': [], 'vip': [],
         'status_excluded': [], 'no_contact': [], 'ofir': [], 'new_biz': []}
    for r in rows:
        if r['brand'] not in ('גאיה', 'ווינר'):
            b['ofir'].append(r); continue
        if r['is_midwife']:
            b['midwife_pending'].append(r); continue
        if r['is_vip']:
            b['vip'].append(r); continue
        idk = re.sub(r'\D', '', str(r['id_number'] or '')).lstrip('0')
        if idk and idk in new_biz_ids:
            b['new_biz'].append(r); continue
        # New business by SOURCE — join_form/harel_proposal leads (no issued policy yet, so the
        # 90-day policy-record check above misses them) + issued new_policy rows. These are never
        # renewals, so they must never get a renewal reminder whatever their work status.
        if ('import_source' in r.keys() and (r['import_source'] or '') in NEW_BUSINESS_SOURCES):
            b['new_biz'].append(r); continue
        # Group-owner customers (e.g. Aviram's therapists) renew via the owner_renewal_confirm
        # button flow to the owner, NOT the standard campaign — exclude from the auto-send.
        if ('group_owner' in r.keys() and (r['group_owner'] or '').strip()):
            b['status_excluded'].append(r); continue
        if (r['status'] or '') in CAMPAIGN_STOP_STATUSES or \
           ('do_not_contact' in r.keys() and r['do_not_contact']):
            b['status_excluded'].append(r); continue
        phone = re.sub(r'\D', '', str(r['phone'] or ''))
        email = _campaign_email_for(conn, r, ins_email)
        touched = False
        if email and '@' in email:
            b['email'].append((r, email)); touched = True
        if phone:
            b['whatsapp'].append(r); touched = True
        if not touched:
            b['no_contact'].append(r)
    return b

def _price_line(is_midwife, premium):
    amt = renewal_amount(is_midwife, premium)
    if amt is None:
        return 'המחיר נשאר כמו שנה שעברה.'
    return f'המחיר לשנה הקרובה: <strong>{amt:,} ₪</strong> — ללא שינוי מהשנה שעברה.'

EMAIL_LOGO = {
    'גאיה': 'https://gaia-ins.co.il/logo.png',
    'ווינר': 'https://winner-ins.co.il/logo-mark.png',
}

def render_renewal_email(cust, month_name):
    import html as _html
    link, label = renewal_link(cust['brand'], cust['is_midwife'])
    name = _html.escape(str(cust['name'] or ''))
    logo = EMAIL_LOGO.get(cust['brand'], '')
    logo_html = (f'<div style="margin-bottom:12px"><img src="{logo}" alt="" '
                 f'style="height:56px;width:auto;border:0"></div>') if logo else ''
    return (f'<table dir="rtl" width="100%" cellpadding="0" cellspacing="0" style="direction:rtl">'
            f'<tr><td align="right" style="padding:0">'
            f'<div style="display:inline-block;max-width:640px;text-align:right;direction:rtl;'
            f'font-family:Arial,sans-serif;font-size:15px;line-height:1.7;color:#222">'
            f'{logo_html}'
            f'<p>שלום, {name}</p>'
            f'<p>הפוליסה המקצועית שלך מסתיימת בסוף חודש {month_name}.</p>'
            f'<p>{_price_line(cust["is_midwife"], cust["premium_last_year"])}</p>'
            f'<p>לחידוש הפוליסה וצפייה בתנאים העדכניים (שלא השתנו), יש להיכנס לקישור:</p>'
            f'<p><a href="{link}" style="background:#0d6efd;color:#fff;padding:10px 22px;'
            f'border-radius:6px;text-decoration:none;display:inline-block">{label}</a></p>'
            f'<p>נשמח להמשיך ולהעניק לך את השירות והליווי המקצועי.</p>'
            + ('<p style="color:#555;margin-top:16px">—<br>שרון דר,<br>'
               'מנהל תחום אחריות מקצועית</p>')
            + (f'<p style="margin-top:14px;padding-top:10px;border-top:1px solid #eee;'
               f'color:#555;font-size:14px">{_html.escape(WINNER_EMAIL_UPDATE)}</p>'
               if cust['brand'] == 'ווינר' else '')
            + '</div></td></tr></table>')

def send_campaign_email(to_email, subject, html_body):
    """Send one renewal email. Prefers Resend (HTTPS API — works on Railway, which blocks
    outbound SMTP); falls back to Gmail SMTP (works when run locally). Never raises."""
    user, pw = EMAIL_CONFIG['username'], EMAIL_CONFIG['password']
    key = os.environ.get('RESEND_API_KEY', '')
    try:
        if key:
            import requests
            frm = os.environ.get('RESEND_FROM', user)  # a verified-domain address once set up
            r = requests.post('https://api.resend.com/emails',
                headers={'Authorization': f'Bearer {key}', 'Content-Type': 'application/json'},
                json={'from': frm, 'to': [to_email], 'subject': subject, 'html': html_body},
                timeout=20)
            if r.status_code >= 300:
                print(f'[campaign] resend error {r.status_code}: {r.text[:200]}')
            return r.status_code < 300
        # Fallback: Gmail SMTP (blocked on Railway; fine when the sender runs on the laptop).
        import smtplib, ssl
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        if not user or not pw:
            return False
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = user
        msg['To'] = to_email
        msg.attach(MIMEText(re.sub('<[^>]+>', '', html_body).strip(), 'plain', 'utf-8'))
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        ctx = ssl.create_default_context()
        with smtplib.SMTP('smtp.gmail.com', 587, timeout=25) as s:
            s.starttls(context=ctx)
            s.login(user, pw)
            s.sendmail(user, [to_email], msg.as_string())
        return True
    except Exception as e:
        print(f'[campaign] email send failed: {e}')
        return False

def within_business_hours(now=None):
    """True only Sun–Thu 08:00–16:00 (no Fri/Sat sends)."""
    now = now or datetime.datetime.now()
    if now.weekday() in (4, 5):  # Fri=4, Sat=5
        return False
    return 8 <= now.hour < 16

# Winner customers were historically messaged from Gaia's number; a closing line on Winner's
# message announces the new Winner number (058-7900009) and reassures it's the same team.
WINNER_WA_TRUST = ("דרך אגב, זה שרון מווינר-אופיר, התחדשנו בטלפון של אלופים - 058-7900009. "
                   "בעבר אולי היינו בקשר גם ממספר אחר שלנו - אבל זה אנחנו, אותו צוות ואותו שירות ללא עלות.")
# Short closing update line for Winner renewal emails (announces the new Winner number).
WINNER_EMAIL_UPDATE = "עדכון - התחדשנו בטלפון של אלופים - 058-7900009"

def render_renewal_whatsapp(cust, month_name):
    """Plain-text renewal message for WhatsApp (concise — link + price + signature).
    Winner messages open with a trust line (see WINNER_WA_TRUST)."""
    link, _ = renewal_link(cust['brand'], cust['is_midwife'])
    amt = renewal_amount(cust['is_midwife'], cust['premium_last_year'])
    price = (f"המחיר לשנה הקרובה: {amt:,} ₪ — ללא שינוי מהשנה שעברה."
             if amt is not None else "המחיר נשאר כמו שנה שעברה.")
    winner_upd = f"\n\n{WINNER_EMAIL_UPDATE}" if cust['brand'] == 'ווינר' else ''
    return (f"שלום, {cust['name']}\n"
            f"הפוליסה המקצועית שלך מסתיימת בסוף חודש {month_name}.\n"
            f"{price}\n\n"
            f"לחידוש הפוליסה וצפייה בתנאים העדכניים (שלא השתנו):\n{link}\n\n"
            f"—\nשרון דר,\nמנהל תחום אחריות מקצועית"
            f"{winner_upd}")

def _wa_api_authed():
    # Accepts the wa-sender token OR a SEPARATE bot token (WA_API_TOKEN_BOT) so the bot project can
    # be authorized + revoked independently. Either valid X-WA-Token passes.
    got = request.headers.get('X-WA-Token') or ''
    valid = {t for t in (os.environ.get('WA_API_TOKEN', ''), os.environ.get('WA_API_TOKEN_BOT', '')) if t}
    return bool(got) and got in valid

@app.route('/api/stage-import', methods=['POST'])
def api_stage_import():
    """Token-authed equivalent of the '/admin/import' STAGE step, so a renewal file already
    on the laptop can be staged into a pending import (making the 'שינוי חידוש לחודש חדש'
    button appear) without re-uploading through the browser. Nothing is applied — commit is
    still a manual button click in the admin UI."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    f = request.files.get('file')
    month_name = (request.form.get('month_name') or '').strip()
    source = request.form.get('source', 'gaia_winner')
    by = (request.form.get('by') or 'טעינה חד-פעמית מהלפטופ').strip()
    if not f or not month_name:
        return jsonify({'error': 'missing file or month_name'}), 400
    try:
        blob = f.read()
        wb = load_workbook(io.BytesIO(blob), data_only=True)
        rows = _extract_renewal_rows(wb.active, source)
        conn = get_db()
        report = _validate_renewal_file(conn, rows)
        conn.execute("""INSERT INTO pending_imports
            (source, month_name, filename, file_blob, report_json, uploaded_at, uploaded_by, status)
            VALUES (?,?,?,?,?,?,?, 'pending')""",
            (source, month_name, f.filename, blob, json.dumps(report, ensure_ascii=False),
             datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), by))
        pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'pending_id': pid, 'report': report})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/fix-id', methods=['POST'])
def api_fix_id():
    """Token-authed correction of a wrong ת"ז in the master, keeping every linked record
    consistent: the insureds master row, any customers rows (so a later month-commit does
    NOT recreate the duplicate), and the client_events activity log (re-keyed so history
    follows the person). Pass dry_run=1 to inspect without changing anything."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    old_id = normalize_id_number(data.get('old_id'))
    new_id = normalize_id_number(data.get('new_id'))
    name_check = (data.get('name') or '').strip()
    dry_run = bool(data.get('dry_run'))
    if not old_id or not new_id or old_id == new_id:
        return jsonify({'error': 'need distinct old_id/new_id'}), 400
    conn = get_db()
    ins = conn.execute("SELECT id, name FROM insureds WHERE id_number=?", (old_id,)).fetchone()
    if not ins:
        conn.close()
        return jsonify({'error': f'no insured with id_number {old_id}'}), 404
    if name_check and name_check not in (ins['name'] or ''):
        conn.close()
        return jsonify({'error': f"name mismatch: master has '{ins['name']}', expected to contain '{name_check}'"}), 409
    collide = conn.execute("SELECT id, name FROM insureds WHERE id_number=? AND id!=?",
                           (new_id, ins['id'])).fetchone()
    if collide:
        conn.close()
        return jsonify({'error': f"new_id {new_id} already exists on insured '{collide['name']}'"}), 409
    old_key, new_key = event_key(old_id, ''), event_key(new_id, '')
    cust_n = conn.execute("SELECT COUNT(*) c FROM customers WHERE id_number=?", (old_id,)).fetchone()['c']
    evt_n = conn.execute("SELECT COUNT(*) c FROM client_events WHERE idkey=?", (old_key,)).fetchone()['c']
    info = {'insured': ins['name'], 'old_id': old_id, 'new_id': new_id,
            'customers_rows': cust_n, 'event_rows': evt_n, 'old_key': old_key, 'new_key': new_key}
    if dry_run:
        conn.close()
        return jsonify({'dry_run': True, **info})
    conn.execute("UPDATE insureds SET id_number=?, updated_at=? WHERE id=?",
                 (new_id, datetime.datetime.now().isoformat(), ins['id']))
    conn.execute("UPDATE customers SET id_number=? WHERE id_number=?", (new_id, old_id))
    if old_key and new_key:
        conn.execute("UPDATE client_events SET idkey=? WHERE idkey=?", (new_key, old_key))
    log_event(conn, new_key, f"תיקון ת\"ז: {old_id} → {new_id} (טעינה חד-פעמית מהלפטופ)",
              'system', kind='id_fix')
    conn.commit()
    conn.close()
    return jsonify({'ok': True, **info})

@app.route('/api/rename-month', methods=['POST'])
def api_rename_month():
    """Token-authed month rename (fixes a garbled name). With no name, just reads months."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    mid = data.get('id')
    conn = get_db()
    if name:
        if mid:
            conn.execute("UPDATE months SET name=? WHERE id=?", (name, mid))
        else:
            row = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
            if row:
                conn.execute("UPDATE months SET name=? WHERE id=?", (name, row['id']))
        conn.commit()
    months = [dict(r) for r in conn.execute("SELECT id, name, is_active FROM months ORDER BY id DESC")]
    conn.close()
    return jsonify({'ok': True, 'months': months})

@app.route('/api/month-stats')
def api_month_stats():
    """Token-authed per-month status breakdown — lets late renewals on an archived
    month be verified over time (renewed = status 'חודש')."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    out = []
    for m in conn.execute("SELECT id, name, is_active FROM months ORDER BY id DESC").fetchall():
        total = conn.execute("SELECT COUNT(*) c FROM customers WHERE month_id=?", (m['id'],)).fetchone()['c']
        renewed = conn.execute("SELECT COUNT(*) c FROM customers WHERE month_id=? AND status IN ('חודש','חודש - בוצעה שיחת מכירה')",
                               (m['id'],)).fetchone()['c']
        out.append({'id': m['id'], 'name': m['name'], 'is_active': m['is_active'],
                    'total': total, 'renewed': renewed, 'not_renewed': total - renewed})
    conn.close()
    return jsonify({'months': out})

@app.route('/api/brand-census')
def api_brand_census():
    """Token-authed READ-ONLY diagnostic: per-month × per-brand customer counts (+ how many are
    'חודש'). Used to see where Ofir/Gaia/Winner rows live before/after any data move."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    out = []
    for m in conn.execute("SELECT id, name, is_active FROM months ORDER BY id DESC").fetchall():
        brands = {}
        for r in conn.execute(
            "SELECT COALESCE(NULLIF(brand,''),'(ריק)') b, COUNT(*) n, "
            "SUM(CASE WHEN status IN ('חודש','חודש - בוצעה שיחת מכירה') THEN 1 ELSE 0 END) renewed "
            "FROM customers WHERE month_id=? GROUP BY b ORDER BY n DESC", (m['id'],)).fetchall():
            brands[r['b']] = {'total': r['n'], 'renewed': r['renewed']}
        out.append({'id': m['id'], 'name': m['name'], 'is_active': m['is_active'], 'brands': brands})
    conn.close()
    return jsonify({'months': out})

@app.route('/api/funnel')
def api_funnel():
    """Token-authed: the exact renewal-funnel numbers the dashboard shows for the active month —
    the combined גאיה+ווינר view plus each present brand. Uses the same _renewal_funnel helper."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    month = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    rows = conn.execute("""SELECT status, brand, sector, form_received_at, import_source,
                           call_status_1, call_status_2, call_status_3
                           FROM customers WHERE month_id=?""", (month['id'],)).fetchall()
    conn.close()
    present = [b for b in ('גאיה', 'ווינר', 'אופיר') if any(r['brand'] == b for r in rows)]
    active = [b for b in ('גאיה', 'ווינר') if b in present]
    out = {'month': month['name'], 'views': {}, 'pending': _pending_split(rows)}
    if len(active) > 1:
        out['views']['גאיה + ווינר'] = _renewal_funnel([r for r in rows if r['brand'] in active])
    for b in present:
        out['views'][b] = _renewal_funnel([r for r in rows if r['brand'] == b])
    return jsonify(out)

@app.route('/api/delivery-audit')
def api_delivery_audit():
    """Token-authed reliability net for the DASHBOARD's post-event individual sends (Sharon's hard
    rule: a customer who renewed/was-issued, or a cert customer, must NEVER silently miss their
    message). Surfaces, for the active cycle: renewed/issued customers whose THIS-CYCLE policy
    arrived but wasn't delivered ('stuck'), renewed/issued with no policy doc at all past a longer
    horizon ('no_doc'), and matched certs not sent ('cert_stuck') + unresolved no_match certs."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    now = datetime.datetime.now()
    cut_sent = (now - datetime.timedelta(hours=24)).strftime('%Y-%m-%d %H:%M')   # stuck: arrived >24h ago, unsent
    cut_nodoc = (now - datetime.timedelta(hours=72)).strftime('%Y-%m-%d %H:%M')  # no-doc: renewed >72h ago, no PDF
    cycle = (now - datetime.timedelta(days=45)).strftime('%Y-%m-%d %H:%M')       # this-cycle policy window
    # Only DELIVERABLE policy documents count — the delivery system sends renewal (חידוש) + new
    # (חדש) PDFs, never proposals ("אינ'") or other doc types. Matching that keeps the audit honest.
    DELIV = "(pr.doc_type_label LIKE '%חדש%' OR pr.doc_type_label LIKE '%חידוש%')"
    conn = get_db()
    month = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    # A deliverable PDF stops auto-retrying once it's older than the 10-day WhatsApp-pending window
    # (see _policy_queue_items). So a doc unsent AND past that window is genuinely stuck (won't self-
    # heal); an unsent doc still inside the window is just pending and must NOT raise an alert.
    retry_window = (now - datetime.timedelta(days=11)).strftime('%Y-%m-%d %H:%M')
    rows = conn.execute(
        f"""SELECT c.id, c.name, c.id_number, c.brand, c.status, c.status_changed_at,
                  (SELECT COUNT(*) FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id
                     WHERE ltrim(COALESCE(pr.insured_id,''),'0')=ltrim(COALESCE(c.id_number,''),'0')
                       AND pd.received_at >= ? AND {DELIV}) AS docs,
                  (SELECT COUNT(*) FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id
                     WHERE ltrim(COALESCE(pr.insured_id,''),'0')=ltrim(COALESCE(c.id_number,''),'0')
                       AND pd.received_at >= ? AND {DELIV}
                       AND (COALESCE(pd.whatsapp_sent_at,'')!='' OR COALESCE(pd.email_sent_at,'')!='')) AS sent,
                  (SELECT MAX(pd.received_at) FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id
                     WHERE ltrim(COALESCE(pr.insured_id,''),'0')=ltrim(COALESCE(c.id_number,''),'0')
                       AND pd.received_at >= ? AND {DELIV}) AS max_recv
           FROM customers c
           WHERE c.month_id=? AND c.status IN ('חודש','חודש - בוצעה שיחת מכירה','הופק')
             AND COALESCE(c.group_owner,'')='' AND COALESCE(c.id_number,'')!=''""",
        (cycle, cycle, cycle, month['id'])).fetchall()
    stuck, no_doc, pending = [], [], 0
    for r in rows:
        if r['sent'] > 0:
            continue  # this-cycle policy delivered — fine
        chg = r['status_changed_at'] or ''
        item = {'id': r['id'], 'name': r['name'], 'id_number': r['id_number'],
                'brand': r['brand'], 'status': r['status'], 'since': chg or '(no date)',
                'doc_received': r['max_recv']}
        if r['docs'] > 0:
            if (r['max_recv'] or '') and r['max_recv'] < retry_window:
                stuck.append(item)           # arrived, not delivered, past retry window → real miss
            else:
                pending += 1                 # still inside the auto-retry window — not a gap yet
        elif not chg or chg <= cut_nodoc:
            no_doc.append(item)              # renewed/issued, no PDF at all → waiting/lost
    cert_stuck = [dict(x) for x in conn.execute(
        "SELECT cust_name, id_number, brand, received_at FROM cert_requests "
        "WHERE match_status='matched' AND COALESCE(wa_sent_at,'')='' AND received_at <= ? ORDER BY received_at",
        (cut_sent,)).fetchall()]
    cert_no_match = conn.execute(
        "SELECT COUNT(*) FROM cert_requests WHERE match_status='no_match' AND COALESCE(wa_sent_at,'')=''").fetchone()[0]
    conn.close()
    return jsonify({'month': month['name'],
                    'policy_stuck': stuck, 'policy_no_doc': no_doc,
                    'policy_pending_in_window': pending,
                    'cert_stuck': cert_stuck, 'cert_no_match': cert_no_match,
                    'total_gaps': len(stuck) + len(no_doc) + len(cert_stuck) + cert_no_match})

@app.route('/api/new-biz-in-renewals')
def api_new_biz_in_renewals():
    """Token-authed diagnostic: active-month rows from a NEW-BUSINESS pipeline whose work status
    isn't a terminal new-business one — i.e. the ones that leaked into the renewal funnel before
    the source-based exclusion. Read-only; shows exactly what the fix removes from the counts."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    month = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not month:
        conn.close(); return jsonify({'error': 'no active month'}), 400
    ph = ','.join('?' * len(NEW_BUSINESS_SOURCES))
    rows = conn.execute(
        f"SELECT id, name, id_number, brand, status, import_source, policy_number "
        f"FROM customers WHERE month_id=? AND import_source IN ({ph}) "
        f"AND COALESCE(status,'') NOT IN ('ממתין להפקה','הופק','בוטל') ORDER BY brand, name",
        [month['id']] + list(NEW_BUSINESS_SOURCES)).fetchall()
    conn.close()
    return jsonify({'month': month['name'], 'count': len(rows), 'items': [dict(r) for r in rows]})

@app.route('/api/ofir-test-copy', methods=['POST'])
def api_ofir_test_copy():
    """Token-authed: make an INERT practice copy of the Ofir book into the ACTIVE month.
    Copied rows are tagged import_source='test_ofir' → excluded from policy auto-delivery
    (_policy_queue_items), and Ofir is already out of the campaign by brand — so these rows
    NEVER send WhatsApp/email or start any process; they're only a customer-file to practice on.
    A ת"ז that already exists in the active month (a REAL customer) is skipped, so reps never
    practice on a live customer. {clear:true} deletes all test_ofir rows — run it before the
    September load so the practice data doesn't persist."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    conn = get_db()
    if data.get('clear'):
        n = conn.execute("SELECT COUNT(*) c FROM customers WHERE import_source='test_ofir'").fetchone()['c']
        conn.execute("DELETE FROM customers WHERE import_source='test_ofir'")
        conn.commit()
        conn.close()
        return jsonify({'cleared': n})
    active = conn.execute("SELECT id, name FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    src = conn.execute("SELECT id, name FROM months WHERE is_active=0 ORDER BY id DESC LIMIT 1").fetchone()
    if not active or not src:
        conn.close()
        return jsonify({'error': 'need an active + an archived month'}), 400
    aug_id, jul_id = active['id'], src['id']
    cols = [r['name'] for r in conn.execute("PRAGMA table_info(customers)").fetchall() if r['name'] != 'id']
    exprs = [str(int(aug_id)) if c == 'month_id'
             else ("'test_ofir'" if c == 'import_source' else '"%s"' % c) for c in cols]
    src_ofir = conn.execute("SELECT COUNT(*) c FROM customers WHERE month_id=? AND brand='אופיר'",
                            (jul_id,)).fetchone()['c']
    before = conn.execute("SELECT COUNT(*) c FROM customers WHERE import_source='test_ofir'").fetchone()['c']
    conn.execute(
        f"INSERT INTO customers ({','.join(cols)}) SELECT {','.join(exprs)} FROM customers "
        "WHERE month_id=? AND brand='אופיר' AND ltrim(COALESCE(id_number,''),'0') NOT IN "
        "(SELECT ltrim(COALESCE(id_number,''),'0') FROM customers WHERE month_id=?)",
        (jul_id, aug_id))
    conn.commit()
    after = conn.execute("SELECT COUNT(*) c FROM customers WHERE import_source='test_ofir'").fetchone()['c']
    conn.close()
    return jsonify({'ok': True, 'source_month': src['name'], 'active_month': active['name'],
                    'source_ofir': src_ofir, 'copied': after - before,
                    'skipped_already_in_active': src_ofir - (after - before), 'test_ofir_total': after})

@app.route('/api/set-occupations', methods=['POST'])
def api_set_occupations():
    """Token-authed bulk fill of the occupation column (עיסוק המבוטח, extracted locally from
    the policy PDFs), matching by ת"ز on customers + insureds. Body: {map: {"<ת"ז>": "<occ>"}}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    m = (request.get_json(silent=True) or {}).get('map') or {}
    conn = get_db()
    cust_n = ins_n = 0
    for idn, occ in m.items():
        z = re.sub(r'\D', '', str(idn or '')).lstrip('0')
        occ = (occ or '').strip()
        if not z or not occ:
            continue
        cust_n += conn.execute("UPDATE customers SET occupation=? WHERE ltrim(COALESCE(id_number,''),'0')=?",
                               (occ, z)).rowcount
        ins_n += conn.execute("UPDATE insureds SET occupation=? WHERE ltrim(COALESCE(id_number,''),'0')=?",
                              (occ, z)).rowcount
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'input': len(m), 'customers_updated': cust_n, 'insureds_updated': ins_n})

@app.route('/api/sync-renewed-active', methods=['POST'])
def api_sync_renewed_active():
    """Token-authed retroactive sync: every customer marked 'חודש' whose master record is not
    'פעיל' is reactivated (status='פעיל', override). Optional {check:name} returns that insured's
    status after the sync, to verify a specific late renewal."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    conn = get_db()
    now = datetime.datetime.now().isoformat()
    updated, seen = [], set()
    for r in conn.execute("SELECT id_number, name FROM customers WHERE status IN ('חודש','חודש - בוצעה שיחת מכירה')").fetchall():
        idn = normalize_id_number(r['id_number'])
        key = (idn or '').lstrip('0')
        if not key or key in seen:
            continue
        seen.add(key)
        res = conn.execute(
            "UPDATE insureds SET status='פעיל', status_override=1, updated_at=? "
            "WHERE ltrim(id_number,'0')=? AND (status IS NULL OR status!='פעיל')", (now, key))
        if res.rowcount:
            updated.append(r['name'])
    conn.commit()
    checked = None
    if data.get('check'):
        checked = [dict(x) for x in conn.execute(
            "SELECT name, id_number, status, status_override FROM insureds WHERE name LIKE ? LIMIT 8",
            (f"%{data['check']}%",)).fetchall()]
    conn.close()
    return jsonify({'ok': True, 'updated': len(updated), 'names': updated, 'checked': checked})

@app.route('/api/backfill-activity-log', methods=['POST'])
def api_backfill_activity_log():
    """Token-authed one-time backfill: log a 'פעילות <month>' summary event (by ת"ז) for every
    customer of an archived month that had activity but no such event yet. Pass dry_run to
    preview. Defaults to the most recent archived (inactive) month."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    dry = bool(data.get('dry_run'))
    mid = data.get('month_id')
    conn = get_db()
    if mid:
        m = conn.execute("SELECT * FROM months WHERE id=?", (mid,)).fetchone()
    else:
        m = conn.execute("SELECT * FROM months WHERE is_active=0 ORDER BY id DESC LIMIT 1").fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'no archived month'}), 404
    logged, skipped, samples = _backfill_activity_for(conn, m['id'], m['name'], dry=dry)
    if not dry:
        conn.commit()
    conn.close()
    result = {'ok': True, 'dry_run': dry, 'month': m['name'],
              'skipped_existing': skipped, 'samples': samples}
    result['to_log' if dry else 'logged'] = logged
    return jsonify(result)

@app.route('/api/email-coverage')
def api_email_coverage():
    """Token-authed email-coverage diagnostic for the active month: per brand, how many
    customers have a resolvable email at all vs the enrichment gap, and how many actually
    land in the campaign send bucket (after midwife/VIP/stop-status exclusions)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'error': 'no active month'}), 404
    conn = get_db()
    ins_email = {re.sub(r'\D', '', str(r['id_number'] or '')).lstrip('0'): (r['email'] or '').strip()
                 for r in conn.execute("SELECT id_number, email FROM insureds")}
    stats = {}
    for c in conn.execute("SELECT * FROM customers WHERE month_id=?", (month['id'],)).fetchall():
        brand = c['brand']
        d = stats.setdefault(brand, {'total': 0, 'with_email': 0, 'no_email': 0,
                                     'midwife': 0, 'vip': 0, 'send_bucket': 0})
        d['total'] += 1
        email = _campaign_email_for(conn, c, ins_email)
        has = bool(email and '@' in email)
        d['with_email'] += 1 if has else 0
        d['no_email'] += 0 if has else 1
        if c['is_midwife']:
            d['midwife'] += 1
        if c['is_vip']:
            d['vip'] += 1
        excluded = (c['is_midwife'] or c['is_vip'] or (c['status'] or '') in CAMPAIGN_STOP_STATUSES
                    or ('do_not_contact' in c.keys() and c['do_not_contact']))
        if has and not excluded:
            d['send_bucket'] += 1
    conn.close()
    return jsonify({'month': month['name'], 'brands': stats})

def _month_state(conn):
    """Snapshot of months + counts, for showing the before/after of a month transition."""
    rows = conn.execute(
        """SELECT id, name, is_active,
                  (SELECT COUNT(*) FROM customers WHERE customers.month_id=months.id) AS customers
           FROM months ORDER BY id DESC""").fetchall()
    insureds = conn.execute("SELECT COUNT(*) c FROM insureds").fetchone()['c']
    active = next((dict(r) for r in rows if r['is_active']), None)
    return {'active_month': active['name'] if active else None,
            'insureds_total': insureds, 'months': [dict(r) for r in rows]}

@app.route('/api/commit-import', methods=['POST'])
def api_commit_import():
    """Token-authed equivalent of the '/admin/import/commit' button — apply a staged
    pending import (the month transition), returning a before/after snapshot so the
    promotion + archival can be inspected. Uses the newest pending import if no pid."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    pid = (request.get_json(silent=True) or {}).get('pid')
    conn = get_db()
    if pid:
        p = conn.execute("SELECT * FROM pending_imports WHERE id=? AND status='pending'", (pid,)).fetchone()
    else:
        p = conn.execute("SELECT * FROM pending_imports WHERE status='pending' ORDER BY id DESC LIMIT 1").fetchone()
    if not p:
        conn.close()
        return jsonify({'error': 'no pending import'}), 404
    # Outgoing month's renewal outcome (what promotion will set: 'חודש'→פעיל, else→לא פעיל).
    prev = conn.execute("SELECT * FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    outgoing = None
    if prev:
        renewed = conn.execute("SELECT COUNT(*) c FROM customers WHERE month_id=? AND status IN ('חודש','חודש - בוצעה שיחת מכירה')",
                               (prev['id'],)).fetchone()['c']
        total = conn.execute("SELECT COUNT(*) c FROM customers WHERE month_id=?",
                             (prev['id'],)).fetchone()['c']
        outgoing = {'name': prev['name'], 'total': total, 'renewed_active': renewed,
                    'not_renewed_inactive': total - renewed}
    before = _month_state(conn)
    _snapshot_db('before_import')  # rollback point BEFORE the month transition is applied
    try:
        wb = load_workbook(io.BytesIO(p['file_blob']), data_only=True)
        count, promoted, label = _apply_import(conn, wb, p['source'], p['month_name'])
        conn.execute("UPDATE pending_imports SET status='committed' WHERE id=?", (p['id'],))
        conn.commit()
        after = _month_state(conn)
        conn.close()
        threading.Thread(target=_run_post_load_enrichment, daemon=True).start()
        return jsonify({'ok': True, 'loaded': count, 'promoted': promoted, 'label': label,
                        'outgoing_month': outgoing, 'before': before, 'after': after,
                        'enrichment': 'running in background'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

def _policy_to972(phone):
    p = re.sub(r'\D', '', str(phone or ''))
    if not p:
        return ''
    if p.startswith('0'):
        return '972' + p[1:]
    if not p.startswith('972'):
        return '972' + p
    return p

_POLICY_FORCE_IDS = set()  # TEST-only: ת"ז forced into the queue regardless of the 48h window
_POLICY_LIVE_IDS = set()   # per-ת"ז LIVE override: real recipient even while the system is in test mode

def _policy_queue_items(conn, brand_key):
    """Documents ready for auto-delivery on `brand_key` ('gaia'|'winner'): a recent
    (≤48h) Harel RENEWAL PDF whose ת"ز matches a customer in ANY month (incl. archived) marked 'חודש',
    still pending on at least one channel. In test mode recipients are forced to Sharon."""
    brands = ['גאיה'] if brand_key == 'gaia' else ['ווינר', 'אופיר']
    custs = {}
    # Match renewed customers across ALL months (not just the active one) so a late renewer
    # from a previous month still gets their policy — the ±48h fresh-PDF window bounds it.
    # Most-recent month wins on duplicate ת"ז.
    # test_ofir rows are inert practice data — never auto-deliver a policy for them.
    # 'הופק' (issued) is treated the same as 'חודש' (renewed) for policy delivery (Sharon: they're
    # the same in the calcs) — so a renewal PDF for an 'הופק' customer is still delivered. A NEW-doc
    # for an 'הופק' customer is skipped here (not a renewal doc) and handled by the new-business path.
    # Group-owner customers (e.g. Aviram's therapists) are EXCLUDED from auto-delivery — their PDFs
    # are prepared manually (price removed) and sent to the group owner by hand.
    q = ("SELECT * FROM customers WHERE status IN ('חודש','חודש - בוצעה שיחת מכירה','הופק') AND brand IN (%s) "
         "AND COALESCE(import_source,'')!='test_ofir' AND COALESCE(group_owner,'')='' "
         "ORDER BY month_id ASC, id ASC"
         % ','.join('?' * len(brands)))
    for c in conn.execute(q, brands).fetchall():
        idn = normalize_id_number(c['id_number'])
        if idn:
            custs[idn] = c
    if not custs:
        return []
    cutoff = (datetime.datetime.now() - datetime.timedelta(hours=POLICY_SEND_WINDOW_HOURS)
              ).strftime('%Y-%m-%d %H:%M')
    doc_sql = """SELECT pd.id AS doc_id, pd.received_at, pd.whatsapp_sent_at, pd.email_sent_at,
                        pd.policy_number, pr.insured_id, pr.doc_type_label
                 FROM policy_documents pd JOIN policy_records pr ON pr.policy_document_id = pd.id"""
    # 48h fresh window, PLUS any WhatsApp-still-pending policy within a wider 10-day window — so a
    # renewal delivered by EMAIL during a WhatsApp hold still gets its WhatsApp once the hold lifts.
    wider = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime('%Y-%m-%d %H:%M')
    rows = list(conn.execute(
        doc_sql + " WHERE pd.received_at >= ? OR (COALESCE(pd.whatsapp_sent_at,'')='' "
        "AND pd.received_at >= ?) ORDER BY pd.received_at DESC, pd.id DESC",
        (cutoff, wider)).fetchall())
    # TEST helper: force specific ת"ז's renewal docs in regardless of the 48h window
    # (they still must be marked 'חודש' to match `custs`). No effect in normal operation.
    if _POLICY_FORCE_IDS:
        ph = ','.join('?' * len(_POLICY_FORCE_IDS))
        forced = conn.execute(
            doc_sql + f" WHERE ltrim(COALESCE(pr.insured_id,''),'0') IN ({ph}) "
            "ORDER BY pd.received_at DESC, pd.id DESC", list(_POLICY_FORCE_IDS)).fetchall()
        rows = list(forced) + rows
    # Only the MOST-RECENT renewal per person (rows are newest-first): once a ת"ז is handled we
    # skip their older documents, so an existing+renewed pair never delivers the old one.
    items, seen_custs = [], set()
    for r in rows:
        if not is_renewal_doc(r['doc_type_label']):
            continue
        idn = normalize_id_number(r['insured_id'])
        c = custs.get(idn)
        if not c or idn in seen_custs:
            continue
        seen_custs.add(idn)
        wa_pending = not r['whatsapp_sent_at']
        em_pending = not r['email_sent_at']
        if not (wa_pending or em_pending):
            continue
        real_phone = _policy_to972(c['phone'])
        real_email = (c['email'] or '').strip() or (_campaign_email_for(conn, c) or '')
        # Live if the system is live, OR this specific ת"ז is a per-customer live override.
        # Compare leading-zero-stripped (the override list stores the stripped form), so IDs
        # that start with 0 (pre-1988) match correctly.
        live = (not POLICY_AUTOSEND_TEST) or ((idn or '').lstrip('0') in _POLICY_LIVE_IDS)
        items.append({
            'doc_id': r['doc_id'],
            'name': c['name'],
            'policy_number': r['policy_number'],
            'brand': brand_key,
            'phone': real_phone if live else _policy_to972(POLICY_TEST_PHONE),
            'email': real_email if live else POLICY_TEST_EMAIL,
            'whatsapp_pending': wa_pending,
            'email_pending': em_pending,
            'wa_text': POLICY_WA_RENEWAL,
            'email_subject': POLICY_EMAIL_SUBJECT,
            'email_body': policy_email_body(c['name']),
            'email_html': policy_email_html(c['name']),
            'pdf_url': f'/api/policy/pdf/{r["doc_id"]}',
            'test_mode': not live,
            'is_midwife': bool(c['is_midwife']),
            'intended': f"{c['name']} · {real_phone or '—'} · {real_email or '—'}",
        })
    # ── New-business policies (WhatsApp only), gated by POLICY_NEW_MODE ('test'→Sharon,'live') ──
    if POLICY_NEW_MODE in ('test', 'live'):
        new_test = (POLICY_NEW_MODE == 'test')
        seen_new = set()
        # Forced ת"ז (via /api/policy/force-test) re-enter regardless of the window — so a stuck
        # new-business policy that aged out of the 10-day window can be re-sent on demand.
        nb_where = "pd.received_at >= ? OR (COALESCE(pd.whatsapp_sent_at,'')='' AND pd.received_at >= ?)"
        nb_params = [cutoff, wider]
        if _POLICY_FORCE_IDS:
            fph = ','.join('?' * len(_POLICY_FORCE_IDS))
            nb_where += f" OR ltrim(COALESCE(pr.insured_id,''),'0') IN ({fph})"
            nb_params += list(_POLICY_FORCE_IDS)
        for r in conn.execute(
            f"""SELECT pd.id AS doc_id, pd.received_at, pd.whatsapp_sent_at, pd.email_sent_at,
                      pd.policy_number, pr.insured_id, pr.insured_name, pr.phone_mobile,
                      pr.agent_number, pr.doc_type_label, pr.email AS pr_email
               FROM policy_documents pd JOIN policy_records pr ON pr.policy_document_id = pd.id
               WHERE {nb_where}
               ORDER BY pd.received_at DESC, pd.id DESC""", nb_params).fetchall():
            if not is_new_doc(r['doc_type_label']):
                continue
            if r['whatsapp_sent_at'] and r['email_sent_at']:
                continue  # already delivered on both channels
            if _new_policy_brand_key(r['agent_number']) != brand_key:
                continue
            key = (normalize_id_number(r['insured_id']) or '').lstrip('0')
            if not key or key in seen_new:
                continue
            # Backlog guard: a website-form lead received on/before 1/8 may already have been
            # issued + sent manually — skip auto-send so the customer isn't messaged twice.
            lead = conn.execute(
                "SELECT form_received_at FROM customers WHERE import_source='join_form' "
                "AND ltrim(COALESCE(id_number,''),'0')=? ORDER BY id DESC LIMIT 1", (key,)).fetchone()
            if lead and (lead['form_received_at'] or '')[:10] <= '2026-08-01' and key not in _POLICY_FORCE_IDS:
                continue  # (a forced ת"ז overrides the backlog guard — explicit re-send)
            real_phone = _policy_to972(r['phone_mobile'])
            real_email = (r['pr_email'] or '').strip()
            if not real_email:
                ce = conn.execute("SELECT email FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? "
                                  "AND COALESCE(email,'')!='' ORDER BY id DESC LIMIT 1", (key,)).fetchone()
                real_email = (ce['email'].strip() if ce and ce['email'] else '')
            has_email = bool(real_email and '@' in real_email)
            # A no-phone/no-email new policy is still queued so the local sender forwards it to
            # Sharon (rule: issued-but-undeliverable → me) rather than dropping it silently.
            seen_new.add(key)
            _mw = conn.execute("SELECT 1 FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? "
                               "AND COALESCE(is_midwife,0)=1 LIMIT 1", (key,)).fetchone()
            items.append({
                'doc_id': r['doc_id'],
                'name': r['insured_name'] or '',
                'policy_number': r['policy_number'],
                'brand': brand_key,
                'is_midwife': bool(_mw),
                'phone': _policy_to972(POLICY_TEST_PHONE) if new_test else (real_phone or ''),
                'email': (POLICY_TEST_EMAIL if new_test else real_email) if has_email else '',
                'whatsapp_pending': not r['whatsapp_sent_at'],
                'email_pending': (not r['email_sent_at']) and has_email,
                'wa_text': POLICY_WA_NEW,
                'email_subject': POLICY_EMAIL_SUBJECT,
                'email_body': new_policy_email_body(r['insured_name'] or ''),
                'email_html': new_policy_email_html(r['insured_name'] or ''),
                'pdf_url': f'/api/policy/pdf/{r["doc_id"]}',
                'kind': 'new',
                'test_mode': new_test,
                'intended': f"{r['insured_name']} · {real_phone or '—'} · {real_email or '—'} (חדש)",
            })
    return items

@app.route('/api/policy/queue')
def policy_queue():
    """Per-brand auto-delivery list for the local sender (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    brand = request.args.get('brand', '')
    if brand not in ('gaia', 'winner'):
        return jsonify({'error': 'bad brand'}), 400
    conn = get_db()
    items = _policy_queue_items(conn, brand)
    conn.close()
    return jsonify({'brand': brand, 'test_mode': POLICY_AUTOSEND_TEST, 'count': len(items), 'items': items})

@app.route('/api/policy/pdf/<int:doc_id>')
def policy_pdf(doc_id):
    """Serve a policy PDF to the local sender so it can attach + save it (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    doc = conn.execute('SELECT filename, filepath FROM policy_documents WHERE id=?', (doc_id,)).fetchone()
    conn.close()
    if not doc or not doc['filepath'] or not os.path.exists(doc['filepath']):
        return jsonify({'error': 'not found'}), 404
    safe_name = re.sub(r'[\r\n]+', ' ', doc['filename']).strip()
    return send_file(doc['filepath'], as_attachment=True, download_name=safe_name)

@app.route('/api/policy/sent', methods=['POST'])
def policy_sent():
    """Mark a policy as delivered on a channel + log it to the client timeline (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    doc_id, channel = data.get('doc_id'), data.get('channel')
    if not doc_id or channel not in ('whatsapp', 'email'):
        return jsonify({'error': 'need doc_id + channel'}), 400
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    col = 'whatsapp_sent_at' if channel == 'whatsapp' else 'email_sent_at'
    conn.execute(f"UPDATE policy_documents SET {col}=? WHERE id=?", (now, doc_id))
    pr = conn.execute(
        """SELECT insured_id, insured_name, phone_mobile, email, agent_number, doc_type_label,
                  policy_number, policy_document_id
           FROM policy_records WHERE policy_document_id=? LIMIT 1""", (doc_id,)).fetchone()
    if pr and pr['insured_id']:
        ch = 'וואטסאפ' if channel == 'whatsapp' else 'מייל'
        idkey = event_key(normalize_id_number(pr['insured_id']), f'doc-{doc_id}')
        if is_new_doc(pr['doc_type_label']):
            # New business — create a serviceable customer record + log without "חידוש".
            _ensure_new_customer(conn, pr)
            tag = ' [בדיקה]' if POLICY_NEW_MODE == 'test' else ''
            note = f"פוליסה חדשה נשלחה ({ch}){tag}"
        else:
            tag = ' [בדיקה]' if POLICY_AUTOSEND_TEST else ''
            note = f"פוליסת חידוש נשלחה אוטומטית ({ch}){tag}"
            # Delivering the renewal policy means the customer renewed → flip an OPEN status to 'חודש'
            # so they drop off the renewal work lists + reminders. Never override an already-settled
            # or declined status (that's a conflict for a human to resolve).
            _idn = normalize_id_number(pr['insured_id'])
            if _idn and not POLICY_AUTOSEND_TEST:
                _c = conn.execute(
                    "SELECT c.id, c.status FROM customers c JOIN months m ON m.id=c.month_id "
                    "WHERE m.is_active=1 AND ltrim(COALESCE(c.id_number,''),'0')=?", (_idn.lstrip('0'),)).fetchone()
                _settled = ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'לא רוצים לחדש', 'לא מחדש', 'בוטל')
                if _c and (_c['status'] or '') not in _settled:
                    conn.execute("UPDATE customers SET status='חודש', status_changed_at=? WHERE id=?", (now, _c['id']))
                    log_event(conn, idkey, "סטטוס עודכן אוטומטית ל-'חודש' (פוליסת חידוש נמסרה)", 'system', kind='status')
        log_event(conn, idkey, note, 'system', kind='policy_send')
    mid = conn.execute("SELECT message_id FROM policy_documents WHERE id=?", (doc_id,)).fetchone()
    conn.commit()
    conn.close()
    # After delivery, label the source Harel email 'טופל-שליחה-אוטומטית' + archive it (background).
    if mid and mid['message_id']:
        threading.Thread(target=_label_email, args=(mid['message_id'],), daemon=True).start()
    return jsonify({'ok': True})

@app.route('/api/policy/redownload', methods=['POST'])
def api_policy_redownload():
    """Recover a policy whose stored PDF went missing (e.g. deleted): drop the policy_documents row +
    its policy_records so the source Harel email becomes re-ingestible, then re-scan —
    check_policy_documents re-downloads + re-saves the PDF and it re-enters the delivery queue. Only
    acts when the file is actually gone. Body {doc_id, days:int=14}. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    doc_id = d.get('doc_id')
    days = int(d.get('days', 14))
    if not doc_id:
        return jsonify({'error': 'need doc_id'}), 400
    conn = get_db()
    row = conn.execute("SELECT filepath, policy_number FROM policy_documents WHERE id=?", (doc_id,)).fetchone()
    if not row:
        conn.close(); return jsonify({'error': 'doc not found'}), 404
    fp = row['filepath']
    if fp and os.path.exists(fp):
        conn.close(); return jsonify({'error': 'PDF exists — not re-downloading', 'filepath': fp}), 400
    pol = row['policy_number']
    conn.execute("DELETE FROM policy_records WHERE policy_document_id=?", (doc_id,))
    conn.execute("DELETE FROM policy_documents WHERE id=?", (doc_id,))
    conn.commit(); conn.close()
    n = check_policy_documents(days_back=days, keep_pdf=True)
    # Did it come back with a real file?
    conn = get_db()
    back = conn.execute("SELECT id, filepath FROM policy_documents WHERE policy_number=? "
                        "ORDER BY id DESC LIMIT 1", (pol,)).fetchone()
    recovered = bool(back and back['filepath'] and os.path.exists(back['filepath']))
    conn.close()
    return jsonify({'ok': True, 'deleted_doc': doc_id, 'policy_number': pol,
                    'rescanned': n, 'recovered': recovered,
                    'new_doc_id': (back['id'] if back else None)})

@app.route('/api/policy/reconcile-renewals', methods=['POST'])
def api_reconcile_renewals():
    """Retroactive fix for Sharon's rule: a recent renewal (חידוש) policy PDF whose customer was NOT
    advanced to a renewed/settled status → flip to 'חודש' so it auto-delivers (delivery needs
    חודש/הופק). Body {days:int=7, dry_run:bool}. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    days = int(d.get('days', 7)); dry = bool(d.get('dry_run'))
    cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%Y-%m-%d %H:%M')
    RSET = ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'בוטל', 'לא רוצים לחדש', 'לא מחדש')
    ph = ','.join('?' * len(RSET))
    conn = get_db()
    rows = conn.execute(
        f"""SELECT c.id, c.name, c.status, c.brand, ltrim(COALESCE(c.id_number,''),'0') AS idn,
                   pr.period_start, pd.received_at
            FROM policy_documents pd
            JOIN policy_records pr ON pr.policy_document_id=pd.id
            JOIN customers c ON ltrim(COALESCE(c.id_number,''),'0')=ltrim(COALESCE(pr.insured_id,''),'0')
            WHERE pd.received_at>=? AND pr.doc_type_label LIKE '%חידוש%'
              AND COALESCE(pd.whatsapp_sent_at,'')='' AND COALESCE(pd.email_sent_at,'')=''
              AND COALESCE(c.status,'') NOT IN ({ph})
              AND COALESCE(c.group_owner,'')='' AND COALESCE(c.import_source,'')!='test_ofir'""",
        [cutoff] + list(RSET)).fetchall()
    # Keep only customers with a CURRENT-cycle renewal doc (period_start ≥ 1st of next month) — never
    # act on a stale policy (Sharon's rule).
    seen = {}
    for r in rows:
        if r['id'] in seen or not _renewal_period_ok(r['period_start']):
            continue
        seen[r['id']] = {'id': r['id'], 'name': r['name'], 'was': r['status'], 'brand': r['brand'],
                         'idn': r['idn'], 'period_start': r['period_start']}
    fixed = list(seen.values())
    if not dry:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        for r in fixed:
            conn.execute("UPDATE customers SET status='חודש', status_changed_at=? WHERE id=?", (now, r['id']))
            try:
                log_event(conn, event_key(r['idn'], f"reconcile-{r['id']}"),
                          "סטטוס עודכן ל-'חודש' (הגיעה פוליסת חידוש)", 'system', kind='status')
            except Exception:
                pass
        conn.commit()
    conn.close()
    return jsonify({'days': days, 'dry_run': dry, 'count': len(fixed),
                    'fixed': [{k: v[k] for k in ('id', 'name', 'was', 'brand', 'period_start')} for v in fixed]})

@app.route('/api/policy/relink', methods=['POST'])
def api_policy_relink():
    """Token: re-run new-business lead→policy linking for a ת"ז (or doc_id). Backfills a missing
    policy number and/or upgrades a stuck 'ממתין להפקה' lead to 'הופק' — idempotent, safe to re-run.
    Body {id_number} or {doc_id}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    idn = re.sub(r'\D', '', str(d.get('id_number', '')))
    doc_id = d.get('doc_id')
    cols = ("insured_id, insured_name, phone_mobile, email, agent_number, "
            "doc_type_label, policy_number, policy_document_id")
    cols_pr = ", ".join("pr." + c.strip() for c in cols.split(","))  # qualified — policy_number exists in both tables
    conn = get_db()
    if doc_id:
        pr = conn.execute(f"SELECT {cols} FROM policy_records WHERE policy_document_id=? LIMIT 1",
                          (doc_id,)).fetchone()
    elif idn:
        pr = conn.execute(
            f"SELECT {cols_pr} FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id "
            "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? ORDER BY pd.id DESC LIMIT 1",
            (idn.lstrip('0'),)).fetchone()
    else:
        conn.close(); return jsonify({'error': 'need id_number or doc_id'}), 400
    if not pr:
        conn.close(); return jsonify({'error': 'no policy_record'}), 404
    _ensure_new_customer(conn, pr)
    conn.commit()
    z = normalize_id_number(pr['insured_id'] or '').lstrip('0')
    row = conn.execute(
        "SELECT c.id, c.status, c.policy_number FROM customers c JOIN months m ON m.id=c.month_id "
        "WHERE m.is_active=1 AND ltrim(COALESCE(c.id_number,''),'0')=? ORDER BY c.id DESC LIMIT 1",
        (z,)).fetchone()
    conn.close()
    return jsonify({'ok': True, 'policy_number': pr['policy_number'],
                    'customer': (dict(row) if row else None)})

@app.route('/api/wa/queue')
def wa_queue():
    """Per-brand WhatsApp send list for the local sender tool (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    brand = request.args.get('brand', '')
    if brand not in ('גאיה', 'ווינר'):
        return jsonify({'error': 'bad brand'}), 400
    month = active_month()
    if not month:
        return jsonify({'brand': brand, 'count': 0, 'items': []})
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    month_name = HEB_MONTHS[datetime.datetime.now().month]
    today = datetime.date.today().isoformat()
    items = []
    for r in buckets['whatsapp']:
        # Skip anyone already messaged THIS month (whatsapp_sent_date is set) — the renewal
        # WhatsApp goes once per customer, so a multi-day drip never re-messages the same person.
        if r['brand'] != brand or (r['whatsapp_sent_date'] or '').strip():
            continue
        phone = re.sub(r'\D', '', str(r['phone'] or ''))
        if phone.startswith('0'):
            phone = '972' + phone[1:]
        elif not phone.startswith('972'):
            phone = '972' + phone
        items.append({'id': r['id'], 'name': r['name'], 'phone': phone,
                      'message': render_renewal_whatsapp(r, month_name)})
    conn.close()
    return jsonify({'brand': brand, 'count': len(items), 'items': items})

_rebuild_master_state = {'running': False, 'last': ''}

def _run_rebuild_master():
    _rebuild_master_state['running'] = True
    try:
        conn = get_db()
        n = rebuild_insureds(conn)
        c = recompute_insured_statuses(conn)
        conn.close()
        _rebuild_master_state['last'] = f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}: upserted {n}, status-updated {c}"
        print('[rebuild-master] ' + _rebuild_master_state['last'])
    except Exception as e:
        _rebuild_master_state['last'] = 'ERROR: ' + str(e)[:200]
        print('[rebuild-master] ' + _rebuild_master_state['last'])
    finally:
        _rebuild_master_state['running'] = False

@app.route('/api/rebuild-master', methods=['POST', 'GET'])
def api_rebuild_master():
    """Rebuild the insureds master ('לקוחות קבוצת אופיר') from ALL policy_records — one row per
    ת"ז, status פעיל/לא-פעיל by policy end date. Loads every scanned policy's person (incl. the
    Gaia/Winner 'orphans'). Non-destructive (preserves activity + admin overrides). Runs in the
    background. Token-authed. GET returns status."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    if request.method == 'GET':
        conn = get_db()
        total = conn.execute("SELECT COUNT(*) FROM insureds").fetchone()[0]
        conn.close()
        return jsonify({'running': _rebuild_master_state['running'],
                        'last_run': _rebuild_master_state['last'], 'insureds_total': total})
    if _rebuild_master_state['running']:
        return jsonify({'ok': False, 'msg': 'כבר רץ'})
    threading.Thread(target=_run_rebuild_master, daemon=True).start()
    return jsonify({'ok': True, 'msg': 'בנייה מחדש של המאסטר החלה ברקע'})

@app.route('/api/policy/coverage')
def api_policy_coverage():
    """Read-only: coverage gap for Gaia/Winner (excl. Ofir). Counts 2026 renewal/new policy
    documents by agent-brand, the unique insureds behind them, how many have NO customer record
    ('orphans' — a scanned policy with nobody loaded), and a sample. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    since = request.args.get('since', '2025-09-01')  # policy year = Sept 2025 → today
    conn = get_db()
    cust_ids = set(
        r[0] for r in conn.execute(
            "SELECT DISTINCT ltrim(COALESCE(id_number,''),'0') FROM customers "
            "WHERE COALESCE(import_source,'')!='test_ofir'").fetchall() if r[0])
    rows = conn.execute(
        """SELECT pr.insured_id, pr.insured_name, pr.doc_type_label, pr.agent_number, pd.received_at
           FROM policy_records pr JOIN policy_documents pd ON pd.id=pr.policy_document_id
           WHERE pd.received_at >= ?
             AND (pr.doc_type_label LIKE '%חידוש%' OR pr.doc_type_label LIKE '%חדש%')""", (since,)).fetchall()
    docs_by_brand = {}
    gw_insureds, gw_orphan_ids = set(), set()
    orphans = []
    for r in rows:
        ag = re.sub(r'\D', '', str(r['agent_number'] or ''))
        brand = NEW_AGENT_BRAND.get(ag, '(לא ידוע)')
        docs_by_brand[brand] = docs_by_brand.get(brand, 0) + 1
        if brand in ('גאיה', 'ווינר'):
            idn = (normalize_id_number(r['insured_id']) or '').lstrip('0')
            if not idn:
                continue
            gw_insureds.add(idn)
            if idn not in cust_ids and idn not in gw_orphan_ids:
                gw_orphan_ids.add(idn)
                orphans.append({'idn': r['insured_id'], 'name': r['insured_name'],
                                'type': r['doc_type_label'], 'brand': brand, 'recv': r['received_at']})
    orphans.sort(key=lambda o: o['recv'] or '', reverse=True)
    return jsonify({
        'since': since,
        'docs_by_brand': docs_by_brand,
        'gaia_winner_unique_insureds': len(gw_insureds),
        'gaia_winner_orphans_no_customer': len(gw_orphan_ids),
        'orphan_sample': orphans[:40],
    })

@app.route('/api/policy/debug')
def api_policy_debug():
    """Read-only: for a ת"ז, list its policy_documents (+ sent timestamps) and all policy_send
    log events — to diagnose duplicate deliveries (multiple docs vs a re-send). Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    idn = re.sub(r'\D', '', request.args.get('id', '')).lstrip('0')
    if not idn:
        return jsonify({'error': 'need id'}), 400
    conn = get_db()
    docs = [dict(r) for r in conn.execute(
        "SELECT pd.id AS doc_id, pd.received_at, pd.whatsapp_sent_at, pd.email_sent_at, "
        "pd.policy_number, pr.doc_type_label, pr.insured_name, pr.period_start, pr.period_end "
        "FROM policy_documents pd JOIN policy_records pr ON pr.policy_document_id=pd.id "
        "WHERE ltrim(COALESCE(pr.insured_id,''),'0')=? ORDER BY pd.received_at DESC", (idn,)).fetchall()]
    events = [dict(r) for r in conn.execute(
        "SELECT created_at, note, created_by FROM client_events "
        "WHERE idkey=? AND kind='policy_send' ORDER BY id DESC", (idn,)).fetchall()]
    conn.close()
    return jsonify({'idn': idn, 'doc_count': len(docs), 'docs': docs,
                    'policy_send_events': events, 'send_event_count': len(events)})

@app.route('/api/campaign/new-biz-check')
def api_campaign_new_biz_check():
    """How many active-month customers are now excluded from the renewal campaign because they
    have a NEW-business policy on file, and how many of them ALREADY got a reminder (the wrong
    sends). Token-authed, read-only."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'new_biz_total': 0, 'already_messaged': 0, 'list': []})
    conn = get_db()
    nb = campaign_eligibility(conn, month['id'])['new_biz']
    already = [{'id': r['id'], 'name': r['name'], 'brand': r['brand'], 'status': r['status'],
                'wa_sent': r['whatsapp_sent_date'], 'email_sent': r['email_sent_date']}
               for r in nb
               if (r['whatsapp_sent_date'] or '').strip() or (r['email_sent_date'] or '').strip()]
    conn.close()
    return jsonify({'new_biz_total': len(nb), 'already_messaged': len(already), 'list': already})

@app.route('/api/wa/template-queue')
def wa_template_queue():
    """Eligible renewal-campaign customers for the OFFICIAL Cloud API template `renewal_reminder`.
    Returns the 4 body params per customer (name, month, price, link) so the sender doesn't have
    to replicate the price/link logic. Same eligibility as /api/wa/queue (skips anyone already
    messaged this month). Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    brand = request.args.get('brand', '')
    if brand not in ('גאיה', 'ווינר'):
        return jsonify({'error': 'bad brand'}), 400
    month = active_month()
    if not month:
        return jsonify({'brand': brand, 'count': 0, 'items': []})
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    month_name = HEB_MONTHS[datetime.datetime.now().month]
    items = []
    for r in buckets['whatsapp']:
        # Skip anyone already messaged this cycle — the campaign whatsapp flag OR either reminder
        # flag (lreom/lr25). This keeps the FIRST notice off customers who already got a reminder
        # (e.g. the 100 who received last_reminder_eom), so nobody gets a double message.
        already = ((r['whatsapp_sent_date'] or '').strip()
                   or (r['lreom_sent_at'] or '').strip()
                   or (r['lr25_sent_at'] or '').strip())
        if r['brand'] != brand or already:
            continue
        phone = re.sub(r'\D', '', str(r['phone'] or ''))
        if phone.startswith('0'):
            phone = '972' + phone[1:]
        elif not phone.startswith('972'):
            phone = '972' + phone
        amt = renewal_amount(r['is_midwife'], r['premium_last_year'])
        price = f"{(amt or 750):,} ₪"          # None (endorsement/no-data) → 750 per Sharon's rule
        link, _ = renewal_link(r['brand'], r['is_midwife'])
        # Template body params can't hold newlines/tabs — name/month/price/link are all safe.
        items.append({'id': r['id'], 'name': r['name'], 'phone': phone,
                      'params': [str(r['name'] or '').strip(), month_name, price, link]})
    conn.close()
    return jsonify({'brand': brand, 'count': len(items), 'items': items})

@app.route('/api/wa/sent', methods=['POST'])
def wa_sent():
    """Mark a WhatsApp message as sent + log it to the client timeline (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    cid = data.get('id')
    conn = get_db()
    r = conn.execute("SELECT id_number FROM customers WHERE id=?", (cid,)).fetchone()
    if r:
        conn.execute("UPDATE customers SET whatsapp_sent_date=? WHERE id=?",
                     (datetime.date.today().isoformat(), cid))
        idkey = event_key(r['id_number'], 'cust-%d' % cid)
        log_event(conn, idkey, f"נשלחה הודעת וואטסאפ ({data.get('brand', '')})",
                  'וואטסאפ אוטומטי', kind='whatsapp_sent')
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/card-update/queue')
def card_update_queue():
    """Customers marked 'התקבל חידוש - כ.א לא תקין' who still need the "update your payment
    method" message on WhatsApp and/or email, for the given brand (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    brand = request.args.get('brand', '')  # 'gaia' | 'winner'
    if brand not in ('gaia', 'winner'):
        return jsonify({'error': 'bad brand'}), 400
    he = 'גאיה' if brand == 'gaia' else 'ווינר'
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM customers WHERE status=? AND brand=? "
        "AND COALESCE(import_source,'')!='test_ofir' ORDER BY month_id DESC, id DESC",
        (CARD_UPDATE_STATUS, he)).fetchall()
    items, seen = [], set()
    for r in rows:
        key = (normalize_id_number(r['id_number']) or '').lstrip('0')
        if key and key in seen:
            continue          # one message per person even if the ת"ז repeats across months
        if key:
            seen.add(key)
        wa_pending = not (r['card_update_wa_at'] or '')
        em_pending = not (r['card_update_email_at'] or '')
        if not (wa_pending or em_pending):
            continue
        phone = _policy_to972(r['phone'])
        email = (r['email'] or '').strip() or (_campaign_email_for(conn, r) or '')
        has_email = bool(email and '@' in email)
        items.append({
            'id': r['id'],
            'name': r['name'],
            'brand': brand,
            'phone': phone,
            'email': email if has_email else '',
            'whatsapp_pending': wa_pending,
            'email_pending': em_pending and has_email,
            'wa_text': card_update_wa_text(r['name'], brand),
            'email_subject': CARD_UPDATE_EMAIL_SUBJECT,
            'email_body': card_update_email_body(r['name'], brand),
            'email_html': card_update_email_html(r['name'], brand),
        })
    conn.close()
    return jsonify({'brand': brand, 'count': len(items), 'items': items})

@app.route('/api/card-update/status-check')
def card_update_status_check():
    """Read-only diagnostic: every customer in the card-update status (ANY brand/month) with
    their brand + per-channel sent markers — to see who's marked and who still needs sending.
    Also returns a fuzzy list of anyone whose status merely CONTAINS 'חידוש' + 'גבי', to catch a
    near-miss status string. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    exact = [dict(r) for r in conn.execute(
        "SELECT id, name, brand, status, month_id, phone, email, "
        "card_update_wa_at, card_update_email_at, import_source "
        "FROM customers WHERE status=? ORDER BY brand, id", (CARD_UPDATE_STATUS,)).fetchall()]
    fuzzy = [dict(r) for r in conn.execute(
        "SELECT id, name, brand, status, status_changed_at, handled_by FROM customers "
        "WHERE status LIKE '%חידוש%' AND status<>? "
        "ORDER BY status_changed_at DESC", (CARD_UPDATE_STATUS,)).fetchall()]
    # Optional ?q= name search — show a specific customer's ACTUAL stored status/brand.
    by_name = []
    q = (request.args.get('q') or '').strip()
    if q:
        by_name = [dict(r) for r in conn.execute(
            "SELECT id, name, brand, status, month_id, card_update_wa_at, card_update_email_at "
            "FROM customers WHERE name LIKE ? ORDER BY id DESC LIMIT 30", ('%' + q + '%',)).fetchall()]
    conn.close()
    return jsonify({'target_status': CARD_UPDATE_STATUS, 'exact_count': len(exact),
                    'exact': exact, 'other_חידוש_statuses': fuzzy, 'by_name': by_name})

@app.route('/api/card-update/sent', methods=['POST'])
def card_update_mark_sent():
    """Mark the payment-method message as sent on a channel + log it (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    cid, channel = data.get('id'), data.get('channel')
    if not cid or channel not in ('whatsapp', 'email'):
        return jsonify({'error': 'need id + channel'}), 400
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    col = 'card_update_wa_at' if channel == 'whatsapp' else 'card_update_email_at'
    conn = get_db()
    r = conn.execute("SELECT id_number FROM customers WHERE id=?", (cid,)).fetchone()
    if r:
        conn.execute(f"UPDATE customers SET {col}=? WHERE id=?", (now, cid))
        ch = 'וואטסאפ' if channel == 'whatsapp' else 'מייל'
        idkey = event_key(r['id_number'], 'cust-%d' % cid)
        log_event(conn, idkey, f"נשלחה הודעת עדכון אמצעי גביה ({ch})", 'system', kind='card_update_send')
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/form-debug')
def api_form_debug():
    """Read-only: a ת"ז's website-form fields on the customer record + any admin-queue
    (unmatched_submissions) row — to see why form details do/don't show. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    idn = re.sub(r'\D', '', request.args.get('id', '')).lstrip('0')
    if not idn:
        return jsonify({'error': 'need id'}), 400
    conn = get_db()
    cust = [dict(r) for r in conn.execute(
        "SELECT id, name, brand, status, form_received_at, form_email, form_installments, "
        "form_payment_method, form_coverage, form_comments, "
        "CASE WHEN COALESCE(form_card_number,'')!='' THEN 'yes' ELSE 'no' END AS has_card "
        "FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? ORDER BY id DESC", (idn,)).fetchall()]
    subs = [dict(r) for r in conn.execute(
        "SELECT id, status, message_id, name, installments, payment_method, coverage, email, comments, "
        "CASE WHEN COALESCE(card_number,'')!='' THEN 'yes' ELSE 'no' END AS has_card "
        "FROM unmatched_submissions WHERE ltrim(COALESCE(id_number,''),'0')=? ORDER BY id DESC", (idn,)).fetchall()]
    conn.close()
    return jsonify({'idn': idn, 'customers': cust, 'submissions': subs})

@app.route('/api/scan-health')
def api_scan_health():
    """Read-only health of the email scanner (the Railway worker): recency of policy documents
    (הפקות/חידושים from Harel/Ofir mail) and website join-form leads, plus a live IMAP connect
    test with the same creds the worker uses. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    now = datetime.datetime.now()
    h24 = (now - datetime.timedelta(hours=24)).strftime('%Y-%m-%d %H:%M')
    d7 = (now - datetime.timedelta(days=7)).strftime('%Y-%m-%d %H:%M')
    d7d = (now - datetime.timedelta(days=7)).strftime('%Y-%m-%d')
    conn = get_db()
    one = lambda sql, *a: conn.execute(sql, a).fetchone()[0]
    out = {
        'now': now.strftime('%Y-%m-%d %H:%M'),
        'last_policy_doc': one("SELECT MAX(received_at) FROM policy_documents"),
        'policy_docs_24h': one("SELECT COUNT(*) FROM policy_documents WHERE received_at >= ?", h24),
        'policy_docs_7d': one("SELECT COUNT(*) FROM policy_documents WHERE received_at >= ?", d7),
        'last_join_lead': one("SELECT MAX(form_received_at) FROM customers WHERE import_source='join_form'"),
        'join_leads_7d': one("SELECT COUNT(*) FROM customers WHERE import_source='join_form' AND form_received_at >= ?", d7d),
    }
    # Scanner heartbeat: when did a scan last COMPLETE. scan_age_minutes is computed server-side
    # (avoids client timezone issues) — the watchdog uses it to detect a stalled scanner.
    row = conn.execute("SELECT v FROM app_kv WHERE k='last_scan_at'").fetchone()
    out['last_scan_at'] = row['v'] if row else None
    out['scan_age_minutes'] = None
    if out['last_scan_at']:
        try:
            last = datetime.datetime.strptime(out['last_scan_at'][:19], '%Y-%m-%d %H:%M:%S')
            out['scan_age_minutes'] = int((now - last).total_seconds() // 60)
        except Exception:
            pass
    conn.close()
    try:
        m = imaplib.IMAP4_SSL(EMAIL_CONFIG['imap_server'], EMAIL_CONFIG['imap_port'], timeout=30)
        m.login(EMAIL_CONFIG['username'], EMAIL_CONFIG['password'])
        m.select('INBOX')
        m.logout()
        out['imap_ok'] = True
    except Exception as e:
        out['imap_ok'] = False
        out['imap_err'] = str(e)[:160]
    return jsonify(out)

@app.route('/api/mask-existing-cards', methods=['POST'])
def api_mask_existing_cards():
    """One-time PCI cleanup: truncate any stored FULL card number to '****<last4>' across the
    card columns. Values already at ≤4 digits are left as-is. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    total, by_col = 0, {}
    for tbl, col in (('customers', 'form_card_number'), ('unmatched_submissions', 'card_number')):
        n = 0
        for r in conn.execute(f"SELECT id, {col} AS v FROM {tbl} WHERE COALESCE({col},'')!=''").fetchall():
            digits = re.sub(r'\D', '', r['v'] or '')
            if len(digits) > 4:  # a full/long card number — not already masked to the last 4
                conn.execute(f"UPDATE {tbl} SET {col}=? WHERE id=?", ('****' + digits[-4:], r['id']))
                n += 1
        by_col[f'{tbl}.{col}'] = n
        total += n
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'masked': total, 'by_column': by_col})

@app.route('/api/backup-db')
def backup_db():
    """Download the live DB for an off-site backup (token-authed)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    return send_file(DB_PATH, as_attachment=True,
                     download_name='renewals_backup_%s.db' % datetime.date.today().isoformat())

@app.route('/api/email/queue')
def email_queue():
    """Rendered renewal emails for the local sender tool (Railway blocks SMTP, so the
    laptop does the actual sending). Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    month = active_month()
    if not month:
        return jsonify({'count': 0, 'items': []})
    brand = request.args.get('brand', '')  # '' = all; else 'גאיה'/'ווינר' for the staggered schedule
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    month_name = HEB_MONTHS[datetime.datetime.now().month]
    today = datetime.date.today().isoformat()
    items = []
    for cust, email in buckets['email']:
        if brand and cust['brand'] != brand:
            continue
        if (cust['email_sent_date'] or '') == today:
            continue
        items.append({'id': cust['id'], 'email': email,
                      'subject': 'חידוש הפוליסה המקצועית שלך',
                      'html': render_renewal_email(cust, month_name)})
    conn.close()
    return jsonify({'brand': brand, 'count': len(items), 'items': items})

@app.route('/api/email/sent', methods=['POST'])
def email_sent():
    """Mark a renewal email as sent + log it (token-authed; called by the local sender)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    cid = data.get('id')
    conn = get_db()
    r = conn.execute("SELECT id_number FROM customers WHERE id=?", (cid,)).fetchone()
    if r:
        conn.execute("UPDATE customers SET email_sent_date=? WHERE id=?",
                     (datetime.date.today().isoformat(), cid))
        idkey = event_key(r['id_number'], 'cust-%d' % cid)
        log_event(conn, idkey, f"נשלח מייל חידוש ל-{data.get('email', '')}",
                  'מייל אוטומטי (לפטופ)', kind='email_sent')
        conn.commit()
    conn.close()
    return jsonify({'ok': True})

def _run_email_blast(app_ctx_month_id, month_name, recipients, who):
    """Background email send: one-by-one with a short delay, each logged to the timeline."""
    sent = 0
    for cust, email in recipients:
        try:
            html_body = render_renewal_email(cust, month_name)
            if send_campaign_email(email, 'חידוש הפוליסה המקצועית שלך', html_body):
                conn = get_db()
                conn.execute("UPDATE customers SET email_sent_date=? WHERE id=?",
                             (datetime.date.today().isoformat(), cust['id']))
                idkey = event_key(cust['id_number'], 'cust-%d' % cust['id'])
                log_event(conn, idkey, f'נשלח מייל חידוש ל-{email}', who, kind='email_sent')
                conn.commit(); conn.close()
                sent += 1
            time.sleep(0.7)
        except Exception as e:
            print(f'[campaign] email failed for {email}: {e}')
    print(f'[campaign] email blast done — {sent}/{len(recipients)} sent')

@app.route('/admin/campaign', methods=['GET', 'POST'])
@login_required
@superadmin_required
def campaign():
    month = active_month()
    if not month:
        flash('אין חודש פעיל', 'danger')
        return redirect(url_for('admin'))
    conn = get_db()
    buckets = campaign_eligibility(conn, month['id'])
    month_name = HEB_MONTHS[datetime.datetime.now().month]
    if request.method == 'POST':
        action = request.form.get('action')
        who = session.get('display_name') or session.get('username', '')
        if action == 'test':
            # render for the first eligible recipient (or a sample) and send to the office inbox
            sample = buckets['email'][0][0] if buckets['email'] else None
            demo = sample or {'name': 'בדיקה', 'brand': 'גאיה', 'is_midwife': 0,
                              'premium_last_year': 750, 'id_number': '', 'email': ''}
            body = render_renewal_email(demo, month_name)
            ok = send_campaign_email(EMAIL_CONFIG['username'], 'בדיקה — מייל חידוש', body)
            conn.close()
            flash('נשלחה בדיקה לתיבת המשרד' if ok else 'שליחת הבדיקה נכשלה', 'success' if ok else 'danger')
            return redirect(url_for('campaign'))
        if action == 'email_send':
            # already-sent-today are skipped (multi-touch safety)
            today = datetime.date.today().isoformat()
            recips = [(c, e) for (c, e) in buckets['email']
                      if (c['email_sent_date'] or '') != today]
            conn.close()
            if not within_business_hours() and request.form.get('force') != '1':
                flash('מחוץ לשעות הפעילות (א׳–ה׳ 8:00–16:00). לשליחה בכל זאת סמן "כפה".', 'warning')
                return redirect(url_for('campaign'))
            threading.Thread(target=_run_email_blast,
                             args=(month['id'], month_name, recips, who), daemon=True).start()
            flash(f'השליחה החלה ברקע — {len(recips)} מיילים. כל שליחה מתועדת ביומן.', 'info')
            return redirect(url_for('campaign'))
        conn.close()
        return redirect(url_for('campaign'))
    counts = {k: len(v) for k, v in buckets.items()}
    conn.close()
    return render_template('campaign.html', counts=counts, month_name=month_name,
                           in_hours=within_business_hours())


# ── SITE123 email enrichment: recover customer-submitted email/phone from Gmail ──
_site123_state = {'running': False, 'done': None, 'report': None, 'started': None}
_SITE123_EXCLUDE = ('gaia-ins.co.il', 'winner-ins.co.il', 'site123.com', 'do-not-reply')

def parse_site123_email(text):
    """Extract the customer-submitted fields from a SITE123 website-form email body."""
    cem = ''
    for l in text.split('\n'):
        if 'דואר' in l or 'אלקטרוני' in l or 'מייל' in l:
            m = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', l)
            if m and not any(x in m.group(0) for x in _SITE123_EXCLUDE):
                cem = m.group(0); break
    if not cem:
        for m in re.finditer(r'[\w.+-]+@[\w-]+\.[\w.]+', text):
            if not any(x in m.group(0) for x in _SITE123_EXCLUDE):
                cem = m.group(0); break
    mid = re.search(r'ת\.?ז[^\d]{0,15}(\d{5,9})', text)
    mph = re.search(r'טלפון[^\d]{0,15}(0\d[\d\- ]{6,})', text)
    mnm = re.search(r'שם מלא\s*:?\s*(.+?)\s*(?:דואר|אלקטרוני|מייל|מספר|טלפון|ת\.?ז|:|$)', text)
    return {'email': cem,
            'id': (mid.group(1).zfill(9) if mid else ''),
            'phone': (re.sub(r'\D', '', mph.group(1)) if mph else ''),
            'name': (mnm.group(1).strip() if mnm else '')}

def _site123_body(msg):
    import html as _html
    txt = ''
    for p in msg.walk():
        ct = p.get_content_type()
        if ct in ('text/plain', 'text/html'):
            try:
                payload = p.get_payload(decode=True).decode(p.get_content_charset() or 'utf-8', 'replace')
            except Exception:
                continue
            if ct == 'text/html':
                payload = re.sub(r'<[^>]+>', ' ', payload)
            txt += '\n' + _html.unescape(payload)
    return txt

def backfill_site123(days_back=400, limit=None):
    """Scan Gmail for SITE123 form emails and enrich insureds/customers by ת"ז. The
    customer-submitted contact details are authoritative: email is filled/overridden, and
    phone is overridden on a mismatch. Read-only on Gmail; writes only to the DB."""
    from email.utils import parsedate_to_datetime
    import imaplib, email as _email, datetime as _dt
    cfg = EMAIL_CONFIG
    rep = {'scanned': 0, 'parsed_ok': 0, 'unique': 0, 'matched': 0,
           'emails_updated': 0, 'phones_updated': 0, 'unmatched': 0}
    if not cfg['username'] or not cfg['password']:
        return rep
    mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
    mail.login(cfg['username'], cfg['password'])
    mail.select('"[Gmail]/All Mail"', readonly=True)
    since = (_dt.datetime.now() - _dt.timedelta(days=days_back)).strftime('%d-%b-%Y')
    typ, data = mail.search(None, f'(FROM "site123.com" SINCE {since})')
    ids = data[0].split() if data and data[0] else []
    if limit:
        ids = ids[-limit:]
    latest = {}
    for num in ids:
        rep['scanned'] += 1
        try:
            typ, d = mail.fetch(num, '(RFC822)')
            msg = _email.message_from_bytes(d[0][1])
            p = parse_site123_email(_site123_body(msg))
        except Exception:
            continue
        if not (p['id'] and is_israeli_id(p['id']) and p['email'] and '@' in p['email']):
            continue
        rep['parsed_ok'] += 1
        try:
            dt = parsedate_to_datetime(msg.get('Date'))
        except Exception:
            dt = None
        k = p['id']
        if k not in latest or (dt and latest[k][0] and dt > latest[k][0]):
            latest[k] = (dt, p['email'], p['phone'], p['name'])
    mail.logout()
    rep['unique'] = len(latest)
    conn = get_db()
    def norm(s):
        return re.sub(r'\D', '', str(s or ''))
    for idn, (dt, em, ph_, nm) in latest.items():
        z = idn.lstrip('0')
        matched = em_upd = ph_upd = False
        for tbl in ('insureds', 'customers'):
            for r in conn.execute(f"SELECT id, email, phone FROM {tbl} WHERE ltrim(COALESCE(id_number,''),'0')=?", (z,)).fetchall():
                matched = True
                sets, vals = [], []
                if em and (r['email'] or '').strip().lower() != em.lower():
                    sets.append('email=?'); vals.append(em); em_upd = True
                if ph_ and norm(r['phone'])[-9:] != norm(ph_)[-9:]:
                    sets.append('phone=?'); vals.append(ph_); ph_upd = True
                if sets:
                    vals.append(r['id'])
                    conn.execute(f"UPDATE {tbl} SET {','.join(sets)} WHERE id=?", vals)
        rep['matched' if matched else 'unmatched'] += 1
        rep['emails_updated'] += 1 if em_upd else 0
        rep['phones_updated'] += 1 if ph_upd else 0
    conn.commit()
    conn.close()
    return rep

def _run_site123_backfill(days_back):
    _site123_state.update(running=True, started=datetime.datetime.now().strftime('%H:%M'), report=None)
    try:
        _site123_state['report'] = backfill_site123(days_back)
    except Exception as e:
        _site123_state['report'] = {'error': str(e)}
    finally:
        _site123_state.update(running=False, done=datetime.datetime.now().strftime('%H:%M'))

@app.route('/admin/backfill-site123', methods=['POST'])
@login_required
@superadmin_required
def admin_backfill_site123():
    if _site123_state['running']:
        flash('סריקת SITE123 כבר רצה', 'warning')
        return redirect(url_for('admin'))
    days = int(request.form.get('days', 400))
    threading.Thread(target=_run_site123_backfill, args=(days,), daemon=True).start()
    flash('סריקת SITE123 החלה ברקע — רענן לעדכון', 'info')
    return redirect(url_for('admin'))


def _enrich_from_bytes(conn, filename, data):
    """Parse a contacts export (CSV/XLSX bytes) and fill MISSING email/phone on
    insureds+customers. Content-based parsing (email by @, ת"ז by checksum, phone by
    pattern), matching by ת"ז then phone (file phone normalised to leading-0). Returns
    (emails_filled, phones_filled); the caller commits."""
    rows = []
    fname = (filename or '').lower()
    if fname.endswith(('.xlsx', '.xls')):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for r in ws.iter_rows(values_only=True):
                rows.append(['' if c is None else str(c) for c in r])
    else:
        import csv as _csv, io as _io
        text = data.decode('utf-8', 'replace') if isinstance(data, (bytes, bytearray)) else str(data)
        for r in _csv.reader(_io.StringIO(text)):
            rows.append(r)

    def z(s):
        return re.sub(r'\D', '', str(s or '')).lstrip('0')

    def ph0(s):
        d = re.sub(r'\D', '', str(s or ''))
        return ('0' + d[-9:]) if len(d) >= 9 else ''

    id2, ph2email = {}, {}
    for r in rows:
        cells = [str(c or '').strip() for c in r]
        email = next((c for c in cells if '@' in c and '.' in c and 'http' not in c), '')
        if '@' not in email:
            email = ''
        idc = next((c for c in cells if 5 <= len(re.sub(r'\D', '', c)) <= 9 and is_israeli_id(c)), '')
        phone = next((c for c in cells if re.fullmatch(r'0?5\d{8}', re.sub(r'\D', '', c))), '')
        zi = z(idc)
        if zi:
            d = id2.setdefault(zi, {})
            if email:
                d.setdefault('email', email)
            if phone:
                d.setdefault('phone', ph0(phone))
        if phone and email:
            ph2email.setdefault(ph0(phone)[-9:], email)

    em_fill = ph_fill = 0
    for tbl in ('insureds', 'customers'):
        for row in conn.execute(f"SELECT id, id_number, email, phone FROM {tbl}").fetchall():
            rec = id2.get(z(row['id_number']), {})
            cur_em = (row['email'] or '').strip()
            cur_ph = (row['phone'] or '').strip()
            if (not cur_em or '@' not in cur_em):
                new_em = rec.get('email') or ph2email.get(re.sub(r'\D', '', cur_ph)[-9:], '')
                if new_em:
                    conn.execute(f"UPDATE {tbl} SET email=? WHERE id=?", (new_em, row['id']))
                    em_fill += 1
            if not cur_ph and rec.get('phone'):
                conn.execute(f"UPDATE {tbl} SET phone=? WHERE id=?", (rec['phone'], row['id']))
                ph_fill += 1
    return em_fill, ph_fill


def _enrich_from_file(conn, f):
    """Thin wrapper: enrich from an uploaded Flask file object."""
    return _enrich_from_bytes(conn, f.filename, f.read())


def _enrich_dir():
    # Lazy (ATTACHMENTS_DIR is defined later in the module).
    return os.path.join(ATTACHMENTS_DIR, 'enrich')

_enrich_state = {'running': False, 'started': None, 'done': None, 'report': None}

def _run_post_load_enrichment(days_back=400):
    """Background job after a month load: cross-reference SITE123 (Gmail) + the stored
    contact files, keyed by ת"ז, to fill emails for the newly-loaded customers."""
    _enrich_state.update(running=True, started=datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), report=None)
    rep = {'site123': None, 'files': []}
    try:
        rep['site123'] = backfill_site123(days_back)
    except Exception as e:
        rep['site123'] = {'error': str(e)}
    try:
        conn = get_db()
        d = _enrich_dir()
        if os.path.isdir(d):
            for fn in sorted(os.listdir(d)):
                fp = os.path.join(d, fn)
                if not os.path.isfile(fp):
                    continue
                try:
                    with open(fp, 'rb') as fh:
                        em, ph = _enrich_from_bytes(conn, fn, fh.read())
                    rep['files'].append({'file': fn, 'emails': em, 'phones': ph})
                except Exception as e:
                    rep['files'].append({'file': fn, 'error': str(e)})
        conn.commit()
        conn.close()
    except Exception as e:
        rep['files_error'] = str(e)
    _enrich_state.update(running=False, done=datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), report=rep)

@app.route('/api/enrich-files', methods=['GET', 'POST'])
def api_enrich_files():
    """Token-authed management of the stored enrichment files that auto-run on each month
    load. POST uploads/refreshes a file; GET lists them."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = _enrich_dir()
    os.makedirs(d, exist_ok=True)
    if request.method == 'POST':
        f = request.files.get('file')
        if not f:
            return jsonify({'error': 'missing file'}), 400
        safe = re.sub(r'[\\/*?:"<>|]', '_', f.filename or 'file.csv')
        f.save(os.path.join(d, safe))
    files = [{'name': fn, 'size': os.path.getsize(os.path.join(d, fn))}
             for fn in sorted(os.listdir(d)) if os.path.isfile(os.path.join(d, fn))]
    return jsonify({'ok': True, 'files': files})

@app.route('/api/enrich-status')
def api_enrich_status():
    """Token-authed status/report of the last auto-enrichment run."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    return jsonify(_enrich_state)

@app.route('/api/policy/force-test', methods=['POST'])
def api_policy_force_test():
    """Token-authed TEST helper: force a specific ת"ז's renewal PDF into the delivery queue
    regardless of the 48h window (the customer must still be marked 'חודש'), to test a
    delivery end-to-end. {q: ת"ז} adds; {clear: true} resets. In-memory (clears on redeploy)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    if data.get('clear'):
        _POLICY_FORCE_IDS.clear()
    q = re.sub(r'\D', '', str(data.get('q') or '')).lstrip('0')
    if q:
        _POLICY_FORCE_IDS.add(q)
    return jsonify({'ok': True, 'forced_ids': sorted(_POLICY_FORCE_IDS)})

@app.route('/api/policy/send-live', methods=['POST'])
def api_policy_send_live():
    """Send ONE specific customer to their REAL contact while the system stays in test mode:
    resets that ת"ז's renewal docs' sent flags (so the latest re-enters the queue) and marks
    the ת"ז as a live override. {q: ת"ז} to arm; {clear: true} to reset the override list."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    if data.get('clear'):
        _POLICY_LIVE_IDS.clear()
        return jsonify({'ok': True, 'live_ids': []})
    q = re.sub(r'\D', '', str(data.get('q') or '')).lstrip('0')
    if not q:
        return jsonify({'error': 'need q'}), 400
    conn = get_db()
    reset = conn.execute(
        """UPDATE policy_documents SET whatsapp_sent_at=NULL, email_sent_at=NULL
           WHERE id IN (SELECT pd.id FROM policy_documents pd
                        JOIN policy_records pr ON pr.policy_document_id=pd.id
                        WHERE ltrim(COALESCE(pr.insured_id,''),'0')=?)""", (q,))
    conn.commit()
    conn.close()
    _POLICY_LIVE_IDS.add(q)
    return jsonify({'ok': True, 'live_ids': sorted(_POLICY_LIVE_IDS), 'docs_reset': reset.rowcount})

@app.route('/api/policy/inbox-search', methods=['POST'])
def api_policy_inbox_search():
    """Token-authed IMAP diagnostic (read-only): find emails whose SUBJECT contains {q}
    (last 14 days) + list recent ComposeDoc subjects — to see the real sender/subject of a
    policy the scanner missed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    q = str((request.get_json(silent=True) or {}).get('q') or '').strip()
    cfg = EMAIL_CONFIG
    out = {'query': q, 'matches': [], 'recent_composedoc': []}
    try:
        import imaplib, email as _email
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=14)).strftime('%d-%b-%Y')
        if q:
            _, data = mail.search(None, f'(SINCE {since} SUBJECT "{q}")')
            for mid in (data[0].split() if data and data[0] else [])[-10:]:
                _, hd = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT DATE)])')
                h = _email.message_from_bytes(hd[0][1])
                out['matches'].append({'from': decode_str(h.get('From', '')),
                                       'subject': decode_str(h.get('Subject', '')), 'date': h.get('Date', '')})
        ids = _search_policy_emails(mail, since)
        out['composedoc_count'] = len(ids)
        for mid in ids[-6:]:
            _, hd = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE)])')
            h = _email.message_from_bytes(hd[0][1])
            out['recent_composedoc'].append({'subject': decode_str(h.get('Subject', '')), 'date': h.get('Date', '')})
        mail.logout()
    except Exception as e:
        out['error'] = str(e)
    return jsonify(out)

@app.route('/api/policy/inspect-email', methods=['POST'])
def api_policy_inspect_email():
    """Token-authed: open recent ComposeDoc emails (optionally filtered by subject substring),
    parse their PDF attachment, and return filename + parsed fields — to see how to extract the
    policy number for subject-less 'new policy' emails."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    subj = str(data.get('subject_contains') or '').strip()
    days = int(data.get('days') or 3)
    cfg = EMAIL_CONFIG
    out = {'emails': []}
    try:
        import imaplib, email as _email
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        ids = _search_policy_emails(mail, since)
        for mid in ids[-10:]:
            _, full = mail.fetch(mid, '(BODY.PEEK[])')
            msg = _email.message_from_bytes(full[0][1])
            subject = decode_str(msg.get('Subject', ''))
            if subj and subj not in subject:
                continue
            info = {'subject': subject, 'date': msg.get('Date', ''), 'attachments': []}
            for part in msg.walk():
                cd = str(part.get('Content-Disposition', ''))
                if 'attachment' not in cd and part.get_content_type() != 'application/octet-stream':
                    continue
                fn = decode_str(part.get_filename() or '')
                if not fn:
                    continue
                dbytes = part.get_payload(decode=True)
                info['attachments'].append({'filename': fn,
                                            'parsed': (parse_harel_policy_pdf(dbytes) if dbytes else {})})
            out['emails'].append(info)
        mail.logout()
    except Exception as e:
        out['error'] = str(e)
    return jsonify(out)

@app.route('/api/policy/pdf-text', methods=['POST'])
def api_policy_pdf_text():
    """Token-authed: find the ComposeDoc email whose PDF matches ת"ز {q} and return its raw
    text lines + parsed fields — to locate fields the parser doesn't extract (e.g. מס' תוספת)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    data = request.get_json(silent=True) or {}
    q = re.sub(r'\D', '', str(data.get('q') or '')).lstrip('0')
    days = int(data.get('days') or 8)
    cfg = EMAIL_CONFIG
    out = {'q': q, 'found': False}
    try:
        import imaplib, email as _email
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        ids = _search_policy_emails(mail, since)
        for mid in reversed(ids):
            _, full = mail.fetch(mid, '(BODY.PEEK[])')
            msg = _email.message_from_bytes(full[0][1])
            for part in msg.walk():
                cd = str(part.get('Content-Disposition', ''))
                if 'attachment' not in cd and part.get_content_type() != 'application/octet-stream':
                    continue
                dbytes = part.get_payload(decode=True)
                if not dbytes:
                    continue
                fields = parse_harel_policy_pdf(dbytes)
                if re.sub(r'\D', '', str(fields.get('insured_id', ''))).lstrip('0') == q:
                    out.update(found=True, parsed=fields, subject=decode_str(msg.get('Subject', '')),
                               date=msg.get('Date', ''), filename=decode_str(part.get_filename() or ''),
                               lines=_policy_pdf_lines(dbytes, 60))
                    break
            if out['found']:
                break
        mail.logout()
    except Exception as e:
        out['error'] = str(e)
    return jsonify(out)

@app.route('/api/policy/scan', methods=['POST'])
def api_policy_scan():
    """Token-authed on-demand scan for new Harel policy PDFs (so a just-arrived renewal is
    processed now instead of waiting for the 5-minute poll). Runs in the background."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    def _scan_all():
        check_policy_documents()
        check_join_forms()
        check_renewal_forms()
    threading.Thread(target=_scan_all, daemon=True).start()
    return jsonify({'ok': True, 'scanning': True})

@app.route('/api/lead/attachments')
def lead_attachments():
    """Leads whose customer-uploaded document is on the server but not yet filed to OneDrive
    (token-authed; the local wa-sender fetches + saves them to the לקוחות folder)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    rows = conn.execute(
        """SELECT id, id_number, name, lead_doc_path FROM customers
           WHERE import_source='join_form' AND lead_doc_path IS NOT NULL AND lead_doc_path!=''
                 AND (lead_doc_saved IS NULL OR lead_doc_saved='')""").fetchall()
    conn.close()
    out = []
    for r in rows:
        ext = os.path.splitext(r['lead_doc_path'])[1] or '.jpg'
        out.append({'cust_id': r['id'], 'id_number': r['id_number'], 'name': r['name'],
                    'filename': f"מסמך {r['id_number']}{ext}", 'url': f"/api/lead/attachment/{r['id']}"})
    return jsonify({'items': out, 'count': len(out)})

@app.route('/api/lead/attachment/<int:cid>')
def lead_attachment(cid):
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    r = conn.execute("SELECT lead_doc_path FROM customers WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not r or not r['lead_doc_path'] or not os.path.exists(r['lead_doc_path']):
        return jsonify({'error': 'not found'}), 404
    return send_file(r['lead_doc_path'], as_attachment=True,
                     download_name=os.path.basename(r['lead_doc_path']))

@app.route('/api/lead/attachment/saved', methods=['POST'])
def lead_attachment_saved():
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    cid = (request.get_json(silent=True) or {}).get('cust_id')
    if not cid:
        return jsonify({'error': 'need cust_id'}), 400
    conn = get_db()
    conn.execute("UPDATE customers SET lead_doc_saved=? WHERE id=?",
                 (datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), cid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/policy/trace', methods=['POST'])
def api_policy_trace():
    """Token-authed diagnosis of why a policy is / isn't in the auto-delivery queue.
    Accepts {q: policy number | ת"ز | name}; returns the matching documents + customers."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    q = str((request.get_json(silent=True) or {}).get('q') or '').strip()
    if not q:
        return jsonify({'error': 'need q'}), 400
    qz = re.sub(r'\D', '', q).lstrip('0')
    like = f'%{q}%'
    conn = get_db()
    cutoff = (datetime.datetime.now() - datetime.timedelta(hours=POLICY_SEND_WINDOW_HOURS)
              ).strftime('%Y-%m-%d %H:%M')
    docs = []
    for d in conn.execute(
        """SELECT pd.id, pd.policy_number, pd.received_at, pd.whatsapp_sent_at, pd.email_sent_at,
                  pr.insured_id, pr.insured_name, pr.doc_type_label
           FROM policy_documents pd LEFT JOIN policy_records pr ON pr.policy_document_id=pd.id
           WHERE ltrim(COALESCE(pd.policy_number,''),'0')=? OR ltrim(COALESCE(pr.insured_id,''),'0')=?
                 OR pr.insured_name LIKE ? ORDER BY pd.id DESC LIMIT 12""", (qz, qz, like)).fetchall():
        docs.append({'doc_id': d['id'], 'policy_number': d['policy_number'], 'received_at': d['received_at'],
                     'insured_id': d['insured_id'], 'insured_name': d['insured_name'],
                     'doc_type': d['doc_type_label'], 'is_renewal': is_renewal_doc(d['doc_type_label']),
                     'within_48h': bool(d['received_at'] and d['received_at'] >= cutoff),
                     'whatsapp_sent_at': d['whatsapp_sent_at'], 'email_sent_at': d['email_sent_at']})
    custs = []
    for c in conn.execute(
        """SELECT c.id, m.name mname, m.is_active, c.name, c.id_number, c.status, c.brand
           FROM customers c LEFT JOIN months m ON m.id=c.month_id
           WHERE ltrim(COALESCE(c.id_number,''),'0')=? OR c.name LIKE ?
           ORDER BY c.month_id DESC LIMIT 12""", (qz, like)).fetchall():
        custs.append({'cust_id': c['id'], 'month': c['mname'], 'active_month': bool(c['is_active']),
                      'name': c['name'], 'id_number': c['id_number'], 'status': c['status'],
                      'marked_renewed': (c['status'] in ('חודש', 'חודש - בוצעה שיחת מכירה')), 'brand': c['brand']})
    conn.close()
    return jsonify({'query': q, 'window_cutoff': cutoff, 'policy_documents': docs, 'customers': custs})


@app.route('/admin/enrich-contacts', methods=['POST'])
@login_required
@superadmin_required
def enrich_contacts():
    """Fill missing email/phone on insureds+customers from an uploaded contacts export."""
    f = request.files.get('file')
    if not f:
        flash('חסר קובץ', 'danger')
        return redirect(url_for('admin'))
    try:
        conn = get_db()
        em_fill, ph_fill = _enrich_from_file(conn, f)
        conn.commit()
        conn.close()
        flash(f'העשרה מ-{f.filename}: מולאו {em_fill} מיילים ו-{ph_fill} טלפונים (לפי ת"ז/טלפון).', 'success')
    except Exception as e:
        flash(f'שגיאה בהעשרה: {e}', 'danger')
    return redirect(url_for('admin'))


@app.route('/api/enrich-contacts', methods=['POST'])
def api_enrich_contacts():
    """Token-authed enrichment (same logic as /admin/enrich-contacts) — returns fill counts."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'missing file'}), 400
    try:
        conn = get_db()
        em_fill, ph_fill = _enrich_from_file(conn, f)
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'file': f.filename, 'emails_filled': em_fill, 'phones_filled': ph_fill})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/renewal', methods=['POST'])
def api_renewal():
    """Receives form submissions from winner-ins.co.il/renew and gaia-ins.co.il/renew"""
    data = request.json or request.form.to_dict()

    id_number = normalize_id_number(data.get('id_number') or data.get('id'))
    phone = str(data.get('phone') or data.get('telephone') or '').strip()
    name = str(data.get('name') or data.get('full_name') or '').strip()
    email = str(data.get('email') or '').strip()
    installments = str(data.get('installments') or data.get('payment_installments') or '').strip()
    payment_method = str(data.get('payment_method') or '').strip()
    comments = str(data.get('comments') or '').strip()
    brand = str(data.get('brand') or '').strip()

    if not id_number and not phone:
        return jsonify({'ok': False, 'error': 'missing id or phone'}), 400

    month = active_month()
    if not month:
        return jsonify({'ok': False, 'error': 'no active month'}), 400

    conn = get_db()
    customer = None
    if id_number:
        norm_id = id_number.lstrip('0')
        customer = conn.execute(
            "SELECT * FROM customers WHERE month_id=? AND ltrim(id_number,'0')=?",
            (month['id'], norm_id)
        ).fetchone()
    if not customer and phone:
        clean_phone = phone.replace('-', '').replace(' ', '')
        customer = conn.execute(
            "SELECT * FROM customers WHERE month_id=? AND replace(replace(phone,'-',''),' ','')=?",
            (month['id'], clean_phone)
        ).fetchone()

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

    if customer:
        conn.execute("""UPDATE customers SET status='טופס התקבל',
                        form_email=?, form_installments=?, form_payment_method=?,
                        form_received_at=?, form_comments=?, status_changed_at=?
                        WHERE id=?""",
                     (email, installments, payment_method, now, comments, now, customer['id']))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'matched': True, 'customer': customer['name']})
    else:
        conn.execute("""INSERT INTO customers
            (month_id, name, id_number, phone, brand, status,
             form_email, form_installments, form_payment_method, form_received_at, form_comments,
             whatsapp_source)
            VALUES (?,?,?,?,?,'טופס התקבל',?,?,?,?,?,?)""",
            (month['id'], name, id_number, phone, brand,
             email, installments, payment_method, now, comments,
             'ווינר' if brand == 'אופיר' else None))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'matched': False, 'note': 'added as new'})


# ── Email parsing helpers ────────────────────────────────────

def decode_str(s):
    """Decode MIME-encoded email header string."""
    parts = decode_header(s)
    result = ''
    for b, enc in parts:
        if isinstance(b, bytes):
            result += b.decode(enc or 'utf-8', errors='replace')
        else:
            result += b
    # MIME header folding can leave embedded \r\n — breaks HTTP headers (Content-Disposition) if left in
    return re.sub(r'[\r\n]+', ' ', result).strip()

def parse_renewal_email(msg_text, subject=''):
    """
    Parse form fields from renewal emails.
    Format: fields and values are space-separated in sequence (no colons).
    e.g. 'שם מלא ארנה אדם מספר ת.ז 056062608 אימייל ...'
    """
    # Known field tokens. Order doesn't matter — the splitter sorts longest-first so
    # that e.g. 'כתובת חונך' wins over 'כתובת' and 'ת.ז המצהיר' over 'ת.ז'.
    FIELDS = [
        'שם מלא', 'מספר ת.ז', 'birth_date', 'אימייל', 'טלפון',
        'coverage_option', 'מספר תשלומים', 'מספר פוליסה',
        'אמצעי גביה', 'מספר כרטיס', 'תוקף כרטיס',
        'ת.ז בעל הכרטיס', 'שם בעל הכרטיס', 'card_holder_name', 'הכרטיס על שם המבוטח',
        'מקצועות נוספים', 'הערות',
        # Join / underwriting forms
        'כתובת חונך', 'ת.ז חונך', 'שם החונך', 'כתובת', 'עיר',
        'תאריך לידה', 'מגדר', 'מצב משפחתי', 'מספר ילדים', 'מקצוע',
        'תאריך תחילת ביטוח', 'מקצועות', 'מוסד / ארגון',
        'עוסק כחברה', 'שם החברה', 'חבר בארגון מקצועי', 'שם הארגון',
        'חבר בקופת חולים', 'שם קופת חולים', 'שכיר', 'שם המעסיק',
        'מבוטח ב-5 שנים האחרונות', 'חברת ביטוח קודמת', 'ביטוח בוטל בעבר',
        'תביעות ב-5 שנים האחרונות', 'תנאים מיוחדים / החרגות',
        'תביעות עתידיות ידועות', 'פירוט היסטוריה', 'מעורב בהונאה',
        'פגיעה בפרטיות', 'הטרדה', 'נמנע מעיסוק במקצוע', 'פירוט האיסור',
        'שם המצהיר', 'ת.ז המצהיר', 'תאריך הצהרה', 'הסכמה לשיווק',
    ]

    # Build regex that splits on any known field name — longest first so a longer
    # field name is never swallowed by a shorter one that prefixes it.
    escaped = [re.escape(f) for f in sorted(FIELDS, key=len, reverse=True)]
    splitter = '(' + '|'.join(escaped) + ')'
    parts = re.split(splitter, msg_text)

    result = {}
    i = 1
    while i < len(parts) - 1:
        key = parts[i].strip()
        val = parts[i + 1].strip() if i + 1 < len(parts) else ''
        # Remove leading/trailing em-dash placeholder
        val = val.strip('— ').strip()
        result[key] = val
        i += 2

    # Brand from subject line: "גאיה | ..." or "ווינר | ..."
    brand = ''
    if 'גאיה' in subject:
        brand = 'גאיה'
    elif 'ווינר' in subject:
        brand = 'ווינר'

    return {
        # Payment-update forms carry no "שם מלא" — fall back to the card holder's name.
        'name': (result.get('שם מלא', '') or result.get('שם בעל הכרטיס', '')
                 or result.get('card_holder_name', '')),
        'id_number': result.get('מספר ת.ז', ''),
        'phone': result.get('טלפון', ''),
        'email': result.get('אימייל', ''),
        'installments': result.get('מספר תשלומים', ''),
        'payment_method': result.get('אמצעי גביה', ''),
        'comments': result.get('הערות', ''),
        'brand': brand,
        'policy_number': result.get('מספר פוליסה', ''),
        'card_number': result.get('מספר כרטיס', ''),
        'card_expiry': result.get('תוקף כרטיס', ''),
        'card_holder_id': result.get('ת.ז בעל הכרטיס', ''),
        'coverage_option': result.get('coverage_option', ''),
        # Every field the form sent, kept verbatim so the UI can show the full
        # submission as a readable table instead of losing it.
        'all_fields': {k: v for k, v in result.items() if v},
    }

def get_email_body(msg):
    """Extract plain text body from an email message."""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            cd = str(part.get('Content-Disposition', ''))
            if ct == 'text/plain' and 'attachment' not in cd:
                charset = part.get_content_charset() or 'utf-8'
                return part.get_payload(decode=True).decode(charset, errors='replace')
        # Fallback: try HTML
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                charset = part.get_content_charset() or 'utf-8'
                html = part.get_payload(decode=True).decode(charset, errors='replace')
                return re.sub(r'<[^>]+>', ' ', html)
    else:
        charset = msg.get_content_charset() or 'utf-8'
        return msg.get_payload(decode=True).decode(charset, errors='replace')
    return ''

NO_NAME = '(ללא שם)'


def name_from_records(conn, idn):
    """The known client name for an ID. Payment-update forms often omit the name, so we
    fill it from anything we already hold — including an earlier form from the same ID —
    rather than filing it as '(ללא שם)'. The placeholder itself is never treated as a name."""
    idn = (idn or '').lstrip('0')
    if not idn:
        return ''
    for sql in ("SELECT name FROM insureds WHERE ltrim(COALESCE(id_number,''),'0')=? "
                "AND COALESCE(name,'') NOT IN ('', ?) LIMIT 1",
                "SELECT name FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? "
                "AND COALESCE(name,'') NOT IN ('', ?) ORDER BY id DESC LIMIT 1",
                "SELECT name FROM unmatched_submissions WHERE ltrim(COALESCE(id_number,''),'0')=? "
                "AND COALESCE(name,'') NOT IN ('', ?) ORDER BY id DESC LIMIT 1"):
        r = conn.execute(sql, (idn, NO_NAME)).fetchone()
        if r:
            return r['name']
    return ''


def process_renewal_data(data, message_id='', subject='', received_at=''):
    """
    Match email form data to a customer in the active month.
    - Matched → update customer, status='טופס התקבל', return customer_id
    - Not matched → save to unmatched_submissions for admin review, return None
    """
    id_number      = normalize_id_number(data.get('id_number'))
    phone          = str(data.get('phone') or '').strip()
    name           = str(data.get('name') or '').strip()
    email_val      = str(data.get('email') or '').strip()
    installments   = str(data.get('installments') or '').strip()
    payment_method = str(data.get('payment_method') or '').strip()
    comments       = str(data.get('comments') or '').strip()
    brand          = str(data.get('brand') or '').strip()
    coverage       = str(data.get('coverage_option') or '').strip()
    # PCI: never store the full card — keep only the last 4 (masked). Matches the join-form path.
    _card_digits   = re.sub(r'\D', '', str(data.get('card_number') or ''))
    card_number    = ('****' + _card_digits[-4:]) if len(_card_digits) >= 4 else ''
    card_expiry    = str(data.get('card_expiry') or '').strip()
    card_holder_id = str(data.get('card_holder_id') or '').strip()
    raw_fields     = json.dumps(data.get('all_fields') or {}, ensure_ascii=False)

    now = received_at or datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

    conn = get_db()
    month = conn.execute("SELECT * FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()

    if not month:
        conn.close()
        print('[email-sync] אין חודש פעיל')
        return None

    # No name on the form (common on payment-update forms) but we know this ID → use
    # the name we already have, so the item is identifiable instead of '(ללא שם)'.
    if not name and id_number:
        name = name_from_records(conn, id_number)
        if name:
            print(f'[email-sync] הושלם שם לפי ת.ז {id_number}: {name}')

    if not id_number and not phone:
        # No identifying info — send to admin
        conn.execute('''INSERT OR IGNORE INTO unmatched_submissions
            (received_at, subject, name, id_number, phone, email, brand, installments,
             payment_method, card_number, card_expiry, card_holder_id, coverage, comments,
             raw_fields, message_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (now, subject, name, id_number, phone, email_val, brand, installments,
             payment_method, card_number, card_expiry, card_holder_id, coverage, comments,
             raw_fields, message_id))
        conn.commit()
        conn.close()
        print('[email-sync] חסר מזהה → unmatched')
        return None

    customer = None
    if id_number:
        customer = conn.execute(
            "SELECT * FROM customers WHERE month_id=? AND ltrim(id_number,'0')=?",
            (month['id'], id_number.lstrip('0'))
        ).fetchone()
    if not customer and phone:
        clean = phone.replace('-', '').replace(' ', '')
        customer = conn.execute(
            "SELECT * FROM customers WHERE month_id=? AND replace(replace(phone,'-',''),' ','')=?",
            (month['id'], clean)
        ).fetchone()

    if customer:
        conn.execute("""UPDATE customers SET status='טופס התקבל',
                        form_email=?, form_installments=?, form_payment_method=?,
                        form_received_at=?, form_coverage=?, form_comments=?,
                        form_card_number=?, form_card_expiry=?, form_id_card_holder=?,
                        status_changed_at=?
                        WHERE id=?""",
                     (email_val, installments, payment_method, now, coverage, comments,
                      card_number, card_expiry, card_holder_id, now, customer['id']))
        conn.commit()
        cid = customer['id']
        conn.close()
        print(f'[email-sync] עודכן: {customer["name"]} → טופס התקבל')
        return cid
    else:
        # No match in current month → admin queue
        conn.execute('''INSERT OR IGNORE INTO unmatched_submissions
            (received_at, subject, name, id_number, phone, email, brand, installments,
             payment_method, card_number, card_expiry, card_holder_id, coverage, comments,
             raw_fields, message_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (now, subject, name, id_number, phone, email_val, brand, installments,
             payment_method, card_number, card_expiry, card_holder_id, coverage, comments,
             raw_fields, message_id))
        conn.commit()
        conn.close()
        print(f'[email-sync] לא זוהה: {name} → תור אדמין')
        return None

ATTACHMENTS_DIR = os.environ.get('ATTACHMENTS_DIR', os.path.join(os.path.dirname(__file__), 'attachments')).strip()
os.makedirs(ATTACHMENTS_DIR, exist_ok=True)

def _save_attachments(msg, customer_id):
    """Extract and save email attachments, record in DB."""
    saved = []
    for part in msg.walk():
        cd = str(part.get('Content-Disposition', ''))
        if 'attachment' not in cd:
            continue
        raw_fn = part.get_filename()
        if not raw_fn:
            continue
        filename = decode_str(raw_fn)
        data = part.get_payload(decode=True)
        if not data:
            continue
        cust_dir = os.path.join(ATTACHMENTS_DIR, str(customer_id))
        os.makedirs(cust_dir, exist_ok=True)
        # Avoid collisions
        safe_fn = re.sub(r'[\\/*?:"<>|]', '_', filename)
        filepath = os.path.join(cust_dir, safe_fn)
        with open(filepath, 'wb') as f:
            f.write(data)
        conn = get_db()
        conn.execute(
            'INSERT INTO customer_attachments (customer_id, filename, filepath, uploaded_at) VALUES (?,?,?,?)',
            (customer_id, filename, filepath, datetime.datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        saved.append(filename)
        print(f'[email-sync] קובץ נשמר: {filename}')
    return saved


_email_check_lock = threading.Lock()

def touch_scan_heartbeat():
    """Record that the scanner is alive/progressing (last_scan_at). Called after EACH scan step
    in email_poll_thread, so a slow-but-progressing cycle stays 'fresh' and the watchdog only
    fires on a genuine per-step stall — not on a long, healthy cycle. Best-effort."""
    try:
        hb = get_db()
        hb.execute("INSERT INTO app_kv(k,v) VALUES('last_scan_at',?) "
                   "ON CONFLICT(k) DO UPDATE SET v=excluded.v",
                   (datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),))
        hb.commit(); hb.close()
    except Exception:
        pass

def check_email_inbox(days_back=30):
    """Connect to IMAP, process renewal emails not yet seen (tracked by Message-ID in DB)."""
    if not _email_check_lock.acquire(blocking=False):
        print('[email-sync] בדיקה כבר רצה — דילוג')
        return 0
    try:
        n = _check_email_inbox_impl(days_back=days_back)
        touch_scan_heartbeat()
        return n
    finally:
        _email_check_lock.release()

def _check_email_inbox_impl(days_back=30):
    cfg = EMAIL_CONFIG
    if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
        return 0

    processed = 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')

        # Search window (routine poll passes a short window to stay fast; processed_emails dedups)
        since_date = (datetime.datetime.now() - datetime.timedelta(days=days_back)).strftime('%d-%b-%Y')
        status, data = mail.search(None, f'FROM "{cfg["sender_filter"]}" SINCE {since_date}')
        if status != 'OK':
            mail.logout()
            return 0

        conn = get_db()
        # Get month load time to filter only emails after that point
        month = conn.execute("SELECT created_at FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
        month_loaded_at = month['created_at'][:16].replace('T', ' ') if month else '2000-01-01 00:00'

        for mid in data[0].split():
            # Peek at headers — avoids marking as read
            _, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hdr_data[0][1])
            message_id = hdr.get('Message-ID', '').strip()
            subject = decode_str(hdr.get('Subject', ''))

            # Parse email date
            raw_date = hdr.get('Date', '')
            try:
                from email.utils import parsedate_to_datetime
                email_dt = parsedate_to_datetime(raw_date)
                # Convert to local system time for comparison with month_loaded_at (also local)
                email_dt_str = email_dt.astimezone().strftime('%Y-%m-%d %H:%M')
            except Exception:
                email_dt_str = '2099-01-01 00:00'

            # Skip emails that arrived before the month was loaded
            if email_dt_str < month_loaded_at:
                continue

            # Skip already processed
            if message_id and conn.execute(
                'SELECT 1 FROM processed_emails WHERE message_id=?', (message_id,)
            ).fetchone():
                continue

            if cfg['subject_filter'] and cfg['subject_filter'] not in subject:
                continue

            # Fetch full email without marking read
            _, full_data = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(full_data[0][1])
            body = get_email_body(msg)
            # Skip the automated morning monitor submissions (uptime check on the website
            # forms). They carry a sentinel in the body, so this catches every form type —
            # even ones with no name/ID (e.g. 'מינוי סוכן').
            if any(m in body for m in ('MONITOR-CHECK-DO-NOT-PROCESS', 'automated-daily-check', 'monitor-check@example.com')):
                if message_id:
                    conn.execute(
                        'INSERT OR IGNORE INTO processed_emails (message_id, processed_at) VALUES (?,?)',
                        (message_id, datetime.datetime.now().isoformat())
                    )
                    conn.commit()
                continue
            fields = parse_renewal_email(body, subject)
            cid = process_renewal_data(fields, message_id=message_id,
                                        subject=subject, received_at=email_dt_str)
            # Mark processed regardless (matched or unmatched)
            if message_id:
                conn.execute(
                    'INSERT OR IGNORE INTO processed_emails (message_id, processed_at) VALUES (?,?)',
                    (message_id, datetime.datetime.now().isoformat())
                )
                conn.commit()
            if cid:
                _save_attachments(msg, cid)
            processed += 1

        conn.close()
        mail.logout()
    except Exception as e:
        print(f'[email-sync] שגיאה: {e}')

    return processed

POLICY_DOCS_DIR = os.path.join(ATTACHMENTS_DIR, 'policies')
LEAD_DOCS_DIR = os.path.join(ATTACHMENTS_DIR, 'leads')

# ── Automatic policy delivery (stage 1: renewals) ────────────────────────────
# When a Harel renewal PDF arrives whose ת"ז matches a customer already marked
# 'חודש' (renewed) in the active month, the local wa-sender picks it up and sends
# it to the customer on BOTH channels (WhatsApp + email) with the PDF attached.
# TEST mode routes every send to Sharon only until the flag is turned off.
POLICY_AUTOSEND_TEST = os.environ.get('POLICY_AUTOSEND_TEST', '1') != '0'
POLICY_TEST_PHONE = os.environ.get('POLICY_TEST_PHONE', '0502030579')
POLICY_TEST_EMAIL = os.environ.get('POLICY_TEST_EMAIL', 'sharon@gaia-ins.co.il')
POLICY_SEND_WINDOW_HOURS = 48  # only auto-send documents received within this window

# Shared opt-in (customer-initiated, marketing-consent friendly) + Harel personal-area line.
POLICY_OPTIN = ('אם מעניין אותך לשמוע איך אפשר להגן גם עליך אישית באופן אישי (אם קורה לך אישית משהו) '
                'ולא רק על העסק, אפשר להשיב להודעה "אשמח" ואדאג שיחזרו אליך בהקדם, בלי שום התחייבות. 😊')
POLICY_HAREL_URL = 'https://www.harel-group.co.il/Pages/login-page/Login.aspx'

POLICY_WA_RENEWAL = (
    "תודה שבחרת להמשיך אצלנו את הביטוח המקצועי!\n\n"
    "*הפוליסה היא החשבונית, אפשר להעביר את זה לרואה החשבון שלך וזה מה שצריך כהוצאה מוכרת*.\n"
    "אני זמין כאן בוואטסאפ לכל שאלה או שירות.\n\n"
    "רק רציתי לציין שאנחנו בקבוצה מציעים גם פגישה אישית בשיחת טלפון, אצלך בבית, "
    "או בכל מקום שנוח לך!\nאחרי שהפתענו אותך במחיר של הביטוח הזה, אולי נצליח להפתיע "
    "אותך גם בביטוחים האחרים 😊\n\n"
    f"{POLICY_OPTIN}"
)

# ── New-business policies (WhatsApp only) ─────────────────────────────────────
# 'off' = disabled · 'test' = send to Sharon · 'live' = send to the real customer.
POLICY_NEW_MODE = os.environ.get('POLICY_NEW_MODE', 'off')
# Agent number printed on the Harel policy → agency. Ofir's new business is sent from Winner.
NEW_AGENT_BRAND = {'50185': 'גאיה', '411998': 'ווינר', '411025': 'אופיר'}

def is_new_doc(label):
    return 'חדש' in (label or '')

def _new_policy_brand_key(agent_number):
    """Agent number on the new-policy PDF → wa-sender brand key (gaia | winner). Ofir→winner."""
    brand = NEW_AGENT_BRAND.get(re.sub(r'\D', '', str(agent_number or '')))
    return None if not brand else ('gaia' if brand == 'גאיה' else 'winner')

POLICY_WA_NEW = (
    "תודה שבחרת בנו לביטוח המקצועי!\n\n"
    "*הפוליסה היא החשבונית, אפשר להעביר את זה לרואה החשבון שלך וזה מה שצריך כהוצאה מוכרת*.\n"
    "אני זמין כאן בוואטסאפ לכל שאלה או שירות.\n\n"
    "רק רציתי לציין שאנחנו בקבוצה מציעים גם פגישה אישית בשיחת טלפון, אצלך בבית, "
    "או בכל מקום שנוח לך!\nאחרי שהפתענו אותך במחיר של הביטוח הזה, אולי נצליח להפתיע "
    "אותך גם בביטוחים האחרים 😊\n\n"
    f"{POLICY_OPTIN}"
)

# Status of a join-form lead awaiting policy issuance (see the join-form pipeline below).
LEAD_STATUS = 'ממתין להפקה'

def _apply_new_occupation(conn, cust_id, pr):
    """At new-business scan time, read the insured's occupation from the policy PDF and, for a
    WINNER midwife (occupation contains 'מיילד'), flag is_midwife — BEFORE the policy is delivered,
    so it auto-files to the midwives folder + shows in the filter without any manual step (Sharon's
    rule: it must be a fully automatic pipeline). Also stores the occupation for everyone."""
    try:
        doc_id = pr['policy_document_id'] if 'policy_document_id' in pr.keys() else None
        if not doc_id:
            return
        d = conn.execute("SELECT filepath FROM policy_documents WHERE id=?", (doc_id,)).fetchone()
        fp = d['filepath'] if d else None
        if not fp or not os.path.exists(fp):
            return
        occ = extract_insured_occupation(fp)
        if not occ:
            return
        row = conn.execute("SELECT brand FROM customers WHERE id=?", (cust_id,)).fetchone()
        mw = 1 if (row and row['brand'] == 'ווינר' and 'מיילד' in occ) else None
        conn.execute("UPDATE customers SET occupation=COALESCE(NULLIF(occupation,''),?), "
                     "is_midwife=COALESCE(is_midwife,?) WHERE id=?", (occ, mw, cust_id))
        if mw:
            _sync_customer_to_insured(conn, cust_id, active=True)
            print(f'[new-midwife] סומנה מיילדת חדשה אוטומטית (cust {cust_id}, {occ[:30]})')
    except Exception as e:
        print(f'[new-midwife] {e}')

def _ensure_new_customer(conn, pr):
    """Create a customers record for a new-business policy (so it's serviceable). If a
    pending-issuance lead (from a website join form) already exists for this ת"ז, upgrade
    it in place — fill the policy number and mark it "הופק" (issued) — instead of duplicating.
    "הופק" drops the record off the work queue and out of the renewal counts."""
    idn = normalize_id_number(pr['insured_id'])
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not idn or not month:
        return
    existing = conn.execute(
        "SELECT id, status FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=?",
        (month['id'], idn.lstrip('0'))).fetchone()
    # No ת"ז match? A Harel proposal lead ("ממתין להפקה") has no ת"ז — it's keyed by the offer
    # number, which becomes the policy number on issuance. Link by policy number and fill the
    # now-known ת"ז, so the lead upgrades in place (→ 'הופק') instead of duplicating.
    if not existing and (pr['policy_number'] or ''):
        existing = conn.execute(
            "SELECT id, status FROM customers WHERE month_id=? AND policy_number=? AND status=?",
            (month['id'], pr['policy_number'], LEAD_STATUS)).fetchone()
        if existing:
            conn.execute("UPDATE customers SET id_number=COALESCE(NULLIF(id_number,''),?) WHERE id=?",
                         (idn, existing['id']))
    brand = NEW_AGENT_BRAND.get(re.sub(r'\D', '', str(pr['agent_number'] or '')), '')
    if existing:
        pn = (pr['policy_number'] or '')
        if (existing['status'] or '') == LEAD_STATUS:
            conn.execute(
                "UPDATE customers SET policy_number=COALESCE(NULLIF(policy_number,''),?), "
                "status='הופק', status_changed_at=? WHERE id=?",
                (pn, datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), existing['id']))
            _sync_customer_to_insured(conn, existing['id'], active=True)
            _apply_new_occupation(conn, existing['id'], pr)
            _resolve_form_queue(conn, idn, escalations=True)
        elif pn:
            # Self-heal: backfill a missing policy number onto an already-issued/serviced row —
            # e.g. the number was blank at first delivery, or a later month re-sync wiped it.
            conn.execute(
                "UPDATE customers SET policy_number=? "
                "WHERE id=? AND COALESCE(NULLIF(policy_number,''),'')=''",
                (pn, existing['id']))
        return
    cur = conn.execute(
        """INSERT INTO customers (month_id, policy_number, name, id_number, phone, email, brand,
                                  status, import_source)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (month['id'], pr['policy_number'], (pr['insured_name'] or ''), idn,
         re.sub(r'\D', '', str(pr['phone_mobile'] or '')), (pr['email'] or ''), brand, 'הופק', 'new_policy'))
    _apply_new_occupation(conn, cur.lastrowid, pr)
    _sync_customer_to_insured(conn, cur.lastrowid, active=True)
    _resolve_form_queue(conn, idn, escalations=True)

POLICY_EMAIL_SUBJECT = "הפוליסה המקצועית שלך"
POLICY_EMAIL_SIGN = ("—\nשרון דר\nמנהל תחום אחריות מקצועית\nגאיה, ווינר ואופיר")

def policy_email_body(name):
    """Plain-text fallback body."""
    greet = f"שלום {name}," if name else "שלום,"
    return (f"{greet}\n\n"
            "תודה שהמשכת איתנו את הביטוח המקצועי לשנה נוספת!\n"
            "מצורפת הפוליסה המחודשת. היא מהווה גם חשבונית להוצאה מוכרת, "
            "אז אפשר להעביר אותה ישירות לרואה החשבון.\n"
            f"ניתן למצוא את הפוליסה גם באזור האישי של חברת הראל - {POLICY_HAREL_URL}\n"
            "אני זמין באופן אישי לכל שאלה או בקשה — בטלפון או בוואטסאפ.\n\n"
            f"{POLICY_OPTIN}\n\n"
            f"{_seasonal_line()}\n"
            f"{POLICY_EMAIL_SIGN}")

def policy_email_html(name):
    """Right-aligned (RTL) HTML body for the policy-delivery email."""
    greet = f"שלום {name}," if name else "שלום,"
    return (
        '<div dir="rtl" style="text-align:right;font-family:Arial,Helvetica,sans-serif;'
        'font-size:15px;line-height:1.6;color:#222;">'
        f'{greet}<br><br>'
        'תודה שהמשכת איתנו את הביטוח המקצועי לשנה נוספת!<br>'
        'מצורפת הפוליסה המחודשת. היא מהווה גם חשבונית להוצאה מוכרת, '
        'אז אפשר להעביר אותה ישירות לרואה החשבון.<br>'
        'ניתן למצוא את הפוליסה גם באזור האישי של חברת הראל - '
        f'<a href="{POLICY_HAREL_URL}">{POLICY_HAREL_URL}</a><br>'
        'אני זמין באופן אישי לכל שאלה או בקשה — בטלפון או בוואטסאפ.<br><br>'
        f'{POLICY_OPTIN}<br><br>'
        f'{CAMPAIGN_CROSS_SELL}{_seasonal_signoff()}'
        '—<br>שרון דר<br>מנהל תחום אחריות מקצועית<br>גאיה, ווינר ואופיר'
        '</div>')

def new_policy_email_body(name):
    """Plain-text email body for a NEW-business policy delivery (not a renewal)."""
    greet = f"שלום {name}," if name else "שלום,"
    return (f"{greet}\n\n"
            "תודה שבחרת בנו לביטוח המקצועי!\n"
            "מצורפת הפוליסה. היא מהווה גם חשבונית להוצאה מוכרת, "
            "אז אפשר להעביר אותה ישירות לרואה החשבון.\n"
            "אני זמין באופן אישי לכל שאלה או בקשה — בטלפון או בוואטסאפ.\n\n"
            f"{POLICY_OPTIN}\n\n"
            f"{_seasonal_line()}\n"
            f"{POLICY_EMAIL_SIGN}")

def new_policy_email_html(name):
    """RTL HTML email body for a NEW-business policy delivery."""
    greet = f"שלום {name}," if name else "שלום,"
    return (
        '<div dir="rtl" style="text-align:right;font-family:Arial,Helvetica,sans-serif;'
        'font-size:15px;line-height:1.6;color:#222;">'
        f'{greet}<br><br>'
        'תודה שבחרת בנו לביטוח המקצועי!<br>'
        'מצורפת הפוליסה. היא מהווה גם חשבונית להוצאה מוכרת, '
        'אז אפשר להעביר אותה ישירות לרואה החשבון.<br>'
        'אני זמין באופן אישי לכל שאלה או בקשה — בטלפון או בוואטסאפ.<br><br>'
        f'{POLICY_OPTIN}<br><br>'
        f'{_seasonal_signoff()}'
        '—<br>שרון דר<br>מנהל תחום אחריות מקצועית<br>גאיה, ווינר ואופיר'
        '</div>')

# ── "Update your payment method" message (status 'התקבל חידוש - כ.א לא תקין') ──
# A renewal form arrived but the card was wrong / blocked / stolen, so the policy can't be
# issued. The customer gets a brand-specific link to fix their payment details; once corrected
# the form is re-received ('טופס התקבל') and the renewal completes ('חודש').
CARD_UPDATE_STATUS = 'התקבל חידוש - כ.א לא תקין'
CARD_UPDATE_LINKS = {
    'gaia':   'https://www.gaia-ins.co.il/card-update',
    'winner': 'https://www.winner-ins.co.il/card-update',
}
CARD_UPDATE_EMAIL_SUBJECT = "עדכון אמצעי גביה — הפוליסה שלך"

def card_update_wa_text(name, brand_key):
    greet = f"שלום {name}," if name else "שלום,"
    link = CARD_UPDATE_LINKS.get(brand_key, '')
    return (f"{greet}\n\n"
            "הטופס התקבל אך אמצעי הגביה אינו תקין.\n"
            "יש לעדכן גביה בקישור הבא:\n"
            f"{link}\n\n"
            "לאחר העדכון הפוליסה תופק ותשלח אליך בהקדם.\n\n"
            f"{_seasonal_line()}")

def card_update_email_body(name, brand_key):
    greet = f"שלום {name}," if name else "שלום,"
    link = CARD_UPDATE_LINKS.get(brand_key, '')
    return (f"{greet}\n\n"
            "הטופס התקבל אך אמצעי הגביה אינו תקין.\n"
            "יש לעדכן גביה בקישור הבא:\n"
            f"{link}\n\n"
            "לאחר העדכון הפוליסה תופק ותשלח אליך בהקדם.\n\n"
            f"{_seasonal_line()}\n"
            f"{POLICY_EMAIL_SIGN}")

def card_update_email_html(name, brand_key):
    greet = f"שלום {name}," if name else "שלום,"
    link = CARD_UPDATE_LINKS.get(brand_key, '')
    return (
        '<div dir="rtl" style="text-align:right;font-family:Arial,Helvetica,sans-serif;'
        'font-size:15px;line-height:1.6;color:#222;">'
        f'{greet}<br><br>'
        'הטופס התקבל אך אמצעי הגביה אינו תקין.<br>'
        'יש לעדכן גביה בקישור הבא:<br>'
        f'<a href="{link}">{link}</a><br><br>'
        'לאחר העדכון הפוליסה תופק ותשלח אליך בהקדם.<br><br>'
        f'{_seasonal_signoff()}'
        '—<br>שרון דר<br>מנהל תחום אחריות מקצועית<br>גאיה, ווינר ואופיר'
        '</div>')

def _wa_brand_key(brand):
    """CRM brand → wa-sender client key. Winner handles ווינר + אופיר numbers."""
    return 'gaia' if brand == 'גאיה' else 'winner'

def is_renewal_doc(doc_type_label):
    return 'חידוש' in (doc_type_label or '')

def _renewal_period_ok(period_start):
    """True if a renewal's coverage START belongs to the CURRENT cycle, not a stale prior-year policy
    being re-scanned. Sharon's rule (2026-09-02): a renewal's coverage begins the 1st of the month
    after its own expiry (Sept renewals → 01/10, late Aug renewals → 01/09), so the safe global guard
    is 'period_start ≥ the 1st of the CURRENT month' — this accepts every current-year renewal
    (01/09/2026, 01/10/2026 …) yet blocks last-year policies (01/09/2025). period_start is DD/MM/YYYY;
    empty/unparseable → False (don't act)."""
    try:
        parts = re.split(r'[/.\-]', str(period_start or '').strip())
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000
        ps = datetime.date(y, m, d)
    except Exception:
        return False
    t = datetime.date.today()
    return ps >= datetime.date(t.year, t.month, 1)

def _policy_pdf_lines(source, limit=60):
    """Raw get_display'd text lines of the policy-schedule page (diagnostics)."""
    try:
        pdf_src = io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
        with pdfplumber.open(pdf_src) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ''
                if any('רשימה' in h for h in t.split('\n')[:2]):
                    return [get_display(l) for l in t.split('\n')][:limit]
    except Exception:
        pass
    return []

def parse_harel_policy_pdf(source):
    """Best-effort field extraction from a Harel policy-schedule ('דף הרשימה') PDF page.
    `source` may be a file path or raw PDF bytes. Layout is consistent across doc types
    (new/renewal/cancellation/change) — same template, different coverage sections.
    Some fields (agent name, rare names with unusual glyphs) may need manual correction."""
    try:
        pdf_src = io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
        with pdfplumber.open(pdf_src) as pdf:
            text = None
            for page in pdf.pages:
                t = page.extract_text() or ''
                if any('רשימה' in h for h in t.split('\n')[:2]):
                    text = t
                    break
        if not text:
            return {}
    except Exception as e:
        print(f'[policy-parse] שגיאת קריאת PDF: {e}')
        return {}

    lines = [get_display(l) for l in text.split('\n')]
    result = {}

    for i, l in enumerate(lines):
        m = re.search(r'\(([^0-9()]+)\s*(\d+)\)', l)
        if m and ('פוליסה' in l or 'תוספת' in l):
            result['doc_type_label'] = m.group(1).strip()
            result['doc_type_code'] = m.group(2)

        if i + 1 < len(lines) and ("מס' הפוליסה" in l or "מספר הפוליסה" in l):
            data_line = lines[i + 1]
            nums = re.findall(r'\d+', data_line.replace('/', ''))
            if nums:
                result['branch'] = nums[0]
            if len(nums) >= 3:
                result['agent_number'] = nums[2]
            agent_name = re.sub(r'[\d/\-]+', '', data_line).strip(' -()"\'')
            agent_name = re.sub(r'^[א-ת]\s+', '', agent_name)
            result['agent_name'] = agent_name.strip()

        if i + 1 < len(lines) and 'שם המבוטח' in l:
            result['insured_name'] = l.split('שם המבוטח וכתובתו')[-1].strip()
            addr_lines = []
            j = i + 1
            while j < len(lines) and 'תקופת' not in lines[j] and 'תאריך תחילת' not in lines[j]:
                addr_lines.append(lines[j].strip())
                j += 1
            result['address'] = ' '.join(addr_lines)

        if i + 1 < len(lines) and 'תקופת הביטוח' in l:
            m5 = re.findall(r'\d{2}/\d{2}/\d{4}', lines[i + 1])
            if len(m5) >= 2:
                result['period_start'] = m5[0]
                result['period_end'] = m5[1]

        if i + 1 < len(lines) and 'e-mail' in l:
            data_line = lines[i + 1]
            m6 = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', data_line)
            result['email'] = m6.group(0) if m6 else ''
            rest = data_line.replace(result['email'], '') if m6 else data_line
            phones = [p.replace(' ', '') for p in re.findall(r'0\d{1,2}-?\s?\d{6,7}', rest)]
            if phones:
                result['phone_mobile'] = phones[0]
            if len(phones) > 1:
                result['phone_home'] = phones[1]

        if i + 1 < len(lines) and 'ת.ז. מבוטח' in l:
            # The data row carries several numbers — ת.ז. מבוטח, optionally ת.ז. בן/בת זוג,
            # and Harel's internal מס' מזהה. The internal id does NOT satisfy the Israeli ID
            # check digit, so we identify the real insured ת.ז by its checksum. When a spouse
            # id is also present (both valid), the insured is the rightmost column, i.e. the
            # last number on the (LTR-extracted) line.
            nums = re.findall(r'\d{5,9}', lines[i + 1])
            valid = [n for n in nums if is_israeli_id(n)]
            if valid:
                result['insured_id'] = valid[-1]
                if len(valid) > 1:
                    result['spouse_id'] = valid[0]
            elif nums:
                # No number passes the checksum (rare OCR/typo case) — best-effort last number.
                result['insured_id'] = nums[-1]

        if i + 1 < len(lines) and 'דמי ביטוח' in l and 'אשראי' in l:
            nums = re.findall(r'-?\d+\.\d{2}', lines[i + 1])
            if nums:
                result['premium'] = nums[0]
            if len(nums) > 1:
                result['total_payment'] = nums[-1]

    # Robust cancellation flag: the doc-type parenthetical can be mis-parsed, but a
    # cancellation reliably carries the "תוספת ביטול לפוליסה" header — trust that.
    full = '\n'.join(lines)
    if 'ביטול לפוליסה' in full or 'תוספת ביטול' in full:
        result['doc_type_label'] = 'ביטול'

    return result

POLICY_SENT_LABEL = 'טופל/שליחה אוטומטית'  # matches Sharon's Gmail filter label (nested: טופל → שליחה אוטומטית)

def _imap_utf7(s):
    """Encode a string to IMAP modified UTF-7 (RFC 3501) — for Gmail label/folder names."""
    import base64
    out, i = [], 0
    while i < len(s):
        if 0x20 <= ord(s[i]) <= 0x7e:
            out.append('&-' if s[i] == '&' else s[i]); i += 1
        else:
            j = i
            while j < len(s) and not (0x20 <= ord(s[j]) <= 0x7e):
                j += 1
            enc = base64.b64encode(s[i:j].encode('utf-16-be')).decode('ascii').rstrip('=').replace('/', ',')
            out.append('&' + enc + '-'); i = j
    return ''.join(out)

def _label_email(message_id, label=POLICY_SENT_LABEL, archive=True):
    """Gmail-label the Harel policy email (by Message-ID) after delivery + archive it out of the
    inbox. Best-effort; never raises. Returns True if a message was found and labelled."""
    cfg = EMAIL_CONFIG
    if not message_id or not cfg.get('password'):
        return False
    ok = False
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        typ, data = mail.search(None, 'HEADER', 'Message-ID', f'"{message_id.strip()}"')
        if typ == 'OK' and data and data[0].split():
            lbl = '"' + _imap_utf7(label) + '"'
            for num in data[0].split():
                mail.store(num, '+X-GM-LABELS', lbl)
                if archive:
                    mail.store(num, '-X-GM-LABELS', '\\Inbox')
                ok = True
        mail.logout()
    except Exception as e:
        print(f'[label] שגיאה: {e}')
    return ok

@app.route('/api/label-email', methods=['POST'])
def api_label_email():
    """Test/backfill: label + archive the Harel email of a given policy doc_id. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    doc_id = (request.get_json(silent=True) or {}).get('doc_id')
    conn = get_db()
    r = conn.execute("SELECT message_id FROM policy_documents WHERE id=?", (doc_id,)).fetchone()
    conn.close()
    if not r or not r['message_id']:
        return jsonify({'error': 'no message_id'})
    return jsonify({'labelled': _label_email(r['message_id'])})

GMAIL_SENT_LABEL = POLICY_SENT_LABEL  # single source of truth — keep both labelers on the same name
_gmail_label_lock = threading.Lock()

def _imap_utf7(s):
    """Encode a string to IMAP modified UTF-7 (for a Hebrew Gmail label name)."""
    import base64
    out, i = [], 0
    while i < len(s):
        o = ord(s[i])
        if 0x20 <= o <= 0x7e:
            out.append('&-' if s[i] == '&' else s[i]); i += 1
        else:
            j = i
            while j < len(s) and not (0x20 <= ord(s[j]) <= 0x7e):
                j += 1
            b = s[i:j].encode('utf-16-be')
            out.append('&' + base64.b64encode(b).decode('ascii').rstrip('=').replace('/', ',') + '-')
            i = j
    return ''.join(out)

def label_sent_policy_emails(limit=None):
    """Gmail-label + archive the Harel emails whose policy was already delivered, so the inbox
    shows only what still needs handling. Searches ALL MAIL (so already-archived ones aren't
    missed), adds GMAIL_SENT_LABEL, removes \\Inbox. Returns {processed, found, not_found, misses}."""
    if not _gmail_label_lock.acquire(blocking=False):
        return {'processed': 0, 'found': 0, 'not_found': 0, 'misses': []}
    try:
        cfg = EMAIL_CONFIG
        if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
            return {'processed': 0, 'found': 0, 'not_found': 0, 'misses': []}
        conn = get_db()
        q = ("SELECT id, message_id, policy_number FROM policy_documents WHERE COALESCE(message_id,'')!='' "
             "AND COALESCE(gmail_labeled,'')='' AND (COALESCE(whatsapp_sent_at,'')!='' "
             "OR COALESCE(email_sent_at,'')!='') ORDER BY id DESC")
        if limit:
            q += f" LIMIT {int(limit)}"
        rows = conn.execute(q).fetchall()
        if not rows:
            conn.close(); return {'processed': 0, 'found': 0, 'not_found': 0, 'misses': []}
        label = '"' + _imap_utf7(GMAIL_SENT_LABEL) + '"'
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        allbox = None                                    # find the \All (All Mail) folder
        try:
            typ, boxes = mail.list()
            for b in (boxes or []):
                line = b.decode('utf-8', 'replace') if isinstance(b, (bytes, bytearray)) else str(b)
                if '\\All' in line:
                    mm = re.findall(r'"([^"]*)"', line)
                    if mm:
                        allbox = mm[-1]; break
        except Exception:
            pass
        mail.select(('"%s"' % allbox) if allbox else 'INBOX')
        found = not_found = 0
        misses = []
        for r in rows:
            mid = (r['message_id'] or '').strip()
            hit = False
            try:
                typ, data = mail.search(None, 'HEADER', 'Message-ID', mid)
                for uid in (data[0].split() if typ == 'OK' else []):
                    mail.store(uid, '+X-GM-LABELS', label)
                    mail.store(uid, '-X-GM-LABELS', '\\Inbox')   # archive out of the inbox
                    hit = True
            except Exception as e:
                print(f'[gmail-label] {mid}: {e}')
            if hit:
                found += 1
                conn.execute("UPDATE policy_documents SET gmail_labeled=? WHERE id=?", (now, r['id']))
            else:
                not_found += 1
                misses.append(r['policy_number'])          # left un-marked → retried next run
        conn.commit(); conn.close()
        mail.logout()
        return {'processed': len(rows), 'found': found, 'not_found': not_found, 'misses': misses[:40]}
    except Exception as e:
        print(f'[gmail-label] שגיאה: {e}')
        return {'processed': 0, 'found': 0, 'not_found': 0, 'misses': [], 'error': str(e)}
    finally:
        _gmail_label_lock.release()

GMAIL_CERT_SENT_LABEL = 'אישורי ביטוח נשלחו'

def label_sent_cert_emails(limit=None, extra_tickets=None):
    """Gmail-label + archive the Harel certificate emails whose cert was delivered, moving them out
    of the inbox into 'אישורי ביטוח נשלחו' (Gmail auto-creates the label) — so the inbox shows only
    certificates still awaiting handling (the operational 'בקרה'). Finds each email by its Harel
    ticket via Gmail search (X-GM-RAW), robust even without a stored Message-ID. `extra_tickets`
    labels emails NOT tracked in cert_requests (e.g. manually-sent ones). Returns counts."""
    if not _gmail_label_lock.acquire(blocking=False):
        return {'processed': 0, 'found': 0, 'not_found': 0}
    try:
        cfg = EMAIL_CONFIG
        if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
            return {'processed': 0, 'found': 0, 'not_found': 0}
        conn = get_db()
        q = ("SELECT id, ticket FROM cert_requests WHERE COALESCE(ticket,'')!='' "
             "AND COALESCE(cert_labeled,'')='' AND COALESCE(wa_sent_at,'')!='' ORDER BY id DESC")
        want = {}                                        # ticket -> cert_requests.id (or None)
        for r in conn.execute(q).fetchall():
            want[r['ticket']] = r['id']
        for t in (extra_tickets or []):
            t = (t or '').strip()
            if t and t not in want:
                want[t] = None
        if not want:
            conn.close(); return {'processed': 0, 'found': 0, 'not_found': 0}
        label = '"' + _imap_utf7(GMAIL_CERT_SENT_LABEL) + '"'
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        allbox = None
        try:
            typ, boxes = mail.list()
            for b in (boxes or []):
                line = b.decode('utf-8', 'replace') if isinstance(b, (bytes, bytearray)) else str(b)
                if '\\All' in line:
                    mm = re.findall(r'"([^"]*)"', line)
                    if mm:
                        allbox = mm[-1]; break
        except Exception:
            pass
        mail.select(('"%s"' % allbox) if allbox else 'INBOX')
        # Scan the actual Harel cert emails and match each by the ticket in its body (Gmail can't
        # reliably search the hex ticket inside a URL, so we read the bodies).
        since = (datetime.date.today() - datetime.timedelta(days=int(limit or 10))).strftime('%d-%b-%Y')
        typ, data = mail.search(None, f'FROM "{HAREL_CERT_SENDER}" SINCE {since}')
        done = set()
        for uid in (data[0].split() if typ == 'OK' and data and data[0] else []):
            try:
                _, fd = mail.fetch(uid, '(BODY.PEEK[])')
                msg = email_lib.message_from_bytes(fd[0][1])
                html = None
                for part in msg.walk():
                    if part.get_content_type() == 'text/html':
                        try: html = part.get_content()
                        except Exception:
                            pl = part.get_payload(decode=True); html = pl.decode('utf-8', 'replace') if pl else None
                        break
                tk = _parse_harel_cert('', html or '')['ticket']
                if tk and tk in want and tk not in done:
                    mail.store(uid, '+X-GM-LABELS', label)
                    mail.store(uid, '-X-GM-LABELS', '\\Inbox')   # archive out of the inbox
                    done.add(tk)
                    if want[tk]:
                        conn.execute("UPDATE cert_requests SET cert_labeled=? WHERE id=?", (now, want[tk]))
            except Exception as e:
                print(f'[cert-label] uid {uid}: {e}')
        conn.commit(); conn.close()
        mail.logout()
        return {'processed': len(want), 'found': len(done), 'not_found': len(want) - len(done)}
    except Exception as e:
        print(f'[cert-label] שגיאה: {e}')
        return {'processed': 0, 'found': 0, 'not_found': 0, 'error': str(e)}
    finally:
        _gmail_label_lock.release()

@app.route('/api/cert/label-sent', methods=['POST', 'GET'])
def api_cert_label_sent():
    """Manual trigger for cert-email labeling. Runs in the BACKGROUND (scanning bodies is slow and
    exceeds the gunicorn worker timeout). Token-authed. ?tickets=t1,t2 labels extra emails not
    tracked in cert_requests (e.g. manually-sent certs)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    extra = [t for t in (request.args.get('tickets', '').split(',')) if t.strip()]
    threading.Thread(target=label_sent_cert_emails, kwargs={'extra_tickets': extra}, daemon=True).start()
    return jsonify({'ok': True, 'started': True, 'extra_tickets': len(extra)})

@app.route('/api/gmail-label-sent', methods=['POST'])
def api_gmail_label_sent():
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    limit = (request.get_json(silent=True) or {}).get('limit')
    return jsonify(label_sent_policy_emails(limit=limit))

_policy_check_lock = threading.Lock()

# Policy-PDF email senders. Harel sends directly (ComposeDoc); the Ofir agency relays some
# renewals from its own address — the attached PDF is still the Harel form (881… numbers),
# so the same parser handles both. Add new relays here to bring them into the scan.
POLICY_EMAIL_SENDERS = ['ComposeDoc@harel-ins.co.il', 'ofirco@ofir-insurance.co.il']

# Some Ofir relay emails declare the Hebrew charset 'iso-8859-8-i' (logical order), which
# Python's codec registry doesn't know — decoding headers/parts would raise. Alias it to
# 'iso-8859-8' so subjects and PDFs from those emails parse instead of crashing the scan.
def _iso88598i(name):
    if name.replace('_', '-').lower() == 'iso-8859-8-i':
        return codecs.lookup('iso-8859-8')
    return None
codecs.register(_iso88598i)

def _search_policy_emails(mail, since_date, extra=''):
    """Search INBOX for policy emails from ANY known policy sender (Harel + Ofir relay),
    returning de-duplicated message sequence numbers (oldest→newest). `extra` appends extra
    IMAP criteria (e.g. a SUBJECT filter)."""
    seen = {}
    for sender in POLICY_EMAIL_SENDERS:
        crit = f'FROM "{sender}" SINCE {since_date}'
        if extra:
            crit = f'{crit} {extra}'
        status, data = mail.search(None, crit)
        if status == 'OK' and data and data[0]:
            for n in data[0].split():
                seen[int(n)] = n
    return [seen[k] for k in sorted(seen)]

def check_policy_documents(days_back=30, keep_pdf=True):
    """Connect to IMAP, look for confirmed-policy emails (Harel ComposeDoc), extract the
    data, and (optionally) save the PDF. `days_back` widens the scan for backfills;
    `keep_pdf=False` parses in memory without storing the file (saves volume space)."""
    if not _policy_check_lock.acquire(blocking=False):
        print('[policy-docs] בדיקה כבר רצה — דילוג')
        return 0
    try:
        return _check_policy_documents_impl(days_back, keep_pdf)
    finally:
        _policy_check_lock.release()

def _check_policy_documents_impl(days_back=30, keep_pdf=True):
    cfg = EMAIL_CONFIG
    if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
        return 0

    from email.utils import parsedate_to_datetime
    processed = 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')

        since_date = (datetime.datetime.now() - datetime.timedelta(days=days_back)).strftime('%d-%b-%Y')
        mids = _search_policy_emails(mail, since_date)

        conn = get_db()
        for mid in mids:
            _, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hdr_data[0][1])
            message_id = hdr.get('Message-ID', '').strip()
            subject = decode_str(hdr.get('Subject', ''))
            try:
                doc_date = parsedate_to_datetime(hdr.get('Date', '')).astimezone().strftime('%Y-%m-%d %H:%M')
            except Exception:
                doc_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

            if message_id and conn.execute(
                'SELECT 1 FROM policy_documents WHERE message_id=?', (message_id,)
            ).fetchone():
                continue

            m = re.search(r'(\d{6,})\s*$', subject.strip())
            subj_policy = m.group(1) if m else None
            # No early skip: new-business policies arrive with a generic subject
            # ("הודעה מהראל…") and carry the policy number in the attachment filename instead.

            _, full_data = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(full_data[0][1])

            saved_any = False
            for part in msg.walk():
                cd = str(part.get('Content-Disposition', ''))
                if 'attachment' not in cd and part.get_content_type() != 'application/octet-stream':
                    continue
                raw_fn = part.get_filename()
                if not raw_fn:
                    continue
                filename = decode_str(raw_fn)
                data_bytes = part.get_payload(decode=True)
                if not data_bytes:
                    continue
                # Policy number: from the subject, else from the filename ("…שמספרה <num>.pdf").
                fnm = re.search(r'(\d{6,})', filename)
                policy_number = subj_policy or (fnm.group(1) if fnm else None)
                if not policy_number:
                    continue
                customer = conn.execute(
                    "SELECT id FROM customers WHERE ltrim(policy_number,'0')=?",
                    (policy_number.lstrip('0'),)).fetchone()
                customer_id = customer['id'] if customer else None
                filepath = ''
                if keep_pdf:
                    folder_key = str(customer_id) if customer_id else f'unmatched_{policy_number}'
                    doc_dir = os.path.join(POLICY_DOCS_DIR, folder_key)
                    os.makedirs(doc_dir, exist_ok=True)
                    safe_fn = re.sub(r'[\\/*?:"<>|]', '_', filename)
                    filepath = os.path.join(doc_dir, safe_fn)
                    with open(filepath, 'wb') as f:
                        f.write(data_bytes)
                cur = conn.execute(
                    '''INSERT OR IGNORE INTO policy_documents
                       (customer_id, policy_number, filename, filepath, received_at, message_id)
                       VALUES (?,?,?,?,?,?)''',
                    (customer_id, policy_number, filename, filepath, doc_date, message_id)
                )
                conn.commit()
                saved_any = True
                status_label = f'ללקוח {customer_id}' if customer_id else 'לא זוהה לקוח'
                print(f'[policy-docs] {"נשמר" if keep_pdf else "עובד"}: {filename} ({policy_number}) {status_label}')

                if cur.lastrowid:
                    fields = parse_harel_policy_pdf(filepath if keep_pdf else data_bytes)
                    if fields:
                        conn.execute(
                            '''INSERT INTO policy_records
                               (policy_document_id, customer_id, policy_number, doc_type_label,
                                doc_type_code, branch, agent_name, agent_number, insured_name,
                                insured_id, spouse_id, address, phone_mobile, phone_home, email,
                                period_start, period_end, premium, total_payment, doc_date, extracted_at)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                            (cur.lastrowid, customer_id, policy_number,
                             fields.get('doc_type_label'), fields.get('doc_type_code'),
                             fields.get('branch'), fields.get('agent_name'), fields.get('agent_number'),
                             fields.get('insured_name'), fields.get('insured_id'), fields.get('spouse_id'),
                             fields.get('address'), fields.get('phone_mobile'), fields.get('phone_home'),
                             fields.get('email'), fields.get('period_start'), fields.get('period_end'),
                             fields.get('premium'), fields.get('total_payment'), doc_date,
                             datetime.datetime.now().isoformat())
                        )
                        conn.commit()
                        # Auto-fill עיסוק from the just-stored PDF onto the customer + insured
                        # (empty-occupation rows only, matched by ת"ז) — no manual batch needed.
                        if keep_pdf and filepath and os.path.exists(filepath):
                            _occ = extract_insured_occupation(filepath)
                            _zid = normalize_id_number(fields.get('insured_id') or '').lstrip('0')
                            if _occ and _zid:
                                conn.execute("UPDATE customers SET occupation=? WHERE "
                                             "ltrim(COALESCE(id_number,''),'0')=? AND COALESCE(occupation,'')=''",
                                             (_occ, _zid))
                                conn.execute("UPDATE insureds SET occupation=? WHERE "
                                             "ltrim(COALESCE(id_number,''),'0')=? AND COALESCE(occupation,'')=''",
                                             (_occ, _zid))
                                conn.commit()
                        # A scanned policy fulfils any pending website form for that ת"ז → resolve it
                        # (drops off /admin/other-forms). Ongoing counterpart of /api/resolve-forms-with-policy.
                        _pid = normalize_id_number(fields.get('insured_id') or '').lstrip('0')
                        if _pid:
                            _resolve_form_queue(conn, _pid)
                            # Align the person's brand to the policy's agent number (source of truth) —
                            # fixes cases where a website-form prefix set the wrong brand.
                            _pbrand = NEW_AGENT_BRAND.get(re.sub(r'\D', '', str(fields.get('agent_number') or '')), '')
                            if _pbrand:
                                conn.execute("UPDATE customers SET brand=? WHERE "
                                             "ltrim(COALESCE(id_number,''),'0')=? AND COALESCE(brand,'')!=?",
                                             (_pbrand, _pid, _pbrand))
                                conn.execute("UPDATE insureds SET brand=? WHERE "
                                             "ltrim(COALESCE(id_number,''),'0')=? AND COALESCE(brand,'')!=?",
                                             (_pbrand, _pid, _pbrand))
                            conn.commit()
                        # Sharon's rule (2026-09-02): a renewal (חידוש) policy PDF arriving IS proof the
                        # customer renewed. If their status wasn't advanced (e.g. 'טופס התקבל', 'ביקשו
                        # לחדש לבד'), flip it to 'חודש' so the policy AUTO-DELIVERS (delivery needs
                        # חודש/הופק). Never override an already-settled/declined status.
                        if (_pid and is_renewal_doc(fields.get('doc_type_label'))
                                and _renewal_period_ok(fields.get('period_start'))):
                            _RSET = ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'בוטל',
                                     'לא רוצים לחדש', 'לא מחדש')
                            _rph = ','.join('?' * len(_RSET))
                            _flipped = conn.execute(
                                f"UPDATE customers SET status='חודש', status_changed_at=? "
                                f"WHERE ltrim(COALESCE(id_number,''),'0')=? "
                                f"AND COALESCE(status,'') NOT IN ({_rph}) "
                                f"AND COALESCE(group_owner,'')='' AND COALESCE(import_source,'')!='test_ofir'",
                                (datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), _pid, *_RSET)).rowcount
                            if _flipped:
                                try:
                                    log_event(conn, event_key(_pid, f'renewpdf-{cur.lastrowid}'),
                                              "סטטוס עודכן ל-'חודש' (הגיעה פוליסת חידוש)", 'system', kind='status')
                                except Exception:
                                    pass
                                conn.commit()

            if saved_any:
                processed += 1

        conn.close()
        mail.logout()
    except Exception as e:
        print(f'[policy-docs] שגיאה: {e}')

    return processed

# ── Join-request forms (new-business leads from the agency website) ───────────
# Customers fill a quote/enrolment form on the Gaia/Winner site; it arrives from
# resend.dev to the inbox with a rich set of underwriting fields NOT present in the
# Harel policy PDF (prior insurer, health fund, professional org, declarations,
# marketing consent, …). We capture them as a lead ("ממתין להפקה") keyed by ת"ז;
# when Sharon later issues the policy in Harel, the arriving PDF upgrades the lead
# (see _ensure_new_customer) and the message-send fires.
JOIN_FORM_SENDER = 'onboarding@resend.dev'
JOIN_FORM_SUBJECT_MARK = 'בקשת הצטרפות חדשה'

def _lead_brand_from_subject(subject):
    """'גאיה | בקשת הצטרפות חדשה' / 'ווינר | …' → brand."""
    s = (subject or '').strip()
    for b in ('גאיה', 'ווינר', 'אופיר'):
        if s.startswith(b):
            return b
    return ''

def parse_join_form(html):
    """Parse the two-column HTML table of a join-request form → {label: value}."""
    import html as _htmlmod
    out = {}
    for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', html or '', re.S | re.I):
        tds = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.S | re.I)
        if len(tds) < 2:
            continue
        label = _htmlmod.unescape(re.sub(r'<[^>]+>', '', tds[0])).strip()
        val = _htmlmod.unescape(re.sub(r'<[^>]+>', '', tds[1])).strip()
        if not label:
            continue
        out[label] = '' if val in ('—', '-') else val
    return out

def _lead_idn(f):
    """Applicant ת"ז from a parsed join form. Try the known labels first, else fall back to ANY
    parsed value that is a valid Israeli ID (checksum) — so a relabelled/reordered form field
    still resolves instead of silently dropping the lead."""
    idn = normalize_id_number(f.get('מספר ת.ז') or f.get('ת.ז המצהיר') or '')
    if idn:
        return idn
    for v in f.values():
        d = re.sub(r'\D', '', str(v or ''))
        if 5 <= len(d) <= 9 and is_israeli_id(normalize_id_number(d)):
            return normalize_id_number(d)
    return ''

def _ingest_join_form(conn, subject, html, received_at, attach_path=None):
    """Store/refresh a join-form lead in the active month, keyed by ת"ז. Returns ת"ז or None."""
    f = parse_join_form(html)
    idn = _lead_idn(f)
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not idn or not month:
        return None
    brand = _lead_brand_from_subject(subject)
    name = f.get('שם מלא') or f.get('שם המצהיר') or ''
    phone = re.sub(r'\D', '', f.get('טלפון') or '')
    email = f.get('אימייל') or ''
    address = ' '.join(x for x in (f.get('כתובת'), f.get('עיר')) if x)
    occupation = f.get('מקצועות') or ''
    installments = f.get('מספר תשלומים') or ''
    pay_method = f.get('אמצעי גביה') or ''
    consent = f.get('הסכמה לשיווק') or ''
    card = re.sub(r'\D', '', f.get('מספר כרטיס') or '')
    card_last4 = ('****' + card[-4:]) if len(card) >= 4 else ''
    # Rich underwriting payload — keep everything EXCEPT the full card number/expiry (PCI).
    safe = dict(f)
    if safe.get('מספר כרטיס'):
        safe['מספר כרטיס'] = card_last4
    safe.pop('תוקף כרטיס', None)
    lead_json = json.dumps(safe, ensure_ascii=False)
    row = conn.execute(
        "SELECT id, status FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=?",
        (month['id'], idn.lstrip('0'))).fetchone()
    if row:
        keep_status = row['status'] if (row['status'] and row['status'] != LEAD_STATUS) else LEAD_STATUS
        conn.execute(
            """UPDATE customers SET name=COALESCE(NULLIF(?,''),name), phone=COALESCE(NULLIF(?,''),phone),
                   email=COALESCE(NULLIF(?,''),email), address=COALESCE(NULLIF(?,''),address),
                   occupation=COALESCE(NULLIF(?,''),occupation), brand=COALESCE(NULLIF(?,''),brand),
                   form_installments=?, form_payment_method=?, form_received_at=?,
                   form_card_number=?, marketing_consent=?, lead_form_json=?, status=?,
                   lead_doc_path=COALESCE(?,lead_doc_path)
               WHERE id=?""",
            (name, phone, email, address, occupation, brand, installments, pay_method, received_at,
             card_last4, consent, lead_json, keep_status, attach_path, row['id']))
        cid = row['id']
    else:
        cur = conn.execute(
            """INSERT INTO customers (month_id, name, id_number, phone, email, address, brand,
                   occupation, status, import_source, form_installments, form_payment_method,
                   form_received_at, form_card_number, marketing_consent, lead_form_json, lead_doc_path)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (month['id'], name, idn, phone, email, address, brand, occupation, LEAD_STATUS,
             'join_form', installments, pay_method, received_at, card_last4, consent, lead_json, attach_path))
        cid = cur.lastrowid
    card_note = f" · אשראי {card_last4}" if card_last4 else ""
    log_event(conn, event_key(idn, 'cust-%d' % cid),
              f"טופס הצטרפות התקבל ({brand or '—'}) · עיסוק: {occupation or '—'}{card_note}",
              'system', kind='join_form')
    _resolve_form_queue(conn, idn)  # dedupe — the lead is tracked as a customer, not a raw form
    return idn

@app.route('/api/lead-debug')
def api_lead_debug():
    """Diagnostic: find the join-form email containing ת"ז {id}, run parse_join_form, and show the
    parsed fields + resolved ת"ז — to see why a lead didn't ingest. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    want = re.sub(r'\D', '', request.args.get('id', '')).lstrip('0')
    days = int(request.args.get('days', 3))
    cfg = EMAIL_CONFIG
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        _, data = mail.search(None, f'FROM "{JOIN_FORM_SENDER}" SINCE {since}')
        for mid in (data[0].split() if data and data[0] else []):
            _, fd = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(fd[0][1])
            html = ''
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    try:
                        html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True)
                        html = pl.decode('utf-8', 'replace') if pl else ''
                    break
            if want and want not in re.sub(r'\D', '', html):
                continue
            f = parse_join_form(html)
            mail.logout()
            return jsonify({'subject': decode_str(msg.get('Subject', '')), 'fields': f,
                            'resolved_idn': _lead_idn(f),
                            'label_idn': normalize_id_number(f.get('מספר ת.ז') or f.get('ת.ז המצהיר') or '')})
        mail.logout()
        return jsonify({'error': 'not found', 'want': want})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/stuck-leads')
def api_stuck_leads():
    """Read-only check: join-form emails (last {days}) whose applicant ת"ז is NOT yet a customer —
    i.e. leads that failed to ingest. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int(request.args.get('days', 5))
    cfg = EMAIL_CONFIG
    stuck, ok = [], 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime('%d-%b-%Y')
        _, data = mail.search(None, f'FROM "{JOIN_FORM_SENDER}" SINCE {since}')
        conn = get_db()
        seen = set()
        for mid in (data[0].split() if data and data[0] else []):
            _, hd = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
            subj = decode_str(email_lib.message_from_bytes(hd[0][1]).get('Subject', ''))
            if JOIN_FORM_SUBJECT_MARK not in subj:
                continue
            _, fd = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(fd[0][1])
            html = ''
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    try:
                        html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True)
                        html = pl.decode('utf-8', 'replace') if pl else ''
                    break
            f = parse_join_form(html)
            idn = _lead_idn(f)
            nm = f.get('שם מלא') or f.get('שם המצהיר') or '?'
            if not idn:
                stuck.append({'name': nm, 'idn': None, 'reason': 'no-id'})
                continue
            z = idn.lstrip('0')
            if z in seen:
                continue
            seen.add(z)
            cust = conn.execute("SELECT status FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? "
                                "ORDER BY id DESC LIMIT 1", (z,)).fetchone()
            if cust:
                ok += 1
            else:
                stuck.append({'name': nm, 'idn': idn})
        conn.close()
        mail.logout()
    except Exception as e:
        return jsonify({'error': str(e)})
    return jsonify({'days': days, 'has_customer': ok, 'stuck_count': len(stuck), 'stuck': stuck})

@app.route('/api/reingest-join-forms', methods=['POST'])
def api_reingest_join_forms():
    """Recover lost leads: clear the processed-dedup and re-scan join forms (idempotent upsert by
    ת"ז), so forms that failed ת"ז extraction before the fix now ingest. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int((request.get_json(silent=True) or {}).get('days') or 14)
    conn = get_db()
    cleared = conn.execute("DELETE FROM processed_leads").rowcount
    conn.commit()
    conn.close()
    n = check_join_forms(days_back=days)
    return jsonify({'ok': True, 'cleared_dedup': cleared, 'leads_ingested': n})

_join_form_lock = threading.Lock()

def check_join_forms(days_back=14):
    """Scan the inbox for agency join-request forms and ingest them as leads."""
    if not _join_form_lock.acquire(blocking=False):
        return 0
    try:
        return _check_join_forms_impl(days_back)
    finally:
        _join_form_lock.release()

def _check_join_forms_impl(days_back=14):
    cfg = EMAIL_CONFIG
    if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
        return 0
    from email.utils import parsedate_to_datetime
    processed = 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since_date = (datetime.datetime.now() - datetime.timedelta(days=days_back)).strftime('%d-%b-%Y')
        status, data = mail.search(None, f'FROM "{JOIN_FORM_SENDER}" SINCE {since_date}')
        if status != 'OK':
            mail.logout(); return 0
        conn = get_db()
        for mid in data[0].split():
            _, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hdr_data[0][1])
            message_id = hdr.get('Message-ID', '').strip()
            subject = decode_str(hdr.get('Subject', ''))
            if JOIN_FORM_SUBJECT_MARK not in subject:
                continue
            if message_id and conn.execute(
                    'SELECT 1 FROM processed_leads WHERE message_id=?', (message_id,)).fetchone():
                continue
            try:
                received_at = parsedate_to_datetime(hdr.get('Date', '')).astimezone().strftime('%Y-%m-%d %H:%M')
            except Exception:
                received_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            _, full_data = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(full_data[0][1])
            html = None
            for part in msg.walk():
                if part.get_content_type() == 'text/html' and 'attachment' not in str(part.get('Content-Disposition', '')):
                    try:
                        html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True)
                        html = pl.decode('utf-8', 'replace') if pl else None
                    break
            if not html:
                continue
            f_preview = parse_join_form(html)
            idn_preview = _lead_idn(f_preview)
            attach_path = None
            for part in msg.walk():
                if 'attachment' not in str(part.get('Content-Disposition', '')):
                    continue
                raw_fn = part.get_filename()
                payload = part.get_payload(decode=True)
                if not raw_fn or not payload:
                    continue
                ext = os.path.splitext(decode_str(raw_fn))[1] or '.jpg'
                folder = os.path.join(LEAD_DOCS_DIR, idn_preview or 'unknown')
                os.makedirs(folder, exist_ok=True)
                attach_path = os.path.join(folder, f'מסמך {idn_preview}{ext}')
                with open(attach_path, 'wb') as fh:
                    fh.write(payload)
                break
            idn = _ingest_join_form(conn, subject, html, received_at, attach_path)
            # Only mark the email processed once a lead was actually created. If ת"ז extraction
            # failed (idn is None), leave it UNprocessed so a later scan retries — otherwise the
            # lead is silently lost forever.
            if message_id and idn:
                conn.execute('INSERT OR IGNORE INTO processed_leads (message_id, processed_at) VALUES (?,?)',
                             (message_id, datetime.datetime.now().isoformat()))
            conn.commit()
            if idn:
                processed += 1
                print(f'[join-forms] ליד נקלט: {idn} ({subject})')
            else:
                print(f'[join-forms] ⚠️ לא חולץ ת"ז — לא סומן מעובד, יינסה שוב: {subject}')
        conn.close()
        mail.logout()
    except Exception as e:
        print(f'[join-forms] שגיאה: {e}')
    return processed

# ── Renewal-request forms (existing customer asks to renew via the website) ───
# Same sender/format as the join forms, but a different subject ("… חידוש פוליסה").
# A match to an active-month customer marks them 'טופס התקבל' → they appear in the work queue.
RENEWAL_FORM_SUBJECT_MARK = 'חידוש פוליסה'
_renewal_form_lock = threading.Lock()

def _ingest_renewal_form(conn, subject, html, received_at):
    """Mark an existing active-month customer 'טופס התקבל' from a website renewal request.
    Returns (ת"ז, matched?). Never downgrades a customer who already renewed/issued."""
    f = parse_join_form(html)
    idn = normalize_id_number(f.get('מספר ת.ז') or f.get('ת.ז המצהיר') or '')
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not idn or not month:
        return (idn or '', False)
    row = conn.execute(
        "SELECT id, status FROM customers WHERE month_id=? AND ltrim(COALESCE(id_number,''),'0')=?",
        (month['id'], idn.lstrip('0'))).fetchone()
    if not row:
        return (idn, False)
    if (row['status'] or '') in ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק'):
        return (idn, True)
    conn.execute(
        "UPDATE customers SET status='טופס התקבל', form_received_at=?, "
        "email=COALESCE(NULLIF(email,''),?), status_changed_at=? WHERE id=?",
        (received_at, (f.get('אימייל') or ''),
         datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), row['id']))
    log_event(conn, event_key(idn, 'cust-%d' % row['id']),
              "טופס בקשת חידוש התקבל מהאתר", 'system', kind='renewal_form')
    return (idn, True)

def check_renewal_forms(days_back=14):
    if not _renewal_form_lock.acquire(blocking=False):
        return (0, 0)
    try:
        return _check_renewal_forms_impl(days_back)
    finally:
        _renewal_form_lock.release()

def _check_renewal_forms_impl(days_back=14):
    """Returns (processed, unmatched) — unmatched are renewal requests whose ת"ז isn't an
    active-month customer (surfaced by the monitor)."""
    cfg = EMAIL_CONFIG
    if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
        return (0, 0)
    from email.utils import parsedate_to_datetime
    processed = unmatched = 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        since_date = (datetime.datetime.now() - datetime.timedelta(days=days_back)).strftime('%d-%b-%Y')
        status, data = mail.search(None, f'FROM "{JOIN_FORM_SENDER}" SINCE {since_date}')
        if status != 'OK':
            mail.logout(); return (0, 0)
        conn = get_db()
        for mid in data[0].split():
            _, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hdr_data[0][1])
            message_id = hdr.get('Message-ID', '').strip()
            subject = decode_str(hdr.get('Subject', ''))
            if RENEWAL_FORM_SUBJECT_MARK not in subject:
                continue
            if message_id and conn.execute(
                    'SELECT 1 FROM processed_leads WHERE message_id=?', (message_id,)).fetchone():
                continue
            try:
                received_at = parsedate_to_datetime(hdr.get('Date', '')).astimezone().strftime('%Y-%m-%d %H:%M')
            except Exception:
                received_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            _, full_data = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(full_data[0][1])
            html = None
            for part in msg.walk():
                if part.get_content_type() == 'text/html' and 'attachment' not in str(part.get('Content-Disposition', '')):
                    try:
                        html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True); html = pl.decode('utf-8', 'replace') if pl else None
                    break
            if not html:
                continue
            idn, matched = _ingest_renewal_form(conn, subject, html, received_at)
            if message_id:
                conn.execute('INSERT OR IGNORE INTO processed_leads (message_id, processed_at) VALUES (?,?)',
                             (message_id, datetime.datetime.now().isoformat()))
            conn.commit()
            processed += 1
            if not matched:
                unmatched += 1
            print(f'[renewal-forms] {"נקלט" if matched else "לא תואם"}: {idn} ({subject})')
        conn.close()
        mail.logout()
    except Exception as e:
        print(f'[renewal-forms] שגיאה: {e}')
    return (processed, unmatched)

# ── Harel "completed details" proposal emails → pending-issuance lead ─────────
# Harel notifies the agent when a customer finishes submitting details for a
# professional-liability proposal ("קיבלנו השלמת פרטים מ<שם> להצעה..."). There is NO ת"ז
# in the email — only name, phone, offer number and the insurance period. We capture it as a
# lead ("ממתין להפקה") keyed by the Harel OFFER number (which becomes the policy number once
# issued). When the policy PDF later arrives, _ensure_new_customer links it by policy number,
# fills the ת"ז and flips the lead to 'הופק' → it becomes 'פעיל' in the master and drops off
# the issuance queue.
HAREL_COMPLETED_SENDER = 'HarelInsurance@harel-group.co.il'
HAREL_COMPLETED_SUBJECT_MARK = 'קיבלנו השלמת פרטים'
_harel_completed_lock = threading.Lock()

def _parse_harel_completed(subject, html):
    """Extract {name, offer, phone, period} from a Harel completed-details email (free text,
    not a table). Robust to HTML — strips tags then regexes the labelled lines."""
    import html as _htmlmod
    text = re.sub(r'<[^>]+>', ' ', html or '')
    text = _htmlmod.unescape(text)
    text = re.sub(r'[ \t‏‎\xa0‌]+', ' ', text)
    out = {}
    m = re.search(r'השלמת פרטים מ(.+?) להצעה', subject or '')
    if not m:
        m = re.search(r'קיבלנו מ(.+?) את המידע', text)
    out['name'] = m.group(1).strip() if m else ''
    m = re.search(r"מס['׳]\s*הצעה\s*:?\s*(\d{5,})", text)
    out['offer'] = m.group(1) if m else ''
    m = re.search(r'מספר טלפון\s*:?\s*([0-9\-]{7,})', text)
    out['phone'] = re.sub(r'\D', '', m.group(1)) if m else ''
    m = re.search(r'תקופת הביטוח\s*:?\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})', text)
    out['period'] = f'{m.group(1)} - {m.group(2)}' if m else ''
    return out

def _ingest_harel_completed(conn, subject, html, received_at):
    """Store/refresh a Harel proposal lead as 'ממתין להפקה', keyed by the offer number.
    Returns (offer, name) — offer is None if nothing could be extracted (so it retries)."""
    f = _parse_harel_completed(subject, html)
    offer, phone = f['offer'], f['phone']
    name = f['name'] or (f'הצעה {offer}' if offer else '')
    month = conn.execute("SELECT id FROM months WHERE is_active=1 ORDER BY id DESC LIMIT 1").fetchone()
    if not offer or not month:
        return (None, None)
    # Already issued? Most proposals from the 1st already became policies — don't create a
    # phantom 'ממתין להפקה' lead for one that's already active. Mark processed (return offer).
    if conn.execute("SELECT 1 FROM policy_records WHERE policy_number=?", (offer,)).fetchone():
        return (offer, name)
    note = f"השלמת פרטים להצעה (הראל · אחריות מקצועית) · תקופה: {f['period'] or '—'}"
    row = conn.execute(
        "SELECT id, status FROM customers WHERE month_id=? AND policy_number=?",
        (month['id'], offer)).fetchone()
    if row:
        # never downgrade a record that already renewed/issued
        if (row['status'] or '') in ('חודש', 'חודש - בוצעה שיחת מכירה', 'הופק', 'פעיל'):
            return (offer, name)
        conn.execute(
            "UPDATE customers SET name=COALESCE(NULLIF(?,''),name), phone=COALESCE(NULLIF(?,''),phone), "
            "form_received_at=?, sharon_notes=COALESCE(NULLIF(sharon_notes,''),?), status=? WHERE id=?",
            (f['name'], phone, received_at, note, LEAD_STATUS, row['id']))
        cid = row['id']
    else:
        cur = conn.execute(
            """INSERT INTO customers (month_id, policy_number, name, phone, status,
                   import_source, form_received_at, sharon_notes)
               VALUES (?,?,?,?,?,?,?,?)""",
            (month['id'], offer, name, phone, LEAD_STATUS, 'harel_proposal', received_at, note))
        cid = cur.lastrowid
    log_event(conn, event_key(offer, 'cust-%d' % cid),
              f"התקבלה השלמת פרטים להצעה {offer} ({name})", 'system', kind='harel_proposal')
    return (offer, name)

def check_harel_completed(days_back=14):
    if not _harel_completed_lock.acquire(blocking=False):
        return 0
    try:
        return _check_harel_completed_impl(days_back)
    finally:
        _harel_completed_lock.release()

def _check_harel_completed_impl(days_back=14):
    cfg = EMAIL_CONFIG
    if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
        return 0
    from email.utils import parsedate_to_datetime
    processed = 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        # Never scan earlier than the 1st of the current month — older proposals from prior
        # months are already issued policies, not pending-issuance leads (Sharon's rule).
        today = datetime.date.today()
        since = max(datetime.date(today.year, today.month, 1),
                    today - datetime.timedelta(days=days_back))
        since_date = since.strftime('%d-%b-%Y')
        status, data = mail.search(None, f'FROM "{HAREL_COMPLETED_SENDER}" SINCE {since_date}')
        if status != 'OK':
            mail.logout(); return 0
        conn = get_db()
        for mid in data[0].split():
            _, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hdr_data[0][1])
            message_id = hdr.get('Message-ID', '').strip()
            subject = decode_str(hdr.get('Subject', ''))
            if HAREL_COMPLETED_SUBJECT_MARK not in subject:
                continue
            if message_id and conn.execute(
                    'SELECT 1 FROM processed_leads WHERE message_id=?', (message_id,)).fetchone():
                continue
            try:
                received_at = parsedate_to_datetime(hdr.get('Date', '')).astimezone().strftime('%Y-%m-%d %H:%M')
            except Exception:
                received_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            _, full_data = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(full_data[0][1])
            html = None
            for part in msg.walk():
                if part.get_content_type() == 'text/html' and 'attachment' not in str(part.get('Content-Disposition', '')):
                    try:
                        html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True); html = pl.decode('utf-8', 'replace') if pl else None
                    break
            if html is None:  # fall back to a plain-text body
                for part in msg.walk():
                    if part.get_content_type() == 'text/plain':
                        pl = part.get_payload(decode=True); html = pl.decode('utf-8', 'replace') if pl else None
                        break
            if not html:
                continue
            offer, name = _ingest_harel_completed(conn, subject, html, received_at)
            if offer and message_id:
                conn.execute('INSERT OR IGNORE INTO processed_leads (message_id, processed_at) VALUES (?,?)',
                             (message_id, datetime.datetime.now().isoformat()))
            conn.commit()
            if offer:
                processed += 1
                print(f'[harel-proposal] ליד נקלט: {offer} {name} ({subject})')
            else:
                print(f"[harel-proposal] ⚠️ לא חולץ מס' הצעה — לא סומן מעובד, יינסה שוב: {subject}")
        conn.close()
        mail.logout()
    except Exception as e:
        print(f'[harel-proposal] שגיאה: {e}')
    return processed

@app.route('/api/harel-proposal-scan')
def api_harel_proposal_scan():
    """Manual trigger for the Harel completed-details scanner. Token-authed. Returns how many
    proposal leads were ingested, plus the current 'ממתין להפקה' harel_proposal leads."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    days = int(request.args.get('days', 14))
    n = check_harel_completed(days_back=days)
    conn = get_db()
    rows = conn.execute(
        "SELECT policy_number, name, phone, status, sharon_notes, form_received_at "
        "FROM customers WHERE import_source='harel_proposal' ORDER BY id DESC LIMIT 50").fetchall()
    conn.close()
    return jsonify({'processed': n, 'leads': [dict(r) for r in rows]})

# ── Harel "certificate of insurance" (אישור קיום ביטוח) emails → auto-fetch + deliver ──
# Harel emails the agent a link (generic-identification/?ticket=...) to a certificate it ISSUED
# (not in our dashboard). No PDF attached — it sits behind the ticket + the insured's ת"ז. We
# extract the ticket + the insured NAME from the email body, match the name to our master to get
# ת"ז+phone+brand, and queue it. The wa-sender then downloads the PDF (Puppeteer types the ת"ז)
# and delivers it (Cloud API primary, Gaia WhatsApp-Web fallback). Same sender as the proposal
# emails — differ by SUBJECT.
HAREL_CERT_SENDER = 'HarelInsurance@harel-group.co.il'
HAREL_CERT_SUBJECT_MARK = 'אישור לביטוח קיים'
# From-today-onward only — never process certificate emails older than go-live (Sharon's rule:
# no retro). The 14-day lookback still catches recent ones if the laptop was off a few days.
CERT_GOLIVE = datetime.date(2026, 8, 17)
_harel_cert_lock = threading.Lock()

def _parse_harel_cert(subject, html):
    """Extract {ticket, name} from a Harel certificate email."""
    import html as _htmlmod
    out = {'ticket': '', 'name': ''}
    m = re.search(r'generic-identification/\?ticket=([0-9a-fA-F]+)', html or '')
    out['ticket'] = m.group(1) if m else ''
    text = re.sub(r'<[^>]+>', ' ', html or '')
    text = _htmlmod.unescape(text)
    text = re.sub(r'[ \t‏‎\xa0‌]+', ' ', text)
    m = re.search(r'אישור קיום ביטוח עבור\s+(.+?)\s+לבקשתך', text)
    if not m:
        m = re.search(r'עבור\s+(.{2,30}?)\s+לבקשתך', text)
    out['name'] = m.group(1).strip() if m else ''
    return out

def _match_insured_by_name(conn, name):
    """Name → (dict{id_number,name,phone,brand}, status). Order-independent match against the
    master (insureds) then customers. status: 'matched' (exactly one ת"ז) | 'ambiguous' | 'no_match'."""
    if not name:
        return (None, 'no_match')
    like = f'%{name}%'
    cond, params = _name_search('name', name, like)
    by_id = {}
    for tbl in ('insureds', 'customers'):
        rows = conn.execute(
            f"SELECT id_number, name, phone, email, brand FROM {tbl} WHERE {cond} "
            f"AND COALESCE(id_number,'')!=''", params).fetchall()
        for r in rows:
            key = (r['id_number'] or '').lstrip('0')
            if key:
                by_id.setdefault(key, dict(r))
        if by_id:
            break  # prefer the master; only fall to customers if the master had nothing
    if len(by_id) == 1:
        return (list(by_id.values())[0], 'matched')
    return (None, 'ambiguous' if len(by_id) > 1 else 'no_match')

def _ingest_harel_cert(conn, subject, html, received_at, message_id=''):
    """Queue a Harel certificate request keyed by ticket. Returns (ticket, name) or (None, None)."""
    f = _parse_harel_cert(subject, html)
    ticket, name = f['ticket'], f['name']
    if not ticket:
        return (None, None)
    if conn.execute("SELECT 1 FROM cert_requests WHERE ticket=?", (ticket,)).fetchone():
        return (ticket, name)  # already queued
    row, mstatus = _match_insured_by_name(conn, name)
    idn = (row['id_number'] if row else '') or ''
    phone = (row['phone'] if row else '') or ''
    email = (row['email'] if row else '') or ''
    brand = (row['brand'] if row else '') or ''
    cust = conn.execute(
        "SELECT id FROM customers WHERE ltrim(COALESCE(id_number,''),'0')=? ORDER BY id DESC LIMIT 1",
        (idn.lstrip('0'),)).fetchone() if idn else None
    conn.execute(
        """INSERT INTO cert_requests (ticket, cust_name, id_number, phone, email, brand, customer_id,
               received_at, match_status, message_id, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (ticket, name, idn, phone, email, brand, (cust['id'] if cust else None), received_at, mstatus,
         (message_id or '').strip(), datetime.datetime.now().isoformat()))
    if idn:
        log_event(conn, event_key(idn, ('cust-%d' % cust['id']) if cust else ('cert-%s' % ticket)),
                  f"התקבלה בקשת אישור קיום ביטוח (הראל) · התאמה: {mstatus}", 'system', kind='cert_request')
    if mstatus != 'matched':
        # unmatched cert → won't auto-send (we don't know who); alert Sharon so nothing falls silently.
        conn.execute("INSERT INTO owner_alerts (text, created_at) VALUES (?,?)",
                     (f"⚠️ אישור קיום ביטוח לא זוהה: \"{name}\" ({mstatus}) — לטיפול ידני",
                      datetime.datetime.now().isoformat()))
    return (ticket, name)

def check_cert_emails(days_back=14):
    if not _harel_cert_lock.acquire(blocking=False):
        return 0
    try:
        return _check_cert_emails_impl(days_back)
    finally:
        _harel_cert_lock.release()

def _check_cert_emails_impl(days_back=14):
    cfg = EMAIL_CONFIG
    if not cfg['enabled'] or not cfg['imap_server'] or not cfg['password']:
        return 0
    from email.utils import parsedate_to_datetime
    processed = 0
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        # Floor at go-live so we never retro-process old certificate emails (Sharon's rule).
        since = max(datetime.date.today() - datetime.timedelta(days=days_back), CERT_GOLIVE)
        since_date = since.strftime('%d-%b-%Y')
        status, data = mail.search(None, f'FROM "{HAREL_CERT_SENDER}" SINCE {since_date}')
        if status != 'OK':
            mail.logout(); return 0
        conn = get_db()
        for mid in data[0].split():
            _, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hdr_data[0][1])
            message_id = hdr.get('Message-ID', '').strip()
            subject = decode_str(hdr.get('Subject', ''))
            if HAREL_CERT_SUBJECT_MARK not in subject:
                continue
            if message_id and conn.execute(
                    'SELECT 1 FROM processed_leads WHERE message_id=?', (message_id,)).fetchone():
                continue
            try:
                received_at = parsedate_to_datetime(hdr.get('Date', '')).astimezone().strftime('%Y-%m-%d %H:%M')
            except Exception:
                received_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            _, full_data = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(full_data[0][1])
            html = None
            for part in msg.walk():
                if part.get_content_type() == 'text/html' and 'attachment' not in str(part.get('Content-Disposition', '')):
                    try:
                        html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True); html = pl.decode('utf-8', 'replace') if pl else None
                    break
            if not html:
                continue
            ticket, name = _ingest_harel_cert(conn, subject, html, received_at, message_id)
            if ticket and message_id:
                conn.execute('INSERT OR IGNORE INTO processed_leads (message_id, processed_at) VALUES (?,?)',
                             (message_id, datetime.datetime.now().isoformat()))
            conn.commit()
            if ticket:
                processed += 1
                print(f'[harel-cert] נקלטה בקשת אישור: {name or "?"} (ticket {ticket[:10]}…)')
            else:
                print(f'[harel-cert] ⚠️ לא חולץ ticket — יינסה שוב: {subject}')
        conn.close()
        mail.logout()
    except Exception as e:
        print(f'[harel-cert] שגיאה: {e}')
    return processed

@app.route('/api/cert/queue')
def api_cert_queue():
    """wa-sender pulls matched, not-yet-sent certificate requests to download + deliver.
    Token-authed. Only rows with a resolved ת"ז are returned (unmatched wait for Sharon)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    conn = get_db()
    rows = conn.execute(
        "SELECT id, ticket, cust_name, id_number, phone, email, brand, customer_id, received_at "
        "FROM cert_requests WHERE match_status='matched' AND wa_sent_at IS NULL "
        "AND COALESCE(id_number,'')!='' AND (COALESCE(phone,'')!='' OR COALESCE(email,'')!='') "
        "ORDER BY id LIMIT 20").fetchall()
    conn.close()
    return jsonify({'count': len(rows), 'items': [dict(r) for r in rows]})

@app.route('/api/cert/sent', methods=['POST'])
def api_cert_sent():
    """wa-sender reports a cert delivered. Marks wa_sent_at + logs it under the customer.
    Body: {ticket, target ('customer'|'sharon'|'sharon-fallback'|'gaia-fallback'), saved:bool}."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(force=True, silent=True) or {}
    ticket = (d.get('ticket') or '').strip()
    if not ticket:
        return jsonify({'error': 'need ticket'}), 400
    conn = get_db()
    row = conn.execute("SELECT id, id_number, customer_id, cust_name FROM cert_requests WHERE ticket=?",
                       (ticket,)).fetchone()
    if not row:
        conn.close(); return jsonify({'error': 'unknown ticket'}), 404
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    conn.execute("UPDATE cert_requests SET wa_sent_at=?, wa_target=?, pdf_saved=?, "
                 "email_sent_at=COALESCE(?,email_sent_at) WHERE ticket=?",
                 (now, d.get('target') or 'customer', 1 if d.get('saved') else 0,
                  (now if d.get('email_ok') else None), ticket))
    if row['id_number']:
        tgt = d.get('target') or 'customer'
        label = {'customer': 'ללקוח', 'sharon': 'לשרון (טסט)',
                 'sharon-fallback': 'לשרון (fallback)', 'gaia-fallback': 'ללקוח דרך גאיה (fallback)'}.get(tgt, tgt)
        chans = '+'.join([c for c, ok in (('וואטסאפ', d.get('wa_ok')), ('מייל', d.get('email_ok'))) if ok]) or '—'
        log_event(conn, event_key(row['id_number'],
                  ('cust-%d' % row['customer_id']) if row['customer_id'] else ('cert-%s' % ticket)),
                  f"אישור קיום ביטוח נשלח {label} ({chans})", 'system', kind='cert_sent')
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/cert/resolve', methods=['POST'])
def api_cert_resolve():
    """Manually resolve a no_match / ambiguous certificate request so it can be delivered.
    Body {ticket, phone?, id_number?, email?}. Priority: an explicit id_number wins; otherwise
    look the phone up in the master (insureds) + active-month customers and, if it points to
    exactly ONE ת"ز, use it. Sets the cert_request's id_number/phone/email/customer_id +
    match_status='matched' so /api/cert/queue picks it up. Returns candidates if ambiguous."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    d = request.get_json(silent=True) or {}
    ticket = (d.get('ticket') or '').strip()
    if not ticket:
        return jsonify({'error': 'need ticket'}), 400
    conn = get_db()
    cr = conn.execute("SELECT id, cust_name FROM cert_requests WHERE ticket=?", (ticket,)).fetchone()
    if not cr:
        conn.close(); return jsonify({'error': 'unknown ticket'}), 404
    idn = re.sub(r'\D', '', str(d.get('id_number') or ''))
    phone = re.sub(r'\D', '', str(d.get('phone') or ''))
    email = (d.get('email') or '').strip()
    # No explicit ת"ז? Resolve it from the phone via the master + active customers.
    if not idn and phone:
        cands = {}
        for r in conn.execute("SELECT id_number, name, phone, email, brand FROM insureds "
                              "WHERE REPLACE(REPLACE(COALESCE(phone,''),'-',''),' ','')=?", (phone,)).fetchall():
            z = re.sub(r'\D', '', r['id_number'] or '')
            if z: cands[z] = {'name': r['name'], 'email': r['email'], 'brand': r['brand'], 'src': 'master'}
        for r in conn.execute("SELECT id_number, name, phone, email, brand FROM customers "
                              "WHERE REPLACE(REPLACE(COALESCE(phone,''),'-',''),' ','')=?", (phone,)).fetchall():
            z = re.sub(r'\D', '', r['id_number'] or '')
            if z: cands.setdefault(z, {'name': r['name'], 'email': r['email'], 'brand': r['brand'], 'src': 'customer'})
        if len(cands) == 1:
            idn = next(iter(cands))
            if not email:
                email = cands[idn].get('email') or ''
        else:
            conn.close()
            return jsonify({'resolved': False, 'reason': ('no phone match' if not cands else 'ambiguous'),
                            'candidates': [{'id_number': k, **v} for k, v in cands.items()]})
    if not idn:
        conn.close(); return jsonify({'error': 'need id_number or a phone that resolves to one'}), 400
    # Pull master contact + link to an active-month customer if present.
    ins = conn.execute("SELECT name, phone, email, brand FROM insureds "
                       "WHERE ltrim(COALESCE(id_number,''),'0')=?", (idn.lstrip('0'),)).fetchone()
    cust = conn.execute("SELECT c.id, c.brand FROM customers c JOIN months m ON m.id=c.month_id "
                        "WHERE m.is_active=1 AND ltrim(COALESCE(c.id_number,''),'0')=?",
                        (idn.lstrip('0'),)).fetchone()
    if not phone:
        phone = re.sub(r'\D', '', str((ins['phone'] if ins else '') or ''))
    if not email and ins:
        email = ins['email'] or ''
    brand = (cust['brand'] if cust else (ins['brand'] if ins else None))
    # Resolving = make it (re-)deliverable: also clear any prior send marks so a cert that went to a
    # wrong number re-enters the queue and is re-sent to the corrected contact.
    conn.execute("UPDATE cert_requests SET id_number=?, phone=?, email=?, brand=COALESCE(?,brand), "
                 "customer_id=COALESCE(?,customer_id), match_status='matched', "
                 "wa_sent_at=NULL, wa_target=NULL, email_sent_at=NULL WHERE ticket=?",
                 (idn, phone, email, brand, (cust['id'] if cust else None), ticket))
    conn.commit()
    conn.close()
    return jsonify({'resolved': True, 'ticket': ticket, 'id_number': idn, 'phone': phone,
                    'email': email, 'brand': brand, 'customer_id': (cust['id'] if cust else None)})

@app.route('/api/cert/preview')
def api_cert_preview():
    """READ-ONLY: list Harel certificate emails since a date WITHOUT queuing or sending — to see
    what arrived (e.g. before go-live). Token-authed. ?since=YYYY-MM-DD (default: 3 days back)."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    cfg = EMAIL_CONFIG
    if not cfg['enabled']:
        return jsonify({'error': 'email not configured'}), 400
    since_s = request.args.get('since')
    try:
        since = datetime.datetime.strptime(since_s, '%Y-%m-%d').date() if since_s \
            else (datetime.date.today() - datetime.timedelta(days=3))
    except ValueError:
        return jsonify({'error': 'bad since (YYYY-MM-DD)'}), 400
    items = []
    try:
        mail = imaplib.IMAP4_SSL(cfg['imap_server'], cfg['imap_port'], timeout=30)
        mail.login(cfg['username'], cfg['password'])
        mail.select('INBOX')
        status, data = mail.search(None,
            f'FROM "{HAREL_CERT_SENDER}" SINCE {since.strftime("%d-%b-%Y")}')
        conn = get_db()
        for mid in (data[0].split() if status == 'OK' and data and data[0] else []):
            _, hd = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE)])')
            hdr = email_lib.message_from_bytes(hd[0][1])
            subject = decode_str(hdr.get('Subject', ''))
            if HAREL_CERT_SUBJECT_MARK not in subject:
                continue
            _, fd = mail.fetch(mid, '(BODY.PEEK[])')
            msg = email_lib.message_from_bytes(fd[0][1])
            html = None
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    try: html = part.get_content()
                    except Exception:
                        pl = part.get_payload(decode=True); html = pl.decode('utf-8', 'replace') if pl else None
                    break
            f = _parse_harel_cert(subject, html or '')
            row, mstatus = _match_insured_by_name(conn, f['name'])
            already = bool(f['ticket'] and conn.execute(
                "SELECT 1 FROM cert_requests WHERE ticket=?", (f['ticket'],)).fetchone())
            items.append({'date': hdr.get('Date', ''), 'name': f['name'], 'ticket': f['ticket'],
                          'match': mstatus, 'id_number': (row['id_number'] if row else ''),
                          'phone': (row['phone'] if row else ''), 'email': (row['email'] if row else ''),
                          'already_in_queue': already})
        conn.close(); mail.logout()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'since': since.isoformat(), 'count': len(items), 'items': items})

@app.route('/api/cert/scan')
def api_cert_scan():
    """Manual trigger + view of the certificate queue. Token-authed."""
    if not _wa_api_authed():
        return jsonify({'error': 'unauthorized'}), 403
    n = check_cert_emails(days_back=int(request.args.get('days', 14)))
    conn = get_db()
    rows = conn.execute(
        "SELECT ticket, cust_name, id_number, phone, brand, match_status, wa_sent_at, wa_target "
        "FROM cert_requests ORDER BY id DESC LIMIT 50").fetchall()
    conn.close()
    return jsonify({'processed': n, 'requests': [dict(r) for r in rows]})

def email_poll_thread():
    """Background thread: check inbox every N seconds. A heartbeat is recorded after EACH step
    so a long-but-progressing cycle keeps the scanner 'alive' for the watchdog."""
    # Routine poll scans a SHORT window so each cycle stays fast (a re-scan of a 14–30-day window
    # re-fetches headers for hundreds of emails every 3 min → the thread stalls and the scanner
    # falls behind). New emails always arrive same/next day, so a few days is plenty; a wider
    # backfill is available on demand via /api/policy/scan?days= etc. + the periodic deep scan below.
    POLL_DAYS = 4
    _cyc = [0]
    while True:
        time.sleep(EMAIL_CONFIG['check_interval'])
        # Once an hour, widen the window as a safety net so nothing is missed after an outage.
        _cyc[0] += 1
        _days = 21 if (_cyc[0] % 20 == 0) else POLL_DAYS   # ~every 20 cycles (≈1h) → 21-day sweep
        # Daily disk cleanup — the /data volume once filled to 99% and took the DB down. Delete PDFs
        # >7 days (delivered + on OneDrive) once/day (~480 cycles) + once shortly after startup.
        if _cyc[0] == 3 or _cyc[0] % 480 == 0:
            try:
                _n, _fb = _free_old_pdfs(7)
                if _n:
                    print(f'[disk-cleanup] נמחקו {_n} PDF ישנים, שוחררו {round(_fb/1024/1024, 1)}MB')
            except Exception as e:
                print(f'[disk-cleanup] {e}')
        try:
            n = check_email_inbox(days_back=_days)
            if n:
                print(f'[email-sync] עובדו {n} מיילים חדשים')
        except Exception as e:
            print(f'[email-sync] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            n2 = check_policy_documents(days_back=_days)
            if n2:
                print(f'[policy-docs] עובדו {n2} פוליסות חדשות')
        except Exception as e:
            print(f'[policy-docs] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            n3 = check_join_forms(days_back=_days)
            if n3:
                print(f'[join-forms] נקלטו {n3} טפסי הצטרפות')
        except Exception as e:
            print(f'[join-forms] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            rp, ru = check_renewal_forms(days_back=_days)
            if rp:
                print(f'[renewal-forms] עובדו {rp} טפסי חידוש ({ru} לא תואמים)')
        except Exception as e:
            print(f'[renewal-forms] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            hp = check_harel_completed(days_back=_days)
            if hp:
                print(f'[harel-proposal] נקלטו {hp} השלמות פרטים להצעה')
        except Exception as e:
            print(f'[harel-proposal] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            hc = check_cert_emails(days_back=_days)
            if hc:
                print(f'[harel-cert] נקלטו {hc} בקשות אישור קיום ביטוח')
        except Exception as e:
            print(f'[harel-cert] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            _mwc = get_db(); mw = auto_mark_midwives(_mwc); _mwc.commit(); _mwc.close()
            if mw['customers'] or mw['insureds']:
                print(f"[midwife] סומנו אוטומטית {mw['customers']} לקוחות + {mw['insureds']} מבוטחים כמיילדות")
        except Exception as e:
            print(f'[midwife] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            lr = label_sent_policy_emails()
            if lr.get('found'):
                print(f"[gmail-label] תויגו ואורכבו {lr['found']} מיילי פוליסות שנשלחו")
        except Exception as e:
            print(f'[gmail-label] שגיאת thread: {e}')
        touch_scan_heartbeat()
        try:
            cl = label_sent_cert_emails()
            if cl.get('found'):
                print(f"[cert-label] תויגו ואורכבו {cl['found']} מיילי אישורי ביטוח שנשלחו")
        except Exception as e:
            print(f'[cert-label] שגיאת thread: {e}')
        touch_scan_heartbeat()

# ── Admin email trigger ──────────────────────────────────────

@app.route('/admin/check-email', methods=['POST'])
@login_required
@admin_required
def admin_check_email():
    if not EMAIL_CONFIG['enabled']:
        flash('סנכרון מייל לא מוגדר עדיין — יש להגדיר IMAP בקובץ app.py', 'warning')
        return redirect(url_for('admin'))
    # Run in background so the page doesn't timeout on large inboxes
    threading.Thread(target=check_email_inbox, daemon=True).start()
    flash('בדיקת מייל הופעלה ברקע — רענן את הדף בעוד 10 שניות', 'info')
    return redirect(url_for('admin'))


@app.route('/refresh', methods=['POST'])
@login_required
def refresh_data():
    """Manual 'refresh' — pull emails + policy PDFs on demand, for when the
    background poll isn't running. Runs in a background thread so the request
    returns immediately (a synchronous IMAP scan exceeds gunicorn's worker
    timeout and gets the worker killed → 500)."""
    if not EMAIL_CONFIG['enabled']:
        flash('סנכרון מייל לא מוגדר עדיין', 'warning')
        return redirect(url_for('index'))

    def _run():
        try:
            check_email_inbox()
            check_policy_documents()
            check_join_forms()
            check_renewal_forms()
            check_harel_completed()
            check_cert_emails()
            conn = get_db()
            rebuild_insureds(conn)
            recompute_insured_statuses(conn)
            conn.close()
        except Exception as e:
            print(f'[refresh] שגיאה: {e}')

    threading.Thread(target=_run, daemon=True).start()
    flash('רענון הופעל — הנתונים יתעדכנו תוך מספר שניות. רענן את הדף.', 'info')
    return redirect(url_for('index'))


_backfill_state = {'running': False, 'done': 0, 'started': None, 'days': 0}

@app.route('/admin/backfill', methods=['POST'])
@login_required
@superadmin_required
def admin_backfill():
    """One-time backfill: scan up to a year of Harel PDFs, extract customer data +
    cancellations into the master. Data-only (keep_pdf=False) to stay within storage.
    Runs in the background; safe to leave and check back. Superadmin-only operational tool."""
    if _backfill_state['running']:
        flash('סריקה כבר רצה ברקע — המתן לסיומה', 'warning')
        return redirect(url_for('admin'))
    try:
        days = int(request.form.get('days', '30'))
    except ValueError:
        days = 30
    days = max(1, min(days, 400))

    def _run(days_back):
        _backfill_state.update(running=True, done=0, started=datetime.datetime.now().strftime('%H:%M'), days=days_back)
        try:
            n = check_policy_documents(days_back=days_back, keep_pdf=False)
            conn = get_db()
            rebuild_insureds(conn)
            recompute_insured_statuses(conn)
            conn.close()
            _backfill_state['done'] = n
            print(f'[backfill] הסתיים — {n} מסמכים חדשים, {days_back} ימים אחורה')
        except Exception as e:
            print(f'[backfill] שגיאה: {e}')
        finally:
            _backfill_state['running'] = False

    threading.Thread(target=_run, args=(days,), daemon=True).start()
    flash(f'סריקה אחורה של {days} ימים הופעלה ברקע — זה עשוי לקחת זמן. רענן את הדף מדי פעם.', 'info')
    return redirect(url_for('admin'))


@app.route('/submit', methods=['POST'])
def form_submit():
    """Direct POST from website forms (gaia-website / winner-website)."""
    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({'ok': False, 'error': 'no data'}), 400

    fields = {
        'name':           str(data.get('name') or '').strip(),
        'id_number':      normalize_id_number(data.get('id_number')),
        'phone':          str(data.get('phone') or '').strip(),
        'email':          str(data.get('email') or '').strip(),
        'installments':   str(data.get('installments') or '').strip(),
        'payment_method': str(data.get('payment_method') or '').strip(),
        'comments':       str(data.get('comments') or '').strip(),
        'brand':          str(data.get('brand') or '').strip(),
        'card_number':    str(data.get('card_number') or '').strip(),
        'card_expiry':    str(data.get('card_expiry') or '').strip(),
        'card_holder_id': str(data.get('card_holder_id') or '').strip(),
        'coverage_option': str(data.get('coverage_option') or '').strip(),
    }

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    # Unique ID based on content + time to prevent exact duplicate submissions
    import hashlib
    unique_id = f"web-{hashlib.md5((fields['id_number']+fields['phone']+now).encode()).hexdigest()[:12]}"

    cid = process_renewal_data(fields, message_id=unique_id, subject=f"טופס חידוש {fields['brand']}", received_at=now)
    print(f'[submit] {fields["name"]} ({fields["id_number"]}) brand={fields["brand"]} → cid={cid}')
    return jsonify({'ok': True})


@app.route('/db-status')
@login_required
@superadmin_required
def db_status():
    """Diagnostic endpoint — shows DB health (super-admin only; was public and leaked
    the schema + user count)."""
    try:
        conn = get_db()
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] if 'users' in tables else 'N/A'
        conn.close()
        return jsonify({
            'db_path': DB_PATH,
            'db_exists': os.path.exists(DB_PATH),
            'tables': tables,
            'user_count': user_count,
        })
    except Exception as e:
        return jsonify({'error': str(e), 'db_path': DB_PATH}), 500


# הפעל DB ו-email thread גם תחת gunicorn
try:
    print(f'[startup] calling init_db() on {DB_PATH}')
    init_db()
    print(f'[startup] init_db() done — db file exists: {os.path.exists(DB_PATH)}')
except Exception as _e:
    print(f'[startup] ERROR in init_db(): {_e}')
    import traceback; traceback.print_exc()

if EMAIL_CONFIG['enabled']:
    try:
        _t = threading.Thread(target=email_poll_thread, daemon=True)
        _t.start()
        print('[email-sync] Thread פעיל — יבדוק כל 5 דקות')
    except Exception as _e:
        print(f'[email-sync] ERROR starting thread: {_e}')

if __name__ == '__main__':
    print("=" * 50)
    print("מערכת שירות לקוחות פועלת!")
    print("כתובת גישה: http://localhost:5000")
    print("=" * 50)
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=False)
